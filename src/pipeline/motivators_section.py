"""Build the Motivators & Anti-Motivators HTML section for a hiring or coaching report.

This module is the baked-in version of the v4 integrated spike: motivator/anti callouts +
alignment block + Standard Map (with bucket-colored L2 borders) + DISC wiring panel.

Public API:
    build_section(respondent, *, mapping_xlsx=None, questions_tsv=None,
                  narratives_json=None) -> str
        Returns an HTML fragment (no <html>/<style> wrapper) ready to drop into the
        {{MOTIVATORS_ANTIMOTIVATORS_SECTION}} token of the hiring or coaching template.

The respondent dict must contain:
    name             : 'Jody Bender'
    first_name       : 'Jody'
    nat_pos          : 34
    nat_label        : 'Relating Supporter'       (TTI label — primary wedge is last word)
    nat_intensity    : 0.85                        (primary DISC score / 100)
    adp_pos          : 17
    adp_label        : 'Supporting Relater'
    adp_intensity    : 0.785
    disc             : [D, I, S, C] Natural scores (0-100)
    l2_scores        : {L2_Short: Z|Algo score}    from L2.dax rollup
Optional:
    shift_note       : 'Adapted shifts from ...'    rendered below the wiring panel
    partial_data     : bool                         renders "partial data" note in hub
    intensity_word   : 'Strong' / 'Moderate' / 'Light' (override; otherwise derived)

Reads these files by default (override with keyword args):
    l2_wedge_map.xlsx                     (workspace root)
    _pipeline/data/questions_full.tsv      (for L2_Num lookup)
    _pipeline/data/wedge_narratives.json   (wedge -> motivator/anti prose)

The speculative-mapping caveat is NOT rendered here. The report's top-level caveat
banner should carry that (see SKILL.md guidance)."""

import csv
import json
import math
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

# ---- Constants ----
WEDGES = ['Implementor','Conductor','Persuader','Promoter','Relater','Supporter','Coordinator','Analyzer']
WEDGE_CENTER = {w: i*45 for i, w in enumerate(WEDGES)}
OPPOSITES = {
    'Conductor':'Supporter','Supporter':'Conductor',
    'Persuader':'Coordinator','Coordinator':'Persuader',
    'Promoter':'Analyzer','Analyzer':'Promoter',
    'Relater':'Implementor','Implementor':'Relater',
}
SECONDARY_MAP = {
    'Conduct':'Conductor','Persuad':'Persuader','Promot':'Promoter','Relat':'Relater',
    'Support':'Supporter','Coordinat':'Coordinator','Analyz':'Analyzer','Implement':'Implementor',
}
# Option 3 palette: the pill FILL carries the 4-bucket signal directly. No more
# continuous z-ramp fill competing with a bucket-color border. Each pill is a
# simple colored badge and the exact z-score is printed as a small number inside.
#
# (fill_color, text_color) for each bucket. Text color chosen for contrast.
BUCKET_FILL = {
    # Four unambiguously distinct hues. Installed-against-grain gets the bright
    # celebratory green (it's the most respect-worthy finding in a file);
    # running-naturally is blue; aligned-but-not-running matches DISC D red.
    'motivator_strong': ('#2563eb', '#ffffff'),   # running naturally — BLUE
    'motivator_weak':   ('#c0392b', '#ffffff'),   # aligned but not running — DISC D RED
    'anti_strong':      ('#22c55e', '#ffffff'),   # installed against grain — BRIGHT VIBRANT GREEN
    'anti_weak':        ('#d4a84b', '#1a2332'),   # routines to install — GOLD
    # Mid-score in-zone — pale sibling of the zone's dominant anchor
    'motivator_neutral':('#dbeafe', '#1a2332'),   # pale blue
    'anti_neutral':     ('#dcfce7', '#1a2332'),   # pale green
    'cross':            ('#f3f4f6', '#5a6773'),   # gray — perpendicular wedges
    'no_score':         ('#e8ecf0', '#5a6773'),
}
# Legacy name kept for the alignment-block text grid (borders on panels).
BUCKET_COLORS = {
    'motivator_strong': '#2563eb',
    'motivator_weak':   '#c0392b',
    'anti_strong':      '#22c55e',
    'anti_weak':        '#d4a84b',
}
CROSS_BORDER_COLOR = '#d0d4da'   # thin neutral pill outline

# Wedge adjacency on the wheel. Used to expand the motivator/anti zones to
# include the two adjacent wedges on each side — so a respondent's wiring
# "honor" reaches beyond the primary wedge.
ADJACENT_WEDGES = {
    'Implementor': ('Analyzer', 'Conductor'),
    'Conductor':   ('Implementor', 'Persuader'),
    'Persuader':   ('Conductor', 'Promoter'),
    'Promoter':    ('Persuader', 'Relater'),
    'Relater':     ('Promoter', 'Supporter'),
    'Supporter':   ('Relater', 'Coordinator'),
    'Coordinator': ('Supporter', 'Analyzer'),
    'Analyzer':    ('Coordinator', 'Implementor'),
}


def _zone_of(respondent_primary: str, l2_primary: str) -> str:
    """Return 'motivator', 'anti', or 'cross' for an L2's primary wedge
    relative to the respondent's primary wedge. Zones include adjacent wedges."""
    anti = OPPOSITES[respondent_primary]
    motivator_set = {respondent_primary, *ADJACENT_WEDGES[respondent_primary]}
    anti_set = {anti, *ADJACENT_WEDGES[anti]}
    if l2_primary in motivator_set: return 'motivator'
    if l2_primary in anti_set:      return 'anti'
    return 'cross'


def _z_text_color(z):
    """Map a Z|Algo score to a text color so the Standard Map pills communicate
    score polarity even when the pill fill is grey (cross-wedge) or pale neutral.
    Dark/bright colors for materially positive/negative, readable mid-tone for
    mild deltas, neutral grey for near-zero."""
    if z is None: return '#5a6773'
    if z >=  1.0: return '#14532d'   # dark green — "great"
    if z >=  0.5: return '#16a34a'   # readable green — "positive"
    if z <= -1.0: return '#7f1d1d'   # dark red — "materially negative"
    if z <= -0.5: return '#dc2626'   # readable red — "negative"
    return '#5a6773'                 # neutral — near zero


# Buckets whose fill color is dark/saturated — keep white text for contrast.
# Every other bucket (cross, pale neutrals, no_score) delegates its text color
# to _z_text_color(z) so the wheel can be scanned for misalignment.
_DARK_FILL_BUCKETS = {'motivator_strong', 'motivator_weak', 'anti_strong'}


def _bucket_for(zone: str, z):
    """Return a bucket key that maps into BUCKET_FILL."""
    if z is None: return 'no_score'
    if zone == 'cross': return 'cross'
    if zone == 'motivator':
        if z >= 0.5:  return 'motivator_strong'
        if z <= -0.5: return 'motivator_weak'
        return 'motivator_neutral'
    if zone == 'anti':
        if z >= 0.5:  return 'anti_strong'
        if z <= -0.5: return 'anti_weak'
        return 'anti_neutral'
    return 'cross'


def _repo_root() -> Path:
    # _pipeline/src/pipeline/motivators_section.py → ROOT is three parents up
    return Path(__file__).resolve().parents[3]


def _load_mapping(path: Path | None = None) -> tuple[list[dict], dict]:
    path = path or (_repo_root() / 'l2_wedge_map.xlsx')
    wb = load_workbook(path, data_only=True)
    ws = wb['L2_Wedge_Mapping']
    # Skip banner row if present (e.g. SPECULATIVE or PARTIALLY VALIDATED captions
    # span the header row). Detection: real header row must contain 'L2_Short'.
    hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    data_start = 2
    if 'L2_Short' not in hdrs:
        hdrs = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
        data_start = 3
    cidx = {h: hdrs.index(h)+1 for h in hdrs}
    rows = []
    for r in range(data_start, ws.max_row+1):
        l2 = ws.cell(row=r, column=cidx['L2_Short']).value
        if not l2: continue
        rows.append({
            'l2_short': l2,
            'primary': ws.cell(row=r, column=cidx['Primary_Wedge']).value,
            'secondary': ws.cell(row=r, column=cidx['Secondary_Wedge']).value,
            'clean_q': ws.cell(row=r, column=cidx['Q_Clean']).value,
            'clean_max': ws.cell(row=r, column=cidx['Pol_Clean_Max']).value,
            'all_max': ws.cell(row=r, column=cidx['Pol_All_Max']).value,
        })
    wedge_map = {r['l2_short']: {'primary':r['primary'], 'secondary':r['secondary']} for r in rows}
    return rows, wedge_map


def _load_l2num_map(path: Path | None = None) -> dict:
    path = path or (_repo_root() / '_pipeline' / 'data' / 'questions_full.tsv')
    m = {}
    with open(path, encoding='utf-8') as f:
        for row in csv.DictReader(f, delimiter='\t'):
            m[row['L2_Short']] = row['L2_Num']
    return m


def _load_narratives(path: Path | None = None) -> dict:
    path = path or (_repo_root() / '_pipeline' / 'data' / 'wedge_narratives.json')
    return json.loads(path.read_text(encoding='utf-8'))


def _parse_label(label: str) -> tuple[str, str | None]:
    parts = label.strip().split()
    primary = parts[-1]
    secondary = None
    if len(parts) > 1:
        mod = parts[0]
        if mod.lower().endswith('ing'):
            mod = mod[:-3]
        secondary = SECONDARY_MAP.get(mod, mod + 'er')
    return primary, secondary


def _angle_delta(a, b):
    return ((b - a) % 360 + 540) % 360 - 180


def _classify_bucket(respondent_primary, l2_info, z):
    """Used by the alignment block (4-bucket grid). Returns bucket name or None.
    Uses the expanded 3-wedge zone (primary + 2 adjacent) for motivator/anti."""
    if z is None: return None
    zone = _zone_of(respondent_primary, l2_info['primary'])
    if zone == 'motivator':
        if z >= 0.5:  return 'motivator_strong'
        if z <= -0.5: return 'motivator_weak'
    if zone == 'anti':
        if z >= 0.5:  return 'anti_strong'
        if z <= -0.5: return 'anti_weak'
    return None


def _intensity_word(intensity):
    if intensity >= 0.80: return 'Strong'
    if intensity >= 0.55: return 'Moderate'
    return 'Light'


def compute_intensity_from_disc(disc):
    """Compute Standard Map marker intensity (0..1) from a [D, I, S, C] DISC array.

    The intensity controls how far from the center of the wheel the N (Natural) and
    A (Adapted) markers are drawn. The right semantic mapping is:

      - Strongly-anchored profiles (one DISC value clearly dominant, others much lower)
        push markers TOWARD the outer edge.
      - ACROSS / center-of-wheel profiles (four DISC values clustered near the 50/50/50/50
        baseline) pull markers TOWARD the center of the wheel.

    Reading any single DISC score in isolation (e.g. C/100) is the wrong model — it gave
    Armstrong_Patrick (Adapted DISC 52/52/35/75 → ACROSS per the TTI wheel page) an
    intensity of 0.75, drawing his marker at 75% of radius and visually contradicting the
    ACROSS / no-strong-anchor narrative. The fix (added 2026-04-27) is to compute the
    mean-absolute-deviation of the four DISC values from a 50/50/50/50 baseline,
    normalized to [0..1]. Floor at 0.10 so center-of-wheel markers stay clear of the very
    inside of the inner ring.

    Examples:
      Armstrong Natural [48, 52, 52, 71] → deviation 27/200 = 0.135 → ~0.14
      Armstrong Adapted [52, 52, 35, 75] → deviation 44/200 = 0.220
      Schott Natural    [63, 72, 72, 18] → deviation 89/200 = 0.445
      Schott Adapted    [72, 66, 25, 42] → deviation 71/200 = 0.355

    The intensity-from-deviation formula is the canonical mapping going forward.
    """
    deviation = sum(abs(v - 50) for v in disc)
    intensity = deviation / 200.0
    return max(0.10, min(intensity, 1.0))


def _fmt_z(z):
    """Format z-score for the small number displayed inside the pill."""
    if z is None: return '—'
    return f'{z:+.2f}'


def _render_alignment_block(buckets: dict) -> str:
    labels = {'motivator_strong':'Routines running naturally','motivator_weak':'Aligned but not running',
              'anti_strong':'Installed against the grain','anti_weak':'Routines to install'}
    descs = {'motivator_strong':'Motivator-aligned + strong. Wiring paying off.',
             'motivator_weak':'Motivator-aligned + weak. Wiring should carry this — probe why it\'s not.',
             'anti_strong':'Anti-aligned + strong. A tremendous positive — installed against the grain.',
             'anti_weak':'Anti-aligned + weak. Coaching frontier: turn into a routine, summon a motivator.'}
    # Each panel gets a pill-shaped title BADGE at the top in the exact bucket
    # fill color from BUCKET_FILL. Panel body bg is a pale-tint sibling of the
    # bucket color so every panel has a clearly distinct hue family.
    bg = {'motivator_strong':'#dbeafe',  # pale blue
          'motivator_weak':  '#fbd8d2',  # pale red (DISC D family)
          'anti_strong':     '#dcfce7',  # pale green
          'anti_weak':       '#fae6a8'}  # pale gold
    def bucket(k):
        items = buckets.get(k, [])
        pill_fill, pill_text = BUCKET_FILL[k]
        html = f'<div class="bucket" style="background:{bg[k]}; border-left:4px solid {BUCKET_COLORS[k]};">'
        # Pill-shaped title badge in the bucket color — matches pills on the map
        html += f'<div class="bucket-pill" style="background:{pill_fill}; color:{pill_text};">{labels[k]}</div>'
        html += f'<div class="bucket-desc">{descs[k]}</div>'
        if not items:
            html += '<div class="bucket-empty">None.</div>'
        else:
            html += '<div class="bucket-items">'
            for it in items[:3]:
                z = it['z']
                # Chip color matches the bucket so each L2 wears its panel color.
                html += f'<div class="l2-row"><span class="z-chip" style="background:{pill_fill}; color:{pill_text};">{z:+.2f}</span><span class="l2-name">{it["l2"]}</span></div>'
            html += '</div>'
        html += '</div>'
        return html
    inner = ''.join(bucket(k) for k in ['motivator_strong','motivator_weak','anti_strong','anti_weak'])
    return f'<div class="alignment-grid">{inner}</div>'


def _render_callouts(primary_wedge, anti_wedge, name_first, narratives, intensity_word):
    m = narratives[primary_wedge]['motivator'].replace('{NAME}', name_first).replace('{INTENSITY_WORD}', intensity_word)
    a = narratives[primary_wedge]['anti'].replace('{NAME}', name_first).replace('{INTENSITY_WORD}', intensity_word)
    return f'''<div class="callouts-pair">
<div class="callout motivator">
  <div class="callout-label">Motivator zone · where {name_first} is energized</div>
  <div class="callout-wedge">{primary_wedge}<span class="intensity-tag">{intensity_word}</span></div>
  <div class="callout-body">{m}</div>
</div>
<div class="callout anti">
  <div class="callout-label">Anti-motivator zone · where {name_first} drains</div>
  <div class="callout-wedge">{anti_wedge} <span class="pos">(directly across)</span><span class="intensity-tag">{intensity_word}</span></div>
  <div class="callout-body">{a}</div>
</div>
</div>'''


def _render_wiring(disc, nat_pos, nat_label, nat_intensity, adp_pos, adp_label, adp_intensity, shift_note=''):
    bars = ''
    # Bars are monochrome (dark navy) — the bar HEIGHT carries the score; the
    # bar COLOR should not encode anything that could be confused with the L2
    # alignment grammar elsewhere in the report (blue=motivator-strong,
    # green=against-the-grain, tan=anti-aligned-weak, red=motivator-weak).
    # The classic DISC palette (D=red, I=orange, S=green, C=blue) creates
    # false visual rhyme with that grammar even though it carries no
    # semantic meaning here. Keep these neutral.
    for axis, val in zip(['D','I','S','C'], disc):
        bars += f'<div class="disc-bar"><span class="disc-label">{axis}</span><div class="disc-track"><div class="disc-fill" style="width:{val}%;background:#1a2332;"></div></div><span class="disc-val">{val}</span></div>'
    adp_row = ''
    if adp_pos != nat_pos or adp_label != nat_label:
        adp_row = f'<tr><th>Adapted</th><td>pos {adp_pos} — {adp_label}</td><td>intensity {adp_intensity:.2f}</td></tr>'
    note_html = f'<div class="wiring-note">{shift_note}</div>' if shift_note else ''
    return f'''<div class="wiring-panel">
<div class="wiring-title">DISC wiring</div>
<div class="wiring-grid">
  <div class="disc-bars">{bars}</div>
  <table class="wiring-table">
    <tr><th>Natural</th><td>pos {nat_pos} — {nat_label}</td><td>intensity {nat_intensity:.2f}</td></tr>
    {adp_row}
  </table>
</div>
{note_html}
</div>'''


def _marker_angle(primary, secondary):
    pc = WEDGE_CENTER[primary]
    if not secondary or secondary not in WEDGE_CENTER: return pc
    d = _angle_delta(pc, WEDGE_CENTER[secondary])
    return (pc + (13 if d >= 0 else -13)) % 360


def _wrap_label(text, max_single=17):
    if len(text) <= max_single or ' ' not in text: return [text]
    words = text.split()
    best_split = 1; best = float('inf')
    for i in range(1, len(words)):
        l1 = ' '.join(words[:i]); l2 = ' '.join(words[i:])
        s = abs(len(l1) - len(l2)) + max(len(l1), len(l2)) * 0.3
        if s < best: best = s; best_split = i
    return [' '.join(words[:best_split]), ' '.join(words[best_split:])]


def _render_standard_map(respondent_name, primary_wedge, scores, nat_label, nat_intensity,
                         adp_label, adp_intensity, mapping_rows, l2num_map, full_data=True):
    anti = OPPOSITES[primary_wedge]
    rows = [{**r, 'l2_num': l2num_map.get(r['l2_short'], '?')} for r in mapping_rows]
    SIZE = 1100; cx = cy = SIZE / 2
    rInner = 70; rOuter = 470
    rUsable_in = rInner + 18; rUsable_out = rOuter - 28

    def polar(r, deg):
        a = (deg - 90) * math.pi / 180
        return (cx + r * math.cos(a), cy + r * math.sin(a))

    def wedge_path(rO, rI, sd, ed):
        p1 = polar(rO, sd); p2 = polar(rO, ed); p3 = polar(rI, ed); p4 = polar(rI, sd)
        la = 1 if (ed - sd) > 180 else 0
        return (f'M {p1[0]} {p1[1]} A {rO} {rO} 0 {la} 1 {p2[0]} {p2[1]} '
                f'L {p3[0]} {p3[1]} A {rI} {rI} 0 {la} 0 {p4[0]} {p4[1]} Z')

    def pill_params(pol):
        if pol >= 0.75: return {'font':12,'pad_h':10,'pad_v':8}
        if pol < 0.25:  return {'font':10,'pad_h':8,'pad_v':6}
        return               {'font':11,'pad_h':9,'pad_v':7}

    # Pre-compute per-L2 geometry
    for r in rows:
        p = r['clean_max'] if r['clean_max'] is not None else r['all_max']
        r['_test_only'] = (r['clean_q'] == 0)
        pc = 0 if p is None else max(0, min(1, p))
        r['_polarity'] = pc
        full_label = f"{r['l2_num']} {r['l2_short']}"
        if len(full_label) <= 22:
            r['_lines'] = [full_label]
        else:
            body = _wrap_label(r['l2_short'], 18)
            r['_lines'] = [full_label] if len(body) == 1 else [f"{r['l2_num']} {body[0]}", body[1]]
        params = pill_params(pc)
        r['_font'] = params['font']
        line_h = params['font'] * 1.15
        max_len = max(len(l) for l in r['_lines'])
        r['_w'] = max_len * params['font'] * 0.58 + params['pad_h'] * 2
        # extra row for the z-score (smaller font, f-2, plus 2px gap)
        z_row_h = (params['font'] - 2) * 1.15 + 2
        r['_h'] = line_h * len(r['_lines']) + z_row_h + params['pad_v'] * 2
        r['_target_radius'] = rUsable_in + pc * (rUsable_out - rUsable_in)

    by_prim = defaultdict(list)
    for r in rows: by_prim[r['primary']].append(r)
    for wedge, items in by_prim.items():
        pc_ang = WEDGE_CENTER[wedge]
        def lean(it, _pc=pc_ang): return _angle_delta(_pc, WEDGE_CENTER[it['secondary']]) if it['secondary'] else 0
        items.sort(key=lambda x: (lean(x), x['l2_num']))
        n = len(items)
        for i, it in enumerate(items):
            offset = 0 if n == 1 else (-20 + (40 / (n - 1)) * i)
            it['_angle'] = (pc_ang + offset) % 360
            it['_radius'] = it['_target_radius']
            x, y = polar(it['_radius'], it['_angle'])
            it['_x'] = x; it['_y'] = y

    def overlap(a, b):
        dx = abs(a['_x'] - b['_x']); dy = abs(a['_y'] - b['_y'])
        mx = (a['_w'] + b['_w']) / 2 + 4; my = (a['_h'] + b['_h']) / 2 + 4
        return (mx - dx, my - dy) if dx < mx and dy < my else None

    # Force-directed non-overlap
    for _ in range(120):
        forces = {id(r): [0.0, 0.0] for r in rows}
        for i, a in enumerate(rows):
            for b in rows[i + 1:]:
                o = overlap(a, b)
                if o:
                    ox, oy = o
                    dx = b['_x'] - a['_x']; dy = b['_y'] - a['_y']
                    d = math.sqrt(dx * dx + dy * dy) or 0.01
                    push = max(ox, oy) * 0.55
                    nx = dx / d; ny = dy / d
                    forces[id(a)][0] -= nx * push; forces[id(a)][1] -= ny * push
                    forces[id(b)][0] += nx * push; forces[id(b)][1] += ny * push
        disp = 0
        for r in rows:
            fx, fy = forces[id(r)]
            if fx == 0 and fy == 0: continue
            r['_x'] += fx; r['_y'] += fy
            dx, dy = r['_x'] - cx, r['_y'] - cy
            rad = math.sqrt(dx * dx + dy * dy)
            ang = (math.atan2(dy, dx) * 180 / math.pi + 90) % 360
            tr = r['_target_radius']
            if rad < rUsable_in: rad = rUsable_in
            if rad > rUsable_out: rad = rUsable_out
            rad = rad * 0.85 + tr * 0.15
            pc_ang = WEDGE_CENTER[r['primary']]
            d_ang = _angle_delta(pc_ang, ang)
            if d_ang > 21:  ang = (pc_ang + 21) % 360
            if d_ang < -21: ang = (pc_ang - 21 + 360) % 360
            nx, ny = polar(rad, ang)
            disp += abs(nx - r['_x']) + abs(ny - r['_y'])
            r['_x'] = nx; r['_y'] = ny; r['_angle'] = ang; r['_radius'] = rad
        if disp < 0.5: break

    parts = [f'<svg viewBox="0 0 {SIZE} {SIZE}" xmlns="http://www.w3.org/2000/svg">']
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rOuter}" fill="#ffffff" stroke="#cfd4da" stroke-width="1"/>')
    for w, c in WEDGE_CENTER.items():
        sd, ed = c - 22.5, c + 22.5
        if w == primary_wedge:   fill, stroke, sw = '#e8f6ed','#27ae60',2
        elif w == anti:          fill, stroke, sw = '#fbe9e6','#c0392b',2
        else:                    fill, stroke, sw = '#ffffff','#d0d4da',1
        parts.append(f'<path d="{wedge_path(rOuter, rInner, sd, ed)}" fill="{fill}" stroke="{stroke}" stroke-width="{sw}"/>')
    for pol, lbl in [(0.25,'weak'),(0.5,'moderate'),(0.75,'strong'),(1.0,'peak')]:
        r_ring = rUsable_in + pol * (rUsable_out - rUsable_in)
        parts.append(f'<circle cx="{cx}" cy="{cy}" r="{r_ring}" fill="none" stroke="#a0a8b0" stroke-width="1.2" stroke-dasharray="6 6" opacity="0.6"/>')
        p = polar(r_ring, 22.5)
        parts.append(f'<rect x="{p[0]-32}" y="{p[1]-9}" width="64" height="16" rx="8" fill="#ffffff" fill-opacity="0.95" stroke="#a0a8b0" stroke-width="0.5"/>')
        parts.append(f'<text x="{p[0]}" y="{p[1]+4}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" font-weight="700" fill="#5a6773">{lbl} ({pol:.2f})</text>')
    for i in range(8):
        a = i * 45 + 22.5
        to = polar(rOuter, a); ti = polar(rInner, a)
        parts.append(f'<line x1="{to[0]}" y1="{to[1]}" x2="{ti[0]}" y2="{ti[1]}" stroke="#cfd4da" stroke-width="1"/>')
    for w, c in WEDGE_CENTER.items():
        lp = polar(rOuter + 28, c)
        fw, col = (700, '#1e6b3f') if w == primary_wedge else ((700, '#8a251a') if w == anti else (600, '#4a5662'))
        parts.append(f'<text x="{lp[0]}" y="{lp[1]}" text-anchor="middle" dominant-baseline="middle" font-family="DM Sans, sans-serif" font-size="15" font-weight="{fw}" fill="{col}">{w.upper()}</text>')
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rInner}" fill="#fafbfc" stroke="#cfd4da" stroke-width="1"/>')
    parts.append(f'<text x="{cx}" y="{cy-10}" text-anchor="middle" font-family="DM Serif Display, serif" font-size="14" fill="#1a2332">{respondent_name}</text>')
    parts.append(f'<text x="{cx}" y="{cy+6}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" fill="#6b7884">{primary_wedge}</text>')
    if not full_data:
        parts.append(f'<text x="{cx}" y="{cy+22}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="8" fill="#c0392b" font-style="italic">partial data</text>')
    for r in rows:
        x, y = r['_x'], r['_y']
        z = scores.get(r['l2_short'])
        zone = _zone_of(primary_wedge, r['primary'])
        bucket = _bucket_for(zone, z)
        fill, textcol_bucket = BUCKET_FILL[bucket]
        # Dark/saturated fill buckets keep white text for contrast. Every other
        # pill (cross, pale neutrals, no_score) encodes the Z score in the text
        # color so the wheel is readable for misalignment at a glance.
        textcol = textcol_bucket if bucket in _DARK_FILL_BUCKETS else _z_text_color(z)
        w, h, f = r['_w'], r['_h'], r['_font']
        dash = ' stroke-dasharray="3 2"' if r['_test_only'] else ''
        parts.append(f'<rect x="{x-w/2}" y="{y-h/2}" width="{w}" height="{h}" rx="{min(h/2, 10)}" fill="{fill}" fill-opacity="0.98" stroke="{CROSS_BORDER_COLOR}" stroke-width="1"{dash}/>')
        n = len(r['_lines']); line_h = f * 1.15
        # Main label (shifted up slightly to leave room for z-score line)
        start_y = y - (n - 1) / 2 * line_h + f / 3 - 5
        for i, line in enumerate(r['_lines']):
            parts.append(f'<text x="{x}" y="{start_y + i*line_h}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="{f}" font-weight="700" fill="{textcol}">{line}</text>')
        # Z-score displayed as a small number below the label
        z_y = start_y + n * line_h + 2
        parts.append(f'<text x="{x}" y="{z_y}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="{f-2}" font-weight="600" fill="{textcol}" fill-opacity="0.80">{_fmt_z(z)}</text>')
    npr, nsc = _parse_label(nat_label)
    na = _marker_angle(npr, nsc); nr = rInner + 14 + max(0, min(1, nat_intensity)) * (rOuter - rInner - 28)
    nx, ny = polar(nr, na)
    parts.append(f'<circle cx="{nx}" cy="{ny}" r="27" fill="none" stroke="#d4a84b" stroke-width="3" stroke-dasharray="4 3"/>')
    parts.append(f'<circle cx="{nx}" cy="{ny}" r="20" fill="#1a2332" stroke="#ffffff" stroke-width="3.5"/>')
    parts.append(f'<text x="{nx}" y="{ny+5}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="14" font-weight="800" fill="#ffffff">N</text>')
    apr, asc = _parse_label(adp_label)
    if (apr != npr) or (asc != nsc) or abs(adp_intensity - nat_intensity) > 0.02:
        aa = _marker_angle(apr, asc); ar = rInner + 14 + max(0, min(1, adp_intensity)) * (rOuter - rInner - 28)
        ax, ay = polar(ar, aa)
        sp = []
        for i in range(10):
            sr = 22 if i % 2 == 0 else 10
            sa = (i * 36 - 90) * math.pi / 180
            sp.append(f'{ax + sr * math.cos(sa)},{ay + sr * math.sin(sa)}')
        parts.append(f'<polygon points="{" ".join(sp)}" fill="#d4a84b" stroke="#1a2332" stroke-width="2.5"/>')
        parts.append(f'<text x="{ax}" y="{ay+5}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="13" font-weight="800" fill="#1a2332">A</text>')
    parts.append('</svg>')
    return ''.join(parts)


SECTION_CSS = """
/* Motivators & Anti-Motivators section — scoped class prefix ma- */
.ma-section { margin: 32px 0; }
.ma-section .callouts-pair { display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:16px; }
.ma-section .callout { border-left:4px solid; padding:14px 18px; border-radius:4px; }
.ma-section .callout.motivator { border-color:#27ae60; background:#f0faf3; }
.ma-section .callout.anti { border-color:#c0392b; background:#fdf3f2; }
.ma-section .callout-label { font-size:10px; font-weight:700; letter-spacing:1.2px; text-transform:uppercase; margin-bottom:4px; }
.ma-section .callout.motivator .callout-label { color:#1e6b3f; }
.ma-section .callout.anti .callout-label { color:#8a251a; }
.ma-section .callout-wedge { font-family:"DM Serif Display",serif; font-size:18px; color:#1a2332; margin-bottom:8px; }
.ma-section .callout-wedge .pos { color:#6b7884; font-size:13px; font-family:"DM Sans",sans-serif; }
.ma-section .callout-wedge .intensity-tag { display:inline-block; font-family:"DM Sans",sans-serif; font-size:10px; font-weight:700; letter-spacing:1px; text-transform:uppercase; padding:2px 7px; border-radius:3px; margin-left:6px; vertical-align:middle; color:white; }
.ma-section .callout.motivator .intensity-tag { background:#27ae60; }
.ma-section .callout.anti .intensity-tag { background:#c0392b; }
.ma-section .callout-body { font-size:13px; color:#3a4652; line-height:1.55; }
.ma-section .alignment-grid { display:grid; grid-template-columns:1fr 1fr 1fr 1fr; gap:10px; margin-bottom:20px; }
.ma-section .bucket { padding:10px 14px; border-radius:4px; }
.ma-section .bucket-label { font-size:10px; font-weight:700; letter-spacing:1.1px; text-transform:uppercase; margin-bottom:3px; }
.ma-section .bucket-pill { display:inline-block; font-size:11px; font-weight:700; padding:4px 10px; border-radius:10px; margin-bottom:6px; letter-spacing:0.3px; }
.ma-section .bucket-desc { font-size:10.5px; color:#5a6773; margin-bottom:8px; line-height:1.4; }
.ma-section .bucket-empty { font-size:11px; color:#8a939c; font-style:italic; }
.ma-section .bucket-items { display:flex; flex-direction:column; gap:5px; }
.ma-section .l2-row { font-size:11.5px; color:#2c3e50; display:flex; align-items:baseline; gap:6px; flex-wrap:wrap; }
.ma-section .z-chip { display:inline-block; font-family:"DM Sans",sans-serif; font-weight:700; font-size:10px; padding:2px 6px; border-radius:3px; min-width:38px; text-align:center; color:white; }
.ma-section .z-chip.pos { background:#27ae60; }
.ma-section .z-chip.neg { background:#c0392b; }
.ma-section .l2-name { font-weight:500; }
.ma-section .map-block { background:#fafbfc; border:1px solid #e5e8ec; border-radius:6px; padding:16px; margin-bottom:16px; }
.ma-section .map-block svg { width:100%; height:auto; display:block; }
.ma-section .map-legend { margin-top:10px; padding:8px 12px; background:#f3f5f8; border-radius:4px; font-size:11px; color:#4a5662; line-height:1.5; }
.ma-section .map-legend strong { color:#1a2332; }
.ma-section .map-legend .border-swatch { display:inline-block; width:16px; height:10px; border-radius:3px; border-width:2.5px; border-style:solid; margin-right:2px; vertical-align:middle; background:#fafbfc; }
.ma-section .map-legend .fill-swatch { display:inline-block; width:18px; height:12px; border-radius:3px; border:1px solid #d0d4da; margin-right:2px; vertical-align:middle; }
.ma-section .wiring-panel { background:#fafbfc; border:1px solid #e5e8ec; border-radius:6px; padding:16px 20px; }
.ma-section .wiring-title { font-family:"DM Serif Display",serif; font-size:15px; color:#1a2332; margin-bottom:10px; }
.ma-section .wiring-grid { display:grid; grid-template-columns:280px 1fr; gap:24px; align-items:center; }
.ma-section .disc-bars { display:flex; flex-direction:column; gap:6px; }
.ma-section .disc-bar { display:flex; align-items:center; gap:10px; font-size:12px; }
.ma-section .disc-label { font-weight:700; width:14px; text-align:right; color:#1a2332; }
.ma-section .disc-track { flex:1; height:14px; background:#e5e8ec; border-radius:2px; overflow:hidden; }
.ma-section .disc-fill { height:100%; }
.ma-section .disc-val { width:30px; text-align:right; font-weight:700; color:#1a2332; }
.ma-section .wiring-table { border-collapse:collapse; font-size:12px; }
.ma-section .wiring-table th { text-align:left; font-weight:700; color:#6b7884; padding:4px 10px 4px 0; font-size:11px; letter-spacing:0.5px; text-transform:uppercase; }
.ma-section .wiring-table td { padding:4px 14px 4px 0; color:#2c3e50; }
.ma-section .wiring-note { margin-top:10px; font-size:11.5px; color:#6b7884; font-style:italic; }
"""


LEGEND_MAP = '''<div class="map-legend">
    <strong>Pill fill = coaching bucket</strong> (exact Z|Algo printed inside each pill):
    <span class="fill-swatch" style="background:#2563eb;"></span> running naturally
    <span class="fill-swatch" style="background:#c0392b;"></span> aligned but not running
    <span class="fill-swatch" style="background:#22c55e;"></span> installed against the grain
    <span class="fill-swatch" style="background:#d4a84b;"></span> routines to install
    <span class="fill-swatch" style="background:#dbeafe;"></span> motivator zone, mid
    <span class="fill-swatch" style="background:#dcfce7;"></span> anti zone, mid
    <span class="fill-swatch" style="background:#f3f4f6;"></span> cross-wedge.
    Zones include the primary wedge plus the two adjacent wedges. <strong>Radius + pill size</strong> = polarity strength.
</div>'''


def build_section(respondent: dict, *, mapping_xlsx=None, questions_tsv=None, narratives_json=None,
                  include_css: bool = True, section_title: str = "Motivators & Anti-Motivators") -> str:
    """Return an HTML fragment for the {{MOTIVATORS_ANTIMOTIVATORS_SECTION}} token.

    If include_css=True, inlines the section's scoped <style> so the fragment works standalone.
    Set include_css=False if the CSS is already present in the template's <style> block.
    """
    mapping_rows, wedge_map = _load_mapping(mapping_xlsx)
    l2num_map = _load_l2num_map(questions_tsv)
    narratives = _load_narratives(narratives_json)

    name = respondent['name']
    first = respondent.get('first_name') or name.split()[0]
    nat_label = respondent['nat_label']
    primary_wedge, _ = _parse_label(nat_label)
    anti_wedge = OPPOSITES[primary_wedge]
    nat_intensity = respondent['nat_intensity']
    adp_label = respondent.get('adp_label', nat_label)
    adp_intensity = respondent.get('adp_intensity', nat_intensity)
    intensity_word = respondent.get('intensity_word') or _intensity_word(nat_intensity)
    scores = respondent['l2_scores']
    disc = respondent['disc']
    full_data = not respondent.get('partial_data', False)
    shift_note = respondent.get('shift_note', '')

    # Alignment buckets
    buckets = {'motivator_strong':[], 'motivator_weak':[], 'anti_strong':[], 'anti_weak':[]}
    for l2, z in scores.items():
        info = wedge_map.get(l2)
        if not info: continue
        b = _classify_bucket(primary_wedge, info, z)
        if b in buckets:
            buckets[b].append({'l2': l2, 'z': z, 'primary': info['primary'], 'secondary': info['secondary']})
    for k in buckets:
        buckets[k].sort(key=lambda x: -abs(x['z']))

    callouts_html = _render_callouts(primary_wedge, anti_wedge, first, narratives, intensity_word)
    alignment_html = _render_alignment_block(buckets)
    svg = _render_standard_map(name, primary_wedge, scores, nat_label, nat_intensity,
                               adp_label, adp_intensity, mapping_rows, l2num_map, full_data=full_data)
    wiring_html = _render_wiring(disc, respondent['nat_pos'], nat_label, nat_intensity,
                                 respondent['adp_pos'], adp_label, adp_intensity, shift_note=shift_note)

    css_block = f'<style>{SECTION_CSS}</style>' if include_css else ''
    return f'''{css_block}
<div class="ma-section">
<h2 class="section-title" style="font-family:'DM Serif Display',serif;font-size:22px;color:#1a2332;margin-bottom:12px;">{section_title}</h2>
{callouts_html}
{alignment_html}
<div class="map-block">{svg}</div>
{LEGEND_MAP}
{wiring_html}
</div>'''
