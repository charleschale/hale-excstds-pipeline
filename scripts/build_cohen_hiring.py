"""Build the Excellence Standards hiring report for Matthew Cohen (CEO, Provable Markets).

End-to-end pipeline:
1. Load respondent data from xlsx (L1, L2, Flags, Metadata, Non-Scorable)
2. Compute distribution chart tokens from Histogram Data
3. Build motivators_section HTML
4. Fill all 60 tokens in the template
5. Save filled HTML

Run from repo root:
    python _pipeline/scripts/build_cohen_hiring.py
Outputs to:
    _reports/Cohen_Matthew_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

RESPONDENT_XLSX = ROOT / '_respondents' / '20250721.matt@provablemarkets.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Cohen_Matthew_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    """Load all respondent data from xlsx sheets."""
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    # L1 scores
    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title,
                'z_algo': z_algo,
                'z_human': z_human,
                'rf_count': rf_count
            }

    # L2 scores (canonical source per SKILL.md)
    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    # Flags and overall metrics
    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    # Extract flag values
    flag_names = []
    flag_row = 1
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(flag_row, col).value
        if flag_name:
            flag_names.append((col, flag_name))

    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    # Non-Scorable questions
    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    # Metadata
    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data,
        'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall,
        'z_human_overall': z_human_overall,
        'rf_num': rf_num,
        'questions_answered': questions_answered,
        'flags_lit': flags_lit,
        'non_scorable': non_scorable,
        'name': name,
        'email': email,
        'date_str': date_str,
    }

# ============================================================================
# LOAD HISTOGRAM DATA FOR DISTRIBUTION CHARTS
# ============================================================================

def load_histogram_data():
    """Load population distribution data."""
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)

    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({
                'z_algo': z_algo,
                'z_human': z_human,
                'sf': sf
            })

    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    flag_rows = rf_values
    return zalgo_rows, flag_rows

# ============================================================================
# BUILD DISTRIBUTION CHART TOKENS
# ============================================================================

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    """Canonical chart token builder — see build_hechler_hiring.py for full commentary.
    Chart 1: 2-row labels, collapse leading/trailing empty bins
    Chart 2: 2-row labels, hide empty columns
    Chart 3: REVERSE axis (high flags LEFT, low flags RIGHT)
    """
    bin_edges = [i * 0.5 for i in range(-8, 9)]

    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo = row['z_algo']
        z_human = row['z_human']
        sf = row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        idx = list(range(left, right + 1))
        return idx

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1
            continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1
                break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    n_flag = len(flag_counts)
    rev_flag_bin = n_flag - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# BUILD NARRATIVE SECTIONS (Three Axes, Concerns, Wiring Fit, etc.)
# ============================================================================

def build_three_axes_narratives():
    """Cohen three-axes — CANDIDATE framing (evaluating for a Series B growth-stage CEO seat,
    NOT incumbent). Badges honestly recalibrated so the hiring manager sees the real diligence
    shape: Talent AMBER (selector-density is real but most of it comes from the seat he is
    currently running; the belief-extension pattern is a material diligence item for a seat that
    requires building the layer below the exec team). Judgment AMBER-GREEN (strong decisional
    engine, in-the-room initiating is the Series-B-altitude frontier). Skills GREEN (domain fit).
    """

    talent_badge = 'badge-amber'
    talent_badge_text = 'AMBER'
    talent_card_body = '''Strong external-validator density. MD-level trading seats at Nomura (Equity Finance) and Jefferies (Securities Lending) before founding Provable Markets; as founder-CEO moved Provable from $18M Seed through a contested co-founder separation out to a $75M Series A Extension with publicly-reported ~40% adoption among top-20 prime-brokerage clients. Selectors at the institutional and venture altitude have consistently picked this operator. For a CEO hire in regulated FinTech, the pattern evidence is meaningful.

The Talent-axis tension the hiring manager should weigh seriously: Flag_UnconditBelief at <strong>Sev</strong> + Flag_DreamTeam &ldquo;Upgrade Team&rdquo; + HoldsAvgDownChain + TTI Indifferent cluster (Altruistic 31, Harmonious 0, Selfless 24, Receptive 25 &mdash; all four bottom-ranked Driving Forces sit in the belief/advocacy wedge). Investing in Others L1 is net-positive (+0.83) driven by Demonstrating Genuine Fanness L2 (+4.38), but Developmental Discipline L2 (&minus;1.26) and Reciprocal Followership L2 (&minus;0.23) tell the shadow: he admires performers who have already delivered, and has not yet invested developmental capital in those who have not. For a Series B growth-stage CEO seat &mdash; where building the next layer below the exec team is a core part of the job &mdash; this is a material diligence item, not a soft probe.'''

    judgment_badge = 'badge-amber'
    judgment_badge_text = 'AMBER-GREEN'
    judgment_card_body = '''Decisional machinery is the strongest part of the file and fits the Series B CEO seat cleanly. L1 #9 Deliberate Urgency +1.60 (L2 Extreme Proactivity +1.97), L1 #8 Organizational Decision Making +1.20 (L2 Clarity of Accountability +1.57), L1 #7 Not Pleasing +1.75 (L2 Discomfort For Self +1.39, L2 Cares About Others Not Their Approval +1.16). Operational evidence: the co-founder separation executed during the scale-up is the most predictive single career-arc data point for decisive judgment under founder-grade personnel pressure.

The watch item that determines Series B fit: <strong>in-the-room initiating accountability</strong>. L1 #3 Conducting &amp; Outvoted &minus;1.38, L2 Dialogue vs. Direction &minus;1.07, L2 Power &amp; Status Management &minus;1.13, L2 Ability To Disappear &minus;1.00, L2 Sublimating Ego &minus;0.67, Flag_InitiatingAccountability at <strong>Sev</strong>. Decisions route through him; the team performs accountability for him rather than owning it themselves. Self-assessment Q9917 names delegation as a specific weakness &mdash; instrument corroborates exactly. At Seed/A this is how the engine works; at Series B it is the specific failure mode that breaks growth-stage CEO transitions. The interview needs to answer whether this is installable-against-the-wiring, or structural.'''

    skills_badge = 'badge-green'
    skills_badge_text = 'GREEN'
    skills_card_body = '''The instrument does not measure domain skill directly &mdash; noted with the usual caveat that strong skills do not compensate for weak talent or weak judgment at executive altitude.

That said: MD-level institutional trading seats at Nomura and Jefferies + founded and scaled Provable Markets in the same regulated-FinTech sector constitute strong domain fit for a capital-markets-infrastructure CEO seat. Regulator interface, counterparty credibility, and institutional legitimacy are all covered without friction. Domain skill is not the frontier of this file &mdash; the Role-Fit section below names the actual frontiers.'''

    return {
        'TALENT_BADGE_CLASS': talent_badge,
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': judgment_badge,
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': skills_badge,
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    """Two Targeted Concerns for Cohen — both grounded in multiple lit flags + L2 scores.
    Concern 1: In-the-room Initiating Accountability (master finding).
    Concern 2: Depth below directs + belief extension (HoldsAvgDownChain + Condit Belief Sev).
    """
    concerns_title = 'Two Targeted Concerns'
    concerns_items = """            <div class="concern-item">
                <div class="concern-number">1</div>
                <div class="concern-text">
                    <strong>In-the-room initiating accountability &mdash; conductor lane, not sheriff lane.</strong>
                    Flag_InitiatingAccountability at <strong>Sev</strong>, L1 #3 Conducting &amp; Outvoted (&minus;1.38), L2 2.1 Dialogue vs. Direction (&minus;1.07), L2 2.2 Power &amp; Status (&minus;1.13), L2 4.1 Ability To Disappear (&minus;1.00), L2 2.4 Sublimating Ego (&minus;0.67) all circle the same signal: the accountability machinery is strong &mdash; Clarity of Accountability L2 +1.57, Extreme Proactivity +1.97, Drives Accountability +0.88, Deliberate Urgency L1 +1.60 &mdash; but decisions route through <em>him</em> and the team performs accountability <em>for</em> him rather than owning it themselves. Q9917 self-assessment names delegation as a weakness; instrument corroborates exactly. TTI Conducting Implementor wiring (D=69 + C=61, low-I, Objective 83 + Structured 68 primary) is the engine that built Provable &mdash; and also the wiring that produces this teammate-facing risk. Probe Form 8 Facilitative Mindset and Leadership Deep-Dive: does he run meetings by asking or by telling, and can he name two directs who raised urgency to him in the last month (rather than the other way)?
                </div>
            </div>
            <div class="concern-item">
                <div class="concern-number">2</div>
                <div class="concern-text">
                    <strong>Depth below directs + conditional belief posture.</strong>
                    Flag_HoldsAvgDownChain + Flag_UnconditBelief at <strong>Sev</strong> (&ldquo;Condit Belief Sev&rdquo;) + Flag_DreamTeam &ldquo;Upgrade Team&rdquo; + Flag_SatisfiedGripes (Low). L2 5.8 Standards = What You Tolerate &minus;0.87, L2 1.6 Developmental Discipline &minus;1.26, L2 1.5 Reciprocal Followership &minus;0.23. TTI Indifferent cluster &mdash; Harmonious 0, Selfless 24, Receptive 25, Altruistic 31 &mdash; is the wiring tell. Pattern reads: he genuinely admires team members who have already delivered (Demonstrating Genuine Fanness L2 +4.38 is top-decile), and the machinery that extends developmental belief <em>before</em> the proof is not a default channel. The shadow: the layer below his directs may be carrying seats he would not hire for today, which routes accountability that should sit at the directs&rsquo; layer back up to him (compounding Concern 1). Probe Form 8 Talent Development: <em>what people over your career have you nurtured who have gone on to do great things?</em> Listen for one pre-proof investment inside 60 seconds, not post-proof admiration.
                </div>
            </div>"""

    return {
        'CONCERNS_TITLE': concerns_title,
        'CONCERNS_ITEMS': concerns_items,
    }

def build_wiring_fit():
    """Wiring-Fit Check: 2 <strong>-led items with <span class="wiring-flag"> pills.
    Cohen's wiring ALIGNS with the instrument findings (no mismatch) — which itself is the
    read: the gap is in-the-wiring-wedge, meaning routines-installed-against-the-grain is
    the correct coaching framing, not a wiring-contradicts-behavior diagnosis.
    """
    wiring_fit_items = (
        '<strong>Conducting Implementor wiring corroborates the initiating-accountability read:</strong> '
        'DISC D=69 + C=61 + I=32 (below midline) + TTI Objective 83 + Structured 68 + Commanding 49 '
        'all point at directive structured execution &mdash; the engine that built Provable. Low-I + '
        'low-Dialogue L2 (&minus;1.07) + low-Power&amp;Status L2 (&minus;1.13) say the in-the-room '
        'conductor routines are not wired in and must be installed deliberately. No wiring-vs-behavior '
        'divergence &mdash; the wiring and the instrument agree.'
        '<span class="wiring-flag">Targeted Concern</span><br>'
        '<strong>Indifferent-cluster aligns with the belief-extension read:</strong> '
        'TTI Harmonious 0, Selfless 24, Receptive 25, Altruistic 31 &mdash; all four bottom-ranked '
        'Driving Forces sit in the belief/advocacy wedge. ExcStds Condit Belief Sev + HoldsAvgDownChain + '
        'Developmental Discipline L2 &minus;1.26 corroborate. The machinery that extends unconditional '
        'belief before proof is not wired in &mdash; Provable&rsquo;s next altitude requires routines '
        'installed against the grain.'
        '<span class="wiring-flag">Diligence Item</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    """Hard-to-Learn Gate count fraction. Cohen: 0/4 lit (Urgency +1.60, Org DM +1.20
    are both top-decile; Conditional Belief is one of the four Hard-to-Learn dimensions
    and IS lit — so 1/4 lit, meaning 3/4 clean)."""
    # Hard-to-Learn dimensions (per METHODOLOGY): Urgency, Org Decision Making,
    # Conditional Belief, Satisfied with Gripes. Cohen's Urgency +1.60 clean, Org DM +1.20 clean,
    # Condit Belief Sev LIT, Satisfied with Gripes | Low LIT. So 2/4 lit.
    hard_to_learn = '2/4'
    return {'HARD_TO_LEARN': hard_to_learn}

#
# Canonical Form 8 question library — SOURCE OF TRUTH.
# These exact strings must appear verbatim in each probe-question block.
#
FORM8_QUESTIONS = [
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]


def build_interview_probes():
    """10 Form 8 probes tailored to Cohen's flag profile and wiring."""
    tailored = [
        # 1 — Two-Sport Athlete (Talent axis)
        ("TWO-SPORT ATHLETE (Talent axis)",
         "Listen for specifics OUTSIDE the trading/FinTech/CEO lane &mdash; sport, craft, service, academic pursuit &mdash; and what the pursuit revealed about grit, coachability, and handling being bad at something before becoming good. Cohen&rsquo;s visible arc is institutional trading &rarr; founder-CEO; a genuine cross-domain pursuit is the rare-behavior signal. If the answer is the Provable raise or the trading career, press for something further afield before accepting."),
        # 2 — Talent Development (Concern 2 · Belief-extension + depth)
        ("TALENT DEVELOPMENT (Concern 2 &middot; Pre-proof belief)",
         "Central probe for Concern 2. Listen for one case where he backed the person <em>before</em> they had proven themselves. Specificity about extending developmental belief ahead of performance proof is the core signal. Low Developmental Discipline L2 (&minus;1.26), Condit Belief Sev flag, and Indifferent Altruistic (31) predict generic answers or post-proof admiration. Strong Talent-Development records produce two names and two specific pre-proof investments (time, assignment, exposure) inside 60 seconds."),
        # 3 — TORC (self-awareness calibration)
        ("TORC &mdash; TRUTH-OVER-COMFORT (Self-awareness)",
         "Anchor the question to a real boss: his MD seat at Jefferies, the senior at Nomura, or the early Provable lead investor. Listen for whether he names real developmental areas or defaults to strengths-reframing (&ldquo;my weakness is I care too much&rdquo;). D=69 + Commanding (49) + C=61 wiring can convert weakness questions into achievement stories &mdash; push past the first answer to get the second. Q9917 already offered temper, patience, and delegation; does he name those under an external boss&rsquo;s voice, or something softer?"),
        # 4 — Emotional Maturity (Ownership pattern)
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "Note whether he chooses a personal adversity or defaults to the co-founder separation or the Series A Extension fundraise. Pattern of ownership vs. external attribution is the signal. A mature answer attributes learning to self; a defensive answer attributes the setback to the co-founder, the market, or the cap table. Q9917 self-report lists temper as a weakness &mdash; listen for whether the adversity answer owns that, or re-frames it."),
        # 5 — Punctuates Differently (Talent axis)
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior probe. Listen for a specific, idiosyncratic practice &mdash; not a general work-ethic claim. &ldquo;I work harder&rdquo; or &ldquo;I care more&rdquo; is the negative signal. A true Punctuates Differently answer names a concrete routine the interviewer has not heard before. Given Objective (83) + Structured (68) primary, the answer may well be a disciplined process / methodology one &mdash; which is the positive read."),
        # 6 — Facilitative Mindset (Concern 1 · Dialogue over direction)
        ("FACILITATIVE MINDSET (Concern 1 &middot; Dialogue over direction)",
         "Central probe for Concern 1. A Facilitative answer names a belief AND the conditions under which he would update it &mdash; evidence of reasoning over direction. A non-Facilitative answer either holds the belief absolute or caves under pressure. Listen for structured exception-logic, not platitudes. Ties directly to L2 2.1 Dialogue vs. Direction (&minus;1.07), L2 2.2 Power &amp; Status (&minus;1.13), and L1 #3 Conducting &amp; Outvoted (&minus;1.38). Follow-up: &ldquo;When was the last time your team convinced you to change your mind on something that mattered?&rdquo;"),
        # 7 — Commitment (Diligence wiring)
        ("COMMITMENT (Diligence wiring &middot; L1 #9 strong)",
         "Daily-routine specificity. Cohen&rsquo;s Deliberate Urgency L1 is +1.60, Extreme Proactivity L2 +1.97 &mdash; top-decile; the habit is wired in. A strong answer names a concrete 7:45am routine around top priorities or a daily end-of-day discipline. Vague answers (&ldquo;I stay on top of things&rdquo;) would contradict the instrument read &mdash; press for the specific artifact (calendar block, journal, list) if the first answer is generic."),
        # 8 — Leadership Deep-Dive (Concern 1 + 2)
        ("LEADERSHIP DEEP-DIVE (Concerns 1 &amp; 2 &middot; Crew not basketball)",
         "Ask him to literally sketch the Provable top two levels. Tests both Targeted Concerns. A leader running the accountability routine names owners, open roles, and upgrade candidates in the same breath. Follow-up: &ldquo;Which of these seats would you not re-hire today?&rdquo; + &ldquo;Name two people two levels down you have invested developmental capital in before they had proved out.&rdquo; HoldsAvgDownChain + Upgrade Team + Developmental Discipline L2 (&minus;1.26) all surface here. Hesitation, multiple owners for the same priority, or post-proof-only names are the flags."),
        # 9 — Passion (mission construction)
        ("PASSION &amp; MISSION CONSTRUCTION",
         "Tests mission-construction ability. With Commanding (49) + Objective (83) + a CEO seat through a co-founder separation, Cohen has multiple paths &mdash; listen for whether he roots passion in customer outcome (prime-brokerage counterparties, regulated market infrastructure) or defaults to generic scale-language. Weak answers invoke &ldquo;any team can be motivated&rdquo;; strong answers name the concrete hook. Bridges into the &ldquo;Why this next seat, why now&rdquo; follow-up."),
        # 10 — Continuous Improvement
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes Concern 2. Listen for whether he treats learning, reflection, and direct-report coaching time as work &mdash; or only meetings and deliverables. A Continuous Improvement operator counts time spent in others&rsquo; development as core work, not overhead. A narrow answer (&ldquo;work = output&rdquo;) reinforces the Condit Belief + Developmental Discipline wiring deficit. A strong answer names a specific coaching-block cadence with directs."),
    ]

    assert len(tailored) == len(FORM8_QUESTIONS) == 10, "Must have exactly 10 probes"
    probes = []
    for (category, coaching), (_form8_name, canonical_q, _substr) in zip(tailored, FORM8_QUESTIONS):
        display_q = f'"{canonical_q}"'
        probes.append((category, display_q, coaching))

    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    interview_probes = "\n".join(parts)

    return {'INTERVIEW_PROBE_CARDS': interview_probes}

def build_teach_items():
    """Teach Items fraction — TeachTop10 sheet has 10 rows for Cohen."""
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    """Dimensional Scorecard with L1 + L2 hierarchy.
    Cohen's chosen L1s for the hiring report:
      - L1 #1 Investing in Others (+0.83, net strength with shadow at L2 1.6)
      - L1 #3 Conducting & Outvoted (−1.38, the master concern)
      - L1 #5 Pushing Extreme Accountability (mixed)
      - L1 #9 Deliberate Urgency (+1.60, top-decile anchor)
    """
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Demonstrating Genuine Fanness',
            'Reciprocal Followership',
            'Developmental Discipline',
        ]),
        (3, 'CONDUCTING RATHER THAN PLAYING LEAD GUITAR', [
            'Dialogue Vs. Direction',
            'Power & Status Management',
            'Sublimating Ego',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Basic Machinery Of Accountability',
            'Drives Accountability',
            'Stds = What Tolerate',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Extreme Proactivity',
            'Action Over Inaction',
        ]),
    ]

    labels = []
    scores = []
    is_l1 = []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found in L2 sheet: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    """Talent Radar for Cohen.
    Canonical labels (fixed order): Two-Sport Athlete, Punctuates Differently, Facilitative
    Mindset, Wit (interview-only), Deep Repertoire, Discipline/Routine, Understands Symbolism
    (interview-only).
    Scores 1-5, null for interview-only.
    - Two-Sport Athlete: 2 (career visible is finance/fintech only — needs interview probe)
    - Punctuates Differently: 4 (Not Pleasing +1.75 top-decile + Discomfort For Self +1.39 —
      genuine rare-behavior signal of conviction over approval)
    - Facilitative Mindset: 2 (L1 #2 mid −0.87; L2 2.1 Dialogue −1.07; L2 2.4 Sublimating Ego −0.67)
    - Wit: null
    - Deep Repertoire: 4 (Nomura MD → Jefferies MD → founded and scaled Provable through
      co-founder separation — operator breadth across regulated FinTech)
    - Discipline/Routine: 5 (Extreme Proactivity +1.97, Deliberate Urgency +1.60, Clarity of
      Accountability +1.57 — top-decile routine machinery; TTI Structured 68 corroborates)
    - Understands Symbolism: null
    """
    talent_radar_scores = json.dumps([2, 4, 2, None, 4, 5, None])

    talent_radar_profile = (
        "Discipline/Routine (5) is the standout &mdash; Extreme Proactivity +1.97, Deliberate "
        "Urgency +1.60, Clarity of Accountability +1.57 are all top-decile, and TTI Structured "
        "(68) + Objective (83) corroborate. Punctuates Differently (4) is a real strength "
        "tied to Not Pleasing +1.75 and Discomfort For Self +1.39 &mdash; genuine conviction "
        "over approval. Deep Repertoire (4) &mdash; MD-level trading at Nomura and Jefferies, "
        "founded and scaled Provable through a co-founder separation. Two-Sport Athlete (2) "
        "must be probed in interview (career visible is finance-only). Facilitative Mindset "
        "(2) is the frontier &mdash; L1 #3 &minus;1.38, Dialogue L2 &minus;1.07, Power &amp; "
        "Status L2 &minus;1.13. Wit (&#8856;) and Understands Symbolism (&#8856;) cannot be "
        "assessed by the instrument and must be evaluated in interview."
    )

    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_role_fit():
    """Role-Fit section — seat-responsive 'What Will Be Easy / What Will Be Hard' split.
    Seat: Series B growth-stage CEO, regulated FinTech / capital-markets infrastructure.
    Scaling exec team already in place; board wants commercial velocity + operating discipline.

    The Easy column maps the top-decile L1/L2 strengths + TTI primary cluster + career evidence
    onto the specific demands of the Series B CEO seat (decisional velocity, regulatory posture,
    commercial compounding).

    The Hard column names the conductor-vs-lead-guitarist transition as the Series-B-specific
    failure mode — at Seed/A the CEO IS the routing node; at Series B the exec team must carry
    distributed accountability. This is the seat-specific tension the hiring manager must resolve
    in diligence.
    """
    role_fit_title = 'Role-Fit Read &mdash; What Will Be Easy, What Will Be Hard'
    role_fit_seat = ('Evaluating for: Series B growth-stage CEO &middot; regulated FinTech / '
                     'capital-markets infrastructure &middot; scaling exec team already in '
                     'place &middot; board wants commercial velocity + operating discipline')

    role_fit_easy = (
        '<strong>Decisional velocity and operating cadence.</strong> Deliberate Urgency '
        '(L1 #9) +1.60, Extreme Proactivity +1.97, Clarity of Accountability +1.57, '
        'Organizational Decision Making (L1 #8) +1.20, Not Pleasing (L1 #7) +1.75, '
        'Discomfort For Self +1.39, Action Over Inaction +1.23 are all top-decile. '
        'Operational proof: he executed a co-founder separation while scaling the business '
        'through to a $75M Series A Extension &mdash; the hardest decision most Seed/A CEOs '
        'ever face, made cleanly. Board will get fast, unsentimental decisions.<br><br>'
        '<strong>Regulated-FinTech legitimacy with counterparties and regulators.</strong> '
        'MD-level seats at Nomura (Equity Finance) and Jefferies (Securities Lending) before '
        'founding Provable: the prime-brokerage / securities-lending counterparty network and '
        'the regulator interface are already wired in. ~40% adoption among top-20 prime-brokerage '
        'clients is commercial evidence of that domain trust.<br><br>'
        '<strong>Structured operating engine.</strong> TTI Objective 83 + Structured 68 + '
        'DISC C=61 + Conducting Implementor wheel produce a disciplined-process CEO. Board '
        'reviews, metric cadences, and commercial-velocity tracking will land naturally.'
    )

    role_fit_hard = (
        '<strong>The conductor-not-lead-guitarist transition.</strong> This is the '
        'Series-B-specific failure mode. At Seed/A, the CEO <em>is</em> the routing node and '
        'directive execution is a feature &mdash; exactly what built Provable. At Series B, '
        'the exec team must carry distributed accountability; the CEO moves from playing lead '
        'guitar to conducting. His wiring resists this: L1 #3 Conducting &amp; Outvoted '
        '&minus;1.38, Dialogue vs. Direction &minus;1.07, Power &amp; Status Management &minus;1.13, '
        'Ability To Disappear &minus;1.00, Sublimating Ego &minus;0.67, Flag_InitiatingAccountability '
        'at <strong>Sev</strong>. Self-assessment Q9917 names delegation as a weakness &mdash; '
        'instrument corroborates.<br><br>'
        '<strong>Building belief in the layer below directs.</strong> HoldsAvgDownChain + '
        'Condit Belief Sev + DreamTeam &ldquo;Upgrade Team&rdquo; + Developmental Discipline '
        '&minus;1.26 + Standards=What You Tolerate &minus;0.87 + TTI Indifferent cluster '
        '(Harmonious 0, Selfless 24, Receptive 25, Altruistic 31) read consistently: pre-proof '
        'developmental capital is not a default channel. At Series B, talent density two levels '
        'down IS the growth machine &mdash; the CEO who cannot extend belief before performance '
        'proof becomes the bottleneck for his own exec team.<br><br>'
        '<strong>The binary diligence question.</strong> Can he install conductor routines '
        'against his wiring, OR does the board need to pair him with a COO/President whose '
        'explicit charter is distributed accountability? Onsite interview must produce evidence '
        'for one of those two paths &mdash; the third path (neither installable nor paired) '
        'is the one that breaks this seat.'
    )

    return {
        'ROLE_FIT_TITLE': role_fit_title,
        'ROLE_FIT_SEAT': role_fit_seat,
        'ROLE_FIT_EASY': role_fit_easy,
        'ROLE_FIT_HARD': role_fit_hard,
    }


def build_career_timeline():
    """Cohen's career timeline — from coaching guide DRAFT.md (LinkedIn sweep)."""
    timeline_title = 'Career Timeline &mdash; Trading Floor to Founder-CEO'

    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Education<br>&middot; Penn State</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">Early Trading<br>Pre-Nomura</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">Nomura &mdash; MD Equity Finance<br>Multi-year institutional seat</div>
                <div class="timeline-block" style="background:#1e8449;color:#fff;">Jefferies &mdash; MD Securities Lending<br>Multi-year institutional seat</div>
                <div class="timeline-block" style="background:#d4a84b;color:#1a2332;">Provable Markets &mdash; Founded<br>Seed $18M</div>
                <div class="timeline-block" style="background:#c0392b;color:#fff;">Provable Markets &mdash; Co-founder separation &rarr; Series A Extension<br>$75M &middot; ~40% top-20 client adoption</div>
                <div class="timeline-block" style="background:#6b7280;color:#fff;">Present &mdash; CEO<br>Scaling to Series B</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item">
                    <div class="legend-dot" style="background:#3498db;"></div>
                    <span><strong>Education:</strong> Undergraduate at Penn State (per LinkedIn).</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#1e40af;"></div>
                    <span><strong>Early Trading:</strong> Pre-Nomura finance roles building into institutional seat.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#16a085;"></div>
                    <span><strong>Nomura &mdash; Managing Director, Equity Finance:</strong> Institutional MD-level seat in prime brokerage / equity finance trading.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#1e8449;"></div>
                    <span><strong>Jefferies &mdash; Managing Director, Securities Lending:</strong> Second MD-level institutional seat. Domain depth in regulated market infrastructure.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#d4a84b;"></div>
                    <span><strong>Provable Markets &mdash; Founded (Seed):</strong> Co-founded regulated FinTech for prime-brokerage / securities-lending infrastructure. Raised $18M Seed.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#c0392b;"></div>
                    <span><strong>Provable Markets &mdash; Series A Extension (through co-founder separation):</strong> Executed a co-founder separation while scaling the business; moved through to $75M Series A Extension; publicly reported ~40% adoption among top-20 prime-brokerage clients.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#6b7280;"></div>
                    <span><strong>Present &mdash; CEO:</strong> Running Provable Markets into the Series B arc.</span>
                </div>
            </div>
            <div class="timeline-banner">
                Career pattern: two institutional MD-level seats at Nomura and Jefferies (domain depth + selector validation), then founded Provable Markets and moved the company from Seed through a contested co-founder separation and out to a Series A Extension at $75M. First CEO seat; first true founder arc; strongest operational evidence is the separation executed while scaling.
            </div>"""

    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

# ============================================================================
# BUILD RESPONDENT DICT FOR MOTIVATORS_SECTION
# ============================================================================

def build_respondent_dict(respondent_data):
    """Build the respondent dict for motivators_section.build_section().
    Cohen TTI:
      DISC Natural:  D=69, I=32, S=48, C=61
      DISC Adapted:  D=78, I=28, S=35, C=64
      Wheel position: (9) Conducting Implementor (both Natural and Adapted)
      DF Primary:  Objective 83, Structured 68, Resourceful 58, Commanding 49
    """
    nat_pos = 9
    nat_label = 'Conducting Implementor'
    # Primary wedge for "Conducting Implementor" is Implementor — C-dominant with D secondary.
    # Intensity = (D + C) / 200 reflects the twin engines of Conducting Implementor.
    nat_intensity = (69 + 61) / 200.0  # 0.65

    adp_pos = 9
    adp_label = 'Conducting Implementor'
    adp_intensity = (78 + 64) / 200.0  # 0.71

    shift_note = 'Adapted shift: +9D, −4I, −13S, +3C. The current seat is dialing up drive and dialing down steadiness/warmth.'

    respondent = {
        'name': 'Matthew Cohen',
        'first_name': 'Matthew',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [69, 32, 48, 61],  # D, I, S, C Natural
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }

    return respondent

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    """Post-render QA gate — every check matches _pipeline/QA_CHECKLIST.md."""
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    n_concern_item = count_class('concern-item')
    n_concern_number = count_class('concern-number')
    n_concern_text = count_class('concern-text')
    if n_concern_item < 2:
        failures.append(f"S1: concern-item count {n_concern_item} < 2 (min)")
    if n_concern_number != n_concern_item:
        failures.append(f"S1: concern-number ({n_concern_number}) != concern-item ({n_concern_item})")
    if n_concern_text != n_concern_item:
        failures.append(f"S1: concern-text ({n_concern_text}) != concern-item ({n_concern_item})")
    concern_nums = re.findall(r'<div class="concern-number">(\d+)</div>', html)
    expected_seq = [str(i) for i in range(1, n_concern_item + 1)]
    if concern_nums != expected_seq:
        failures.append(f"S1: concern-number sequence {concern_nums} != expected 1..{n_concern_item}")

    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: class='{cls}' expected exact {expected}, got {got}")

    probe_questions = re.findall(
        r'<div class="probe-question">(.*?)</div>', html, re.DOTALL
    )
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched_probes = []
        for idx, q_html in enumerate(probe_questions, start=1):
            q_lower = q_html.lower()
            matched_here = False
            for (_name, _canonical, substr) in FORM8_QUESTIONS:
                if substr.lower() in q_lower:
                    matched_substrs.add(substr)
                    matched_here = True
                    break
            if not matched_here:
                unmatched_probes.append(idx)
        if len(matched_substrs) < 10:
            failures.append(
                f"S9-Form8: Only {len(matched_substrs)}/10 canonical Form 8 strings present. "
                f"Unmatched: {unmatched_probes}."
            )

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: class='{cls}' expected min {expected}, got {got}")

    # Teach Items fraction
    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value block not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be fraction N/M, got '{teach_val[:60]}'")
        if 'To be populated' in html:
            failures.append("S2: Forbidden placeholder text 'To be populated'")

    # Chart 1 zLabels2 nested
    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:
        failures.append("S4-C1: zLabels2 not a nested 2-row array")

    # Chart 2 — empty columns hidden + last label has "+"
    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/failData/successData")
    else:
        sf_labels = json.loads(m_sf_l.group(1))
        fail_data = json.loads(m_fail.group(1))
        succ_data = json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append(f"S4-C2: sfLabels({len(sf_labels)}) / fail({len(fail_data)}) / success({len(succ_data)}) length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with '+', got {sf_labels[-1]}")

    # Chart 3 — reversed (high flags on LEFT)
    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first_label = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first_label:
            first_val = int(first_label.group(1))
            if first_val < 30:
                failures.append(f"S4-C3: flagLabels3 first label = {first_val} — axis not reversed")

    # Dimensional Scorecard — L1+L2, L1 UPPERCASE, L2 indented
    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        exc_labels = json.loads(m_exc_labels.group(1))
        is_l1_arr = json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: Scorecard has only {len(exc_labels)} rows (min 12)")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: Scorecard has only {l2_rows} L2 rows (min 6)")
        for lbl, is_l1 in zip(exc_labels, is_l1_arr):
            if is_l1 and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' must be UPPERCASE")
                break
            if not is_l1 and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' must be indented 4 spaces")
                break
    else:
        failures.append("S6: excLabels / isL1 not found in output")

    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    # Role-Fit section — seat-responsive two-column split (new canonical Variant 2 block)
    for cls, min_count in [('role-fit-box', 1), ('role-fit-grid', 1),
                            ('role-fit-col easy', 1), ('role-fit-col hard', 1)]:
        got = count_class(cls)
        if got < min_count:
            failures.append(f"S-RF: class='{cls}' expected min {min_count}, got {got}")
    if 'What Will Be Easy' not in html or 'What Will Be Hard' not in html:
        failures.append("S-RF: Role-Fit column labels missing")
    if 'Series B' not in html:
        failures.append("S-RF: Series B seat reference missing from Role-Fit")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} issue(s) ***")
        for f in failures:
            print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)} issue(s)")
    else:
        print("*** QA GATE PASSED ***")


def main():
    print("Loading respondent data...")
    respondent_data = load_respondent_data()
    print(f"  name={respondent_data['name']}, Z|Algo={respondent_data['z_algo_overall']:+.3f}, RF={respondent_data['rf_num']}")
    print(f"  flags: {list(respondent_data['flags_lit'].keys())}")

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution chart tokens...")
    dist_tokens = build_distribution_tokens(
        zalgo_rows,
        flag_rows,
        respondent_data['z_algo_overall'],
        respondent_data['z_human_overall'],
        respondent_data['rf_num']
    )

    print("Building respondent dict for motivators section...")
    respondent = build_respondent_dict(respondent_data)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()
    role_fit = build_role_fit()

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    replacements = {
        'CANDIDATE_NAME': 'Matthew Cohen',
        'CANDIDATE_CREDS': 'Candidate &mdash; Series B Growth-Stage CEO Seat',
        'CANDIDATE_ROLE': 'Candidate for Series B Growth-Stage CEO &mdash; Regulated FinTech / Capital-Markets Infrastructure',
        'REPORT_DATE': 'April 22, 2026',

        'ZALGO_OVERALL': f'{respondent_data["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{respondent_data["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(respondent_data['rf_num']),
        'FLAGS_LIT': f'{len(respondent_data["flags_lit"])}/20',
        'TEACH_ITEMS': '10/10',
        'COHORT_AVG': '+0.24',
        'COHORT_AVG_NUM': '0.2440',

        **axes,
        **concerns,
        **wiring,
        **htl,
        **probes,
        **teach,
        **radar,
        **timeline,
        **role_fit,
        **dist_tokens,

        # DISC scores (Natural)
        'DISC_D_NAT': '69',
        'DISC_I_NAT': '32',
        'DISC_S_NAT': '48',
        'DISC_C_NAT': '61',
        # DISC scores (Adapted)
        'DISC_D_ADP': '78',
        'DISC_I_ADP': '28',
        'DISC_S_ADP': '35',
        'DISC_C_ADP': '64',
        'DISC_NOTE_TEXT': 'Conducting Implementor (D + C twin engines, I below midline). Adapted shift: +9D, -4I, -13S, +3C.',
        'DISC_NOTE_DETAIL': 'Current environment is dialing up drive and compressing steadiness/warmth. TTI Primary cluster &mdash; Objective 83, Structured 68, Resourceful 58, Commanding 49 &mdash; corroborates the structured-directive execution signature. Indifferent cluster (Harmonious 0, Selfless 24, Receptive 25, Altruistic 31) sits directly opposite the Investing-in-Others wedge.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        **build_excstds_scorecard(respondent_data),

        'RECOMMENDATION_TEXT': 'CONDITIONAL HIRE &middot; DILIGENCE-REQUIRED &mdash; the Provable founder arc is strong evidence of a CEO operating engine; the initiating-accountability + depth-below-directs pattern is the material liability that determines Series B fit. Advance to onsite conditional on validating (a) conductor routines installable against the wiring, and (b) the existing exec team can carry distributed accountability &mdash; or plan to pair with a COO/President whose explicit charter is that distribution.',
    }

    # Neutralize EXCSTDS_COLOR_OVERRIDES in comment
    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    # Inject motivators section first
    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    # Fix title
    html = re.sub(
        r'<title>.*?</title>',
        '<title>Matthew Cohen &mdash; CEO, Provable Markets | HALE GLOBAL SUCCESS DIAGNOSTICS</title>',
        html,
        count=1
    )

    # Fix header brand lockup
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    qa_gate(html)

    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")


if __name__ == '__main__':
    main()
