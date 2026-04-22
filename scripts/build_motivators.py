"""Build the Motivators & Anti-Motivators HTML fragment for a respondent.

This is the ONE, AUTHORITATIVE entry point used by both the hiring-report and
the talent-coaching-guide workflows. It encapsulates the full routine:

  1. Load respondent L2 scores from the pipeline xlsx (`L2` sheet).
  2. Load the wedge map universe (32 L2s in `l2_wedge_map.xlsx`).
  3. Enforce the L2-coverage gate — abort if fewer than N-4 L2s are present.
  4. Call `pipeline.motivators_section.build_section(respondent)`.
  5. Return both the fragment and a verification report the caller can assert on.

Usage (from report skill):

    from pathlib import Path
    import sys; sys.path.insert(0, str(Path('_pipeline')/'src'))
    sys.path.insert(0, str(Path('_pipeline')/'scripts'))
    from build_motivators import build_motivators_fragment

    frag, report = build_motivators_fragment(
        respondent_xlsx='respondents/<key3>/data.xlsx',
        name='First Last', first_name='First',
        nat_pos=7, nat_label='Analyzing Implementor', nat_intensity=0.81,
        adp_pos=7, adp_label='Implementing Analyzer', adp_intensity=0.84,
        disc=[75, 34, 18, 81],
        shift_note='...',
    )
    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', frag)

    # Build-time verification — run AFTER all token substitutions
    verify_motivators_injected(html, report)  # raises AssertionError on any failure

The verify_motivators_injected() helper runs the four mandatory build-time checks
documented in SKILL.md Step 5.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# Max L2s in the wedge map that can be legitimately unscored by the pipeline
# (Q_Count = 0). A budget of 4 accommodates the current "Execs can do the job"
# and "CEO gets outside exec" etc. — any gap larger than this means the caller
# passed a scorecard subset instead of the full pipeline L2 set.
COVERAGE_BUDGET = 4


def _load_l2_scores_from_xlsx(xlsx_path: str | Path) -> dict[str, float]:
    """Read all scored L2s from the pipeline respondent xlsx.

    The pipeline's `L2` sheet has columns:
        A: L1_Num, B: L1_Title, C: L2_Num, D: L2_Short, E: Score5_filtered, F: Q_Count
    We key by L2_Short (column D) and value by Score5_filtered (column E).
    L2s with Question_Count = 0 are returned as None in the xlsx — we skip them.
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    if 'L2' not in wb.sheetnames:
        raise ValueError(f'{xlsx_path}: no L2 sheet found (got {wb.sheetnames})')
    ws = wb['L2']
    scores: dict[str, float] = {}
    for r in range(2, ws.max_row + 1):
        short = ws.cell(row=r, column=4).value
        score = ws.cell(row=r, column=5).value
        if short and isinstance(score, (int, float)):
            scores[short] = round(float(score), 3)
    return scores


def _load_wedge_map_universe(mapping_xlsx: str | Path = 'l2_wedge_map.xlsx') -> set[str]:
    """Return the full set of L2_Short names in the wedge map (the 32-L2 universe)."""
    wb = load_workbook(str(mapping_xlsx), data_only=True)
    if 'L2_Wedge_Mapping' not in wb.sheetnames:
        raise ValueError(f'{mapping_xlsx}: no L2_Wedge_Mapping sheet found')
    ws = wb['L2_Wedge_Mapping']
    # Row 1 is the caveat banner; headers on row 2; data row 3+.
    # L2_Short is column C.
    universe = {
        ws.cell(row=r, column=3).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=3).value
    }
    return universe


def _enforce_coverage_gate(l2_scores: dict, universe: set) -> tuple[set, set]:
    """Raise AssertionError if coverage is outside budget; else return (covered, missing)."""
    covered = set(l2_scores) & universe
    missing = universe - set(l2_scores)
    if len(covered) < len(universe) - COVERAGE_BUDGET:
        raise AssertionError(
            f'L2 coverage FAILED: only {len(covered)} of {len(universe)} L2s present. '
            f'Missing: {sorted(missing)}. '
            f'The Standard-Map wheel would render as a near-empty circle. '
            f'Rebuild l2_scores from the pipeline xlsx `L2` sheet '
            f'(use _load_l2_scores_from_xlsx), NOT from the scorecard/Impact subset.'
        )
    return covered, missing


def build_motivators_fragment(
    *,
    respondent_xlsx: str | Path,
    name: str,
    first_name: str,
    nat_pos: int,
    nat_label: str,
    nat_intensity: float,
    adp_pos: int,
    adp_label: str,
    adp_intensity: float,
    disc: list[int],
    shift_note: str = '',
    partial_data: Optional[bool] = None,
    mapping_xlsx: str | Path = 'l2_wedge_map.xlsx',
) -> tuple[str, dict]:
    """Return (fragment_html, verification_report).

    verification_report contains:
        coverage       : 'N of M L2s covered'
        missing        : sorted list of L2_Short names absent from l2_scores
        fragment_size  : len(fragment)
        rect_count     : number of <rect> inside the embedded SVG
        gray_pills     : number of no-score gray pills in the SVG
        l2_row_count   : number of class="l2-row" in the alignment block (capped at 12)
        trace          : the HTML comment the caller should also inject (optional)
    """
    # Late import to keep this script importable even if pipeline/ isn't on path yet.
    import sys
    src = Path(__file__).resolve().parents[1] / 'src'
    if str(src) not in sys.path:
        sys.path.insert(0, str(src))
    from pipeline.motivators_section import build_section

    l2_scores = _load_l2_scores_from_xlsx(respondent_xlsx)
    universe = _load_wedge_map_universe(mapping_xlsx)
    covered, missing = _enforce_coverage_gate(l2_scores, universe)

    if partial_data is None:
        # Pipeline legitimately couldn't score a few L2s (Q_Count=0) — this is
        # NOT "partial data" in the user-facing sense. Only flag partial_data=True
        # if the pipeline ITSELF returned a subset (which is rare).
        partial_data = len(missing) > COVERAGE_BUDGET

    respondent = {
        'name': name,
        'first_name': first_name,
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': list(disc),
        'l2_scores': l2_scores,
        'shift_note': shift_note,
        'partial_data': partial_data,
    }
    fragment = build_section(respondent)

    # Analyse the rendered SVG block for verification report.
    svg_s = fragment.find('<svg viewBox')
    svg_e = fragment.find('</svg>', svg_s) + 6
    svg = fragment[svg_s:svg_e] if svg_s >= 0 else ''
    l2_row_class = 'class="l2-row"'
    report = {
        'coverage': f'{len(covered)} of {len(universe)} L2s',
        'missing': sorted(missing),
        'fragment_size': len(fragment),
        'rect_count': svg.count('<rect '),
        'gray_pills': svg.count('#e8ecf0'),
        'l2_row_count': fragment.count(l2_row_class),
        'trace': (
            '<!-- Motivators & Anti-Motivators — built by '
            '_pipeline/scripts/build_motivators.py via motivators_section.build_section() '
            f'on {len(covered)} of {len(universe)} L2s. '
            'See SKILL.md: Motivators is a REQUIRED section with L2-coverage gate. -->'
        ),
    }
    return fragment, report


def inject_motivators_into_template(
    template_html: str,
    fragment: str,
    report: dict,
) -> str:
    """Replace the `{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}` token with the fragment
    prefixed by a trace comment. Raises AssertionError if the token isn't present
    (which would mean the template was already filled or is malformed).
    """
    token = '{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}'
    if token not in template_html:
        raise AssertionError(
            f'Template does NOT contain {token} — either it was already stripped '
            '(most common regression: upstream .replace call with empty string) '
            'or the template is wrong. Rebuild from _templates/.'
        )
    return template_html.replace(token, report['trace'] + '\n        ' + fragment)


def verify_motivators_injected(html: str, report: dict) -> None:
    """Run the four mandatory build-time checks. Raises AssertionError on any fail.

    These MUST be run AFTER all token substitutions, BEFORE saving the report.
    """
    checks = [
        ('placeholder replaced',
         '{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}' not in html),
        ('ma-section scoped CSS class present',
         'class="ma-section"' in html),
        ('trace comment survived',
         '<!-- Motivators & Anti-Motivators' in html),
        ('SVG rect coverage (>= 30 of 32 L2 pills + 4 ring labels, tolerated budget)',
         html[html.find('<svg viewBox'): html.find('</svg>', html.find('<svg viewBox'))].count('<rect ')
            >= 34 - COVERAGE_BUDGET),
    ]
    fails = [name for name, ok in checks if not ok]
    if fails:
        raise AssertionError(
            f'Motivators build-time verification FAILED: {fails}. '
            f'Fragment report: {report}. '
            f'Do NOT save the file — fix and re-inject.'
        )


# ---- CLI entry point for quick one-off rebuilds ----
if __name__ == '__main__':
    import argparse
    import json
    p = argparse.ArgumentParser()
    p.add_argument('--xlsx', required=True, help='respondent data.xlsx path')
    p.add_argument('--name', required=True)
    p.add_argument('--first-name', required=True)
    p.add_argument('--nat-pos', type=int, required=True)
    p.add_argument('--nat-label', required=True)
    p.add_argument('--nat-intensity', type=float, required=True)
    p.add_argument('--adp-pos', type=int, required=True)
    p.add_argument('--adp-label', required=True)
    p.add_argument('--adp-intensity', type=float, required=True)
    p.add_argument('--disc', required=True, help='"D,I,S,C" e.g. "75,34,18,81"')
    p.add_argument('--shift-note', default='')
    p.add_argument('--out', required=True, help='output .html fragment file')
    args = p.parse_args()
    disc = [int(x) for x in args.disc.split(',')]
    frag, rep = build_motivators_fragment(
        respondent_xlsx=args.xlsx,
        name=args.name, first_name=args.first_name,
        nat_pos=args.nat_pos, nat_label=args.nat_label, nat_intensity=args.nat_intensity,
        adp_pos=args.adp_pos, adp_label=args.adp_label, adp_intensity=args.adp_intensity,
        disc=disc, shift_note=args.shift_note,
    )
    Path(args.out).write_text(frag, encoding='utf-8')
    print(json.dumps(rep, indent=2))
    print(f'\nWrote fragment -> {args.out} ({len(frag):,} bytes)')
