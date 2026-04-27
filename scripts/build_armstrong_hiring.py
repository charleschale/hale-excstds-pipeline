"""Build the Excellence Standards hiring report for Patrick Armstrong (CFO).

Forked from build_schott_hiring.py 2026-04-27. Self-reported as CFO at Provable Markets
in the survey (Q109), but LinkedIn shows current employer as Avolve (May 2025-present).
Evaluated for a Series B growth-stage CFO seat. LinkedIn full-profile sweep complete
(Charles supplied 2026-04-27): 22-year career starting at CDI Corp Sr Analyst (2004),
through Exelon, Lincoln Financial, Comcast Treasury (2010-2013), In-Recruit MD,
Jornaya Director of Finance, then LoanLogics Director of Finance promoted to CFO
(2018-2022, divestiture to Sun Capital + bolt-on of LoanBeam), brief consultant stint,
SoLo Funds CFO (2023-2025, 1y7m), Avolve CFO (May 2025-present). MBA Villanova Finance.
No varsity sports, no awards, no Big-4-Partner level signal.

Run from repo root:
    python _pipeline/scripts/build_armstrong_hiring.py
Outputs to:
    _reports/Armstrong_Patrick_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section, compute_intensity_from_disc

RESPONDENT_XLSX = ROOT / '_respondents' / '20260417.armstrongpatrick@live.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Armstrong_Patrick_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title, 'z_algo': z_algo, 'z_human': z_human, 'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flag_names = []
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        if flag_name:
            flag_names.append((col, flag_name))
    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data, 'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall, 'z_human_overall': z_human_overall,
        'rf_num': rf_num, 'questions_answered': questions_answered,
        'flags_lit': flags_lit, 'non_scorable': non_scorable,
        'name': name, 'email': email, 'date_str': date_str,
    }

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]
    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo, z_human, sf = row['z_algo'], row['z_human'], row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1; continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1; break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    rev_flag_bin = len(flag_counts) - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# NARRATIVE
# ============================================================================

def build_three_axes_narratives():
    talent_badge = 'badge-amber'
    talent_badge_text = 'AMBER'
    talent_card_body = '''Talent is the capacity to grow, read holistically across the career record. The right questions: <strong>has this person been promoted consistently?</strong> <strong>Does the file show greatness in more than one area?</strong> <strong>Does the arc suggest capacity to grow into the next altitude?</strong> On Armstrong&rsquo;s file the answers are, on balance, <em>no</em>.

The strongest line is the LoanLogics arc: <strong>internal promotion from Director of Finance (Dec 2018) to CFO and Treasurer (Jul 2019), 3y2m total</strong>, with documented accomplishments &mdash; led divestiture to Sun Capital Partners and executed a bolt-on acquisition of LoanBeam. That is genuine CFO operating evidence and one real internal-promotion event. But it is the only one on the file. Earlier career was Big-Co individual-contributor finance roles &mdash; CDI Sr Analyst, Exelon, Lincoln Financial, Comcast Treasury &mdash; without internal-promotion advancement at any of them. The In-Recruit MD seat (1y11m) and Jornaya Director of Finance (3y7m) were lateral entries, not promotions earned within an existing organization. The two CFO seats since LoanLogics &mdash; SoLo Funds (1y7m) and Avolve (1y on file) &mdash; are also lateral entries.

The greatness-in-more-than-one-area read is also light: the entire career sits inside the finance lane. Big-Co finance, then Director of Finance at a private company, then CFO &mdash; one craft, executed across multiple seats. No documented adjacent excellence (advisory, board, athletic, civic, intellectual-public-output, founder). MBA Villanova is a solid mid-tier credential without the elite-selector signal density (Big-4 Partner, NYSE-listed CFO, top-firm leadership, awards) that usually overrides instrument-negative reads at this severity.

<strong>What pulls Talent to amber rather than amber-red:</strong> the LoanLogics divestiture-plus-acquisition is a real CFO accomplishment that requires both deal-execution and operating-discipline skill, and the LoanLogics internal-promotion event is genuine. <strong>What keeps it from green:</strong> the rest of the file does not show the consistent-promotion / cross-domain-greatness / capacity-to-grow pattern, and the instrument is reading deeply negative on the dimensions Talent evidence usually overrides &mdash; Conducting &amp; Outvoted L1 &minus;2.11, Replacing Self L1 &minus;2.47, Pushing Extreme Accountability L1 &minus;2.20, four Sev flags.'''

    judgment_badge = 'badge-red'
    judgment_badge_text = 'RED'
    judgment_card_body = '''The instrument is unambiguous. <strong>Z|Algo overall &minus;0.84 sits below the cohort average of +0.24</strong>, and the negative L1 mass is concentrated in the dimensions that constitute executive judgment under pressure. Conducting &amp; Outvoted L1 <strong>&minus;2.11</strong>. Replacing Self L1 <strong>&minus;2.47</strong>. Pushing Extreme Accountability L1 <strong>&minus;2.20</strong> with <strong>9 reverse flags out of 15 questions answered</strong> (60% wrong-direction answers on accountability mechanics). Facilitative Mindset L1 &minus;1.08. The L2 evidence converges: Dialogue vs. Direction L2 &minus;1.53, Respects Collective Wisdom L2 &minus;1.50, CEO gets outside exec L2 &minus;1.47, Drives Accountability L2 &minus;1.08, Standards = What You Tolerate L2 &minus;1.04.

<strong>Four Sev flags + one Hi flag + an &ldquo;Upgrade Team&rdquo; DreamTeam read.</strong> Conditional Belief Sev. HandsOn Sev. <strong>Initiating Accountability Sev</strong>. Driving Accountability Sev. HoldsAvgDownChain Hi. The accountability architecture is not just weak &mdash; it is broken at multiple levels: he does not install single-point ownership (Clarity flag lit), does not drive consequences (Driving Accountability Sev), does not initiate accountability through dialogue (Initiating Accountability Sev), does not hold the standard at depth (HoldsAvgDownChain Hi), and does not extend belief in others (Conditional Belief Sev).

Counter-evidence on the same file: Deliberate Urgency L1 +2.51 (top decile) with Action Over Inaction L2 +2.16, Personal Reliability L1 +0.89, Organizational Decision Making L1 +0.91 with Simplification Methods L2 +2.79. He moves fast, owns his own work, simplifies decisions when he&rsquo;s the one deciding. The judgment failure mode is not analytical or velocity-related; it is <em>structural and team-facing</em> &mdash; the moment the work involves running it through a team, the architecture collapses.

TTI conflict section corroborates: <em>&ldquo;may micromanage and unintentionally alienate others,&rdquo; &ldquo;may feel his view is the only way and not see the reasoning behind other viewpoints,&rdquo; &ldquo;sets unreachable standards for himself and others.&rdquo;</em> No wiring-vs-behavior divergence on this read.'''

    skills_badge = 'badge-amber'
    skills_badge_text = 'AMBER'
    skills_card_body = '''The instrument does not measure domain skill directly. Domain credentials are appropriate for a Series B CFO seat: 22-year finance career, MBA Finance (Villanova), three CFO seats (LoanLogics CFO with divestiture + acquisition execution, SoLo Funds, Avolve), and Big-Co training (Comcast Treasury, Lincoln Financial Strategy &amp; Profitability). FP&amp;A, banking and investor relations, internal controls, divestiture-side deal experience, and bolt-on acquisition experience are all on the file.

<strong>The wiring read does NOT support a canonical CFO fit.</strong> Both Natural (position 60, Promoting Analyzer) and Adapted (position 56, Analyzing Implementor) are explicitly marked <strong>&ldquo;ACROSS&rdquo;</strong> in the TTI &mdash; that is wheel-speak for <em>in the center, between wedges, no strong anchor</em>. Natural DISC scores (D=48, I=52, S=52, C=71) and Adapted (D=52, I=52, S=35, C=75) describe an operator whose three lower DISC values sit within 17 points of each other; only C is meaningfully elevated. <strong>Center-of-wheel profiles correlate with lower success rates across all positions</strong> &mdash; the strongly-anchored profiles aligned to a role&rsquo;s demands are what predict success, and Armstrong is not strongly anchored anywhere. C=75 alone is not enough; without a clear anchor in the Implementor or Conductor wedge, the wiring is not the CFO-canonical pattern. TTI Driving Forces &mdash; Intellectual 85 (Passionate, 2&sigma; above mean), Intentional 72, Objective 69, Commanding 64 (Indifferent: Altruistic 12, Harmonious 12, Instinctive 7) &mdash; carry genuine intellectual curiosity but do not change the wheel-position read.

For a Series B growth-stage CFO seat, the wiring picture and the instrument picture point the same direction: the file does not show a wiring fit that compensates for the four Sev flags and the below-cohort Z|Algo. Self-reported revenue trajectory ($20MM May 2025 &rarr; $23MM May 2026, ~15% YoY) is modest for Series B growth-stage.'''

    return {
        'TALENT_BADGE_CLASS': talent_badge,
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': judgment_badge,
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': skills_badge,
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    concerns_title = 'Two Targeted Concerns'
    concerns_items = """            <div class="concern-item">
                <div class="concern-number">1</div>
                <div class="concern-text">
                    <strong>Accountability architecture is broken at every level &mdash; Initiating, Driving, Clarity, and Holding all fail.</strong>
                    Four flags lit on accountability: Initiating Accountability <strong>Sev</strong>, Driving Accountability <strong>Sev</strong>, Clarity of Accountability lit, HoldsAvgDownChain <strong>Hi</strong>. The L1 evidence: L1 #5 Pushing Extreme Accountability &minus;2.20 with <strong>9 reverse flags out of 15 questions</strong> (60% wrong-direction). L1 #3 Conducting &amp; Outvoted &minus;2.11. The L2 evidence: Dialogue vs. Direction &minus;1.53, Respects Collective Wisdom &minus;1.50, Drives Accountability &minus;1.08, Standards = What You Tolerate &minus;1.04, Sublimating Ego &minus;0.91, Power &amp; Status Management +0.81 (only mid). This is not a single-mechanism failure; it is the entire conducting-through-a-team-without-being-the-bottleneck loop. Probe Form 8 Leadership Deep-Dive: ask him to sketch his most recent finance org chart and name single-point owners for each priority. Listen for whether he describes a team that runs without him or one where everything routes back to him. The TTI conflict section names the same pattern in his own wiring: <em>&ldquo;may feel his view is the only way and not see the reasoning behind other viewpoints,&rdquo; &ldquo;may micromanage and unintentionally alienate others.&rdquo;</em>
                </div>
            </div>
            <div class="concern-item">
                <div class="concern-number">2</div>
                <div class="concern-text">
                    <strong>Replacing Self &minus;2.47 + HandsOn Sev + Conditional Belief Sev = the can&rsquo;t-disappear, won&rsquo;t-extend-belief CFO.</strong>
                    The worst L1 score in his profile is Replacing Self at &minus;2.47, with L2 4.4 CEO gets outside exec &minus;1.47 and L2 4.3 Urgency Down Chain &minus;0.06. HandsOn lit at <strong>Sev</strong>. Conditional Belief lit at <strong>Sev</strong>. DreamTeam reads as &ldquo;Upgrade Team&rdquo; (the strong form, not just &ldquo;Close&rdquo;). Lower Standards for Others than Self flag lit. Pattern: he holds himself to extreme standards (TTI: <em>&ldquo;sets unreachable standards for himself and others&rdquo;</em>), sees the team as not measuring up, and absorbs the work back. The tenure shape on the LinkedIn record corroborates &mdash; SoLo Funds 1y7m, Avolve 1y on file, the modest $20MM&rarr;$23MM revenue trajectory at the current company. Probe Form 8 Talent Development: <em>&ldquo;What people over your career have you nurtured who have gone on to do great things?&rdquo;</em> Listen for one specific pre-proof developmental investment, not retrospective admiration of the ones who already delivered. Then Continuous Improvement: ask whether he counts time spent coaching the FP&amp;A bench and the Controller as core finance work or as overhead.
                </div>
            </div>"""
    return {'CONCERNS_TITLE': concerns_title, 'CONCERNS_ITEMS': concerns_items}

def build_wiring_fit():
    wiring_fit_items = (
        '<strong>Center-of-wheel wiring &mdash; no strong anchor in any wedge.</strong> '
        'Natural position 60 (Promoting Analyzer) and Adapted position 56 (Analyzing Implementor) '
        'are both marked <strong>&ldquo;ACROSS&rdquo;</strong> in the TTI &mdash; wheel-speak for between-wedges, in the center, '
        'no specialized natural orientation. Natural DISC (D=48, I=52, S=52, C=71) and Adapted (D=52, I=52, S=35, C=75) '
        'have only C meaningfully elevated; the other three values cluster within 17 points. <em>Center-of-wheel profiles '
        'correlate with lower success rates across all positions</em> &mdash; specialization in role-aligned wedges is what '
        'predicts role success. The C=75 controls-leaning element is real but does not by itself produce the canonical CFO '
        'Implementor anchor that the seat rewards.'
        '<span class="wiring-flag">Targeted Concern</span><br>'
        '<strong>The TTI Conflict section names the in-the-room failure modes directly:</strong> '
        '<em>&ldquo;May micromanage and unintentionally alienate others,&rdquo; &ldquo;May feel his view is the only way and not see '
        'the reasoning behind other viewpoints,&rdquo; &ldquo;Sets unreachable standards for himself and others,&rdquo; &ldquo;Can come '
        'across as cool and distant because he wants to do everything his way.&rdquo;</em> Each line corroborates a different lit '
        'flag (HandsOn Sev, Conducting &amp; Outvoted, Lower Standards for Others than Self, Conditional Belief Sev). No wiring-vs-'
        'behavior divergence on this read.'
        '<span class="wiring-flag">Targeted Concern</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    # Hard-to-Learn signals: Low Urgency? NO (+2.51 top decile). Low Org DM? NO (+0.91).
    # Conditional Belief Sev? YES. Satisfied with Gripes? YES (Low). Total: 2/4 lit.
    # But the lit ones are heavy — both Conditional Belief and SatVsGripes are diagnostic.
    return {'HARD_TO_LEARN': '2/4'}

FORM8_QUESTIONS = [
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]

def build_interview_probes():
    tailored = [
        ("TWO-SPORT ATHLETE (Talent axis)",
         "No varsity-sport signal on the LinkedIn record (no athletics listed at Villanova) and no documented awards. Listen for whether the proudest result is a finance accomplishment (the LoanLogics divestiture or bolt-on acquisition would be the cleanest reach) or something further afield. The Talent-axis goal is to find a pursuit OUTSIDE the finance lane that revealed grit, coachability, or being bad at something before becoming good. If the answer goes immediately to the LoanLogics deals or the SoLo Funds raise, press once for a non-career pursuit before accepting."),
        ("TALENT DEVELOPMENT (Concern 2 &middot; Pre-proof belief)",
         "Central probe for Concern 2. Conditional Belief <strong>Sev</strong> + L2 1.6 Developmental Discipline &minus;0.10 + L2 1.5 Reciprocal Followership &minus;0.04 + Lower Standards for Others than Self flag + Upgrade Team flag = the strongest predictor in his profile that he extends belief AFTER proof, not before. Listen for one specific case where he backed someone (gave them a stretch role, defended a risky hire, kept investing in someone struggling) <em>before</em> they had earned it. Strong Talent-Development records produce two such names inside 60 seconds. Generic answers, retrospective-only stories, or &ldquo;I challenge people every day&rdquo; without specific names are negative signals."),
        ("TORC &mdash; TRUTH-OVER-COMFORT (Self-awareness)",
         "Anchor to a real boss: the LoanLogics CEO during the divestiture window, the SoLo Funds CEO, or the Sun Capital Partners deal-side counterparty. TTI says he <em>&ldquo;takes criticism personally but will simultaneously move to another aspect of the project&rdquo;</em> &mdash; classic deflection-via-task-shift. Listen for whether he names a real developmental area his most recent boss would identify, or pivots to strengths-reframing. Q9913 self-rating answered '4' &mdash; the second-from-best self-grade &mdash; suggests low self-criticism even before the interview begins; press past the first answer."),
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "Three CFO seats over five years (LoanLogics 3y2m, SoLo Funds 1y7m, Avolve 1y on file) and a 5-month consultant gap (Feb-Jun 2023) put real adversity material on the table. Listen for whether he names a personal adversity or a professional setback &mdash; and if professional, where the ownership lands. The healthy answer attributes learning to self even when the situation was genuinely outside his control; the unhealthy answer accumulates external attributions across the seat transitions."),
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior probe. With TTI Intellectual Primary 85 (Passionate, top of the file), Objective 69, and Adapted Analyzing Implementor wiring, the cleanest strong answer is a specific intellectual / analytical pursuit &mdash; a research routine, a technical-finance deep-dive habit, a reading or model-building discipline that compounds. Generic answers (&ldquo;I work harder,&rdquo; &ldquo;I outprepare people&rdquo;) are negative signals. Press for the artifact: the model template, the deep-work block, the reading list."),
        ("FACILITATIVE MINDSET (Concern 1 &middot; Dialogue and dissent)",
         "Central probe for Concern 1. The L2 evidence is severe here: Dialogue vs. Direction <strong>&minus;1.53</strong>, Respects Collective Wisdom &minus;1.50, Sublimating Ego &minus;0.91. The TTI conflict section: <em>&ldquo;may feel his view is the only way and not see the reasoning behind other viewpoints.&rdquo;</em> A Facilitative answer names a belief AND the conditions under which he would update it AND ideally cites a recent specific case where his team changed his mind on something material. A non-Facilitative answer holds the belief absolute. Follow-up: <em>&ldquo;When was the last time someone on your team convinced you to change your mind on something that mattered?&rdquo;</em> &mdash; pause and listen for specifics."),
        ("COMMITMENT (Diligence wiring &middot; one of the strongest dimensions on file)",
         "Personal Reliability L1 +0.89 + L2 6.3 Commitment to Routine +1.08 + Q24 (daily routines for years) + Adapted DISC C=75 = the routine-machinery is wired in cleanly. Strong answer names a concrete daily practice (a 6 am routine, a specific reading or modeling block, a calendar discipline). Vague answers would contradict the instrument; press for the artifact."),
        ("LEADERSHIP DEEP-DIVE (Concerns 1 &amp; 2 &middot; Org-chart conversation)",
         "Ask him to sketch the Avolve (or LoanLogics looking back) finance org. Tests both Targeted Concerns directly. A leader running the accountability routine names single-point owners for each function and identifies open seats / upgrade candidates in the same breath. Self-graded peer dimensions in the Non-Scorable tab were not filled in (only his own CFO row), so this question is doing the diagnostic work the instrument can&rsquo;t. Follow-ups: <em>&ldquo;Which of these seats would you not re-hire today?&rdquo;</em> + <em>&ldquo;Name one person two levels down you backed before they had proven themselves.&rdquo;</em> Hesitation, multiple owners for the same priority, or post-proof-only names are the flags."),
        ("PASSION &amp; MISSION CONSTRUCTION",
         "Tests mission-construction. The career arc is mostly technical accounting / FP&amp;A / treasury &mdash; substantively serious work without obvious passion-narrative hook. Listen for whether he roots passion in a specific kind of finance problem (tech-enabled lending, fintech operations, deal execution) or defaults to scale-language. The TTI Strengths section reads <em>&ldquo;asks many questions to find the correct answers&rdquo;</em> + <em>&ldquo;takes pride in finding flaws&rdquo;</em> &mdash; his passion-engine may be intellectual/diagnostic. Bridges into &ldquo;Why this seat, why now?&rdquo; given the recent SoLo Funds exit and Avolve tenure."),
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes Concern 2. Listen for whether he counts time spent coaching the FP&amp;A bench, the Controller, or finance-systems owner as core work or only metrics, deliverables, and board prep. A narrow answer (&ldquo;work = output, close, board materials&rdquo;) reinforces the Conditional Belief Sev + Replacing Self &minus;2.47 wiring. A strong answer names a specific coaching cadence with directs &mdash; weekly 1:1 development time, team-development blocks, mentoring routines."),
    ]
    assert len(tailored) == len(FORM8_QUESTIONS) == 10
    probes = []
    for (category, coaching), (_n, canonical_q, _s) in zip(tailored, FORM8_QUESTIONS):
        probes.append((category, f'"{canonical_q}"', coaching))
    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    return {'INTERVIEW_PROBE_CARDS': "\n".join(parts)}

def build_teach_items():
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (3, 'CONDUCTING & OUTVOTED', [
            'Dialogue Vs. Direction',
            'Sublimating Ego',
            'Power & Status Management',
        ]),
        (4, 'RISKING BY REPLACING SELF', [
            'CEO gets outside exec',
            'Urgency Down Chain Of Command',
            'Ability To Disappear',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Drives Accountability',
            'Stds = What Tolerate',
            'Jrs Extreme Proactivity',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Action Over Inaction',
            'Extreme Proactivity',
        ]),
    ]

    labels, scores, is_l1 = [], [], []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    # Two-Sport Athlete: 1 (no varsity sports listed; no athletics signal on file)
    # Punctuates Differently: 2 (Personal Reliability +0.89 mid; Discomfort For Self +0.73 mid;
    #     no rare-behavior signal jumps off the file)
    # Facilitative Mindset: 1 (Dialogue -1.53, Sublimating Ego -0.91, Conducting & Outvoted
    #     L1 -2.11; the worst-rated dimension on the radar)
    # Wit: null
    # Deep Repertoire: 3 (22-yr finance career, MBA Villanova, Comcast Treasury, LoanLogics
    #     CFO with divestiture+acquisition; real but not top-tier)
    # Discipline/Routine: 4 (Personal Reliability +0.89, Commitment to Routine +1.08,
    #     Adapted DISC C=75 = high routine machinery)
    # Understands Symbolism: null
    talent_radar_scores = json.dumps([1, 2, 1, None, 3, 4, None])
    talent_radar_profile = (
        "<strong>Discipline/Routine (4)</strong> is the standout dimension &mdash; Personal "
        "Reliability L1 +0.89, Commitment to Routine L2 +1.08, Adapted C=75 high-routine "
        "wiring; the routine-machinery is wired in cleanly. <strong>Deep Repertoire (3)</strong> "
        "&mdash; 22-year finance career, MBA Villanova, Big-Co training (Comcast Treasury, "
        "Lincoln Financial), LoanLogics CFO with documented divestiture-plus-acquisition "
        "execution; real breadth without elite-firm prestige. <strong>Punctuates Differently "
        "(2)</strong> &mdash; no rare-behavior signal jumps off the file; interview probe "
        "needed. <strong>Two-Sport Athlete (1)</strong> &mdash; read holistically as <em>capacity "
        "to grow</em>, not a varsity-sport checkbox: consistent promotions across roles? (one "
        "internal at LoanLogics; otherwise lateral entries.) Greatness in more than one area? "
        "(One craft, the finance lane.) Cross-domain excellence? (None visible on the file.) "
        "Each lens lands light. <strong>Facilitative Mindset (1)</strong> &mdash; the worst-rated "
        "radar dimension. Dialogue vs. Direction L2 &minus;1.53, Sublimating Ego L2 &minus;0.91, "
        "Conducting &amp; Outvoted L1 &minus;2.11, and TTI <em>&ldquo;may feel his view is the "
        "only way&rdquo;</em> all converge on the same read. <strong>Wit (&#8856;)</strong> and "
        "<strong>Understands Symbolism (&#8856;)</strong> interview-only."
    )
    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_role_fit():
    role_fit_title = 'Role-Fit Read &mdash; What Will Be Easy, What Will Be Hard'
    role_fit_seat = ('Evaluating for: Series B growth-stage CFO seat &middot; same functional '
                     'altitude as today &middot; ~50&ndash;300 employees &middot; building out '
                     'the finance org under the CFO')
    role_fit_easy = (
        '<strong>Personal velocity and reliability are top-decile.</strong> Deliberate Urgency '
        'L1 +2.51 (top decile of the cohort), Action Over Inaction L2 +2.16, Extreme Proactivity '
        'L2 +0.80. Personal Reliability L1 +0.89. Simplification Methods L2 +2.79 (the strongest '
        'single L2 score on the file). Organizational Decision Making L1 +0.91. He moves fast, '
        'simplifies decisions when he&rsquo;s the one deciding, and follows through on his own '
        'commitments.<br><br>'
        '<strong>One genuine CFO accomplishment on the record.</strong> LoanLogics CFO 2019&ndash;'
        '2022: led divestiture to Sun Capital Partners + executed bolt-on acquisition of LoanBeam. '
        'That is real deal-execution and operating-discipline evidence. The internal promotion '
        'path at LoanLogics (Director of Finance &rarr; CFO) is positive selector-validation.'
    )
    role_fit_hard = (
        '<strong>Step back: the wiring is center-of-wheel and the instrument is below cohort.</strong> '
        'Strong CFOs typically anchor in the Implementor or Conductor wedge of the TTI wheel (high-C, '
        'low-I, clearly outside the center). Armstrong&rsquo;s wheel positions are <strong>Natural 60 '
        'Promoting Analyzer (ACROSS) and Adapted 56 Analyzing Implementor (ACROSS)</strong> &mdash; both '
        'marked <em>across-the-center</em>, meaning no strong anchor in any wedge. C=75 alone is not '
        'enough; without specialization in a role-aligned wedge, the wiring is the generalist / lower-'
        'base-rate-success pattern. Z|Algo overall <strong>&minus;0.84 sits below cohort average +0.24</strong>. RF=25 '
        'reverse flags out of 93 questions. <strong>Four Sev flags + one Hi flag + Upgrade Team:</strong> '
        'Conditional Belief Sev, HandsOn Sev, Initiating Accountability Sev, Driving Accountability Sev, '
        'HoldsAvgDownChain Hi. This is the worst flag-basket I have seen in a Hale Global hiring profile.<br><br>'
        '<strong>The accountability architecture is broken at every level.</strong> L1 #5 Pushing Extreme '
        'Accountability &minus;2.20 with <strong>9 reverse flags out of 15 questions answered</strong> '
        '(60% wrong-direction). L1 #3 Conducting &amp; Outvoted &minus;2.11. L2 evidence: Dialogue vs. '
        'Direction &minus;1.53, Respects Collective Wisdom &minus;1.50, Drives Accountability &minus;1.08, '
        'Standards = What You Tolerate &minus;1.04. He does not initiate accountability through dialogue, '
        'does not drive consequences, does not hold the standard at depth, and does not install single-'
        'point ownership. For a Series B CFO seat where the Controller, FP&amp;A lead, and finance-systems '
        'owner below the CFO need to operate without him as the routing node, this is a structural '
        'mismatch with the role.<br><br>'
        '<strong>Replacing Self &minus;2.47 + HandsOn Sev + Conditional Belief Sev = the can&rsquo;t-'
        'disappear, won&rsquo;t-extend-belief CFO.</strong> The TTI Conflict section names the in-the-room '
        'failure modes explicitly: <em>&ldquo;may micromanage and unintentionally alienate others,&rdquo; '
        '&ldquo;may feel his view is the only way and not see the reasoning behind other viewpoints,&rdquo; '
        '&ldquo;sets unreachable standards for himself and others.&rdquo;</em> Self-reported revenue '
        'trajectory at the current company ($20MM May 2025 &rarr; $23MM May 2026, ~15% YoY) is modest for '
        'Series B growth-stage.<br><br>'
        '<strong>The recommendation is therefore not the standard binary diligence question</strong> '
        '(strong-Controller-below or no?). The default lean is <em>no</em>. The amount of onsite evidence '
        'required to override 4 Sev flags + Hi flag + below-cohort Z|Algo + a TTI conflict section that '
        'agrees with the instrument on every line is high. If the hiring company decides to advance, '
        'onsite must produce: (a) the candidate&rsquo;s own framing of the SoLo Funds exit and the '
        'Avolve trajectory, (b) at least one specific pre-proof developmental investment in a current '
        'direct report, (c) a concrete answer to the Form 8 Facilitative Mindset probe that names a '
        'recent case where his team changed his mind on something material. Without all three, default '
        'to no.'
    )
    return {
        'ROLE_FIT_TITLE': role_fit_title,
        'ROLE_FIT_SEAT': role_fit_seat,
        'ROLE_FIT_EASY': role_fit_easy,
        'ROLE_FIT_HARD': role_fit_hard,
    }

def build_career_timeline():
    timeline_title = 'Career Timeline &mdash; Big-Co Finance to Mid-Cap CFO'
    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Education<br>Villanova MBA Finance<br>+ ECNU practicum</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">CDI Corp &mdash; Sr Financial Analyst<br>Aug 2004 - Jul 2007</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">Exelon &mdash; Sr Analyst<br>Aug 2007 - Apr 2008</div>
                <div class="timeline-block" style="background:#1e8449;color:#fff;">Lincoln Financial &mdash; Strategy &amp; Profitability<br>May 2008 - Nov 2009</div>
                <div class="timeline-block" style="background:#0e7490;color:#fff;">Comcast &mdash; Treasury<br>May 2010 - May 2013 &middot; 3y1m</div>
                <div class="timeline-block" style="background:#7c3aed;color:#fff;">In-Recruit &mdash; MD<br>May 2013 - Mar 2015 &middot; 1y11m</div>
                <div class="timeline-block" style="background:#d4a84b;color:#1a2332;">Jornaya &mdash; Director of Finance<br>Mar 2015 - Sep 2018 &middot; 3y7m</div>
                <div class="timeline-block" style="background:#c0392b;color:#fff;">LoanLogics &mdash; Dir Finance &rarr; CFO &amp; Treasurer<br>Dec 2018 - Aug 2022 &middot; 3y9m total<br>Divestiture + bolt-on acquisition</div>
                <div class="timeline-block" style="background:#9333ea;color:#fff;">Consultant (private)<br>Feb 2023 - Jun 2023 &middot; 5mo</div>
                <div class="timeline-block" style="background:#6b7280;color:#fff;">SoLo Funds &mdash; CFO<br>Nov 2023 - May 2025 &middot; 1y7m</div>
                <div class="timeline-block" style="background:#ef4444;color:#fff;">Avolve &mdash; CFO (current)<br>May 2025 - Present &middot; 1y on file<br>Provable Markets per survey self-report</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item"><div class="legend-dot" style="background:#3498db;"></div><span><strong>Education:</strong> Villanova University, MBA Finance. East China Normal University, International Business Practicum (short-term study abroad). No undergraduate institution listed on LinkedIn; no varsity-sport / scholar-athlete signal.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e40af;"></div><span><strong>CDI Corporation &mdash; Sr Financial Analyst (Aug 2004 - Jul 2007):</strong> 3y. Career-start Sr Analyst seat.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#16a085;"></div><span><strong>Exelon &mdash; Sr Analyst (Aug 2007 - Apr 2008):</strong> 9mo. Monthly financial reporting + senior-management presentations using scorecards, bridges, risks/opportunities.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e8449;"></div><span><strong>Lincoln Financial Group &mdash; Strategy &amp; Profitability Management (May 2008 - Nov 2009):</strong> 1y7m. Monthly strategy and profitability reports for executive decision-making.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#0e7490;"></div><span><strong>Comcast &mdash; Treasury (May 2010 - May 2013):</strong> 3y1m. Highest-prestige seat on the file. Daily/monthly/quarterly valuations and reporting.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#7c3aed;"></div><span><strong>In-Recruit &mdash; Managing Director (May 2013 - Mar 2015):</strong> 1y11m. First leadership role. Financial models / projections for investor presentations; investment tracking.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#d4a84b;"></div><span><strong>Jornaya &mdash; Director of Finance (Mar 2015 - Sep 2018):</strong> 3y7m. Built models for Series B funding; 409(a) valuation; FP&amp;A, banking + investor relations.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#c0392b;"></div><span><strong>LoanLogics, Inc. &mdash; Director of Finance (Dec 2018 - Jun 2019, 7mo) &rarr; CFO and Treasurer (Jul 2019 - Aug 2022, 3y2m):</strong> Strongest line on the file. Internal promotion track. <strong>Led divestiture to Sun Capital Partners + bolt-on acquisition of LoanBeam.</strong> Reported to BoD and CEO; oversaw all day-to-day finance functions. Disciplined expense management and aggressive cost cutting.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#9333ea;"></div><span><strong>Consultant (private engagements) (Feb 2023 - Jun 2023):</strong> 5mo. Brought in by LoanLogics investor to review cash management / liquidity for a potential Series A raise. Cyber-security M&amp;A advisory work.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#6b7280;"></div><span><strong>SoLo Funds &mdash; CFO (Nov 2023 - May 2025):</strong> 1y7m. Fintech consumer-lending platform.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#ef4444;"></div><span><strong>Avolve &mdash; CFO (May 2025 - Present, 1y on file):</strong> Current per LinkedIn. Survey self-report (Q109) names Provable Markets as current company &mdash; possible new / concurrent role not yet reflected on LinkedIn. Self-reported revenue: $20MM May 2025 &rarr; $23MM May 2026 (~15% YoY).</span></div>
            </div>
            <div class="timeline-banner">
                Career pattern: 22 years of finance starting at Big-Co Sr Analyst seats (CDI, Exelon, Lincoln Financial, Comcast Treasury), into Director of Finance roles, then a CFO arc starting at LoanLogics. Strongest line is the LoanLogics CFO seat with documented divestiture + bolt-on acquisition. <strong>Three CFO seats over the last seven years</strong> (LoanLogics 3y2m, SoLo Funds 1y7m, Avolve 1y on file), with only LoanLogics crossing three years. No Big-4 Partner credential; no NYSE-listed CFO seat; no awards; no varsity-sport / Two-Sport Athlete signal. MBA Villanova is solid mid-tier credential.
            </div>"""
    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

def build_respondent_dict(respondent_data):
    """TTI: Natural D=48 I=52 S=52 C=71 → Promoting Analyzer (60);
            Adapted D=52 I=52 S=35 C=75 → Analyzing Implementor (56).
       DF Primary: Intellectual 85, Intentional 72, Objective 69, Commanding 64.
       DF Indifferent: Collaborative 21, Altruistic 12, Harmonious 12, Instinctive 7."""
    nat_pos = 60
    nat_label = 'Promoting Analyzer'
    nat_disc = [48, 52, 52, 71]
    nat_intensity = compute_intensity_from_disc(nat_disc)  # 27/200 → ~0.14 (ACROSS — near center)

    adp_pos = 56
    adp_label = 'Analyzing Implementor'
    adp_disc = [52, 52, 35, 75]
    adp_intensity = compute_intensity_from_disc(adp_disc)  # 44/200 → 0.22 (ACROSS — near center)

    shift_note = ('Adapted shift: +4D, 0I, &minus;17S, +4C. Modest adaptation overall &mdash; the '
                  'wiring is naturally close to what the seat needs (high-C, top-left of the wheel). '
                  'The S compression (52&rarr;35) is meaningful; the D and C shifts are minor.')

    respondent = {
        'name': 'Patrick Armstrong',
        'first_name': 'Patrick',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [48, 52, 52, 71],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }
    return respondent

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    n_concern_item = count_class('concern-item')
    n_concern_number = count_class('concern-number')
    n_concern_text = count_class('concern-text')
    if n_concern_item < 2:
        failures.append(f"S1: concern-item count {n_concern_item} < 2")
    if n_concern_number != n_concern_item:
        failures.append(f"S1: concern-number != concern-item")
    if n_concern_text != n_concern_item:
        failures.append(f"S1: concern-text != concern-item")
    concern_nums = re.findall(r'<div class="concern-number">(\d+)</div>', html)
    expected_seq = [str(i) for i in range(1, n_concern_item + 1)]
    if concern_nums != expected_seq:
        failures.append(f"S1: concern sequence {concern_nums} != {expected_seq}")

    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: {cls} expected {expected}, got {got}")

    probe_questions = re.findall(r'<div class="probe-question">(.*?)</div>', html, re.DOTALL)
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched = []
        for idx, q in enumerate(probe_questions, 1):
            qlow = q.lower()
            hit = False
            for (_n, _c, substr) in FORM8_QUESTIONS:
                if substr.lower() in qlow:
                    matched_substrs.add(substr); hit = True; break
            if not hit: unmatched.append(idx)
        if len(matched_substrs) < 10:
            failures.append(f"S9-Form8: {len(matched_substrs)}/10 canonical strings; unmatched probes {unmatched}")

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: {cls} expected min {expected}, got {got}")

    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be N/M, got '{teach_val}'")
    if 'To be populated' in html:
        failures.append("S2: Forbidden placeholder 'To be populated'")

    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:
        failures.append("S4-C1: zLabels2 not nested 2-row")

    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/fail/success")
    else:
        sf_labels = json.loads(m_sf_l.group(1))
        fail_data = json.loads(m_fail.group(1))
        succ_data = json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append("S4-C2: length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with +, got {sf_labels[-1]}")

    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first:
            v = int(first.group(1))
            if v < 30:
                failures.append(f"S4-C3: flagLabels3 first = {v} — axis not reversed")

    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        exc_labels = json.loads(m_exc_labels.group(1))
        is_l1_arr = json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: scorecard rows {len(exc_labels)} < 12")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: L2 rows {l2_rows} < 6")
        for lbl, l1f in zip(exc_labels, is_l1_arr):
            if l1f and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' not UPPERCASE"); break
            if not l1f and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' not indented"); break
    else:
        failures.append("S6: excLabels/isL1 not found")

    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    for cls, mn in [('role-fit-box', 1), ('role-fit-grid', 1),
                     ('role-fit-col easy', 1), ('role-fit-col hard', 1)]:
        if count_class(cls) < mn:
            failures.append(f"S-RF: {cls} expected min {mn}")
    if 'What Will Be Easy' not in html or 'What Will Be Hard' not in html:
        failures.append("S-RF: column labels missing")
    if 'Series B' not in html:
        failures.append("S-RF: Series B reference missing")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    # S10 — Recommendation badge length (added 2026-04-27, Schott build)
    rec_m = re.search(r'<div class="recommendation-badge">(.*?)</div>', html, re.S)
    if rec_m:
        rec_text = re.sub(r'&[a-zA-Z]+;', 'x', rec_m.group(1))
        rec_text = re.sub(r'<[^>]+>', '', rec_text).strip()
        if len(rec_text) > 300:
            failures.append(f"S10: recommendation-badge too long ({len(rec_text)} > 300 chars)")

    # S11b — Role-Fit Hard step-back content (added 2026-04-27, Schott build)
    rfh_m = re.search(r'<div class="role-fit-col hard">(.*?)</div>\s*</div>\s*</div>', html, re.S)
    if rfh_m:
        rfh = rfh_m.group(1).lower()
        has_quadrant = any(q in rfh for q in ['top-left', 'top-right', 'bottom-left', 'bottom-right',
                                              'across', 'center-of-wheel', 'center of wheel'])
        wedges = ['implementor', 'conductor', 'persuader', 'promoter',
                  'relater', 'supporter', 'coordinator', 'analyzer']
        has_wedge = any(w in rfh for w in wedges)
        if not has_quadrant:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wiring-shape reference (quadrant or ACROSS)")
        if not has_wedge:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wedge name")

    # S11b.1 — Wheel-position-required rule (added 2026-04-27, Armstrong build):
    # any narrative that claims a canonical-CFO / top-left / right-wiring / Implementor-quadrant
    # read MUST also cite the actual wheel position number. Reading DISC scores in isolation
    # without checking the wheel has produced shipping-quality errors. The check looks across
    # all wiring-relevant sections (Role-Fit, Wiring-Fit, Skills card, DISC notes) for any
    # canonical-CFO claim and verifies a wheel-position number (1-60) is also present nearby.
    wiring_claim_phrases = [
        'top-left', 'canonical strong-cfo', 'canonical cfo wiring', 'right wiring',
        'wiring is right', 'wiring fits the seat', 'implementor quadrant',
    ]
    html_lower = html.lower()
    has_wiring_claim = any(p in html_lower for p in wiring_claim_phrases)
    has_position_number = bool(re.search(r'\bposition\s+\d{1,2}\b', html_lower))
    if has_wiring_claim and not has_position_number:
        failures.append(
            "S11b.1: wiring narrative makes a canonical-CFO / top-left / right-wiring claim "
            "but does NOT cite a TTI wheel position number. Reading DISC scores in isolation "
            "is insufficient; cite the actual wheel position (Natural / Adapted) before making "
            "wiring-fit claims. See METHODOLOGY 'TTI wheel position is REQUIRED reading'."
        )

    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} ***")
        for f in failures: print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)}")
    print("*** QA GATE PASSED ***")


def main():
    print("Loading respondent data...")
    rd = load_respondent_data()
    print(f"  name={rd['name']}, Z|Algo={rd['z_algo_overall']:+.3f}, RF={rd['rf_num']}")
    print(f"  flags: {list(rd['flags_lit'].keys())}")

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution tokens...")
    dist = build_distribution_tokens(zalgo_rows, flag_rows,
                                     rd['z_algo_overall'], rd['z_human_overall'], rd['rf_num'])

    print("Building respondent dict for motivators_section...")
    respondent = build_respondent_dict(rd)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()
    role_fit = build_role_fit()
    scorecard = build_excstds_scorecard(rd)

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    replacements = {
        'CANDIDATE_NAME': 'Patrick Armstrong',
        'CANDIDATE_CREDS': 'Candidate &mdash; Series B Growth-Stage CFO Seat',
        'CANDIDATE_ROLE': 'Candidate for Series B Growth-Stage CFO &mdash; Same Functional Altitude as Today',
        'REPORT_DATE': 'April 27, 2026',

        'ZALGO_OVERALL': f'{rd["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{rd["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(rd['rf_num']),
        'FLAGS_LIT': f'{len(rd["flags_lit"])}/16',
        'TEACH_ITEMS': '10/10',
        'COHORT_AVG': '+0.24',
        'COHORT_AVG_NUM': '0.2440',

        **axes, **concerns, **wiring, **htl, **probes, **teach,
        **radar, **timeline, **role_fit, **dist, **scorecard,

        'DISC_D_NAT': '48',
        'DISC_I_NAT': '52',
        'DISC_S_NAT': '52',
        'DISC_C_NAT': '71',
        'DISC_D_ADP': '52',
        'DISC_I_ADP': '52',
        'DISC_S_ADP': '35',
        'DISC_C_ADP': '75',
                'DISC_NOTE_TEXT': 'Natural position 60 (Promoting Analyzer, ACROSS) &rarr; Adapted position 56 (Analyzing Implementor, ACROSS). <strong>Both positions marked ACROSS</strong> &mdash; center-of-wheel, no strong anchor in any wedge. Center-of-wheel profiles correlate with lower success rates across all positions. Adaptation is modest; only the S compression (52&rarr;35) is meaningful.',
        'DISC_NOTE_DETAIL': 'TTI Driving Forces &mdash; Primary: Intellectual 85 (Passionate, 2&sigma; above mean), Intentional 72, Objective 69, Commanding 64. Indifferent: Collaborative 21, Altruistic 12, Harmonious 12, Instinctive 7. C=75 carries some controls-and-detail orientation, but with D=52, I=52, S=35 clustered close together the four-axis profile reads as a generalist, not a specialized strong-CFO Implementor anchor. The wheel positions confirm: both Natural and Adapted are ACROSS &mdash; transitional / between-wedges, not anchored in the canonical Implementor or Conductor wedge that strong CFOs typically inhabit.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        'RECOMMENDATION_TEXT': 'NO HIRE &middot; HIGH BAR FOR ANY RECONSIDERATION &mdash; Z|Algo &minus;0.84 below cohort average; 4 Sev flags + 1 Hi flag + Upgrade Team; center-of-wheel wiring with no strong anchor; default to no.',
    }

    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    html = re.sub(
        r'<title>.*?</title>',
        '<title>Patrick Armstrong &mdash; CFO Candidate | HALE GLOBAL SUCCESS DIAGNOSTICS</title>',
        html, count=1
    )
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    qa_gate(html)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")


if __name__ == '__main__':
    main()
