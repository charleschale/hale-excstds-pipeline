"""Build the Excellence Standards Coaching Guide for Alba Quintas Núñez.

Respondent: 20260425.alba.quintasnunez@yale.edu
Role: Yale College student · Head Counselor for Orientation for International Students,
      Co-Director Yale CS Career Fair, MUNTY Alumni Coordinator, Producer/Actress at
      Yale Dramatic Association, Coordinator of US Applications at Programa Aditus,
      Student Tech Coordinator at STC, Member of the European Youth Parliament.
Deliverable: Integrated Coaching Guide (Graphical HTML + PDF) — Variant 1, candidate-facing.

Cohort framing: STUDENT cohort (early-career; learning to lead frame per METHODOLOGY).
Top-decile Z|Algo +1.40. Wiring: Conducting Persuader (Natural #12) → Promoting Persuader
(Adapted #13). High-D + High-I + Low-S + Low-mid-C (Promoter wedge — drive-through-people,
NOT the Implementor / killer-doer pattern). Critical Driving-Force split: Altruistic
EXTREME (81, 3SD above) paired with Intentional EXTREME (6, 3SD below) — mission-framed
investment, not specific-person-framed. Severe flag: HandsOn|Sev.

Run from repo root:
    python _pipeline/scripts/build_alba_coaching.py
Outputs:
    _reports/Quintas-Nunez_Alba_coaching_guide.html
    (PDF written separately by make_pdf_alba.js)
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

RESPONDENT_XLSX = ROOT / '_respondents' / '20260425.alba.quintasnunez@yale.edu' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'coaching_guide_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Quintas-Nunez_Alba_coaching_guide.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title,
                'z_algo': z_algo,
                'z_human': z_human,
                'rf_count': rf_count,
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flags_lit = {}
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        flag_val = ws_flags.cell(2, col).value
        if flag_name and flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[str(q_num)] = answer

    return {
        'l1_data': l1_data,
        'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall,
        'z_human_overall': z_human_overall,
        'rf_num': rf_num,
        'questions_answered': questions_answered,
        'flags_lit': flags_lit,
        'non_scorable': non_scorable,
    }

# ============================================================================
# LOAD HISTOGRAM DATA FOR DISTRIBUTION CHARTS
# ============================================================================

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        rf_val = ws_zalgo.cell(row, 7).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

# ============================================================================
# BUILD DISTRIBUTION CHART TOKENS (orientation: WORSE ON LEFT for all 3 charts)
# ============================================================================

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]

    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo = row['z_algo']
        z_human = row['z_human']
        sf = row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    # Chart 1 collapse — trim leading/trailing zeros, keep marker bins
    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    # Chart 2: keep bins where either cohort has counts OR respondent's bin is here
    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    # Chart 3 — flag bins, REVERSED (high flags LEFT, low flags RIGHT, worse-on-left)
    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1
            continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1
                break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]; upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    n_flag = len(flag_counts)
    rev_flag_bin = n_flag - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# RESPONDENT DICT FOR MOTIVATORS_SECTION
# ============================================================================

def build_respondent_dict(respondent_data):
    # TTI Natural DISC: D=75, I=73, S=22, C=34 (Wheel #12 — Conducting Persuader)
    # TTI Adapted DISC: D=62, I=68, S=28, C=48 (Wheel #13 — Promoting Persuader)
    # Primary wedge: Persuader (D + I twin engines)
    nat_pos = 12
    nat_label = 'Conducting Persuader'
    nat_intensity = (75 + 73) / 200.0  # 0.74

    adp_pos = 13
    adp_label = 'Promoting Persuader'
    adp_intensity = (62 + 68) / 200.0  # 0.65

    shift_note = ('Adapted shift: −13D, −5I, +6S, +14C. Current environment is asking for a slightly'
                  ' less commanding, slightly more steady, slightly more compliant posture — moving'
                  ' from Conducting Persuader (drive-and-influence at very high tempo) toward'
                  ' Promoting Persuader (influence with a touch more pace-match and structure).')

    return {
        'name': 'Alba Quintas Núñez',
        'first_name': 'Alba',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [75, 73, 22, 34],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }

# ============================================================================
# NARRATIVE — SIGNATURE PATTERN
# ============================================================================

SIGNATURE_PATTERN = (
    "<p><strong>Leadership is making people and situations better.</strong> That is the scoreboard."
    " The standards in this guide apply universally — not just in a CEO seat, but in every sphere"
    " where you have influence today: the OIS counselors you lead, the CS Career Fair sub-leads,"
    " the cast of a play, the Aditus mentees, and the people closest to you. The instrument was"
    " calibrated on senior leaders, but the behaviors it surfaces are leadership behaviors at any"
    " altitude. The practice ground varies; the standards do not.</p>"
    "<p>You took this survey as a <strong>pretend CEO</strong> — answering as if you were running a"
    " company with direct reports, even though the actual seat is a decade or two away. That is the"
    " right way for someone in college to engage the instrument: it surfaces your <em>current"
    " instincts</em> about what good leadership looks like, before you have held the role. Your"
    " record at twenty is real (selectors, internal Yale promotion, CPA-Fellowship-winning play,"
    " four-year EYP arc, two national-level FLL finishes, a regional journalism prize). The"
    " instrument adds something different — it shows the version of leadership <em>you currently"
    " imagine yourself becoming</em>.</p>"
    "<p>The picture is striking. <strong>Z|Algo +1.40, top decile</strong> — unusually mature for a"
    " college respondent. Five of nine L1 areas a full standard deviation above the mean: investing"
    " in others, replacing self, deliberate urgency, pushing accountability, and facilitative"
    " mindset. These are <strong>good instincts</strong>. The leader you imagine being is one who"
    " develops her people, builds teams that run without supervision, drives at pace, holds"
    " standards, and leaves room for dialogue. That is most of the operating system of an admired"
    " leader, and you already have the instinct for it.</p>"
    "<p>Where the instincts go <em>against</em> the standard is unambiguous in the data, and worth"
    " knowing now — because untrained instincts harden into habits when the seat actually arrives."
    " The instrument is showing four against-standard instincts that are worth refining"
    " <strong>before</strong> you live them in a real seat:</p>"
    "<ol style='margin:8px 0 12px 18px; padding:0; line-height:1.65;'>"
    "<li><strong>Holding yourself to a higher standard than others.</strong> The flag <em>Lower"
    " Standards for Others than Self</em> is lit at Medium. The admired pattern runs the opposite"
    " way — admired leaders expect <em>more</em> of others than of themselves, because they admire"
    " the people around them. Holding yourself to a higher standard sounds humble; in practice it"
    " means the leader fills the gap by doing the work, and the team never gets the chance to rise."
    " This is the most common against-standard instinct in your cohort, and one of the most"
    " important to refine.</li>"
    "<li><strong>A leader cannot be outvoted.</strong> Your answer to <em>my team can outvote me</em>"
    " was on the &ldquo;cannot&rdquo; side. A leader the team cannot overrule is the boss. A leader"
    " the team can overrule is the conductor. The boss model scales to the size of one person&rsquo;s"
    " judgment; the conductor model scales to the size of the team&rsquo;s collective judgment. Most"
    " early-career imagined-leaders hold the boss model and call it leadership.</li>"
    "<li><strong>The team can be at fault independently of the leader.</strong> Your answer to"
    " <em>there are no bad teams, only bad leaders</em> was on the disagreement side — meaning you"
    " partly hold that a team can be bad in spite of its leader. The admired stance is extreme"
    " ownership: when the team is not delivering, the work to do sits on the leader side first.</li>"
    "<li><strong>Approval from others matters at the threshold.</strong> Your answer to Q131 (the"
    " &ldquo;I require excellence vs. I require respect/approval&rdquo; question) was middle-of-the-"
    " scale, not at the admired end. And the deepest L2 finding in your entire file — by a"
    " meaningful margin — is <em>Cares About Others Not Their Approval</em>. The pattern is wanting"
    " to be liked-and-effective at once, and softening on the &ldquo;effective&rdquo; side when"
    " discomfort would be the more developmental act.</li>"
    "</ol>"
    "<p><strong>One nuance on the fourth instinct that is particular to your cohort.</strong> The"
    " pleasing pattern is not only about wanting personal popularity — that is the easier version"
    " to spot. The harder version, especially common in Yale and similar academic environments, is"
    " <em>pleasing the prevailing ideological or cultural moment</em>: deferring to the positions"
    " the in-group expects, because taking those positions feels like the right thing rather than"
    " like pleasing. Your answer to Q119 (whether leaders should take public stances on issues"
    " unconnected to the core mission of their organization) is one expression of this. Smart young"
    " people in academic environments often answer the way you did; the rationale is unpacked in"
    " Impact item #3 below, and it is worth the careful walk-through.</p>"
    "<p>Combined, the four against-standard instincts produce the loudest single signal in your"
    " file: <strong>HandsOn at the Severe level</strong>. HandsOn|Sev is not a separate finding;"
    " it is the symptom of the four bullets above. When you imagine yourself as a CEO who holds"
    " herself to a higher standard than the team, who cannot be overruled, who would be wronged"
    " by team failure, and who needs others to approve, the predictable result is the leader doing"
    " the work — because the alternative requires the leader&rsquo;s burden the imagined version"
    " of you has not yet imagined picking up.</p>"
    "<p>Two further threads compound the central read. <strong>Conditional Belief at Mixed —"
    " carrying full weight even at the Mixed level.</strong> Your answer to Q125 (&ldquo;I believe"
    " in people in advance of them believing in themselves&rdquo;) was 2 on a 1-TRUE / 5-FALSE"
    " scale — almost ideal, not ideal. Conditional Belief is one of the gateway-belief tests in"
    " the instrument where there is effectively no B-grade: either you believe in people in"
    " advance, or you do not, and the team feels the difference. Mixed-level Conditional Belief"
    " still constrains followership depth and your ability to develop people you do not yet fully"
    " believe in. Treat it like the Severe-level finding it functions as. <strong>Satisfied with"
    " Gripes at Low</strong> is the second compounding thread — a small standing list of things"
    " that bother you about the team that have not yet been converted into decisions. The admired"
    " anchor is <em>easy to please, hard to satisfy, no standing gripes</em>, and the ratio your"
    " answers describe is mildly off in the opposite direction.</p>"
    "<p>The wiring makes the picture coherent. You are a Conducting Persuader — Dominance and"
    " Influence as the twin engines, low Steadiness and low-mid Conscientiousness — drive-through-"
    " warmth, mission-framed, very high tempo. The chart on the right carries the specific scores;"
    " the shape is what matters at this read. The wiring rewards being the highest-energy and"
    " most-influential person in any room. It also makes the leader&rsquo;s burden harder to pick"
    " up, because the engine that runs the show through the leader&rsquo;s own energy is the engine"
    " that resists handing the show over.</p>"
    "<p><strong>The single highest-leverage thing to work on, given that you have not yet held the"
    " seat:</strong> refining each of the four against-standard instincts now, while the cost of"
    " practice is low — peer-team contexts, student-leadership operations, the rooms you already"
    " run — so that when the role arrives (and it will), your starting point is closer to the"
    " admired pattern than the average first-time leader&rsquo;s. The flags and Impact items in"
    " the section ahead are the specific levers.</p>"
    "<p>Your activity record is exceptional, your cohort-relative score is exceptional, and the"
    " framework that produced both will keep producing through the next decade in whatever you"
    " choose. The instrument is showing where the next layer of leadership effectiveness lives —"
    " and the advantage of doing this exercise at twenty is that you can refine the instincts"
    " <em>before</em> you have to live them. <em>Especially the best get better.</em></p>"
)

# ============================================================================
# NARRATIVE — FINGERPRINT
# ============================================================================

FINGERPRINT_NARRATIVE = (
    "<p>Your wiring is <strong>Conducting Persuader</strong> — the Persuader wedge of the DISC"
    " wheel, with Dominance and Influence as the twin engines and Steadiness and Conscientiousness"
    " both well below the midline. You drive results through people and through your own energy,"
    " not through structure or pace-matching. The chart on the right carries the specific scores."
    " The behavioral signature reads: high tempo, multiple simultaneous projects, visible front-of-"
    " room leadership, urgency as a default. The shadow side reads: low patience for routine,"
    " low default attention to policy and structure, limited natural pace-match with people who"
    " move slower. <em>What you build, you build at speed; finishing in the way the structure-and-"
    " routine wirings do is not the natural channel.</em></p>"
    "<p>The Natural-to-Adapted shift slides the wedge one position on the wheel — from Conducting"
    " Persuader to Promoting Persuader. The current environment is asking for a slightly less"
    " commanding posture, a slightly more steady tempo, and noticeably more attention to structure."
    " Read this honestly: the seat you are in right now — concurrent leadership across OIS, the CS"
    " Career Fair, the Dramatic Association production, MUNTY, Aditus, and STC — is asking you to"
    " add operational discipline on top of the persuader engine. The shift is doing some of that"
    " work; the practice section ahead is doing the rest.</p>"
    "<p>The Driving Forces composite corroborates and adds the most coaching-relevant nuance in"
    " your profile. The Primary cluster — <strong>Commanding</strong>, <strong>Objective</strong>,"
    " <strong>Altruistic</strong>, <strong>Receptive</strong> — describes a leader who wants to"
    " control her own destiny, sort functionally rather than aesthetically, fuel the work through a"
    " broader cause, and stay open to new methods. The Indifferent cluster is where the diagnostic"
    " shape lives. <strong>Intentional sits at the bottom of the scale</strong> — three standard"
    " deviations <em>below</em> the mean — and it is the direct opposite of Altruistic. The split"
    " matters: Altruistic says <em>I am moved to help, broadly, for a cause</em>; Intentional says"
    " <em>I am moved to invest specifically in this person, for this person&rsquo;s growth</em>. You"
    " are wired the first way, not the second. Every developmental routine in this guide must be"
    " framed through cause and broader good to land. Specific-individual framings will not stick.</p>"
    "<p>Harmonious, Collaborative, and Intellectual round out the Indifferent cluster — experiential"
    " balance is not a default channel, supporting-role contribution is not a default channel, and"
    " learning for its own sake is not a default channel. None of these is a deficit. They are the"
    " mechanics of leadership that have to be <em>installed</em> as routines, not waited on as"
    " preferences. The <em>What to Work On</em> section names which routines belong on which fuel"
    " line.</p>"
    "<p><strong>The wiring at this altitude — what it produces and what it makes harder.</strong>"
    " The Persuader wedge is recognized for producing exceptional early-career leaders: visible,"
    " energetic, trusted by selectors, able to rally rooms. The activity record is the predictable"
    " expression of that. The wiring is also recognized for making the leader&rsquo;s burden harder"
    " than for many other wirings, because the engine that runs the show through the leader&rsquo;s"
    " own energy is the same engine that resists handing the show over to the team. The flags lit in"
    " your file — HandsOn at Severe, HoldsAvgDownChain, Lower Standards for Others than Self at"
    " Medium, Mixed Conditional Belief, and Satisfied with Gripes at Low — are the classic places"
    " where the Persuader engine&rsquo;s strengths tip into the leader&rsquo;s-burden gaps. Naming the pattern"
    " is the operating manual for the wiring; the routines in the next section are counter-routines"
    " to it.</p>"
)

DRIVING_FORCES_PRIMARY_HTML = (
    "<div style='font-size:12px; color:#333; line-height:1.6;'>"
    "<strong>Commanding</strong> (rank 1, 88 · Passionate) · <strong>Objective</strong> (rank 2, 82 · Passionate)"
    " · <strong>Altruistic</strong> (rank 3, 81 · <em>Extreme</em>) · <strong>Receptive</strong> (rank 4, 65 · Primary)"
    "<p style='margin-top:6px; font-size:11px; color:#5a6773;'>Drive-from-the-front, decide-on-the-functional, fuelled by"
    " a broader cause, with openness to new methods. Three of the four Primaries pull in the same"
    " direction — make things happen, with influence and impact, in service of something larger than"
    " yourself. Receptive is the one that gives you the door into the practice work below: dialogue"
    " over direction, ownership reflection, empowering the team's methodology.</p></div>"
)

DRIVING_FORCES_INDIFFERENT_HTML = (
    "<div style='font-size:12px; color:#333; line-height:1.6;'>"
    "<strong>Intentional</strong> (rank 12, 6 · <em>Extreme</em>) · <strong>Harmonious</strong> (rank 11, 7 · Indifferent)"
    " · <strong>Collaborative</strong> (rank 10, 7 · Indifferent) · <strong>Intellectual</strong> (rank 9, 17 · Indifferent)"
    "<p style='margin-top:6px; font-size:11px; color:#5a6773;'>The Intentional reading at"
    " <em>Extreme</em> is the diagnostic feature — it is the direct opposite of your Altruistic"
    " driver. You are moved by mission and broader good; you are not by-default moved to invest"
    " specifically in this person for this person's growth. Harmonious 7 says you are not moved by"
    " experiential balance; Collaborative 7 says supporting-role contribution is not a default"
    " channel; Intellectual 17 says knowledge for its own sake is not motivating. Each becomes"
    " available, but as a routine installed against the grain.</p></div>"
)

DRIVING_FORCES_IMPLICATIONS_HTML = (
    "<p><strong>What this composite predicts.</strong> A leader who will drive a hard plan with"
    " energy and persuasion, mission-frame the work, decide on functional reasoning, and remain"
    " open to new methods. The Intentional=6 / Harmonious=7 / Collaborative=7 / Intellectual=17"
    " cluster predicts exactly the coaching frontier the ExcStds file reads: <em>specific-person"
    " developmental investment, supporting-role behaviors, and the patient mechanics of a team"
    " rising up rather than the leader pulling them forward</em> are not self-generating. This is"
    " not a call to change the wiring — the wiring is what produces the record. It is a call to"
    " install specific routines against the grain, with one specific framing rule that matters at"
    " your particular DF shape: <strong>frame developmental investment through mission and broader"
    " good, not through specific individuals</strong>. Altruistic Extreme will carry the routine."
    " Intentional Extreme below will not. A statement like <em>this team needs to be capable of"
    " running OIS without me so the next 100 international students get the experience they need</em>"
    " will land. A statement like <em>I want to invest specifically in Person X's growth</em> will"
    " not, and the routine will fade. The same practice, framed twice — only one frame sticks.</p>"
)

# ============================================================================
# WIRING-FIT — colored bold headers per SKILL.md Rule 1b
# Blue #2563eb = motivator-aligned strong
# Green #22c55e = installed against the grain
# Tan #b8862e = anti-aligned weak (practice frontier)
# Red #c0392b = motivator-aligned weak
# ============================================================================

WIRING_FIT_ITEMS = (
    "<p><strong style=\"color:#b8862e;\">Not Pleasing — the deepest landing not yet being made.</strong>"
    " The Persuader wiring defaults to wanting to be liked-and-effective at once. The file shows"
    " where the &ldquo;effective&rdquo; side gets softened to keep the &ldquo;liked&rdquo; side. The"
    " sharpest single L2 finding in your entire file sits here — <em>caring about others more than"
    " their approval of you</em> — and it is at the bottom of the L2 list by a meaningful margin."
    " Both Discomfort For Team and Discomfort For Self also sit below the line. The frontier is"
    " specific: are you willing to make the people around you uncomfortable in service of their own"
    " growth, when comfort is not what they need? The Impact items in this section that make this"
    " concrete are Q63 (peak performance lives in discomfort) and Q3 (stay hard to satisfy with the"
    " team).</p>"
    "<p><strong style=\"color:#b8862e;\">Conducting &amp; Outvoted — the in-the-room conductor"
    " frontier.</strong> Two L2 sub-dimensions carry this L1: <em>conductor over lead guitarist</em>"
    " and <em>empowering team authority</em>. The pattern matches the wiring exactly — the Persuader"
    " engine pulls toward playing the solo rather than drawing the team forward. The strongest single"
    " per-answer Impact item in your top-ten sits inside this L1: <em>can the team outvote me?</em>"
    " &mdash; you answered &ldquo;cannot.&rdquo; That single answer is the lead-guitarist tell, and"
    " it is also the practice ground that unlocks most of the rest of the file.</p>"
    "<p><strong style=\"color:#b8862e;\">Risking by Pushing Extreme Accountability — strong on the"
    " sheriff lane, gap on the in-the-room enforcement.</strong> The L1 itself is well above the"
    " mean — Dominance plus the Commanding driver plus Competitive at Behavioral Hierarchy give you"
    " the engine for accountability work, and proactive ownership reads strongly. The within-cluster"
    " watch-item is the sub-dimension of <em>driving accountability</em> — making consequences land"
    " for others when standards are missed — which sits below the rest of the cluster. The flag"
    " corroboration is the Medium-level <em>Lower Standards for Others than Self</em>: a higher"
    " standard applied to your own work than to the work that comes back from the team, with the"
    " gap closed by you doing the work yourself rather than holding the team accountable to it.</p>"
    "<p><strong style=\"color:#22c55e;\">Risking by Replacing Self — answered like a delegator,"
    " HandsOn at Severe says otherwise.</strong> The L1 sits at the top of your file, and Ability To"
    " Disappear scores positively. This is the structural belief in delegation, and it is real."
    " The seam is the gap between the answer pattern and the day-to-day pattern — the HandsOn|Sev"
    " flag says the actual depth-of-disappearance is not yet matching the structural answer. The"
    " practice ground is closing that gap, which is the focus of Impact items #4 (resist running"
    " point) and #8 (build a team that can outvote you).</p>"
    "<p><strong style=\"color:#22c55e;\">Risking by Investing in Others — installed against the"
    " grain, with one important seam.</strong> Your Intentional Driving Force is at the bottom of"
    " the scale, three standard deviations below the mean. The wiring is not by default oriented"
    " toward specific-individual developmental investment. And yet the L1 sits in the top decile —"
    " <em>demonstrating genuine fanness</em>, <em>developmental mindset</em>, and <em>handling daily"
    " difficulties with dignity</em> all score positively. The seam shows up in <em>reciprocal"
    " followership</em> at the bottom of this cluster, and the Mixed Conditional Belief flag: belief"
    " that lands richly with some people and unevenly with others. The framing rule that matters —"
    " mission-frame the developmental investment, do not specific-individual-frame it.</p>"
    "<p><strong style=\"color:#2563eb;\">Deliberate Urgency — wiring perfectly aligned.</strong>"
    " Urgency Behavioral Hierarchy at the top of the scale, the Commanding driver at the top of the"
    " Driving Forces cluster, Dominance high — the engine is wired for tempo and proactivity. The"
    " one within-cluster watch-item is <em>proactive belief in people</em>, slightly off the rest of"
    " the cluster — the proactivity is about <em>you</em> being ahead of the situation more than"
    " <em>the team</em> being ahead of it. That is the seam that connects to HoldsAvgDownChain.</p>"
    "<p><strong style=\"color:#b8862e;\">Risking by Facilitative Mindset — restraint structure built,"
    " in-the-room dialogue still the gap.</strong> Per the methodology, high-Influence does not"
    " correlate with facilitation; high-I leaders typically fill the room with influence rather than"
    " drawing the room out. Your file shows a more interesting picture — <em>power and status"
    " management</em> and <em>sublimating ego</em> both score strongly. Those are the structural"
    " restraint dimensions and you have them. The frontier is in-the-room: <em>dialogue versus"
    " direction</em> sits below the line. You direct more than you dialogue, and the room knows. The"
    " good news is the foundation behaviors are already there, so the dialogue practice has clean"
    " soil to land in.</p>"
)

# ============================================================================
# EXCSTDS SCORECARD
# ============================================================================

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Reciprocal Followership',
            'Handling Daily Difficulties With Dignity',
            'Demonstrating Genuine Fanness',
            'Developmental Mindset',
            'Developmental Discipline',
        ]),
        (2, 'RISKING BY FACILITATIVE MINDSET', [
            'Dialogue Vs. Direction',
            'Power & Status Management',
            'Sublimating Ego',
        ]),
        (3, 'CONDUCTING & OUTVOTED', [
            'Conductor > Lead Guitarist',
            'Empower Team Authority',
        ]),
        (4, 'RISKING BY REPLACING SELF', [
            'Ability To Disappear',
            'Urgency Down Chain Of Command',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Basic Machinery Of Accountability',
            'Jrs Extreme Proactivity',
            'Drives Accountability',
            'Stds = What Tolerate',
        ]),
        (7, 'NOT PLEASING', [
            'Cares About Others Not Their Approval',
            'Discomfort For Self',
            'Discomfort For Team',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Simplification Methods',
            'Clarity Of Accountability',
            'Respects Collective Wisdom',
            'Facts Over Feelings',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Fire In Belly',
            'Extreme Proactivity',
            'Proactive Belief In People',
            'Action Over Inaction',
        ]),
    ]

    labels, scores, is_l1, missing_l2 = [], [], [], []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = respondent_data['z_algo_overall']
    cohort_avg = -0.115  # Q114.1 ICs cohort static norm — Alba is student/IC

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'ZALGO_OVERALL_NUM': f"{overall_z:.4f}",
        'COHORT_AVG_NUM': f"{cohort_avg:.3f}",
    }

# ============================================================================
# IMPACT ITEMS — 4 flag-driven + 9 algorithm-ranked per-answer
# ============================================================================

def build_impact_items_html():
    flag_subheader = (
        '<div class="practice-subsection-hdr">'
        '<div class="practice-subsection-hdr-title">Flag-Driven Items</div>'
        '<div class="practice-subsection-hdr-blurb">Lit flags from the cohort research &mdash;'
        ' pattern-level effectiveness levers that sit across multiple L1s, which is why they are'
        ' flags and not single-L2 routines. Addressed before per-answer lifts.</div>'
        '</div>'
    )

    flag_cards = []

    # FLAG 1 — HandsOn|Sev (master finding for Alba)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">HandsOn | Sev</div>'
        '<div class="practice-qref">Master finding &middot; the room when you are not in it &middot; severity Severe</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>The thesis.</strong> The HandsOn flag watches whether the leader is the routing'
        ' node for the work, or whether the team is. It does not measure how busy you are; it'
        ' measures whether the operation can run without you in it for a sustained stretch. Admired'
        ' leaders disappear by design — they install routines that let the team carry the plan when'
        ' the leader is out, and they treat <em>can the team run without me for a week</em> as the'
        ' diagnostic test for whether the next altitude of leadership is being practiced. Leaders'
        ' who fire HandsOn at <em>Sev</em> are running point themselves on too much of the work that'
        ' the team should be carrying. The rooms reward it in the short term — work gets done — and'
        ' it caps the scale of the operation in the long term, because the team learns to wait for'
        ' the leader rather than rising.</p>'
        '<p><strong>Why this fires at Severe in your file.</strong> The wiring sets the condition'
        ' (Persuader engine, Urgency at the top of Behavioral Hierarchy, Commanding at the top of'
        ' Driving Forces — you are the highest-energy node in every room you enter). Five concurrent'
        ' leadership roles at Yale plus a four-year EYP arc plus the CS Career Fair plus the Yale'
        ' Dramatic Association production set the load. The combination is predictable: when there'
        ' is something to be done, you do it. The structural answers in your file say you delegate'
        ' (Replacing Self at the top of the file). The HandsOn flag at Sev says the actual day-to-day'
        ' is not yet matching the structural answer.</p>'
        '<p><strong>Why this connects to the Conducting frontier.</strong> If the day-to-day pattern'
        ' is the leader running point on the work, the team does not get to develop the muscle of'
        ' running point themselves. Conductor &gt; Lead Guitarist sits barely above zero in your'
        ' file; Empower Team Authority sits below the line. These are not separate findings from the'
        ' HandsOn flag — they are the in-the-room mechanism that produces the flag.</p>'
        '</div>'
        '<div class="practice-fuel">Gateway routine: pick one workstream — OIS counselor coordination'
        ' is the obvious candidate — and design <em>one full week</em> where you are formally"'
        ' off-deck. Tell the team in advance. Name the decisions they will need to make. Resist the'
        ' impulse to check in. After the week, debrief: what did they do well, what broke, what they'
        ' learned. Repeat in another workstream the following month. The flag goes quiet when the'
        ' team can carry a sustained stretch without you.</div>'
        '</div>'
    )

    # FLAG 2 — HoldsAvgDownChain
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">HoldsAvgDownChain</div>'
        '<div class="practice-qref">Crew, not basketball &middot; depth-of-organization pattern</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p>The flag reads whether the standard you hold at your direct level is the same standard'
        ' enforced one and two levels below. In your current seats, that means: are the OIS'
        ' counselors holding the same standard for the international students they support that you'
        ' would hold? Are the CS Career Fair sub-leads running their pieces the way you would run'
        ' them? When the team is delivering at A-minus while you are delivering at A, the gap is'
        ' usually not effort — it is that you have not yet converted the standard into a routine the'
        ' team can hold without your presence.</p>'
        '<p><strong>The frame is crew, not basketball.</strong> Basketball is about trading for'
        ' better talent at the top of the roster. Business is more like crew: the boat moves at the'
        ' pace of the slowest rower. Admired leaders continuously upgrade the routines and the seats'
        ' at depth, so the layer below the leader is rising rather than holding steady. Unsuccessful'
        ' leaders tolerate average at depth and wait for culture to transform on its own.</p>'
        '<p><strong>Why this compounds with the master finding.</strong> If the layer below your"'
        ' immediate team is holding A-minus, you cannot fully step out — because the work would land'
        ' at A-minus rather than at A, and you would not be willing to release. The HandsOn flag and'
        ' the HoldsAvgDownChain flag are the same finding read through two different lenses: <em>the'
        ' standard is not yet held by anyone else in the room with the same fidelity you hold it.</em></p>'
        '</div>'
        '<div class="practice-fuel">Routine: pick one role two layers below you (a counselor in OIS,'
        ' a sub-team lead in the Career Fair, an Aditus mentee leading their cohort group). For each'
        ' one, ask: <em>if I were selecting today, would I select this person again?</em> If the'
        ' answer is no, name the specific gap, name the development you owe them, and set a date by'
        ' which the seat clears the bar or transitions. Repeat at the next layer. The flag goes quiet'
        ' when the leader stops finding seats they would not select for today.</div>'
        '</div>'
    )

    # FLAG 3 — Mixed Condit Belief — treat with FULL weight; B-grade on gateway-belief = F-grade
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Conditional Belief in Others &middot; Mixed</div>'
        '<div class="practice-qref">Gateway-belief flag &middot; treat at full weight, not soft-pedal</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>The instrument names this flag at three severity levels — Mixed, Severe, and"'
        ' the un-lit baseline. The Mixed level is not a soft version of the Severe level.</strong>'
        ' It is a different reading of the same fundamental gap, and it carries effectively the same'
        ' coaching weight. Conditional Belief is one of the gateway-belief tests in the entire'
        ' instrument: there is effectively no B-grade. Either you believe in people in advance of'
        ' them believing in themselves, or you do not — and the team feels the difference."'
        ' Mixed-level Conditional Belief is not &ldquo;mostly believing in advance.&rdquo; It is"'
        ' belief that lands richly with the team members who have proved themselves in your frame,'
        ' and unevenly with the ones who have not yet proved themselves to you. The team members on'
        ' the second side feel the gap.</p>'
        '<p><strong>Where it lands in your file — the specific question.</strong> Your answer to'
        ' Q125 (<em>I believe in people in advance of them believing in themselves</em>, scale"'
        ' 1-TRUE / 5-FALSE) was a 2. The admired answer is 1. A 2 reads as &ldquo;mostly, but not'
        ' always&rdquo; — and on this question the &ldquo;not always&rdquo; is the part that"'
        ' produces the flag. The supporting evidence sits in the Reciprocal Followership L2 at the'
        ' bottom of your Investing in Others cluster: people follow leaders who believe in them"'
        ' unconditionally, and your wiring does not yet extend belief that way by default.</p>'
        '<p><strong>What the flag constrains.</strong> Two things directly. First, followership depth"'
        ' — the team members who have not yet earned full belief from you sense it, and they ration'
        ' their effort accordingly. Followership compounds when belief is extended in advance; it"'
        ' rationalizes when belief is extended only after proof. Second, the leader&rsquo;s ability to'
        ' develop replacements — you do not fully build up someone you do not fully believe in, and"'
        ' under-developed replacements perpetuate the HandsOn|Sev pattern (because you, the leader,"'
        ' end up doing the work the under-developed replacement was supposed to grow into).</p>'
        '<p><strong>The framing rule that matters at your particular Driving Forces shape.</strong>'
        ' Your DF Altruistic at <em>Extreme</em> paired with Intentional at <em>Extreme</em> below"'
        ' (see the Driving Forces panel above) is the most diagnostic single feature of your'
        ' motivational profile. Coaching routines that rely on <em>specific-individual</em>'
        ' investment will fade — the channel that fuels them is not on by default. The same"'
        ' routines, framed through <em>cause and broader good</em>, will land. <em>Believe in this'
        ' person early because the OIS program needs them to grow into the role we need to fill'
        ' next year</em> works. <em>Believe in this person early because I want to invest in their'
        ' growth</em> will not stick.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: for each direct report (in OIS, CS Career Fair, or'
        ' Aditus), write one sentence naming a capability you would like them to grow into that they'
        ' have <em>not</em> yet demonstrated. Tell them. Frame the development through the mission —'
        ' what the program needs them to be able to do, who they will support by becoming that. Budget'
        ' a specific investment of your time and exposure against it this term. Belief becomes'
        ' observable when it is named, mission-framed, and resourced before the proof.</div>'
        '</div>'
    )

    # FLAG 4 — Lower Standards for Others than Self | Med (the leader's-burden flag)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Lower Standards for Others than Self | Med</div>'
        '<div class="practice-qref">The leader&rsquo;s-burden flag &middot; expecting MORE of others, not less &middot; severity Medium</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> The Lower-Standards flag watches the gap between'
        ' the standard a leader holds for her own work and the standard she holds for the team&rsquo;s"'
        ' work. Most people assume the admired pattern is <em>I hold myself to a higher standard than'
        ' I hold others</em>, because that sounds humble and effort-forward. The instrument is sharp'
        ' on the opposite finding: <strong>admired leaders hold others to a higher standard than'
        ' themselves, because they admire the people around them.</strong> Holding others to a lower'
        ' standard than yourself is not humility; it is the leader filling the gap by doing the work'
        ' herself. It is also the most reliable single signal of an extraordinary individual'
        ' contributor with leadership-shaped activity, who has not yet stepped into the leader&rsquo;s'
        ' burden.</p>'
        '<p><strong>Where your file lands.</strong> Medium severity. The flag is corroborated by"'
        ' three other signals in your file: HandsOn at Severe (the work routes to you), the'
        ' below-line <em>drives accountability</em> sub-dimension (you do not hold others to'
        ' consequences when standards are missed), and the Cares-About-Others-Not-Their-Approval"'
        ' L2 at the bottom of your file (you would rather do it yourself than ask the team to rise"'
        ' to a standard that would discomfort them). These are not separate findings; they are the'
        ' same pattern read four ways. The pattern is leader&rsquo;s-burden-not-yet-picked-up.</p>'
        '<p><strong>Why this flag is the most diagnostic single finding in your file.</strong> The'
        ' admired-leader frame is counter-intuitive and costs something every time you live it. To"'
        ' hold others to a higher standard than yourself, you have to (a) say so out loud, (b) let'
        ' them feel the weight of the standard, (c) refuse to fill the gap by doing the work, and"'
        ' (d) hold the discomfort while the team rises to it or transitions out. Each step is the"'
        ' opposite of the activity-rich, energy-forward, please-everyone-by-doing-everything pattern'
        ' the wiring rewards. This is the single highest-leverage frontier in the file.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: pick one specific standard you hold for your own work in'
        ' an OIS counselor coordination, or a Career Fair sub-team, or an Aditus mentee. Tell the'
        ' team you hold them to that standard and that you will <em>not</em> fill the gap by doing'
        ' the work yourself if it comes back below standard. Mean it. The first time the team'
        ' delivers below the standard and you do not fill the gap — and the work goes back for'
        ' rework or the role transitions — the leader&rsquo;s-burden practice has started.</div>'
        '</div>'
    )

    # FLAG 5 — Satisfied with Gripes | Low (anchor: easy to please, hard to satisfy)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Satisfied with Gripes | Low</div>'
        '<div class="practice-qref">Ratio flag &middot; easy-to-please / hard-to-satisfy anchor &middot; severity Low</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> Satisfied with Gripes is a <em>ratio</em> flag.'
        ' It looks at two self-reports side by side: how satisfied the leader is with the team, and'
        ' how many gripes the leader is holding about the team. The flag trips when the leader reports'
        ' <strong>both</strong> &mdash; some level of satisfaction and a working list of gripes at the'
        ' same time. That combination is the tell.</p>'
        '<p><strong>The admired-leader anchor: easy to please, hard to satisfy.</strong> Admired'
        ' leaders are <em>easy to please</em> &mdash; quick to notice when a team member does good'
        ' work, quick to say so. They are <em>hard to satisfy</em> &mdash; never fully satisfied,'
        ' because the team should always be developing further than today; satisfaction is the word'
        ' that closes the development loop. And they run <em>no standing gripes</em> &mdash; because'
        ' a gripe is a standard the leader has not yet converted into a decision, and admired leaders'
        ' move gripes to decisions the week they surface.</p>'
        '<p><strong>Where your file lands.</strong> Low severity &mdash; the mildest level &mdash; so'
        ' this is not a primary frontier. It is, however, sharing soil with HandsOn|Sev and'
        ' HoldsAvgDownChain: the same root pattern shows up here as standing gripes about pace,'
        ' fidelity, and follow-through that you have not yet converted into decisions or development'
        ' conversations. At your altitude this often shows up as the small everyday frustration with'
        ' a counselor who needed to be reminded twice, the sub-lead who delivered late, the cast'
        ' member who did not run the line cleanly. Each one is a gripe. Each one is a standard not'
        ' yet converted into a routine.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: two rules. (1) <em>No standing gripes.</em> The week a'
        ' gripe surfaces, name the standard the gripe is actually about and move it to a decision —'
        ' a conversation, a role clarification, or a transition plan. (2) <em>Stay hard to satisfy.</em>'
        ' After a strong stretch, ask what the team could be doing a term from now that it cannot do'
        ' today, and start the development routine that gets there. Avoid the word <em>satisfied</em>'
        ' as a description of where the team is — it closes the development loop.</div>'
        '</div>'
    )

    # Per-answer subsection
    peranswer_subheader = (
        '<div style="height:22px;"></div>'
        '<div class="practice-subtitle">Nine standards where deliberate practice will produce the largest lift.</div>'
    )

    impact_cards = []

    # IMPACT 1 — Q123 (decision timing — don't decide too early)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Wait for the right moment to decide</div>'
        '<div class="practice-qref">Q123 &middot; Organizational Decision Making (Impact Rank 1)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#5a6773; border-color:#5a6773;" data-l2="8.7">8.7 Facts Over Feelings</div></div>'
        '<div class="practice-body">'
        '<p>You answered &lsquo;4&rsquo; on a 1-TRUE / 5-FALSE scale, leaning FALSE — meaning you'
        ' tend to decide before you have to, rather than holding optionality open until the timing'
        ' is right. The research is specific: experienced leaders know decision <em>timing</em> can'
        ' matter as much as the choice itself. Deciding too early limits options, narrows what you'
        ' know, and forecloses possibilities that might have surfaced if you had let the situation'
        ' develop another week. The reason leaders make poor decisions most often is something they'
        ' did not yet know.</p>'
        '<p>The wiring read corroborates: D=75, Urgency BH 86, Commanding 88. You are wired to'
        ' decide and move, and that is most of the time the right wiring for the situations you are'
        ' in. The practice is a discipline on top of the wiring: <em>which class of decisions actually'
        ' benefit from waiting?</em> Major hires, role transitions, multi-month commitments — these'
        ' are the irreversible ones where the optionality is the point.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: for any decision you would describe as irreversible — a'
        ' role change, a major leadership commitment, a public position the organization will own —'
        ' write down on one line <em>what would I learn between now and the deadline if I held this'
        ' open?</em> If the answer is non-trivial, the decision waits.</div>'
        '</div>'
    )

    # IMPACT 2 — Q63 (peak performance / discomfort) — with the excellence-produces-respect reframe
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Peak performance lives in discomfort &mdash; requiring excellence is how respect is produced</div>'
        '<div class="practice-qref">Q63 &middot; Not Pleasing (Impact Rank 2)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="7.4">7.4 Discomfort For Team</div></div>'
        '<div class="practice-body">'
        '<p>The question asks where peak performance happens; the scale runs 1-Comfortable to'
        ' 5-Uncomfortable. You answered &lsquo;1&rsquo; — peak performance happens when teams are'
        ' comfortable. The admired answer is &lsquo;5&rsquo; — peak performance requires productive'
        ' discomfort. Nothing great happens until people and teams are pushed out of their comfort'
        ' zones.</p>'
        '<p><strong>The reframe that matters most at your altitude.</strong> A common pattern at"'
        ' twenty — and especially in respectful, internationally-mobile, civically-engaged students'
        ' — is to read &ldquo;requiring excellence&rdquo; and &ldquo;showing respect&rdquo; as'
        ' opposite ends of the same axis. The implicit logic: if I push them, I am not respecting'
        ' them; if I respect them, I should ease the demand. <strong>The admired-leader frame"'
        ' inverts that logic exactly.</strong> Requiring excellence <em>is</em> the highest form of"'
        ' respect. To require excellence of someone, you have to (a) believe they are capable of it,'
        ' (b) care about their growth more than their comfort, and (c) be willing to hold the'
        ' discomfort while they rise to the standard. Easing the demand looks like respect on the'
        ' surface; on a longer time-frame the team experiences it as the leader not believing they'
        ' can rise. Easing the demand is what produces the very approval-anxiety it is trying to"'
        ' soothe.</p>'
        '<p><strong>What this looks like in your file.</strong> The L2 <em>Cares About Others Not"'
        ' Their Approval</em> sits at the bottom of your entire L2 list — by a meaningful margin —'
        ' and Discomfort For Team also sits below the line. The wiring root is the Persuader engine'
        ' and the Altruistic-Extreme driver: your default move is to <em>welcome</em> people in, not"'
        ' to make them productively uncomfortable. The reframe that lets the practice land —"'
        ' <em>discomfort for team is the highest form of caring</em>, and <em>requiring excellence"'
        ' is the highest form of respect</em> — is the bridge from how the wiring wants to operate'
        ' to how an admired leader actually operates.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in your next OIS counselor coordination meeting, name'
        ' one stretch — for one specific counselor — that you have been holding back on because it'
        ' would feel uncomfortable to ask. Ask. The discomfort is the practice; the underlying'
        ' message is <em>I believe you can rise to this</em>, which is the most respectful thing"'
        ' a leader can communicate.</div>'
        '</div>'
    )

    # IMPACT 3 — Q119 — the woke-test walk-through (Yale-cohort failure mode)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Mission discipline &mdash; the Yale-cohort walk-through</div>'
        '<div class="practice-qref">Q119 &middot; Organizational Decision Making (Impact Rank 3)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#5a6773; border-color:#5a6773;" data-l2="8.1">8.1 Simplification Methods</div></div>'
        '<div class="practice-body">'
        '<p>The standard, in full: <em>When leaders take a stance and attempt to do the &ldquo;right&rdquo;'
        ' thing about issues unconnected to the core mission of the organization, everyone loses —'
        ' especially the leader.</em> The scale runs 1-TRUE / 5-FALSE; the admired answer is 1 (TRUE)."'
        ' You answered 4. <strong>This is a common answer pattern in your cohort, and it is worth"'
        ' a careful walk-through</strong> — because the rationale for the admired answer is not'
        ' obvious from twenty, and it runs counter to the prevailing instinct in the academic'
        ' environments most Yale undergraduates have lived in.</p>'
        '<p><strong>The instinct the answer comes from.</strong> Smart, civically-engaged students at'
        ' Yale (and similar campuses) are taught — explicitly and implicitly — that platforms come"'
        ' with responsibility, that silence on important social and political issues is itself a"'
        ' position, and that leaders should use their visibility to advance the causes they believe'
        ' in. There is integrity in that instinct. The question is not whether the issues themselves'
        ' are important. The question is what happens to the <em>organization</em> when the leader'
        ' uses the leadership platform to advance positions outside the organization&rsquo;s mission.</p>'
        '<p><strong>The three predictable consequences when leaders speak outside the lane.</strong></p>'
        '<ol style="margin:6px 0 8px 18px; padding:0;">'
        '<li><strong>The leader becomes a lightning rod.</strong> Whichever side of the issue the'
        ' leader takes, the other side reads the position as official organizational endorsement and"'
        ' feels disenfranchised. The side the leader supports demands a stronger statement. Both'
        ' groups end up dissatisfied; the leader spends energy managing the controversy rather than'
        ' the work.</li>'
        '<li><strong>The team fractures internally.</strong> Teammates with different views on the'
        ' issue — and there will be many — feel they cannot speak honestly without consequence. The"'
        ' collaborative spirit erodes. People who hold competing positions disengage rather than'
        ' debate. The team&rsquo;s ability to do the work the mission requires gets quietly damaged.</li>'
        '<li><strong>The mission gets diluted.</strong> Energy and attention that should go to the"'
        ' organization&rsquo;s actual work get reallocated to managing public-statement"'
        ' controversies. Customers, members, beneficiaries — the people the organization exists to'
        ' serve — get less of the leader&rsquo;s attention.</li>'
        '</ol>'
        '<p><strong>What the discipline is and is not.</strong> The discipline is not silencing the'
        ' leader&rsquo;s personal views. You may hold strong positions on social, political, and'
        ' ideological questions, and you should hold them. The discipline is choosing not to use the'
        ' <em>leadership platform</em> for views unconnected to the role&rsquo;s mission. The Yale"'
        ' Dramatic Association exists to make great theater. Programa Aditus exists to help'
        ' high-potential, low-resource Spanish students reach top US universities. OIS exists to'
        ' support international Yale students. The Yale CS Career Fair exists to connect CS students'
        ' to industry. Each mission is the lane. Personal positions can be voiced from your own name'
        ' and your own platform — not from the org&rsquo;s.</p>'
        '<p><strong>Why this answer pattern reads as part of the broader Pleasing thread.</strong>'
        ' The pleasing instinct is not only about wanting personal popularity. It also includes'
        ' deferring to the positions the in-group expects, because taking those positions feels'
        ' like the right thing rather than like pleasing. Your file&rsquo;s <em>Cares About Others'
        ' Not Their Approval</em> at the bottom of the L2 list, combined with this answer, suggests'
        ' the ideological-conformity version of pleasing is at least partly in play. The discipline"'
        ' Q119 is asking for is mission-loyalty over in-group-loyalty.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: write a one-sentence mission-lane test for each'
        ' leadership role you hold. Tape it to your laptop. If a proposed public statement under that"'
        ' role&rsquo;s name does not serve that sentence, it does not go out. Hold your personal'
        ' positions on your own time and your own platform.</div>'
        '</div>'
    )

    # IMPACT 4 — Q31 (resist getting involved in operations)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Resist running point until the moment requires it</div>'
        '<div class="practice-qref">Q31 &middot; Risking by Pushing Extreme Accountability (Impact Rank 4)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="5.3">5.3 Drives Accountability</div></div>'
        '<div class="practice-body">'
        '<p>The standard: <em>resist being involved in the actual operation or process of getting'
        ' things done until you absolutely have to be — then go all-in</em>. You answered &lsquo;5&rsquo;'
        ' (FALSE) on a 1-TRUE / 5-FALSE scale, meaning your default posture is involvement, not'
        ' restraint. The admired pattern is the paired discipline: stay <em>out</em> by default, but'
        ' when the moment requires it, step <em>in</em> all the way. Most early-career leaders get'
        ' it wrong in both directions — either always involved (the trap your wiring will pull you'
        ' toward) or always absent (the over-correction).</p>'
        '<p>This is the per-answer expression of the HandsOn|Sev flag. The wiring is calibrated for'
        ' involvement. The practice is calibrating an alternative posture: <em>I am not the routing'
        ' node here unless the moment requires me to be</em>.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: when you go all-in on something the team should be'
        ' running, tell them explicitly — &ldquo;I am stepping in on this one, here is why, here is'
        ' when I will step back out.&rdquo; That way the hand-on is a tool, not a reflex, and the team'
        ' learns when to expect you in versus out.</div>'
        '</div>'
    )

    # IMPACT 5 — Q95 (small changes, not big)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Develop people through small, specific changes</div>'
        '<div class="practice-qref">Q95 &middot; Risking by Investing in Others (Impact Rank 5)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="1.6">1.6 Developmental Discipline</div></div>'
        '<div class="practice-body">'
        '<p>The question asks how you generally help people improve — small changes or big ones."'
        ' You answered &lsquo;4&rsquo; (toward big changes). The admired pattern is the opposite:'
        ' <em>small, specific, observable</em> changes that compound. Big-change framings create"'
        ' overwhelm and rarely stick; small-specific framings let the person practice the new habit'
        ' until it becomes natural, then add the next small change.</p>'
        '<p>This pairs with Developmental Discipline at the negative end of your file. The reading"'
        ' fits the wiring: Persuader engine + Commanding driver + Altruistic Extreme says <em>let me'
        ' tell you the big change you need to make</em>. The practice is patience with the increment.'
        ' Mission-frame the increment so the Altruistic driver carries it: <em>this small change'
        ' helps you serve the mission better</em>.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in your next development conversation with a counselor'
        ' or mentee, say one specific behavior — observable, this-week-actionable — and stop there."'
        ' Do not stack it with three other things you have noticed. One change at a time. Revisit it'
        ' next week before adding anything new.</div>'
        '</div>'
    )

    # IMPACT 6 — Q381 (single-point accountability)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">6</span><div style="flex:1;">'
        '<div class="practice-item-title">One owner per outcome &mdash; no shared accountability</div>'
        '<div class="practice-qref">Q381 &middot; Organizational Decision Making (Impact Rank 6)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#5a6773; border-color:#5a6773;" data-l2="8.2">8.2 Clarity Of Accountability</div></div>'
        '<div class="practice-body">'
        '<p>The standard: <em>when multiple parties are accountable, no one truly is.</em> You'
        ' answered &lsquo;4&rsquo; — leaning toward disagreement with the standard, suggesting you'
        ' are still comfortable with shared accountability arrangements. The research is consistent:'
        ' diffusion of accountability across multiple parties produces worse outcomes than singular'
        ' accountability with broad consultation. The Clarity of Accountability flag at Medium'
        ' severity in your file corroborates — the architecture of who-owns-what is not yet razor-'
        ' sharp in the operations you run.</p>'
        '<p>For the Yale CS Career Fair, the OIS counselor program, and the Aditus US Applications'
        ' coordination, this is concrete: <em>which decisions in each operation are yours alone'
        ' versus shared, and where in the org is the work-product owner singular?</em> The collaborative'
        ' architecture you are wired for (Altruistic Extreme + Persuader engine) loves shared work."'
        ' The accountability architecture has to be different — collaborate widely on the doing,'
        ' singular on the owning.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in every project kickoff, the last question is'
        ' &ldquo;who is the single accountable person?&rdquo; Not a group. Not two co-leads. One name.'
        ' If two people both think they are accountable, neither is.</div>'
        '</div>'
    )

    # IMPACT 7 — Q3 (don't be satisfied with the team)
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">7</span><div style="flex:1;">'
        '<div class="practice-item-title">Stay hard to satisfy &mdash; with the team, with yourself</div>'
        '<div class="practice-qref">Q3 &middot; Not Pleasing (Impact Rank 7)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="7.4">7.4 Discomfort For Team</div></div>'
        '<div class="practice-body">'
        '<p>The standard: <em>I am satisfied with my team</em> — scale 1-True to 5-False, ideal is'
        ' 5 (FALSE). You answered &lsquo;1&rsquo; (TRUE). This is the Satisfied-with-Gripes anchor"'
        ' showing up at the per-answer level. Admired leaders are easy to <em>please</em> (quick to'
        ' notice good work) and hard to <em>satisfy</em> (never closing the development loop)."'
        ' Reporting satisfaction with the team is reporting that you have stopped developing them."'
        ' The team is reading you for that: when you say satisfied, you signal that today is the'
        ' ceiling.</p>'
        '<p>The reframe that lets the practice land is the same one that makes Discomfort for Team'
        ' work: staying hard-to-satisfy <em>is</em> the act of admiration. When you believe the OIS'
        ' counselor team can do more next term than they did this term, you do not say satisfied —'
        ' you say <em>here is what we are practicing next</em>. Pleased and proud and not-yet-satisfied"'
        ' are not in conflict; they are the admired posture.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: after a strong stretch (a great Career Fair, a successful'
        ' OIS week, a play that lands well), ask the team: &ldquo;what could we do at this same level"'
        ' a year from now that we cannot do today?&rdquo; And start the development routine that gets'
        ' there. Pleasure is allowed; satisfaction closes the loop.</div>'
        '</div>'
    )

    # IMPACT 8 — Q5 (team can outvote me) — the lead-guitarist-vs-conductor anchor
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">8</span><div style="flex:1;">'
        '<div class="practice-item-title">Build a team that can outvote you</div>'
        '<div class="practice-qref">Q5 &middot; Conducting &amp; Outvoted (Impact Rank 8)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#fef3c7; color:#1a2332; border-color:#b8862e;" data-l2="3.2">3.2 Empower Team Authority</div></div>'
        '<div class="practice-body">'
        '<p>The standard: <em>my team can outvote me.</em> You answered &ldquo;cannot.&rdquo; This is'
        ' one of the two or three sharpest signals in your entire file — and one of the answers the"'
        ' Signature Pattern called out as part of the leader&rsquo;s burden not yet picked up.</p>'
        '<p><strong>Why this answer is the lead-guitarist tell.</strong> A leader the team cannot'
        ' overrule is the boss. A leader the team can overrule is the conductor. The boss makes good'
        ' decisions and the team executes them. The conductor draws the team forward and the team"'
        ' makes the music. The boss model scales to the size of one person&rsquo;s judgment; the conductor'
        ' model scales to the size of the team&rsquo;s collective judgment. Most early-career leaders"'
        ' (and most leaders, period) hold the boss model and call it leadership. The conductor model'
        ' is harder. It requires giving up something — control, certainty, the comfort of being the'
        ' tiebreaker — every time it is lived.</p>'
        '<p><strong>What the answer is doing in your file.</strong> Combined with HandsOn|Sev, the"'
        ' below-line <em>empower team authority</em> sub-dimension, and the Lower-Standards flag,"'
        ' this answer is the cleanest single read of where the practice has not yet started. You are'
        ' running the operations you lead. The teams under you are not yet structurally capable of'
        ' overruling you. This is not a verdict on the teams; it is a verdict on the architecture."'
        ' The architecture is the leader&rsquo;s job to design.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in your next OIS Head Counselor meeting, name one'
        ' category of decisions where the counselors collectively have authority to overrule your'
        ' position. Tell them what the override mechanism is. Then live with it. The first time they'
        ' actually outvote you on a real decision and the world does not end, the conductor practice'
        ' has started.</div>'
        '</div>'
    )

    # IMPACT 9 — Q61 (no bad teams, only bad leaders) — the extreme-ownership anchor
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">9</span><div style="flex:1;">'
        '<div class="practice-item-title">No bad teams, only leaders not yet leading</div>'
        '<div class="practice-qref">Q61 &middot; Personal Reliability (Impact Rank 9)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#5a6773; border-color:#5a6773;" data-l2="6.1">6.1 Extreme Ownership</div></div>'
        '<div class="practice-body">'
        '<p>The standard: <em>there are no bad teams, only bad leaders.</em> You answered toward'
        ' disagreement — meaning you partly hold that a team can be at fault independently of its'
        ' leader. The admired stance is the opposite — extreme ownership: when the team is not'
        ' delivering, the work to do sits on the leader side first.</p>'
        '<p><strong>Why this answer is the leader&rsquo;s-burden test.</strong> The position you'
        ' partly hold is, on the surface, fair-minded — <em>everyone has agency, everyone is'
        ' responsible for their own performance.</em> The admired stance does not deny that. It"'
        ' adds a layer above it: <strong>the leader is responsible for what the team is capable of'
        ' doing</strong> — for the standards held, for the routines installed, for the role-clarity'
        ' given, for the belief extended, for the discomfort that develops them, and for the"'
        ' transitions out of seats that no longer fit. When the team is not delivering, the question"'
        ' is not <em>what is wrong with the team</em>; it is <em>what part of my leadership work"'
        ' have I not yet done</em>. Holding that frame costs something every time you hold it. It"'
        ' is the practice of taking on the burden the boss model assigns to the team.</p>'
        '<p><strong>Where this connects.</strong> Combined with the Q5 answer above (cannot be"'
        ' overruled), the Q33 / Q34 answers below (would not change the bottom of the team), and the'
        ' Lower-Standards flag, this answer paints a consistent picture: when team performance is'
        ' uneven, your default is to do the work yourself rather than to take responsibility for"'
        ' developing or transitioning the team into being capable of doing it. The encouraging signal"'
        ' is Extreme Ownership at the L2 level scoring positive — the structural belief is partly"'
        ' there. The practice is widening it.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: when you catch yourself frustrated with one of your'
        ' teams — OIS counselors, CS Career Fair leads, the cast of the play, Aditus mentees — write'
        ' down on one line: <em>what is one piece of leadership work I, as the leader, have not yet"'
        ' done that would move this team forward?</em> Then do that thing this week. Repeat until'
        ' the frustration is no longer a frustration with the team but a list of leadership routines"'
        ' you are still installing.</div>'
        '</div>'
    )

    return flag_subheader + '\n'.join(flag_cards) + peranswer_subheader + '\n'.join(impact_cards)

# ============================================================================
# TEACH ITEMS — 10 algorithm-ranked
# ============================================================================

def build_teach_items_html():
    teach_cards = []

    # TEACH 1 — Q137 (enable leaders to reach full potential)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Enable the leaders around you to reach their full potential</div>'
        '<div class="practice-qref">Q137 &middot; Risking by Investing in Others (Teach Rank 1)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.5">1.5 Developmental Mindset</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE. Developmental Mindset is at the top of your file. The very best'
        ' leaders share an unwavering belief that people can change and grow far beyond what seems"'
        ' possible — they describe in vivid detail who they see in front of them, why they believe'
        ' in this person, and what success they expect. You already do this; the visible record of'
        ' the OIS counselors you developed into the team that promoted you to Head, the Aditus'
        ' mentees you guided into Ivy admissions, the cast you produced into a CPA-Fellowship-winning'
        ' play. Teach this to peers who treat people as fixed.</p>'
        '</div>'
        '<div class="practice-fuel">Zen-master level: extend this belief specifically to people who'
        ' have not yet earned it in your frame. That is the Conditional Belief frontier showing up"'
        ' as its growth direction.</div>'
        '</div>'
    )

    # TEACH 2 — Q138 (each direct report can do their job without supervision)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Build a team that runs without supervision</div>'
        '<div class="practice-qref">Q138 &middot; Risking by Replacing Self (Teach Rank 2)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="4.1">4.1 Ability To Disappear</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE — your direct reports can do their jobs without supervision or'
        ' guidance. The structural belief is in place; Replacing Self at the top of the file confirms'
        ' it. The teaching frame: building this is not about backing off — it is about <em>investing'
        ' early</em> in the routines and the role-clarity that let the team carry the work. Teach"'
        ' this to peers who confuse delegation with abdication.</p>'
        '</div>'
        '<div class="practice-fuel">Watch-item: the HandsOn|Sev flag in the file says the day-to-day"'
        ' is not yet matching this answer. Teach the belief and practice the gap-closing routine"'
        ' simultaneously.</div>'
        '</div>'
    )

    # TEACH 3 — Q81 (urgency is not dependent on you)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Urgency lives in the team, not in you alone</div>'
        '<div class="practice-qref">Q81 &middot; Risking by Replacing Self (Teach Rank 3)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="9.2">9.2 Extreme Proactivity</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE — urgency in the operations you run does not depend on you to ignite'
        ' it. This is a high bar and it is one of the patterns that distinguishes admired leaders'
        ' at every altitude. Hire and select for intrinsic urgency; do not try to teach it. When the'
        ' team owns the pace, the leader can step back and the work still happens.</p>'
        '</div>'
        '<div class="practice-fuel">Teach this as the selection criterion for who you bring into'
        ' OIS counselor roles next year, who you tap for Aditus, who you cast in productions: do they'
        ' have their own fire, or do they need you to bring it?</div>'
        '</div>'
    )

    # TEACH 4 — Q49 (proactive enough that you don't get tasked)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Proactive enough that you almost never get tasked</div>'
        '<div class="practice-qref">Q49 &middot; Deliberate Urgency (Teach Rank 4)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="9.2">9.2 Extreme Proactivity</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE. Extreme Proactivity at +1.97 in your L2 list confirms — this is the'
        ' top of the file. The reason you have been promoted, selected, awarded, and trusted at"'
        ' twenty is in large part this single behavior: you see what needs doing and you do it before'
        ' anyone has to ask. Teach this to peers as the leadership minimum, not the leadership max.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in every conversation with a direct report, ask one'
        ' question — <em>what are you working on that I would not yet know to ask about?</em>'
        ' Proactivity you can see is one thing; proactivity they lead with is the level-up.</div>'
        '</div>'
    )

    # TEACH 5 — Q124 (100% of direct reports are dream team)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Build teams of franchise-player teammates</div>'
        '<div class="practice-qref">Q124 &middot; Risking by Investing in Others (Teach Rank 5)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.4">1.4 Demonstrating Genuine Fanness</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE — your direct reports are a dream team. Dream Team flag corroborates."'
        ' This is the admired stance: hold yourself to a standard of selecting (and developing) only"'
        ' people who could go anywhere and chose to be with you. The ones who would lift any team they'
        ' joined. Teach this as the selection bar for every leadership role you ever staff.</p>'
        '</div>'
        '<div class="practice-fuel">Watch-item: the Mixed Conditional Belief flag is the gap. The'
        ' &ldquo;dream team&rdquo; status today reflects belief in the people who have already proven'
        ' themselves; the practice is extending the same belief earlier, before the proof.</div>'
        '</div>'
    )

    # TEACH 6 — Q33 (trade bottom 10%)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">6</span><div style="flex:1;">'
        '<div class="practice-item-title">No tolerable bottom 10% &mdash; develop or transition</div>'
        '<div class="practice-qref">Q33 &middot; Risking by Pushing Extreme Accountability (Teach Rank 6)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="5.8">5.8 Stds = What Tolerate</div></div>'
        '<div class="practice-body">'
        '<p>You answered FALSE — you would <em>not</em> trade the bottom 10% of performers for new'
        ' hires. The admired stance is: the leader should be so aggressive on continuous up-or-out'
        ' through development that there is no obvious bottom 10% to trade. The crew, not basketball,'
        ' frame: you are upgrading at depth all the time, so the bottom keeps rising rather than'
        ' staying replaceable. Teach this to peers who confuse the trade-them-out reflex with the'
        ' develop-them-up discipline.</p>'
        '</div>'
        '<div class="practice-fuel">Watch-item: HoldsAvgDownChain is lit. The stance you teach is'
        ' admired; the practice on the ground is the gap. Teaching it is one thing; living it through'
        ' the development conversations and the role transitions is the other.</div>'
        '</div>'
    )

    # TEACH 7 — Q34 (won't fight to keep some)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">7</span><div style="flex:1;">'
        '<div class="practice-item-title">If they leave, fight to keep them &mdash; or move them out</div>'
        '<div class="practice-qref">Q34 &middot; Risking by Pushing Extreme Accountability (Teach Rank 7)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="5.8">5.8 Stds = What Tolerate</div></div>'
        '<div class="practice-body">'
        '<p>You answered FALSE — there is no one on your team you would not fight to keep. The'
        ' admired posture: every seat is a seat you would fight to keep. If the answer is no, the'
        ' conversation or the transition is overdue. Teach this as the personnel discipline.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: each term, for each role under your leadership, ask'
        ' yourself privately — <em>would I fight to keep this person?</em> If the answer is no, name'
        ' the gap, name the development you owe them, and set a date.</div>'
        '</div>'
    )

    # TEACH 8 — Q113 (direct reports lead with strategic thinking)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">8</span><div style="flex:1;">'
        '<div class="practice-item-title">Build teams of leaders, not just executors</div>'
        '<div class="practice-qref">Q113 &middot; Risking by Pushing Extreme Accountability (Teach Rank 8)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="5.2">5.2 Jrs Extreme Proactivity</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE — your direct reports are leaders, not just task-executors. Jrs"'
        ' Extreme Proactivity sits at the top of your L2 list at +2.0. Build teams of true leaders;"'
        ' hold yourself to high talent-selection standards; create teams that could run without you'
        ' for sustained stretches. Teach this to peers who staff for capacity rather than for'
        ' leadership-in-development.</p>'
        '</div>'
        '<div class="practice-fuel">Zen-master level: develop one of your strong counselors into'
        ' someone who could plausibly take your seat as Head Counselor when you graduate. The exit'
        ' from a leadership role is its highest test.</div>'
        '</div>'
    )

    # TEACH 9 — Q50 (zero voluntary attrition)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">9</span><div style="flex:1;">'
        '<div class="practice-item-title">Retention through growth, not through comfort</div>'
        '<div class="practice-qref">Q50 &middot; Risking by Investing in Others (Teach Rank 9)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.5">1.5 Developmental Mindset</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE — voluntary attrition on your teams is essentially zero. People stay'
        ' because they are growing, not because they are comfortable. The admired retention pattern"'
        ' is built on continuous development paths, opportunities to lead, and meaningful'
        ' contribution to mission. Teach this as the alternative to retention-through-perks.</p>'
        '</div>'
        '<div class="practice-fuel">The diagnostic: when someone stays, ask why. If the answer is'
        ' &ldquo;I&rsquo;m growing into things I never thought I could,&rdquo; the retention is real."'
        ' If the answer is &ldquo;I like the team and the schedule works,&rdquo; the retention is'
        ' fragile.</div>'
        '</div>'
    )

    # TEACH 10 — Q98 (decision quality goes down after this much data)
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">10</span><div style="flex:1;">'
        '<div class="practice-item-title">Action over inaction &mdash; decide on the essential signal</div>'
        '<div class="practice-qref">Q98 &middot; Deliberate Urgency (Teach Rank 10)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="9.6">9.6 Action Over Inaction</div></div>'
        '<div class="practice-body">'
        '<p>You answered at the admired end — decision quality does not go up indefinitely with more'
        ' data. The L2 Action Over Inaction sits at +1.50 in your file. Teach the discipline: define'
        ' the critical questions first, identify what data actually answers them, and ignore the'
        ' rest. Most decisions stall in over-research; admired leaders decide on the essential"'
        ' signal and move.</p>'
        '</div>'
        '<div class="practice-fuel">Pair this with the Impact #1 watch-item: <em>action over inaction</em>'
        ' is the discipline for everyday decisions; <em>wait for the right moment</em> is the'
        ' discipline for irreversible ones. Knowing which class a decision is in is the senior craft.</div>'
        '</div>'
    )

    return '\n'.join(teach_cards)

# ============================================================================
# CONNECTION + CAREER TIMELINE + CLOSING
# ============================================================================

CONNECTION_NARRATIVE_HTML = (
    "<p>Read the two lists together. Your <strong>Teach</strong> items are the leader fluencies you"
    " already practice — developmental mindset, building teams that run without supervision, urgency"
    " that lives in the team, extreme proactivity, dream-team selection bar, no-tolerable-bottom"
    " standard, fight-to-keep retention, leaders-not-executors team-building, retention-through-"
    " growth, and the action-over-inaction discipline. This is a remarkable list at twenty: ten"
    " admired-leader fluencies already in working order, anchoring a top-decile Z|Algo and explaining"
    " why selectors keep choosing you.</p>"
    "<p>Your <strong>What-to-Work-On</strong> section leads with <strong>flag-driven items</strong>"
    " — HandsOn at the Severe level, HoldsAvgDownChain, Mixed Conditional Belief, and Satisfied with"
    " Gripes — because flags are pattern-level signals that sit across multiple L1s, not per-answer"
    " lifts. The algorithm-ranked items follow, and they cluster around a single root: <em>building"
    " operations that execute when you are not the routing node</em>. Wait for the right moment to"
    " decide (#1), peak performance lives in discomfort (#2), stay in your lane (#3), resist running"
    " point until the moment requires it (#4), develop through small specific changes (#5), one"
    " owner per outcome (#6), stay hard to satisfy (#7), build a team that can outvote you (#8), no"
    " bad teams only leaders not yet leading (#9) — all expressions of the same habit: <em>moving"
    " from the lead-guitarist posture to the conductor posture</em>.</p>"
    "<p>The connective tissue is in your <strong>anti-motivator wedge</strong>. Your wiring runs"
    " hard on Commanding, Objective, Altruistic Extreme, and Receptive Primary — drive-from-the-"
    " front, decide-on-the-functional, fueled by mission. The directly-opposing wedge is Intentional"
    " (Extreme below) plus Harmonious, Collaborative, and Intellectual at the indifferent end. The"
    " Impact items all sit in that opposing wedge — restraint, dialogue, specific-individual"
    " investment, the patient mechanics of a team rising up rather than the leader pulling them"
    " forward. Which is exactly why the coaching move is <em>routines installed</em>, not <em>"
    " personality changed</em>. The Impact list is not a call to soften the leader you already are;"
    " it is a short list of specific, observable behaviors installed against the grain.</p>"
    "<p>One synthesis. The reason your record at twenty looks the way it does is the <em>engine</em> —"
    " your tempo, your warmth, your willingness to take the lead, your readiness to do the right"
    " thing rather than the popular thing. The practice frontier is the <em>conductor</em> — the"
    " ability to step back from the baton while the engine still runs, and to install belief in the"
    " team before they have proved it out. Same leader, larger reach. Especially the best get better.</p>"
)

CAREER_TIMELINE_TITLE = "Career Timeline &mdash; Alba Quintas Núñez"

CAREER_TIMELINE_HTML = """
            <div class="timeline">
                <div class="timeline-block" style="flex: 1.4; background: #95a5a6;">Earlier life &middot; Galicia / Madrid<br><span style="font-size:9px; opacity:0.8;">Pre-2024 &middot; FLL Galicia 1st (2021), FLL Spain Robot Design 2nd (2021), Best Young Journalist La Voz de Galicia (2023)</span></div>
                <div class="timeline-block" style="flex: 1.6; background: #34495e;">Programa Aditus &middot; US Apps Coordinator<br><span style="font-size:9px; opacity:0.8;">Jan 2024&ndash;present &middot; Volunteer mentor for 25 high-potential limited-resource Spanish students applying to Ivies</span></div>
                <div class="timeline-block" style="flex: 1.8; background: #2c3e50;">Yale &middot; OIS Counselor &amp; Design Lead<br><span style="font-size:9px; opacity:0.8;">Apr 2024&ndash;Dec 2025 &middot; International Student Orientation</span></div>
                <div class="timeline-block" style="flex: 2.4; background: #1a2332;">Yale &middot; Head Counselor for OIS <span style="background:#d4a84b; color:#1a2332; padding:0 4px; border-radius:2px; font-size:9px;">Internal promotion</span><br><span style="font-size:9px; opacity:0.8;">Dec 2025&ndash;present &middot; Directs logistics &amp; ops, coordinates 20+ counselors, supports 100+ incoming students &middot; <strong>current seat</strong></span></div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item">
                    <div class="legend-dot" style="background: #95a5a6;"></div>
                    <span><strong>Pre-Yale (Galicia / Madrid):</strong> First Lego League Galicia 1st Place (2021), First Lego League Spain 2nd Place Robot Design (2021) — two national-level finishes in robotics. Best Young Journalist 2023 (Prensa-Escola Program of La Voz de Galicia, regional opinion-writing tournament). Mother is novelist Alba Quintas Garciandia; the literary-record selectors are not a parental inheritance — they are the respondent&rsquo;s own.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #34495e;"></div>
                    <span><strong>Programa Aditus &middot; Coordinator of US Applications (Jan 2024&ndash;present, volunteer):</strong> Mentoring a cohort of 25 selected high-potential, limited-resource Spanish students through Ivy / top-US admissions. Workshop design, individualized mentorship, communications coordination. Two years and counting.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #2c3e50;"></div>
                    <span><strong>Yale &middot; OIS Counselor &amp; Design Lead (Apr 2024&ndash;Dec 2025):</strong> Counselor for the international students of the Yale Class of 2029. Designed visual materials and merchandise for all participants and workers. The seat from which the internal promotion happened.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #1a2332;"></div>
                    <span><strong>Yale &middot; Head Counselor for OIS (Dec 2025&ndash;present):</strong> Directs logistics and operations for the program, coordinates 20+ counselors, supports 100+ incoming international students. Concurrent: Co-Director of the annual Yale CS Career Fair (May 2025&ndash;Feb 2026, 300+ undergrad and postgrad CS students), Producer &amp; Actress at Yale Dramatic Association (Oct 2024&ndash;Feb 2026 — produced and starred in two plurilingual Galician/English/Spanish plays; second won the CPA Fellowship), MUNTY Alumni Coordinator (May 2025&ndash;present), Student Tech Coordinator at the Student Technology Collaborative (Oct 2024&ndash;present), Member of the European Youth Parliament (Mar 2022&ndash;Feb 2026; Head Organizer Ourense Regional Conference 2023, Jury Vigo Regional Conference 2024).</span>
                </div>
            </div>
            <div class="timeline-banner" style="background:#e8f4ea; color:#1e5a30;">
                Two years at Yale, one internal promotion, five concurrent leadership roles, one CPA-Fellowship-winning theatrical production, and a four-year EYP arc layered on top of three pre-Yale national-level recognitions across journalism and STEM. The signature is a leader chosen repeatedly by independent selectors across multiple domains. <em>Note: Education, About, Skills (full), Recommendations, and Activity sections of LinkedIn were not pulled in this build — the talent-axis read is anchored on the instrument file, the career-history Non-Scorable answers, and the Honors and Experience sections of the public profile.</em>
            </div>
"""

CLOSING_NOTE_HTML = (
    "<p>One last frame. You are already one of the best. Z|Algo +1.40 puts you in the top decile;"
    " five of nine L1 areas a full standard deviation above the mean or more; an internal Yale"
    " promotion in eighteen months; multiple independent selectors across journalism, STEM, theater,"
    " civic leadership, and Yale&rsquo;s own ladder. This guide is simply how to get even better —"
    " because <em>especially the best get better</em>, and the people you will lead in the next"
    " decade deserve the version of you that has installed the conductor on top of the engine you"
    " already have.</p>"
    "<p>Two effectiveness levers, in sequence.</p>"
    "<p><strong>First, the prerequisite: HandsOn|Sev + HoldsAvgDownChain — the room when you are"
    " not in it.</strong> Before the in-the-room conducting work fully pays out, the layer below"
    " your immediate team needs to be at a standard where the team can carry the work without you"
    " in it for sustained stretches. The Impact items #4 (resist running point), #5 (small specific"
    " changes), #8 (build a team that can outvote you), and #9 (no bad teams only leaders not yet"
    " leading) are the operating expressions of this lever. The gateway routine is the one named in"
    " the HandsOn|Sev card: design <em>one full week off-deck</em> in one of your operations, name"
    " the decisions the team will need to make, and resist the impulse to check in.</p>"
    "<p><strong>Second, on that foundation: the Conducting frontier — moving from lead guitarist to"
    " conductor.</strong> The structural pieces are already built (Replacing Self at the top of the"
    " file, Pushing Accountability strong, Deliberate Urgency at the top). The layer on top is the"
    " in-the-room conductor posture: dialogue over direction, team-can-outvote-me as a real"
    " arrangement, peak-performance-lives-in-discomfort as the daily practice with the team. Impact"
    " #8 (build a team that can outvote you) is the gateway routine. The first time the team uses"
    " the override — and the world does not end — the conductor posture starts to feel possible.</p>"
    "<p>Three notes on the framing rules that matter at your particular wiring. (1) <em>Mission-frame"
    " developmental routines.</em> Altruistic Extreme will carry mission-framed investment in people;"
    " Intentional Extreme below will not carry specific-individual-framed investment. The same"
    " practice, framed twice — only one frame sticks. (2) <em>The conductor practice is about"
    " creating space, not adding warmth.</em> The warmth is already there; the practice is making"
    " room for the team to step in. (3) <em>Effectiveness, not good-versus-bad.</em> Neither frontier"
    " is a verdict on the leader you are. They are the levers that move this file further into the"
    " territory of the admired and successful leader you are already on track to become.</p>"
    "<p>Especially the best get better.</p>"
)

# ============================================================================
# DISC NOTES
# ============================================================================

DISC_NOTE_TEXT = (
    "Conducting Persuader (D=75, I=73, S=22, C=34). Adapted shift to Promoting Persuader: −13D, −5I, +6S, +14C."
)
DISC_NOTE_DETAIL = (
    "Adapted profile asks for slightly less commanding posture, slightly more steady tempo,"
    " noticeably more attention to structure (C +14). Commanding 88 + Objective 82 + Altruistic 81"
    " Extreme + Receptive 65 Primary is the wiring of an early-career leader who drives results"
    " through energy and persuasion in service of a broader cause."
)
DISC_ANNOTATION_CODE = ""

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    leaks = re.findall(r'\{\{([A-Z_0-9]+)\}\}', html)
    if leaks:
        failures.append(f"Unreplaced tokens: {sorted(set(leaks))}")

    for cid in ['distChart1', 'distChart2', 'distChart3', 'discChart', 'excstdsChart']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("Brand lockup count < 2")

    teach_items = html.split('Part 1 — What You Teach')[1].split('Part 2 — What to Work On')[0]
    impact_items = html.split('Part 2 — What to Work On')[1].split('How the Two Lists Connect')[0]
    n_teach = teach_items.count('class="practice-item"')
    n_impact_std = impact_items.count('class="practice-item"')
    n_impact_flag = impact_items.count('class="practice-item flag-driven"')
    if n_teach < 10:
        failures.append(f"Teach items: got {n_teach}, need >=10")
    if n_impact_flag < 3:
        failures.append(f"Flag-driven impact items: got {n_impact_flag}, need >=3")
    if n_impact_std < 9:
        failures.append(f"Per-answer impact items: got {n_impact_std}, need >=9")

    m_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_isl1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_labels and m_isl1:
        labels = json.loads(m_labels.group(1))
        isl1 = json.loads(m_isl1.group(1))
        if len(labels) < 12:
            failures.append(f"Scorecard has only {len(labels)} rows (min 12)")
        l2_rows = sum(1 for v in isl1 if not v)
        if l2_rows < 6:
            failures.append(f"Scorecard has only {l2_rows} L2 rows (min 6)")
        for lbl, is_l1 in zip(labels, isl1):
            if is_l1 and lbl != lbl.upper():
                failures.append(f"L1 label '{lbl}' must be UPPERCASE"); break
            if not is_l1 and not lbl.startswith('    '):
                failures.append(f"L2 label '{lbl}' must be 4-space indented"); break

    # Chart 3 reversed gate — first label must be a high flag count, last label must be 0
    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first and int(first.group(1)) < 30:
            failures.append(f"Chart 3 axis not reversed — first label starts with {first.group(1)}")

    if 'class="ma-section"' not in html:
        failures.append("Motivators section not injected")

    if 'Alba' not in html:
        failures.append("Candidate name (Alba) missing")
    if 'Yale' not in html:
        failures.append("Candidate context (Yale) missing")

    if 'Conducting Persuader' not in html:
        failures.append("Fingerprint narrative missing wedge label (Conducting Persuader)")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} issue(s) ***")
        for f in failures:
            print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)} issue(s)")
    else:
        print("*** QA GATE PASSED ***")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("Loading respondent data...")
    respondent_data = load_respondent_data()

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution chart tokens...")
    dist_tokens = build_distribution_tokens(
        zalgo_rows, flag_rows,
        respondent_data['z_algo_overall'],
        respondent_data['z_human_overall'],
        respondent_data['rf_num'],
    )

    print("Building motivators section...")
    respondent = build_respondent_dict(respondent_data)
    motivators_html = build_section(respondent, include_css=True)

    print("Building scorecard...")
    scorecard = build_excstds_scorecard(respondent_data)

    print("Building impact / teach items...")
    impact_html = build_impact_items_html()
    teach_html = build_teach_items_html()

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    # Neutralize in-comment {{EXCSTDS_COLOR_OVERRIDES}} reference if present
    template_html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}}",
        "Substitute EXCSTDS_COLOR_OVERRIDES"
    )

    # Inject motivators section (contains raw braces — do this before token-replace loop)
    template_html = template_html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    replacements = {
        'CANDIDATE_NAME': 'Alba Quintas Núñez',
        'CANDIDATE_CREDS': 'Yale College Student &middot; Head Counselor for OIS',
        'CANDIDATE_ROLE': 'Yale University &middot; concurrent leadership across OIS, CS Career Fair, MUNTY, Aditus, Yale Dramatic Association, Student Tech Collaborative',
        'REPORT_DATE': 'April 25, 2026',

        'SIGNATURE_PATTERN': SIGNATURE_PATTERN,
        'FINGERPRINT_NARRATIVE': FINGERPRINT_NARRATIVE,
        'DRIVING_FORCES_PRIMARY_HTML': DRIVING_FORCES_PRIMARY_HTML,
        'DRIVING_FORCES_INDIFFERENT_HTML': DRIVING_FORCES_INDIFFERENT_HTML,
        'DRIVING_FORCES_IMPLICATIONS_HTML': DRIVING_FORCES_IMPLICATIONS_HTML,

        'ZALGO_OVERALL': f'{respondent_data["z_algo_overall"]:+.2f}',
        'COHORT_AVG': '-0.12',
        'TEACH_ITEMS': '10/10',
        'IMPACT_ITEMS': '14',
        'FLAGS_LIT': str(len(respondent_data['flags_lit'])),
        'REVERSE_FLAGS': str(respondent_data['rf_num']),

        'DISC_D_NAT': '75', 'DISC_I_NAT': '73', 'DISC_S_NAT': '22', 'DISC_C_NAT': '34',
        'DISC_D_ADP': '62', 'DISC_I_ADP': '68', 'DISC_S_ADP': '28', 'DISC_C_ADP': '48',
        'DISC_NOTE_TEXT': DISC_NOTE_TEXT,
        'DISC_NOTE_DETAIL': DISC_NOTE_DETAIL,
        'DISC_ANNOTATION_CODE': DISC_ANNOTATION_CODE,

        'WIRING_FIT_ITEMS': WIRING_FIT_ITEMS,

        **scorecard,
        **dist_tokens,

        'TEACH_ITEMS_HTML': teach_html,
        'IMPACT_ITEMS_HTML': impact_html,

        'CONNECTION_NARRATIVE_HTML': CONNECTION_NARRATIVE_HTML,
        'CAREER_TIMELINE_TITLE': CAREER_TIMELINE_TITLE,
        'CAREER_TIMELINE_HTML': CAREER_TIMELINE_HTML,
        'CLOSING_NOTE_HTML': CLOSING_NOTE_HTML,
    }

    html = template_html
    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    qa_gate(html)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")

if __name__ == '__main__':
    main()
