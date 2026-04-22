"""Rebuild the integrated spike (wheel + alignment + standard map) with the v4 restructure.

Layout per respondent (top to bottom):
  1. Motivator & anti-motivator callouts (2 columns)
  2. Alignment block — 4-bucket grid
  3. Standard Map (big circle). L2 pill FILL = Z|Algo; L2 pill BORDER = 4-bucket color
     - green   = motivator-aligned + strong (running naturally)
     - amber   = motivator-aligned + weak   (aligned but not running)
     - purple  = anti-aligned + strong      (installed against the grain)
     - red     = anti-aligned + weak        (routines to install)
     - gray    = cross-wedge or neutral score (no bucket assignment)
  4. DISC wiring panel (scores, Natural/Adapted positions, intensity notes)

Run from repo root:
    python _pipeline/scripts/build_spike_v4.py

Outputs:
    _reports/spike_wheel_comparison.html
"""
import json, math, csv, shutil, os
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict

ROOT = Path(__file__).resolve().parents[2]
MAP_XLSX = ROOT / 'l2_wedge_map.xlsx'
QUESTIONS_TSV = ROOT / '_pipeline' / 'data' / 'questions_full.tsv'
BILL_XLSX = ROOT / 'bill_george' / 'L1 L2 Skinny Flags.xlsx'
OUT_HTML = ROOT / '_reports' / 'spike_wheel_comparison.html'

# --- Load mapping (skipping SPECULATIVE banner row) ---
wb = load_workbook(MAP_XLSX, data_only=True)
ws = wb['L2_Wedge_Mapping']
hdrs = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
data_start = 2
if hdrs[0] and 'SPECULATIVE' in str(hdrs[0]):
    hdrs = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
    data_start = 3
cidx = {h: hdrs.index(h)+1 for h in hdrs}

l2num_map = {}
with open(QUESTIONS_TSV, encoding='utf-8') as f:
    for row in csv.DictReader(f, delimiter='\t'):
        l2num_map[row['L2_Short']] = row['L2_Num']

mapping_rows = []
for r in range(data_start, ws.max_row+1):
    l2 = ws.cell(row=r, column=cidx['L2_Short']).value
    if not l2: continue
    mapping_rows.append({
        'l2_num': l2num_map.get(l2,'?'), 'l2_short': l2,
        'primary': ws.cell(row=r, column=cidx['Primary_Wedge']).value,
        'secondary': ws.cell(row=r, column=cidx['Secondary_Wedge']).value,
        'clean_q': ws.cell(row=r, column=cidx['Q_Clean']).value,
        'clean_max': ws.cell(row=r, column=cidx['Pol_Clean_Max']).value,
        'all_max': ws.cell(row=r, column=cidx['Pol_All_Max']).value,
    })
wedge_map = {r['l2_short']: {'primary':r['primary'], 'secondary':r['secondary']} for r in mapping_rows}

WEDGES = ['Implementor','Conductor','Persuader','Promoter','Relater','Supporter','Coordinator','Analyzer']
WEDGE_CENTER = {w: i*45 for i, w in enumerate(WEDGES)}
OPPOSITES = {'Conductor':'Supporter','Supporter':'Conductor','Persuader':'Coordinator','Coordinator':'Persuader',
             'Promoter':'Analyzer','Analyzer':'Promoter','Relater':'Implementor','Implementor':'Relater'}

# Option 3 palette: pill FILL carries the 4-bucket signal.
BUCKET_FILL = {
    'motivator_strong': ('#2563eb', '#ffffff'),   # running naturally — BLUE
    'motivator_weak':   ('#c0392b', '#ffffff'),   # aligned but not running — DISC D RED
    'anti_strong':      ('#22c55e', '#ffffff'),   # installed against grain — BRIGHT GREEN (celebrate)
    'anti_weak':        ('#d4a84b', '#1a2332'),   # routines to install — GOLD
    'motivator_neutral':('#dbeafe', '#1a2332'),   # pale blue
    'anti_neutral':     ('#dcfce7', '#1a2332'),   # pale green
    'cross':            ('#f3f4f6', '#5a6773'),
    'no_score':         ('#e8ecf0', '#5a6773'),
}
BUCKET_COLORS = {
    'motivator_strong': '#2563eb',
    'motivator_weak':   '#c0392b',
    'anti_strong':      '#22c55e',
    'anti_weak':        '#d4a84b',
}
PILL_BORDER_COLOR = '#d0d4da'

ADJACENT_WEDGES = {
    'Implementor': ('Analyzer', 'Conductor'),
    'Conductor':   ('Implementor', 'Persuader'),
    'Persuader':   ('Conductor', 'Promoter'),
    'Promoter':    ('Persuader', 'Relater'),
    'Relater':     ('Promoter', 'Supporter'),
    'Supporter':   ('Relater', 'Coordinator'),
    'Coordinator': ('Supporter', 'Analyzer'),
    'Analyzer':    ('Coordinator', 'Implementor'),
}

def zone_of(respondent_primary, l2_primary):
    anti = OPPOSITES[respondent_primary]
    mot = {respondent_primary, *ADJACENT_WEDGES[respondent_primary]}
    ant = {anti, *ADJACENT_WEDGES[anti]}
    if l2_primary in mot: return 'motivator'
    if l2_primary in ant: return 'anti'
    return 'cross'

def bucket_for(zone, z):
    if z is None: return 'no_score'
    if zone == 'cross': return 'cross'
    if zone == 'motivator':
        if z >= 0.5:  return 'motivator_strong'
        if z <= -0.5: return 'motivator_weak'
        return 'motivator_neutral'
    if zone == 'anti':
        if z >= 0.5:  return 'anti_strong'
        if z <= -0.5: return 'anti_weak'
        return 'anti_neutral'
    return 'cross'

def fmt_z(z):
    if z is None: return '—'
    return f'{z:+.2f}'

def classify_bucket(respondent_primary, l2_info, z):
    if z is None: return None
    zone = zone_of(respondent_primary, l2_info['primary'])
    if zone == 'motivator':
        if z >= 0.5:  return 'motivator_strong'
        if z <= -0.5: return 'motivator_weak'
    if zone == 'anti':
        if z >= 0.5:  return 'anti_strong'
        if z <= -0.5: return 'anti_weak'
    return None

# --- Respondent data ---
wb2 = load_workbook(BILL_XLSX, data_only=True)
ws2 = wb2['L2']
bill_scores = {ws2.cell(row=r, column=4).value: ws2.cell(row=r, column=5).value
               for r in range(2, ws2.max_row+1)
               if ws2.cell(row=r, column=4).value and isinstance(ws2.cell(row=r, column=5).value, (int,float))}
jody_scores = {
    'Demonstrating Genuine Fanness':0.89,'Developmental Mindset':0.56,
    'Handling Daily Difficulties With Dignity':-0.12,'Developmental Discipline':-0.05,
    'Reciprocal Followership':-0.04,'Clarity Of Accountability':0.62,
    'Respects Collective Wisdom':0.19,'Simplification Methods':-0.96,
    'Facts Over Feelings':-0.006,'Cares About Others Not Their Approval':-1.47,
    'Discomfort For Team':0.79,
}

def compute_buckets(primary_wedge, scores):
    buckets = {'motivator_strong':[], 'motivator_weak':[], 'anti_strong':[], 'anti_weak':[]}
    for l2, z in scores.items():
        info = wedge_map.get(l2)
        if not info: continue
        b = classify_bucket(primary_wedge, info, z)
        if b in buckets:
            buckets[b].append({'l2':l2, 'z':z, 'primary':info['primary'], 'secondary':info['secondary']})
    for b in buckets:
        buckets[b].sort(key=lambda x: -abs(x['z']))
    return buckets

def render_alignment_block(buckets):
    labels = {'motivator_strong':'Routines running naturally','motivator_weak':'Aligned but not running',
              'anti_strong':'Installed against the grain','anti_weak':'Routines to install'}
    descs = {'motivator_strong':'Motivator-aligned + strong. Wiring paying off.',
             'motivator_weak':'Motivator-aligned + weak. Wiring should carry this — probe why it\'s not.',
             'anti_strong':'Anti-aligned + strong. A tremendous positive — installed against the grain.',
             'anti_weak':'Anti-aligned + weak. Coaching frontier: turn into a routine, summon a motivator.'}
    bg = {'motivator_strong':'#dbeafe',  # pale blue
          'motivator_weak':  '#fbd8d2',  # pale red (DISC D family)
          'anti_strong':     '#dcfce7',  # pale green
          'anti_weak':       '#fae6a8'}  # pale gold
    def bucket(k):
        items = buckets.get(k, [])
        pill_fill, pill_text = BUCKET_FILL[k]
        html = f'<div class="bucket" style="background:{bg[k]}; border-left:4px solid {BUCKET_COLORS[k]};">'
        html += f'<div class="bucket-pill" style="background:{pill_fill}; color:{pill_text};">{labels[k]}</div>'
        html += f'<div class="bucket-desc">{descs[k]}</div>'
        if not items:
            html += '<div class="bucket-empty">None.</div>'
        else:
            html += '<div class="bucket-items">'
            for it in items[:3]:
                z = it['z']
                html += f'<div class="l2-row"><span class="z-chip" style="background:{pill_fill}; color:{pill_text};">{z:+.2f}</span><span class="l2-name">{it["l2"]}</span></div>'
            html += '</div>'
        html += '</div>'
        return html
    inner = ''.join(bucket(k) for k in ['motivator_strong','motivator_weak','anti_strong','anti_weak'])
    return f'<div class="alignment-grid">{inner}</div>'

def render_callouts(primary_wedge, anti_wedge, name_first, narrative_m, narrative_a):
    return f'''<div class="callouts-pair">
<div class="callout motivator">
  <div class="callout-label">Motivator zone &middot; where {name_first} is energized</div>
  <div class="callout-wedge">{primary_wedge}<span class="intensity-tag">Strong</span></div>
  <div class="callout-body">{narrative_m}</div>
</div>
<div class="callout anti">
  <div class="callout-label">Anti-motivator zone &middot; where {name_first} drains</div>
  <div class="callout-wedge">{anti_wedge} <span class="pos">(directly across)</span><span class="intensity-tag">Strong</span></div>
  <div class="callout-body">{narrative_a}</div>
</div>
</div>'''

def render_wiring(nat_disc, nat_pos, nat_label, nat_intensity, adp_pos, adp_label, adp_intensity, notes=''):
    adp_row = ''
    if adp_pos != nat_pos or adp_label != nat_label:
        adp_row = f'<tr><th>Adapted</th><td>pos {adp_pos} — {adp_label}</td><td>intensity {adp_intensity:.2f}</td></tr>'
    bars = ''
    for axis, val, color in zip(['D','I','S','C'], nat_disc, ['#c0392b','#f39c12','#27ae60','#2980b9']):
        bars += f'<div class="disc-bar"><span class="disc-label">{axis}</span><div class="disc-track"><div class="disc-fill" style="width:{val}%;background:{color};"></div></div><span class="disc-val">{val}</span></div>'
    note_html = f'<div class="wiring-note">{notes}</div>' if notes else ''
    return f'''<div class="wiring-panel">
<div class="wiring-title">DISC wiring</div>
<div class="wiring-grid">
  <div class="disc-bars">{bars}</div>
  <table class="wiring-table">
    <tr><th>Natural</th><td>pos {nat_pos} — {nat_label}</td><td>intensity {nat_intensity:.2f}</td></tr>
    {adp_row}
  </table>
</div>
{note_html}
</div>'''

def render_standard_map(respondent_name, primary_wedge, scores, nat_label, nat_intensity, adp_label, adp_intensity, full_data):
    anti = OPPOSITES[primary_wedge]
    rows = [dict(r) for r in mapping_rows]
    SIZE=1100; cx=cy=SIZE/2; rInner=70; rOuter=470
    rUsable_in=rInner+18; rUsable_out=rOuter-28
    def polar(r, deg):
        a = (deg-90)*math.pi/180; return (cx+r*math.cos(a), cy+r*math.sin(a))
    def wedge_path(rO, rI, sd, ed):
        p1=polar(rO,sd); p2=polar(rO,ed); p3=polar(rI,ed); p4=polar(rI,sd)
        la = 1 if (ed-sd)>180 else 0
        return f'M {p1[0]} {p1[1]} A {rO} {rO} 0 {la} 1 {p2[0]} {p2[1]} L {p3[0]} {p3[1]} A {rI} {rI} 0 {la} 0 {p4[0]} {p4[1]} Z'
    def angle_delta(a,b): return ((b-a)%360+540)%360-180
    def wrap_label(text, max_single=17):
        if len(text) <= max_single or ' ' not in text: return [text]
        words = text.split()
        best = float('inf'); best_split = 1
        for i in range(1, len(words)):
            l1=' '.join(words[:i]); l2=' '.join(words[i:])
            s = abs(len(l1)-len(l2)) + max(len(l1),len(l2))*0.3
            if s < best: best = s; best_split = i
        return [' '.join(words[:best_split]), ' '.join(words[best_split:])]
    def pill_params(pol):
        if pol >= 0.75: return {'font':12,'pad_h':10,'pad_v':8}
        if pol < 0.25:  return {'font':10,'pad_h':8,'pad_v':6}
        return               {'font':11,'pad_h':9,'pad_v':7}
    for r in rows:
        p = r['clean_max'] if r['clean_max'] is not None else r['all_max']
        r['_test_only'] = (r['clean_q'] == 0)
        pc = 0 if p is None else max(0, min(1, p))
        r['_polarity'] = pc
        full_label = f"{r['l2_num']} {r['l2_short']}"
        if len(full_label) <= 22: r['_lines'] = [full_label]
        else:
            body = wrap_label(r['l2_short'], 18)
            r['_lines'] = [full_label] if len(body)==1 else [f"{r['l2_num']} {body[0]}", body[1]]
        params = pill_params(pc)
        r['_font'] = params['font']
        line_h = params['font']*1.15
        max_len = max(len(l) for l in r['_lines'])
        r['_w'] = max_len*params['font']*0.58 + params['pad_h']*2
        # extra row for the z-score line (font-2, plus 2px gap)
        z_row_h = (params['font'] - 2) * 1.15 + 2
        r['_h'] = line_h*len(r['_lines']) + z_row_h + params['pad_v']*2
        r['_target_radius'] = rUsable_in + pc*(rUsable_out - rUsable_in)
    by_prim = defaultdict(list)
    for r in rows: by_prim[r['primary']].append(r)
    for wedge, items in by_prim.items():
        pc_ang = WEDGE_CENTER[wedge]
        def lean(it): return angle_delta(pc_ang, WEDGE_CENTER[it['secondary']]) if it['secondary'] else 0
        items.sort(key=lambda x:(lean(x), x['l2_num']))
        n = len(items)
        for i, it in enumerate(items):
            offset = 0 if n == 1 else (-20 + (40/(n-1))*i)
            it['_angle'] = (pc_ang + offset) % 360
            it['_radius'] = it['_target_radius']
            x, y = polar(it['_radius'], it['_angle']); it['_x']=x; it['_y']=y
    def overlap(a, b):
        dx=abs(a['_x']-b['_x']); dy=abs(a['_y']-b['_y'])
        mx=(a['_w']+b['_w'])/2+4; my=(a['_h']+b['_h'])/2+4
        return (mx-dx, my-dy) if dx<mx and dy<my else None
    for _ in range(120):
        forces = {id(r):[0.0,0.0] for r in rows}
        for i, a in enumerate(rows):
            for b in rows[i+1:]:
                o = overlap(a, b)
                if o:
                    ox, oy = o
                    dx=b['_x']-a['_x']; dy=b['_y']-a['_y']
                    d = math.sqrt(dx*dx+dy*dy) or 0.01
                    push = max(ox,oy)*0.55
                    nx=dx/d; ny=dy/d
                    forces[id(a)][0]-=nx*push; forces[id(a)][1]-=ny*push
                    forces[id(b)][0]+=nx*push; forces[id(b)][1]+=ny*push
        disp = 0
        for r in rows:
            fx, fy = forces[id(r)]
            if fx==0 and fy==0: continue
            r['_x']+=fx; r['_y']+=fy
            dx, dy = r['_x']-cx, r['_y']-cy
            rad = math.sqrt(dx*dx+dy*dy)
            ang = (math.atan2(dy,dx)*180/math.pi+90)%360
            tr = r['_target_radius']
            if rad<rUsable_in: rad=rUsable_in
            if rad>rUsable_out: rad=rUsable_out
            rad = rad*0.85 + tr*0.15
            pc_ang = WEDGE_CENTER[r['primary']]
            d_ang = angle_delta(pc_ang, ang)
            if d_ang>21: ang=(pc_ang+21)%360
            if d_ang<-21: ang=(pc_ang-21+360)%360
            nx, ny = polar(rad, ang)
            disp += abs(nx-r['_x'])+abs(ny-r['_y'])
            r['_x']=nx; r['_y']=ny; r['_angle']=ang; r['_radius']=rad
        if disp < 0.5: break
    # z_fill removed — pill fills now come from BUCKET_FILL via bucket_for(zone, z)
    def intensity_r(i): i=max(0,min(1,i)); return rInner+14+i*(rOuter-rInner-28)
    SECONDARY_MAP = {'Conduct':'Conductor','Persuad':'Persuader','Promot':'Promoter','Relat':'Relater',
        'Support':'Supporter','Coordinat':'Coordinator','Analyz':'Analyzer','Implement':'Implementor'}
    def parse_label(label):
        parts = label.strip().split()
        primary = parts[-1]; secondary = None
        if len(parts) > 1:
            mod = parts[0]
            if mod.lower().endswith('ing'): mod = mod[:-3]
            secondary = SECONDARY_MAP.get(mod, mod+'er')
        return primary, secondary
    def marker_angle(p, s):
        pc = WEDGE_CENTER[p]
        if not s or s not in WEDGE_CENTER: return pc
        d = angle_delta(pc, WEDGE_CENTER[s])
        return (pc + (13 if d>=0 else -13)) % 360
    parts = [f'<svg viewBox="0 0 {SIZE} {SIZE}" xmlns="http://www.w3.org/2000/svg">']
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rOuter}" fill="#ffffff" stroke="#cfd4da" stroke-width="1"/>')
    for w, c in WEDGE_CENTER.items():
        sd, ed = c-22.5, c+22.5
        if w==primary_wedge: fill,stroke,sw = '#e8f6ed','#27ae60',2
        elif w==anti:        fill,stroke,sw = '#fbe9e6','#c0392b',2
        else:                fill,stroke,sw = '#ffffff','#d0d4da',1
        parts.append(f'<path d="{wedge_path(rOuter, rInner, sd, ed)}" fill="{fill}" stroke="{stroke}" stroke-width="{sw}"/>')
    for pol, lbl in [(0.25,'weak'),(0.5,'moderate'),(0.75,'strong'),(1.0,'peak')]:
        r = rUsable_in + pol*(rUsable_out - rUsable_in)
        parts.append(f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="#a0a8b0" stroke-width="1.2" stroke-dasharray="6 6" opacity="0.6"/>')
        p = polar(r, 22.5)
        parts.append(f'<rect x="{p[0]-32}" y="{p[1]-9}" width="64" height="16" rx="8" fill="#ffffff" fill-opacity="0.95" stroke="#a0a8b0" stroke-width="0.5"/>')
        parts.append(f'<text x="{p[0]}" y="{p[1]+4}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" font-weight="700" fill="#5a6773">{lbl} ({pol:.2f})</text>')
    for i in range(8):
        a = i*45+22.5; to=polar(rOuter,a); ti=polar(rInner,a)
        parts.append(f'<line x1="{to[0]}" y1="{to[1]}" x2="{ti[0]}" y2="{ti[1]}" stroke="#cfd4da" stroke-width="1"/>')
    for w, c in WEDGE_CENTER.items():
        lp = polar(rOuter+28, c)
        fw,col = (700,'#1e6b3f') if w==primary_wedge else ((700,'#8a251a') if w==anti else (600,'#4a5662'))
        parts.append(f'<text x="{lp[0]}" y="{lp[1]}" text-anchor="middle" dominant-baseline="middle" font-family="DM Sans, sans-serif" font-size="15" font-weight="{fw}" fill="{col}">{w.upper()}</text>')
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{rInner}" fill="#fafbfc" stroke="#cfd4da" stroke-width="1"/>')
    parts.append(f'<text x="{cx}" y="{cy-10}" text-anchor="middle" font-family="DM Serif Display, serif" font-size="14" fill="#1a2332">{respondent_name}</text>')
    parts.append(f'<text x="{cx}" y="{cy+6}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="10" fill="#6b7884">{primary_wedge}</text>')
    if not full_data:
        parts.append(f'<text x="{cx}" y="{cy+22}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="8" fill="#c0392b" font-style="italic">partial data</text>')
    for r in rows:
        x, y = r['_x'], r['_y']
        z = scores.get(r['l2_short'])
        zone = zone_of(primary_wedge, r['primary'])
        bucket = bucket_for(zone, z)
        fill, textcol = BUCKET_FILL[bucket]
        w, h, f = r['_w'], r['_h'], r['_font']
        dash = ' stroke-dasharray="3 2"' if r['_test_only'] else ''
        parts.append(f'<rect x="{x-w/2}" y="{y-h/2}" width="{w}" height="{h}" rx="{min(h/2,10)}" fill="{fill}" fill-opacity="0.98" stroke="{PILL_BORDER_COLOR}" stroke-width="1"{dash}/>')
        n = len(r['_lines']); line_h = f*1.15
        # Main label (shifted up to leave room for z-score line)
        start_y = y - (n-1)/2*line_h + f/3 - 5
        for i, line in enumerate(r['_lines']):
            parts.append(f'<text x="{x}" y="{start_y+i*line_h}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="{f}" font-weight="700" fill="{textcol}">{line}</text>')
        # Z-score number below the label
        z_y = start_y + n*line_h + 2
        parts.append(f'<text x="{x}" y="{z_y}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="{f-2}" font-weight="600" fill="{textcol}" fill-opacity="0.80">{fmt_z(z)}</text>')
    npr, nsc = parse_label(nat_label)
    na = marker_angle(npr, nsc); nr = intensity_r(nat_intensity)
    nx, ny = polar(nr, na)
    parts.append(f'<circle cx="{nx}" cy="{ny}" r="27" fill="none" stroke="#d4a84b" stroke-width="3" stroke-dasharray="4 3"/>')
    parts.append(f'<circle cx="{nx}" cy="{ny}" r="20" fill="#1a2332" stroke="#ffffff" stroke-width="3.5"/>')
    parts.append(f'<text x="{nx}" y="{ny+5}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="14" font-weight="800" fill="#ffffff">N</text>')
    apr, asc = parse_label(adp_label)
    if (apr!=npr) or (asc!=nsc) or abs(adp_intensity-nat_intensity)>0.02:
        aa = marker_angle(apr, asc); ar = intensity_r(adp_intensity)
        ax, ay = polar(ar, aa)
        sp = []
        for i in range(10):
            sr = 22 if i%2==0 else 10
            sa = (i*36-90)*math.pi/180
            sp.append(f'{ax+sr*math.cos(sa)},{ay+sr*math.sin(sa)}')
        parts.append(f'<polygon points="{" ".join(sp)}" fill="#d4a84b" stroke="#1a2332" stroke-width="2.5"/>')
        parts.append(f'<text x="{ax}" y="{ay+5}" text-anchor="middle" font-family="DM Sans, sans-serif" font-size="13" font-weight="800" fill="#1a2332">A</text>')
    parts.append('</svg>')
    return ''.join(parts)

bill_nar = {
    'motivator': 'Bill lights up in environments that reward <strong>decisive action, visible results, and forward motion</strong>. The Persuader overlay means he wins rooms, closes deals, and rallies teams. <strong>Naturally motivated by:</strong> a clear mandate, fast tempo, real stakes, and proportional authority. With D=92 the pull is <em>strong</em>.',
    'anti': 'Opposite wedge = <strong>Supporter / Coordinator</strong>: patience, consensus-building, careful procedural compliance. Not a deficit — drain. When harmony outranks direction, he burns energy unsustainably.',
}
jody_nar = {
    'motivator': 'Jody thrives in environments that reward <strong>steady loyalty, warm relationships, and sustained contribution</strong>. The Relater overlay adds genuine social warmth. <strong>Naturally motivated by:</strong> stable expectations, a cohesive team, a respected role, and time to build trust.',
    'anti': 'Opposite wedge = <strong>Conductor / Persuader</strong>: unilateral decisions against resistance, confrontational challenge, driving urgency. Watch for deferred hard conversations and over-accommodation.',
}

bill_b = compute_buckets('Conductor', bill_scores)
jody_b = compute_buckets('Supporter', jody_scores)
bill_align = render_alignment_block(bill_b)
jody_align = render_alignment_block(jody_b)
bill_callouts = render_callouts('Conductor','Supporter','Bill', bill_nar['motivator'], bill_nar['anti'])
jody_callouts = render_callouts('Supporter','Conductor','Jody', jody_nar['motivator'], jody_nar['anti'])
bill_map = render_standard_map('Bill George','Conductor', bill_scores, 'Persuading Conductor',0.92,'Persuading Conductor',0.85, True)
jody_map = render_standard_map('Jody Bender','Supporter', jody_scores, 'Relating Supporter',0.85,'Supporting Relater',0.785, False)
bill_wiring = render_wiring([92,72,32,14], 11, 'Persuading Conductor', 0.92, 11, 'Persuading Conductor', 0.85,
    'No Natural→Adapted shift — no stress-masking signal.')
jody_wiring = render_wiring([15,66,85,58], 34, 'Relating Supporter', 0.85, 17, 'Supporting Relater', 0.785,
    'Adapted shifts from Supporter into Relater (intensifying strengths, not borrowing from anti-motivator).')

CSS = '''
* { margin:0; padding:0; box-sizing:border-box; }
body { font-family: "DM Sans", sans-serif; color:#2c3e50; background:#f5f7fa; line-height:1.55; padding:32px 0; }
.container { max-width:1240px; margin:0 auto; background:white; box-shadow:0 2px 8px rgba(0,0,0,0.08); padding:40px; }
.caveat { background:#fff3cd; border-left:4px solid #d4a84b; padding:14px 20px; margin-bottom:24px; font-size:13px; color:#6b5518; line-height:1.55; }
.caveat strong { color:#8a4500; }
h1 { font-family:"DM Serif Display", serif; font-size:28px; font-weight:400; color:#1a2332; margin-bottom:6px; }
h1 + .subtitle { font-size:13px; color:#6b7884; letter-spacing:0.4px; margin-bottom:28px; }
.respondent-block { margin:40px 0; padding-top:32px; border-top:1px solid #e5e8ec; }
.respondent-block:first-of-type { border-top:none; padding-top:0; }
.section-title { font-family:"DM Serif Display", serif; font-size:24px; color:#1a2332; margin-bottom:4px; }
.section-title .name { color:#d4a84b; }
.respondent-meta { font-size:12px; color:#6b7884; letter-spacing:0.5px; text-transform:uppercase; margin-bottom:20px; }
.callouts-pair { display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:16px; }
.callout { border-left:4px solid; padding:14px 18px; border-radius:4px; }
.callout.motivator { border-color:#27ae60; background:#f0faf3; }
.callout.anti { border-color:#c0392b; background:#fdf3f2; }
.callout-label { font-size:10px; font-weight:700; letter-spacing:1.2px; text-transform:uppercase; margin-bottom:4px; }
.callout.motivator .callout-label { color:#1e6b3f; }
.callout.anti .callout-label { color:#8a251a; }
.callout-wedge { font-family:"DM Serif Display", serif; font-size:18px; color:#1a2332; margin-bottom:8px; }
.callout-wedge .pos { color:#6b7884; font-size:13px; font-family:"DM Sans", sans-serif; }
.callout-wedge .intensity-tag { display:inline-block; font-family:"DM Sans", sans-serif; font-size:10px; font-weight:700; letter-spacing:1px; text-transform:uppercase; padding:2px 7px; border-radius:3px; margin-left:6px; vertical-align:middle; color:white; }
.callout.motivator .intensity-tag { background:#27ae60; }
.callout.anti .intensity-tag { background:#c0392b; }
.callout-body { font-size:13px; color:#3a4652; }
.alignment-grid { display:grid; grid-template-columns:1fr 1fr 1fr 1fr; gap:10px; margin-bottom:20px; }
.bucket { padding:10px 14px; border-radius:4px; }
.bucket-label { font-size:10px; font-weight:700; letter-spacing:1.1px; text-transform:uppercase; margin-bottom:3px; }
.bucket-pill { display:inline-block; font-size:11px; font-weight:700; padding:4px 10px; border-radius:10px; margin-bottom:6px; letter-spacing:0.3px; }
.bucket-desc { font-size:10.5px; color:#5a6773; margin-bottom:8px; line-height:1.4; }
.bucket-empty { font-size:11px; color:#8a939c; font-style:italic; }
.bucket-items { display:flex; flex-direction:column; gap:5px; }
.l2-row { font-size:11.5px; color:#2c3e50; display:flex; align-items:baseline; gap:6px; flex-wrap:wrap; }
.z-chip { display:inline-block; font-family:"DM Sans", sans-serif; font-weight:700; font-size:10px; padding:2px 6px; border-radius:3px; min-width:38px; text-align:center; color:white; }
.z-chip.pos { background:#27ae60; }
.z-chip.neg { background:#c0392b; }
.map-block { background:#fafbfc; border:1px solid #e5e8ec; border-radius:6px; padding:16px; margin-bottom:16px; }
.map-block svg { width:100%; height:auto; display:block; }
.map-legend { margin-top:10px; padding:8px 12px; background:#f3f5f8; border-radius:4px; font-size:11px; color:#4a5662; line-height:1.5; }
.map-legend strong { color:#1a2332; }
.map-legend .border-swatch { display:inline-block; width:16px; height:10px; border-radius:3px; border-width:2.5px; border-style:solid; margin-right:2px; vertical-align:middle; background:#fafbfc; }
.map-legend .fill-swatch { display:inline-block; width:18px; height:12px; border-radius:3px; border:1px solid #d0d4da; margin-right:2px; vertical-align:middle; }
.wiring-panel { background:#fafbfc; border:1px solid #e5e8ec; border-radius:6px; padding:16px 20px; }
.wiring-title { font-family:"DM Serif Display", serif; font-size:15px; color:#1a2332; margin-bottom:10px; }
.wiring-grid { display:grid; grid-template-columns:280px 1fr; gap:24px; align-items:center; }
.disc-bars { display:flex; flex-direction:column; gap:6px; }
.disc-bar { display:flex; align-items:center; gap:10px; font-size:12px; }
.disc-label { font-weight:700; width:14px; text-align:right; color:#1a2332; }
.disc-track { flex:1; height:14px; background:#e5e8ec; border-radius:2px; overflow:hidden; }
.disc-fill { height:100%; }
.disc-val { width:30px; text-align:right; font-weight:700; color:#1a2332; }
.wiring-table { border-collapse:collapse; font-size:12px; }
.wiring-table th { text-align:left; font-weight:700; color:#6b7884; padding:4px 10px 4px 0; font-size:11px; letter-spacing:0.5px; text-transform:uppercase; }
.wiring-table td { padding:4px 14px 4px 0; color:#2c3e50; }
.wiring-note { margin-top:10px; font-size:11.5px; color:#6b7884; font-style:italic; }
'''

LEGEND_MAP = '''
<div class="map-legend">
    <strong>Pill fill = coaching bucket</strong> (exact Z|Algo printed inside each pill):
    <span class="fill-swatch" style="background:#2563eb;"></span> running naturally
    <span class="fill-swatch" style="background:#c0392b;"></span> aligned but not running
    <span class="fill-swatch" style="background:#22c55e;"></span> installed against the grain
    <span class="fill-swatch" style="background:#d4a84b;"></span> routines to install
    <span class="fill-swatch" style="background:#dbeafe;"></span> motivator zone, mid
    <span class="fill-swatch" style="background:#dcfce7;"></span> anti zone, mid
    <span class="fill-swatch" style="background:#f3f4f6;"></span> cross-wedge.
    Zones include the primary wedge plus the two adjacent wedges. <strong>Radius + pill size</strong> = polarity.
</div>
'''

def respondent_block(name, label, meta, callouts, alignment, stdmap, wiring):
    return f'''<div class="respondent-block">
<div class="section-title">{name} <span class="name">— {label}</span></div>
<div class="respondent-meta">{meta}</div>

{callouts}
{alignment}

<div class="map-block">{stdmap}</div>
{LEGEND_MAP}

{wiring}
</div>'''

HTML = f'''<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<title>Integrated Spike v4</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:ital,opsz,wght@0,9..146,100..900;1,9..146,100..900&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head><body>
<div class="container">
<div class="caveat"><strong>⚠ Speculative mapping.</strong> The L2 → DISC wedge assignments are our best judgment, not empirically validated. Treat patterns as hypotheses to verify. See <code>l2_wedge_map.xlsx</code> and <code>_archive/SPIKE_WHEEL_NOTES.md</code>.</div>

<h1>Motivators &amp; Anti-Motivators — integrated spike</h1>
<div class="subtitle">Per respondent: motivator/anti callouts → alignment buckets → Standard Map with bucket-colored borders → DISC wiring.</div>

{respondent_block('Bill George','Persuading Conductor',
    'Natural pos 11 · Adapted pos 11 · D 92 · I 72 · S 32 · C 14 · full 32-L2 data',
    bill_callouts, bill_align, bill_map, bill_wiring)}

{respondent_block('Jody Bender','Relating Supporter',
    'Natural pos 34 · Adapted pos 17 (Supporting Relater) · D 15 · I 66 · S 85 · C 58 · partial data (11 of 32)',
    jody_callouts, jody_align, jody_map, jody_wiring)}

</div></body></html>
'''

OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
OUT_HTML.write_text(HTML, encoding='utf-8')
print(f'Wrote: {OUT_HTML} ({OUT_HTML.stat().st_size} bytes)')
