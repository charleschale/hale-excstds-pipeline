"""Build the Excellence Standards Coaching Guide for Matthew Cohen.

Respondent: 20250721.matt@provablemarkets.com
Role: CEO & Co-Founder, Provable Markets (FinTech — securities lending platform)
Deliverable: Integrated Coaching Guide (Graphical HTML + PDF)

Run from repo root:
    python _pipeline/scripts/build_cohen_coaching.py
Outputs:
    _reports/cohen_coaching/Cohen_Matthew_coaching_guide.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

RESPONDENT_XLSX = ROOT / '_respondents' / '20250721.matt@provablemarkets.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'coaching_guide_TEMPLATE.html'
OUT_DIR = ROOT / '_reports' / 'cohen_coaching'
OUT = OUT_DIR / 'Cohen_Matthew_coaching_guide.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title,
                'z_algo': z_algo,
                'z_human': z_human,
                'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flags_lit = {}
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        flag_val = ws_flags.cell(2, col).value
        if flag_name and flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[str(q_num)] = answer

    return {
        'l1_data': l1_data,
        'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall,
        'z_human_overall': z_human_overall,
        'rf_num': rf_num,
        'questions_answered': questions_answered,
        'flags_lit': flags_lit,
        'non_scorable': non_scorable,
    }

# ============================================================================
# LOAD HISTOGRAM DATA FOR DISTRIBUTION CHARTS
# ============================================================================

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        rf_val = ws_zalgo.cell(row, 7).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

# ============================================================================
# BUILD DISTRIBUTION CHART TOKENS
# ============================================================================

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]

    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo = row['z_algo']
        z_human = row['z_human']
        sf = row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    # Chart 1 collapse
    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    # Chart 2: hide empty bins
    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    # Chart 3: flag bins — reversed axis (high flags LEFT, low flags RIGHT)
    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1
            continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1
                break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]; upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    n_flag = len(flag_counts)
    rev_flag_bin = n_flag - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# RESPONDENT DICT FOR MOTIVATORS_SECTION
# ============================================================================

def build_respondent_dict(respondent_data):
    # TTI Natural DISC: D=69, I=32, S=48, C=61 (Conducting Implementor, wedge #9)
    # TTI Adapted DISC: D=78, I=28, S=35, C=64 (same wedge; slightly more D, lower I/S)
    nat_pos = 9
    nat_label = 'Conducting Implementor'
    nat_intensity = (69 + 32) / 200.0  # 0.505

    adp_pos = 9
    adp_label = 'Conducting Implementor'
    adp_intensity = (78 + 28) / 200.0  # 0.530

    shift_note = 'Adapted shift: +9D, −4I, −13S, +3C. Environment pulling even harder toward command-and-control; withdrawing steadiness and warmth.'

    return {
        'name': 'Matthew Cohen',
        'first_name': 'Matthew',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [69, 32, 48, 61],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }

# ============================================================================
# NARRATIVE SECTIONS — hand-authored prose for Cohen
# ============================================================================

SIGNATURE_PATTERN = (
    "<p><strong>Leadership is making people and situations better.</strong> That is the scoreboard."
    " Not the cap table, not the valuation, not the title. A leader who delivers results while"
    " shrinking the people around them has not led — they have performed. A leader who succeeds"
    " at the office while their family meets the residue of the build has paid a cost the W&#8209;2"
    " does not measure. This guide is written against that frame.</p>"
    "<p>Your career arc reads a disciplined, fact-before-feeling operator who stepped from MD-level"
    " trading seats at Nomura and Jefferies into founding and running Provable Markets — and moved"
    " the company from Seed to Series A Extension through a co-founder separation the operating"
    " standard required. The fingerprint card below carries the numbers; what matters at this"
    " altitude is the pattern they describe — a data-driven, disciplined builder who decides on"
    " evidence, holds the plan, and will not be pulled off it. Four genuine strengths carry this"
    " file: requiring excellence over popularity, deliberate urgency, clean organizational"
    " decision-making, and real investment in the people who have already delivered. That is the"
    " operating system that built Provable to where it is.</p>"
    "<p><strong>And — this is the frame that has to be honest — a pilot can be great at"
    " everything except landing the plane and still be a bad pilot.</strong> Landing is the move"
    " that defines the job. At the CEO altitude, every L1 area in the Excellence Standards is a"
    " landing. Your file shows four areas where the current expression sits more than a standard"
    " deviation below the mean: <em>conducting (the in-the-room conductor lane), replacing yourself"
    " (the ability to disappear), facilitative mindset (dialogue over direction), and pushing"
    " extreme accountability (standards as high as what you tolerate)</em>. Four landings not yet"
    " being made. The strengths are real — they are why Provable exists. The four landings are the"
    " specific work, and the strengths in the rest of the file do not substitute for them; they"
    " camouflage them until the moment the missing competence is exactly what the situation"
    " requires.</p>"
    "<p>Read together with the wiring, the four landings name a coaching shape we have seen many"
    " times. The Implementor wiring (drive plus precision, warmth and steadiness below the midline)"
    " is a high-performance engine. Uncorrected — meaning paired with the four 1&#963; landings above and"
    " with belief that is extended to those who have already delivered but not yet to those who"
    " have not — it resolves to a specific outcome: <em>a leader who, often without intending it,"
    " keeps the people around them one rung below full capability so the leader stays"
    " indispensable.</em> The prize is indispensability. The cost is structural — nothing scales,"
    " the company hits a ceiling because no one has been built up high enough to take the next"
    " layer, and the leader exhausts themselves carrying it. The cost is also personal — the"
    " people who care most about you (partner, children, parents, you yourself) meet the residue"
    " of the build rather than the person. <strong>You may win; you win at great cost.</strong>"
    " The standards in this guide exist because the alternative — leadership that makes people,"
    " situations, <em>and you yourself</em> better than when you started — is not automatic, not"
    " natural for this wiring, and entirely learnable.</p>"
    "<p><strong>The single highest-leverage thing to work on is in-the-room initiating accountability.</strong></p>"
    "<p>There are two accountability lanes. <em>Pushing</em> accountability is the sheriff lane —"
    " enforcing standards after the fact, driving consequences, holding people to what was"
    " committed. That machinery is strong in your file. <em>Initiating</em> accountability is the"
    " conductor lane — creating the conditions where the team brings accountability <em>up</em>"
    " to you before it has to be enforced <em>down</em> from you. The structural pieces are built:"
    " proactivity, clarity of ownership, willingness to drive consequences. The in-the-room"
    " installation is where the work is — conductor-over-lead-guitarist, dialogue-over-direction,"
    " and the ability to disappear and have the organization carry the plan. The engine that"
    " built Provable was you running point. The next altitude requires the routines that let"
    " the team run point while you conduct — because especially the best get better, and you"
    " are already one of the best.</p>"
    "<p class='sig-pattern-seam'>A <strong>second challenge</strong> is lit that does not derive"
    " causally from the first, but compounds with it. A small cluster of findings points at one"
    " operational tell — <em>standing gripes held at the same time as reported satisfaction</em>,"
    " with an instinct to upgrade that may not yet have been converted into action. The anchor"
    " admired leaders run is the opposite: <em>easy to please, hard to satisfy</em>, and <em>no"
    " standing gripes</em> — because a gripe is a standard that has not yet been converted into"
    " a decision. The specific work is that conversion: gripe → decision → routine.</p>"
)

FINGERPRINT_NARRATIVE = (
    "<p>Your wiring is <strong>Conducting Implementor</strong> — DISC Natural D=69, I=32, S=48, C=61."
    " This is an unusual profile at the CEO altitude: <em>D (drive) plus C (precision)</em> as the"
    " twin engines, with an S that is neither low (fast) nor high (steady-only), and the I sitting"
    " below the midline. The signature is structured execution — build the plan, hold the plan, drive"
    " the plan, move fast but always with a map. TTI corroborates exactly where you would expect it to:"
    " <strong>Objective #1 Primary (83)</strong> · <strong>Structured #2 Primary (68)</strong>"
    " · <strong>Resourceful #3 Primary (58)</strong> · <strong>Commanding #4 Primary (49)</strong>"
    " · <strong>Intentional #5 Situational (47)</strong>. Four Primary driving forces all pointing"
    " at the same operating signature: the data-driven, disciplined builder who knows exactly why each"
    " move matters and will not be pulled off-plan.</p>"
    "<p>The Natural→Adapted shift tells you something important about the current seat: D moves"
    " +9 (69→78), I drops −4, S drops −13 (48→35), C holds roughly flat at 64. The current environment"
    " is dialing <em>up</em> the commanding posture and dialing <em>down</em> steadiness and warmth."
    " For anyone around you this reads the same signal twice: the drive is real, the discipline is"
    " real, and the warmth / pace-matching dimensions are not where you spend your energy. That is an"
    " asset for clarity — which is why your Clarity of Accountability L2 is +1.57 — and precisely why"
    " the growth frontier below (Facilitative Mindset and in-the-room dialogue) has to be installed"
    " as a deliberate routine, not waited on as personality development.</p>"
    "<p><strong>A word on the pattern itself — the Implementor wiring and its known risk.</strong>"
    " Conducting Implementor is a recognized, high-performance wiring. In our population these"
    " profiles are often the strongest <em>doers</em> in the room: they ship, they hold the plan,"
    " they make the hard call. It is also a wiring that carries a specific, well-documented risk"
    " — the same engine that produces the execution can alienate the team it is driving.  "
    " D (drive) plus C (precision) read as standards-pressure; low S reads as no pace-match;"
    " I below the midline reads as limited warmth / third-party advocacy.  Stacked, the signal"
    " teammates receive is <em>the plan matters more than the people on it</em> — even when"
    " that is not what the operator believes. Internally we name this the"
    " <em>killer-doer</em> pattern for a reason: the instrument is unusually good at catching it,"
    " and the exposure points are consistent across Implementor-wired CEOs we have coached."
    " Your file corroborates the familiar seams:"
    " <strong>HoldsAvgDownChain</strong> (a compromise held at the same time the standard is"
    " being carried up the chain), <strong>Condit Belief Sev</strong> (belief extended to those"
    " who have already delivered, not yet to those who have not), <strong>Initiating Accountability"
    " Sev</strong> (the in-the-room conductor lane where teammates bring accountability up to you"
    " before it has to be enforced down from you), and <strong>Satisfied with Gripes | Low</strong>"
    " (standing gripes held at the same time as reported satisfaction). These are the classic"
    " Implementor teammate-alienation signals — the specific places where the profile's strengths"
    " tip into its risks. Naming the pattern is not a verdict on you; it is the operating manual"
    " for the wiring. Every routine in the <em>What to Work On</em> section below is a counter-routine"
    " to this specific risk — the moves that keep the team <em>with</em> you while the plan stays"
    " on the rails.</p>"
)

DRIVING_FORCES_PRIMARY_HTML = (
    "<div style='font-size:12px; color:#333; line-height:1.6;'>"
    "<strong>Objective</strong> (rank 1, 83 · Primary) · <strong>Structured</strong> (rank 2, 68 · Primary)"
    " · <strong>Resourceful</strong> (rank 3, 58 · Primary) · <strong>Commanding</strong> (rank 4, 49 · Primary)"
    "<p style='margin-top:6px; font-size:11px; color:#5a6773;'>Data-driven decisions, structured"
    " execution, practical resourcefulness, and the need to lead from the front. Four Primary forces"
    " pointing at the same engine — the disciplined builder who decides on evidence, holds the plan,"
    " and moves against obstacles without hand-holding.</p></div>"
)

DRIVING_FORCES_INDIFFERENT_HTML = (
    "<div style='font-size:12px; color:#333; line-height:1.6;'>"
    "<strong>Harmonious</strong> (rank 12, score 0 · Indifferent) · <strong>Selfless</strong> (rank 11, score 24 · Indifferent)"
    " · <strong>Receptive</strong> (rank 10, score 25 · Indifferent) · <strong>Altruistic</strong> (rank 9, score 31 · Indifferent)"
    "<p style='margin-top:6px; font-size:11px; color:#5a6773;'>Four consecutive Indifferent forces at"
    " the bottom of the wheel — Harmonious scores <em>zero</em>. Aesthetic harmony is not a default"
    " channel; <em>what I owe this person for their growth</em> is not a default channel; openness to"
    " the unconventional is not a default channel. Each of these becomes available, but as a"
    " <em>routine installed against the grain</em>, not a preference followed.</p></div>"
)

DRIVING_FORCES_IMPLICATIONS_HTML = (
    "<p><strong>What this composite predicts.</strong> A CEO who will drive a hard plan with rigor,"
    " make fact-based calls under pressure, and carry the organization through a ten-x build by"
    " sheer capability and tempo. The Harmonious=0 / Selfless=24 / Receptive=25 / Altruistic=31"
    " cluster predicts exactly the coaching frontier the ExcStds file reads: <em>the mechanics of"
    " warmth, deference, and genuine belief-extension to the team</em> are not self-generating. This"
    " is not a call to change the wiring — the wiring is what built Provable Markets. It is a call"
    " to install specific routines (third-party praise, question-first dialogue, develop-depth below)"
    " against the grain, because those are the routines that let the next layer of the team show up"
    " at full strength when you are not in the room.</p>"
)

WIRING_FIT_ITEMS = (
    "<p><strong style=\"color:#2563eb;\">Not Pleasing (+1.75) — rare top-decile strength.</strong>"
    " L2 Discomfort For Self +1.39, L2 Cares About Others Not Their Approval +1.16, L2 Discomfort For Team +1.07."
    " You are wired to do the right thing rather than the popular thing — the core Admired anchor"
    " for founder-CEOs. Teach this to your senior team. Most operators do not get this right at"
    " scale.</p>"
    "<p><strong style=\"color:#2563eb;\">Deliberate Urgency (+1.60) — confirmed strength.</strong>"
    " L2 Extreme Proactivity +1.97, L2 Action Over Inaction +0.11, L2 Fire In Belly +0.08. Tempo"
    " is wired in. The proactive belief in people is the one L2 that sits below the L1 average —"
    " the pattern is <em>you</em> being proactive rather than <em>you trusting the team</em> to be"
    " proactive. That is the seam that connects to the Initiating Accountability flag.</p>"
    "<p><strong style=\"color:#2563eb;\">Org Decision Making (+1.20) — strong, with one watch-item.</strong>"
    " L2 Clarity Of Accountability +1.57, L2 Simplification Methods +1.22. Decisions get made on"
    " facts, clarity about ownership is strong. L2 Respects Collective Wisdom −1.09 is the watch-item:"
    " you decide well, the room's wisdom may not be getting invited in before you resolve.</p>"
    "<p><strong style=\"color:#b8862e;\">Conducting &amp; Outvoted (&minus;1.38) — the deepest"
    " in-the-room frontier.</strong> L2 Conductor &gt; Lead Guitarist −0.38, L2 Empower Team Authority"
    " −0.08. The pattern says the lead-guitar pull (you playing the solo) is winning over the conductor"
    " posture (drawing the team forward). Combined with D=69 + Commanding #4 Primary, this is the"
    " wiring tell that explains the Initiating Accountability flag — you run point by instinct.</p>"
    "<p><strong style=\"color:#b8862e;\">Facilitative Mindset (&minus;0.87) — the dialogue frontier.</strong>"
    " L2 Dialogue Vs. Direction −1.07, L2 Power &amp; Status Management −1.13, L2 Sublimating Ego"
    " +0.14. You direct more than you dialogue, and status gradient closes the room. The Sublimating"
    " Ego L2 is actually positive — you are not ego-defended, you just fill the room with"
    " direction by default. The practice frontier is a specific routine (question-first), not a"
    " personality change.</p>"
    "<p><strong style=\"color:#b8862e;\">Replacing Self (&minus;0.91) — the soil for scale.</strong>"
    " L2 Ability To Disappear −1.00. The Q9917 self-assessment names delegation directly as a"
    " weakness, and the instrument matches. At Provable's current altitude this is a critical seam:"
    " the next $75M→$200M+ run requires an organization that executes at pace when you are not the"
    " node it routes through.</p>"
)

# ============================================================================
# EXCSTDS SCORECARD
# ============================================================================

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Handling Daily Difficulties With Dignity',
            'Developmental Mindset',
            'Developmental Discipline',
        ]),
        (2, 'RISKING BY FACILITATIVE MINDSET', [
            'Dialogue Vs. Direction',
            'Power & Status Management',
            'Sublimating Ego',
        ]),
        (3, 'CONDUCTING & OUTVOTED', [
            'Conductor > Lead Guitarist',
            'Empower Team Authority',
        ]),
        (4, 'RISKING BY REPLACING SELF', [
            'Ability To Disappear',
            'Urgency Down Chain Of Command',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Basic Machinery Of Accountability',
            'Drives Accountability',
            'Stds = What Tolerate',
        ]),
        (7, 'NOT PLEASING', [
            'Cares About Others Not Their Approval',
            'Discomfort For Self',
            'Discomfort For Team',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Simplification Methods',
            'Clarity Of Accountability',
            'Respects Collective Wisdom',
            'Facts Over Feelings',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Extreme Proactivity',
            'Proactive Belief In People',
            'Action Over Inaction',
        ]),
    ]

    labels, scores, is_l1, missing_l2 = [], [], [], []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = respondent_data['z_algo_overall']
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'ZALGO_OVERALL_NUM': f"{overall_z:.4f}",
        'COHORT_AVG_NUM': f"{cohort_avg:.3f}",
    }

# ============================================================================
# IMPACT ITEMS — 4 flag-driven + 9 algorithm-ranked per-answer
# ============================================================================

def build_impact_items_html():
    flag_subheader = (
        '<div class="practice-subsection-hdr">'
        '<div class="practice-subsection-hdr-title">Flag-Driven Items</div>'
        '<div class="practice-subsection-hdr-blurb">Lit flags from the cohort research &mdash;'
        ' pattern-level effectiveness levers that sit across multiple L1s, which is why they are'
        ' flags and not single-L2 routines. Addressed before per-answer lifts.</div>'
        '</div>'
    )

    flag_cards = []

    # FLAG 1 — Initiating Accountability Sev (master finding)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Initiating Accountability Sev</div>'
        '<div class="practice-qref">Master finding &middot; the conductor lane, not the sheriff lane &middot; severity Severe</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>The thesis.</strong> Pushing Accountability and Initiating Accountability are'
        ' mirror images. <em>Pushing</em> is the sheriff lane &mdash; enforcing standards after the'
        ' fact, holding people to commitments, driving consequences. <em>Initiating</em> is the'
        ' conductor lane &mdash; creating the conditions in which the team brings accountability up'
        ' to the leader before it has to be enforced down from the leader. Most strong operators have'
        ' the sheriff lane; the conductor lane is scarcer and is the one that compounds at scale.</p>'
        '<p><strong>Why this flag is unusually reliable.</strong> It does not live in a single L2 or'
        ' a single L1. It samples across Facilitative Mindset (how dialogue runs in the room),'
        ' Replacing Self (whether authority is actually transferred down), Conducting &amp; Outvoted'
        ' (whether the team can run when the leader is outvoted), and Org Decision Making. Because the'
        ' basket samples four different lanes, no single-lane explanation makes the pattern go away.</p>'
        '<p><strong>Where your file lands.</strong> The <em>structural</em> initiating is there &mdash;'
        ' Extreme Proactivity +1.97, Clarity of Accountability +1.57, Drives Accountability +0.88. The'
        ' machinery is built. Where the flag trips at <em>Sev</em> is the <em>in-the-room</em>'
        ' initiating: Conductor &gt; Lead Guitarist &minus;0.38, Empower Team Authority &minus;0.08,'
        ' Dialogue vs. Direction &minus;1.07, Power &amp; Status &minus;1.13, Ability To Disappear'
        ' &minus;1.00. The pattern is honest: <em>you</em> drive the accountability, and the team'
        ' performs accountability for you rather than owning it themselves. Your Q9917 self-assessment'
        ' names delegation directly as a weakness &mdash; the instrument corroborates exactly.</p>'
        '</div>'
        '<div class="practice-fuel">Gateway routine: answer questions with questions. Run one staff'
        ' meeting this month where every statement from you is reframed as a question. It is'
        ' uncomfortable and eye-opening, and the room will show you what it actually thinks. The'
        ' dialogue, power/status, and conductor behaviors all shift together when the question-first'
        ' routine gets installed; the Impact-ranked items ahead are the supporting routines that flow'
        ' from it.</div>'
        '</div>'
    )

    # FLAG 2 — Condit Belief Sev
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Condit Belief Sev</div>'
        '<div class="practice-qref">Conditional belief-in-others &middot; belief earned by proof, not extended in advance &middot; severity Severe</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> Conditional Belief watches whether a leader'
        ' extends belief <em>before</em> team members have proven themselves, or only <em>after</em>.'
        ' Admired leaders extend belief early &mdash; they back the person before the evidence is in,'
        ' because the belief itself is often what enables the evidence. Leaders who fire the Condit'
        ' Belief flag at <em>Sev</em> tend to withhold belief until the proof lands, which keeps the'
        ' team in a permanent audition posture and caps the scale of the organization.</p>'
        '<p><strong>Where your file lands.</strong> TTI Altruistic rank 9 Indifferent (score 31),'
        ' Harmonious rank 12 (score 0), Selfless rank 11 Indifferent (24), Receptive rank 10'
        ' Indifferent (25) is the wiring tell &mdash; the machinery that extends goodwill without'
        ' requiring it to be earned is simply not a Primary driver. Investing in Others L1 is a net'
        ' positive (+0.83) and Demonstrating Genuine Fanness L2 is +4.38 (a top-decile strength on the'
        ' surface), but Developmental Discipline L2 is &minus;1.26 and Reciprocal Followership L2 is'
        ' &minus;0.23. The pattern: you genuinely admire team members who deliver, and you struggle'
        ' to invest developmental capital in the ones who have not yet. That is a conditional-belief'
        ' posture.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: for each direct report, write one sentence naming a'
        ' capability you would like them to grow into that they have not yet demonstrated. Tell them.'
        ' Budget a specific coaching investment (time, assignment, exposure) against it this quarter.'
        ' Belief becomes observable when it is named and resourced before the proof.</div>'
        '</div>'
    )

    # FLAG 3 — HoldsAvgDownChain
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">HoldsAvgDownChain</div>'
        '<div class="practice-qref">Crew, not basketball &middot; depth-of-organization pattern</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p>The flag reads whether the standard held at the exec table is the same standard enforced'
        ' one, two, and three levels down. Your file pattern: the bottom of the organization would'
        ' improve if traded (Impact Q33, Q34, Q82 all cluster here), and depth below the exec layer'
        ' is not yet running at the same tempo as the top.</p>'
        '<p><strong>The frame is crew, not basketball.</strong> Basketball is about trading for'
        ' better talent at the top of the roster. Business is more like crew: the boat moves at the'
        ' pace of the slowest rower. A leader can have a strong immediate team and still be carrying'
        ' average seats two and three levels down that keep the whole organization at the pace of its'
        ' slowest layer. Admired and successful leaders continuously upgrade at depth; unsuccessful'
        ' leaders tolerate averageness at depth and wait for culture to transform on its own.</p>'
        '<p><strong>Why this compounds with the master finding.</strong> If the layer two levels down'
        ' carries average seats, your exec team cannot push decisions down and have them held &mdash;'
        ' some of the accountability that should sit at their layer has to keep routing up to you,'
        ' because the layer below cannot yet hold it. Fix the depth and Initiating Accountability'
        ' becomes fixable. Leave it and the dialogue routines alone will not be enough.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: pick two roles two levels down where, if you were hiring'
        ' today, you would not hire the current incumbent. Name them out loud to yourself. Set a date'
        ' by which the seat either clears the bar or is in a transition plan. Then repeat at the next'
        ' level down. The flag goes quiet when the leader stops finding seats he would not hire for today.</div>'
        '</div>'
    )

    # FLAG 4 — Satisfied with Gripes | Low
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Satisfied with Gripes | Low</div>'
        '<div class="practice-qref">Ratio flag &middot; easy-to-please / hard-to-satisfy anchor &middot; severity Low</div>'
        '</div></div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> Satisfied with Gripes is a <em>ratio</em> flag.'
        ' It looks at two self-reports side by side: how satisfied the leader is with the team, and'
        ' how many gripes the leader is holding about the team. The flag trips when the leader reports'
        ' <strong>both</strong> &mdash; some level of satisfaction and a working list of gripes at the'
        ' same time. That combination is the tell.</p>'
        '<p><strong>The admired-leader anchor: easy to please, hard to satisfy.</strong> Admired'
        ' leaders are <em>easy to please</em> &mdash; quick to notice when a team member does good'
        ' work, quick to say so. They are <em>hard to satisfy</em> &mdash; never fully satisfied,'
        ' because the team should always be developing further than today. And they run <em>no standing'
        ' gripes</em> &mdash; because a gripe is a standard the leader has not yet converted into a'
        ' decision, and admired leaders move gripes to decisions the week they surface.</p>'
        '<p><strong>Where your file lands.</strong> The flag trips at <em>Low</em> severity &mdash; the'
        ' mildest level &mdash; which means the ratio is only mildly off. But it is not inert,'
        ' especially in combination: HoldsAvgDownChain is lit and Condit Belief is at Sev, and the'
        ' three flags share a root &mdash; tolerating what you have already noticed is not yet good'
        ' enough. Your Q9917 self-assessment names temper, patience, and delegation as personal'
        ' weaknesses; those three often show up as standing gripes about the team (they are slow, they'
        ' require too much handholding, they do not execute with the fidelity you expect). The seam'
        ' is honest: move the gripes to decisions.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: two rules. (1) No standing gripes. The week a gripe'
        ' surfaces, name the standard the gripe is actually about and move it to a decision &mdash;'
        ' a conversation, a role clarification, or a transition plan. (2) Stay hard to satisfy. After'
        ' a strong quarter, ask what the team could be doing a year from now that it cannot do today,'
        ' and start the development routine that gets there. Satisfied is the word that closes the'
        ' development loop; avoid it.</div>'
        '</div>'
    )

    # Per-answer subsection
    peranswer_subheader = (
        '<div style="height:22px;"></div>'
        '<div class="practice-subtitle">Nine standards where deliberate practice will produce the largest lift.</div>'
    )

    # 9 impact cards from ImpactTop10 (ranks 1-9)
    impact_cards = []

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Stay in your lane &mdash; mission over moral posturing</div>'
        '<div class="practice-qref">Q119 &middot; Organizational Decision Making (Impact Rank 1)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#d4a84b; color:#7f1d1d; border-color:#7f1d1d;" data-l2="8.1">8.1 Simplification Methods</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward FALSE on a <em>1-TRUE &lt;&gt; 5-FALSE</em> scale &mdash; suggesting you'
        ' believe leaders <em>can</em> responsibly take public stances on issues unconnected to the'
        ' core mission. The research is specific: leaders and organizations that speak out publicly on'
        ' political or ideological issues outside the core mission become lightning rods &mdash; one'
        ' group wants more, one wants different, and the collaborative spirit inside the team erodes.'
        ' For a founder-CEO at a FinTech rebuilding trust with regulated counterparties, the expected'
        ' value of mission-drift is almost always negative.</p>'
        '<p><strong>Target behavior.</strong> Keep the organization focused inward on mission. Decline'
        ' public positions on debates unrelated to the core work. When pressure arrives from inside or'
        ' outside to take a stance, answer by naming the mission and the promise you owe customers.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: write a one-sentence mission-lane test. If a proposed'
        ' public statement does not serve that sentence, it does not go out under the company name.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Fight to keep everyone &mdash; or move them out</div>'
        '<div class="practice-qref">Q34 &middot; Risking by Pushing Extreme Accountability (Impact Rank 2)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#7f1d1d; border-color:#7f1d1d;" data-l2="5.8">5.8 Stds = What Tolerate</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE &mdash; there are people on your team you would not fight to keep. The'
        ' research frame: your reaction to someone leaving tells you whether they should have already'
        ' left. If your first reaction is relief rather than concern, you have been tolerating'
        ' performance or behavior that does not meet your standards. Every day you keep someone you'
        ' would not fight to keep, you are establishing mediocrity as an acceptable standard for'
        ' everyone. Your strongest performers see who you tolerate and calibrate their standards'
        ' accordingly.</p>'
        '<p>This item is the direct per-answer expression of HoldsAvgDownChain + Satisfied with Gripes.'
        ' It is the specific lift that closes the two flags simultaneously.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: each quarter, for each direct and skip-level seat,'
        ' privately answer: <em>would I fight to keep this person?</em> If the answer is no, there is'
        ' a conversation or a transition that is overdue.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Crew, not basketball &mdash; upgrade through development</div>'
        '<div class="practice-qref">Q33 &middot; Risking by Pushing Extreme Accountability (Impact Rank 3)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#7f1d1d; border-color:#7f1d1d;" data-l2="5.1">5.1 Basic Machinery Of Accountability</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE to &ldquo;trade the bottom 10% and the organization improves.&rdquo; The'
        ' research reframes the question: the goal is not to have no bottom performers, it is to have'
        ' no one satisfied with their current performance. The commitment to improvement is the'
        ' standard, not any specific performance level. A crew approach (everyone committed to getting'
        ' better daily) compounds; a basketball approach (trade for better talent) plateaus because it'
        ' never builds depth.</p>'
        '<p>For a Conducting Implementor wiring with Objective #1 and Structured #2 Primary, the'
        ' natural instinct is the roster move. The practice is building the development machinery so'
        ' the current team raises itself &mdash; which is also how you grow beyond 40% top-20-client'
        ' adoption without blowing out hiring cost.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: two questions for every current employee you are'
        ' considering replacing. (1) Has this person had a real coaching investment against the'
        ' specific behavior I want to see? (2) Is the gap a mindset gap (fix via coaching) or a ceiling'
        ' gap (fix via role change)? Most of the time the answer to #1 is no.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Urgency lives below you, not in you</div>'
        '<div class="practice-qref">Q81 &middot; Risking by Replacing Self (Impact Rank 4)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#d4a84b; color:#7f1d1d; border-color:#7f1d1d;" data-l2="4.3">4.3 Urgency Down Chain Of Command</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward FALSE &mdash; acknowledging that you <em>are</em> the one creating'
        ' urgency in departments, and areas of the business are dependent on your drive. This is the'
        ' literal behavioral tell on the Replacing Self and Initiating Accountability frontiers.'
        ' Urgency that routes through the CEO is a dependency; urgency that lives intrinsically in'
        ' each department is a scalable operating system.</p>'
        '<p>For a Deliberate Urgency score of +1.60 paired with Urgency Down Chain of Command at'
        ' &minus;0.06, the diagnosis is clean: your tempo is real, the cascade is not yet installed.'
        ' Hire and develop lieutenants who carry their own fire-in-belly; measure them on urgency-without-pressure.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: for each of your direct reports, write one sentence'
        ' naming the last time <em>they</em> raised urgency to you (rather than the other way).'
        ' If you cannot write it, that is the seat where the urgency cascade is broken.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Build the company so you can disappear</div>'
        '<div class="practice-qref">Q2 &middot; Risking by Replacing Self (Impact Rank 5)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#c0392b; color:#ffffff; border-color:#c0392b;" data-l2="4.1">4.1 Ability To Disappear</div></div>'
        '<div class="practice-body">'
        '<p>You answered at the most-like-me end of &ldquo;mission critical enough that I can&rsquo;t'
        ' easily go on vacation.&rdquo; For a founder-CEO at a FinTech Series A Extension, that is'
        ' understandable &mdash; you <em>have</em> been the node. The research frame: the measure of'
        ' your success is not whether the company can run without you today, it is whether you are'
        ' building it to run without you a year from now. Ability to Disappear L2 at &minus;1.00 is the'
        ' literal translation of this answer into a score.</p>'
        '<p>Your Q9917 self-assessment names delegation as a specific weakness. The master finding on'
        ' Initiating Accountability is the operating expression of the same pattern. The lift is not'
        ' delegate-more-in-general; it is build-an-org-that-executes-at-pace-without-you, one routine at a time.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: block a one-week stretch six months out where you are'
        ' entirely unreachable. Announce it to the team now. Spend the six months building every'
        ' missing piece that surfaces when you imagine that week.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">6</span><div style="flex:1;">'
        '<div class="practice-item-title">No irreplaceable seats &mdash; character over knowledge</div>'
        '<div class="practice-qref">Q82 &middot; Risking by Pushing Extreme Accountability (Impact Rank 6)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#f3f4f6; color:#7f1d1d; border-color:#7f1d1d;" data-l2="5.8">5.8 Stds = What Tolerate</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE &mdash; there are people you would replace except their experience and'
        ' knowledge are too critical to risk it. The research is specific: if someone has incompetence'
        ' or judgment issues, holding them in a critical role amplifies the damage across every'
        ' decision they make. &ldquo;Too critical to replace&rdquo; is the rationalization that keeps'
        ' the pattern alive &mdash; in a startup, it usually means the seat has not been documented'
        ' and the replacement has not been developed.</p>'
        '<p>Given Provable is moving from the Series A Extension toward a Series B, the seats that'
        ' felt irreplaceable at Seed almost never are at the next stage. The lift is making the'
        ' critical-knowledge claim explicit, scoping the risk honestly, and starting the succession'
        ' work now rather than after the performance call becomes unavoidable.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: name one &ldquo;irreplaceable&rdquo; person on your team.'
        ' Write down exactly what would break if they left tomorrow. Start documenting or distributing'
        ' each item on the list this quarter. Irreplaceability dissolves under light.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">7</span><div style="flex:1;">'
        '<div class="practice-item-title">Small changes, not big ones &mdash; for habits</div>'
        '<div class="practice-qref">Q95 &middot; Risking by Investing in Others (Impact Rank 7)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#c0392b; color:#ffffff; border-color:#c0392b;" data-l2="1.6">1.6 Developmental Discipline</div></div>'
        '<div class="practice-body">'
        '<p>You leaned toward highlighting <em>big</em> changes when coaching people to improve. The'
        ' research is specific: stretch goals work for teams, but they <em>fail</em> for personal'
        ' habit change. Audacious habit goals set people up for the first-setback failure that'
        ' reinforces the old behavior. A D=69, Commanding-Primary wiring will instinctively reach for'
        ' the big lever; the coaching craft is naming the small lever.</p>'
        '<p>Developmental Discipline L2 at &minus;1.26 is the literal score on this pattern. Apply it'
        ' to your own Impact work too: not &ldquo;become a fundamentally more facilitative CEO&rdquo;'
        ' &mdash; but &ldquo;one question-first staff meeting a month&rdquo; or &ldquo;one third-party'
        ' compliment per direct report per week.&rdquo;</p>'
        '</div>'
        '<div class="practice-fuel">The reframe: end-state ambition stays large; weekly targets get'
        ' small enough that the person keeps succeeding. Log the wins. The compounding does the work.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">8</span><div style="flex:1;">'
        '<div class="practice-item-title">Never send messages when upset</div>'
        '<div class="practice-qref">Q44 &middot; Personal Reliability (Impact Rank 8)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#c0392b; color:#ffffff; border-color:#c0392b;" data-l2="6.1">6.1 Extreme Ownership</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward FALSE &mdash; acknowledging that upset communications do sometimes go'
        ' out. Your Q9917 self-assessment names <em>temper</em> directly as a weakness, and the'
        ' instrument corroborates. There is a big difference between reacting (emotional, reflexive,'
        ' fast) and responding (thoughtful, deliberate, slower). As a founder-CEO in a high-stakes'
        ' regulated environment, written communications are permanent artifacts &mdash; they get'
        ' screenshotted, forwarded, cited back. The downside asymmetry is severe.</p>'
        '<p>The practice frontier for this one is not self-insight (you already have it) &mdash; it is'
        ' the circuit-breaker routine.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: one hard rule. <em>No send button when upset.</em> Draft'
        ' the message, save it, walk away. If it still reads right an hour later, send it. If it does'
        ' not, rewrite or delete. Most of them will get rewritten. The ones that survive are the ones'
        ' worth sending.</div>'
        '</div>'
    )

    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">9</span><div style="flex:1;">'
        '<div class="practice-item-title">Do not task below your direct reports</div>'
        '<div class="practice-qref">Q7 &middot; Risking by Pushing Extreme Accountability (Impact Rank 9)</div>'
        '</div><div class="practice-l2-tag" title="L2" style="background:#d4a84b; color:#7f1d1d; border-color:#7f1d1d;" data-l2="3.2">3.2 Empower Team Authority</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward FALSE &mdash; acknowledging that tasking below your direct reports'
        ' does happen. This is the literal behavioral tell on the Conducting &amp; Outvoted frontier'
        ' (L1 &minus;1.38). When a CEO tasks two layers down, the direct report loses authority, the'
        ' person below gets crossed signals, and the org chart stops functioning as a decision-rights'
        ' map. The pattern typically starts from good intent &mdash; adding value, moving fast &mdash;'
        ' and ends in an org where nobody below the CEO fully owns their area.</p>'
        '<p>For your file specifically, this connects directly to the master finding. The routine'
        ' that rebuilds the conductor lane is not &ldquo;delegate more&rdquo; &mdash; it is'
        ' <em>task only your direct reports, ever</em>, and let them task their people.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: one-month experiment. Any time you catch yourself about'
        ' to direct someone below your direct reports, stop, and route it through the direct report'
        ' instead. At the end of the month, review which tasks slowed down, which improved, and what'
        ' the direct report learned about their own area.</div>'
        '</div>'
    )

    return flag_subheader + '\n'.join(flag_cards) + peranswer_subheader + '\n'.join(impact_cards)

# ============================================================================
# TEACH ITEMS — 10 standards he's already effective at
# ============================================================================

def build_teach_items_html():
    teach_cards = []

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Third-party compliments over direct praise</div>'
        '<div class="practice-qref">Q132 &middot; Risking by Investing in Others (Teach Rank 1)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.4">1.4 Demonstrating Genuine Fanness</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward third-party compliments &mdash; the admired pattern. This is an'
        ' unusual strength given your Altruistic rank 9 Indifferent wiring, and it corroborates'
        ' Demonstrating Genuine Fanness L2 at +4.38 (a top-decile strength). Teach this to your'
        ' senior team directly: the admiration they route <em>around</em> their people reads as'
        ' deeply sincere because it has no audience-management motive.</p>'
        '</div>'
        '<div class="practice-fuel">The zen-master level: do this for people two levels down, not just'
        ' your directs. The compound effect is huge and almost nobody does it.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Leadership admired for requiring excellence, not respect</div>'
        '<div class="practice-qref">Q131 &middot; Not Pleasing (Teach Rank 2)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="7.3">7.3 Discomfort For Self</div></div>'
        '<div class="practice-body">'
        '<p>You answered toward &ldquo;I require excellence.&rdquo; This is the admired CEO stance'
        ' and your file demonstrates it cleanly &mdash; Discomfort For Self +1.39, Discomfort For'
        ' Team +1.07, Cares About Others Not Their Approval +1.16. For a founder-CEO who removed a'
        ' co-founder when the operating standard was not being held, this is the credibility anchor.'
        ' Teach it to operators who confuse &ldquo;popular&rdquo; with &ldquo;effective.&rdquo;</p>'
        '</div>'
        '<div class="practice-fuel">Routine: tell the story of a specific decision you made that was'
        ' unpopular in the short term and right in the long term. Name the decision, the pushback,'
        ' and the outcome. The team needs to know what excellence looks like under pressure.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Proactive enough that you almost never get tasked</div>'
        '<div class="practice-qref">Q49 &middot; Deliberate Urgency (Teach Rank 3)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="9.2">9.2 Extreme Proactivity</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE. The score corroborates: Extreme Proactivity L2 +1.97, Deliberate'
        ' Urgency L1 +1.60. This is the top-of-the-file strength &mdash; the reason Provable got from'
        ' Seed to Series A Extension at $75MM while you were simultaneously running the co-founder'
        ' separation. Teach it as the CEO-minimum, not the CEO-max.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in every 1:1 with a direct report, ask one question:'
        ' <em>what are you working on that I would not yet know to ask about?</em> Proactivity you'
        ' can see is one thing; proactivity they lead with is the level-up.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Hands off until all-in &mdash; trust the team to execute</div>'
        '<div class="practice-qref">Q31 &middot; Risking by Pushing Extreme Accountability (Teach Rank 4)</div>'
        '</div><div class="practice-l2-tag" style="background:#dcfce7; color:#5a6773; border-color:#5a6773;" data-l2="5.3">5.3 Drives Accountability</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE &mdash; resist being involved in the operation until you absolutely'
        ' must be, then go all-in. This is the experienced operator&rsquo;s frame &mdash; most'
        ' founder-CEOs get it wrong in both directions (either always involved or always absent).'
        ' The admired pattern is the paired discipline: stay out by default, but be willing to put'
        ' hands all the way on when the moment requires it. Teach this to operators who either'
        ' micromanage or abdicate.</p>'
        '</div>'
        '<div class="practice-fuel">Zen-master level: when you go all-in, tell the team explicitly'
        ' &mdash; &ldquo;I am stepping in on this one, here is why, here is when I will step back'
        ' out.&rdquo; That way the hand-on is a tool, not a reflex.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Choose expected value, not perceived safety</div>'
        '<div class="practice-qref">Q68 &middot; Organizational Decision Making (Teach Rank 5)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="8.7">8.7 Facts Over Feelings</div></div>'
        '<div class="practice-body">'
        '<p>You chose 50% odds of 50% growth over 80% odds of 25% growth &mdash; the higher-expected-value'
        ' option. This is the admired decision. Objective rank 1 Primary (83) + Structured rank 2'
        ' Primary (68) is exactly the wiring that produces this answer instinctively. For a FinTech'
        ' CEO navigating a Series A Extension at $75MM, this is the habit that keeps the company'
        ' moving at the right pace even when the comfortable choice is smaller growth.</p>'
        '</div>'
        '<div class="practice-fuel">Teach it as the rule: for any material decision, write the expected'
        ' value math on one line before the argument starts. Math above, debate below.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">6</span><div style="flex:1;">'
        '<div class="practice-item-title">Difficult conversations, every day</div>'
        '<div class="practice-qref">Q83 &middot; Risking by Investing in Others (Teach Rank 6)</div>'
        '</div><div class="practice-l2-tag" style="background:#dcfce7; color:#5a6773; border-color:#5a6773;" data-l2="1.3">1.3 Handling Daily Difficulties With Dignity</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE &mdash; difficult conversations are a daily practice. This is a'
        ' founder-CEO hallmark that scales. The co-founder separation is the extreme case of this'
        ' skill; the everyday version is the five-minute truth-telling conversations you have across'
        ' the week. Handling Daily Difficulties With Dignity L2 +0.18 is positive, and the bigger'
        ' signal is you already do this at the threshold level.</p>'
        '</div>'
        '<div class="practice-fuel">Zen-master level: when the difficult conversation is not feeling'
        ' difficult to <em>you</em>, check whether you are thinking about how it lands for the'
        ' <em>other person</em>. Empathy is the second gear of this skill.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">7</span><div style="flex:1;">'
        '<div class="practice-item-title">One owner per outcome &mdash; no shared accountability</div>'
        '<div class="practice-qref">Q384 &middot; Organizational Decision Making (Teach Rank 7)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="8.2">8.2 Clarity Of Accountability</div></div>'
        '<div class="practice-body">'
        '<p>You answered at the admired end &mdash; results are worse when multiple people are'
        ' accountable to the same outcome. Clarity of Accountability L2 +1.57 corroborates. Teach'
        ' this as the RACI / RAPID discipline: collaborate widely, but name one person accountable'
        ' for every deliverable. Especially important at Series A&rarr;B scale where org-chart'
        ' ambiguity becomes expensive quickly.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in every project kickoff, the last question is'
        ' &ldquo;who is the single accountable person?&rdquo; Not a group. Not two co-leads. One name.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">8</span><div style="flex:1;">'
        '<div class="practice-item-title">Discomfort is your comfort zone</div>'
        '<div class="practice-qref">Q69 &middot; Not Pleasing (Teach Rank 8)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="7.3">7.3 Discomfort For Self</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE. Discomfort For Self +1.39, Discomfort For Team +1.07. This is the'
        ' mindset that lets you remove a co-founder when the company required it, and keep the'
        ' org&rsquo;s tempo through turbulent periods. Teach it to leaders who are still waiting for'
        ' the comfortable version of the right decision to arrive.</p>'
        '</div>'
        '<div class="practice-fuel">Zen-master level: discomfort is the signal you are at the growth'
        ' edge. When the week feels comfortable, ask what you are not yet confronting.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">9</span><div style="flex:1;">'
        '<div class="practice-item-title">Shared accountability = no accountability</div>'
        '<div class="practice-qref">Q381 &middot; Organizational Decision Making (Teach Rank 9)</div>'
        '</div><div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="8.2">8.2 Clarity Of Accountability</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE &mdash; when multiple parties are accountable, no one truly is. This is'
        ' the paired truth to Q384. Together they form the clearest test in the instrument for'
        ' accountability craft. RACI / RAPID exists exactly to resolve this pattern &mdash; consult'
        ' widely, inform broadly, but accountable is always singular.</p>'
        '</div>'
        '<div class="practice-fuel">Teach it as the rule: any time you hear &ldquo;we are jointly'
        ' accountable for this,&rdquo; ask who will be fired if it does not happen. Silence = the'
        ' accountability is not real yet.</div>'
        '</div>'
    )

    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">10</span><div style="flex:1;">'
        '<div class="practice-item-title">Top performers love being pushed hard</div>'
        '<div class="practice-qref">Q79 &middot; Not Pleasing (Teach Rank 10)</div>'
        '</div><div class="practice-l2-tag" style="background:#dcfce7; color:#5a6773; border-color:#5a6773;" data-l2="7.4">7.4 Discomfort For Team</div></div>'
        '<div class="practice-body">'
        '<p>You answered TRUE. The coaching craft embedded in this answer is knowing the difference'
        ' between <em>pushing</em> (stretching, investing in growth, setting high expectations with'
        ' support) and <em>pressuring</em> (applying stress and anxiety). Top performers thrive on'
        ' the former and disengage under the latter. Teach the distinction explicitly &mdash; too many'
        ' operators confuse the two.</p>'
        '</div>'
        '<div class="practice-fuel">The diagnostic: after a difficult conversation, ask the person'
        ' &ldquo;did that feel like I was investing in you or testing you?&rdquo; The answer tells you'
        ' whether you were pushing or pressuring.</div>'
        '</div>'
    )

    return '\n'.join(teach_cards)

# ============================================================================
# CONNECTION NARRATIVE + CAREER TIMELINE + CLOSING
# ============================================================================

CONNECTION_NARRATIVE_HTML = (
    "<p>Read the two lists together. Your <strong>Teach</strong> items are the operator fluencies"
    " &mdash; what you already do naturally enough to coach others: third-party praise, requiring"
    " excellence over popularity, extreme proactivity, hands-off-until-all-in, expected-value"
    " decisions, daily difficult conversations, single-point accountability, discomfort-as-default,"
    " and knowing the push-vs-pressure distinction. This is the operating system of a founder-CEO"
    " who has already built a FinTech from $18MM Seed to $75MM Series A Extension, and it is"
    " running.</p>"
    "<p>Your <strong>What-to-Work-On</strong> section leads with <strong>flag-driven items</strong>"
    " &mdash; Initiating Accountability Sev, Condit Belief Sev, HoldsAvgDownChain, Satisfied with"
    " Gripes | Low &mdash; because flags are pattern-level signals, not per-answer lifts. The"
    " algorithm-ranked items follow, and they cluster around a single root: <em>building an organization"
    " that executes at pace without you as the routing node</em>. Mission-lane discipline (#1), fight-to-keep"
    " (#2), develop-not-trade (#3), urgency-lives-below (#4), disappear-by-design (#5), no-irreplaceables"
    " (#6), small-changes (#7), no-send-when-upset (#8), task-only-your-directs (#9) &mdash; all expressions"
    " of the same habit: <em>moving from the lead-guitarist posture to the conductor posture</em>. The"
    " L2s that name this behavior &mdash; Ability To Disappear, Dialogue vs. Direction, Empower Team"
    " Authority, Urgency Down Chain &mdash; sit across Conducting &amp; Outvoted, Facilitative Mindset,"
    " and Replacing Self. That is why it is a flag, not a single-L2 lift.</p>"
    "<p>The connective tissue shows up in your <strong>anti-motivator wedge</strong>. Your wiring"
    " (Objective #1 Primary &middot; Structured #2 Primary &middot; Resourceful #3 Primary &middot;"
    " Commanding #4 Primary) runs hard on data-driven, structured, mission-first execution; the"
    " directly-opposing wedge is Altruistic / Harmonious / Selfless / Receptive (ranks 9&ndash;12,"
    " all Indifferent). The Impact items all sit in that opposing wedge. Which is exactly why the"
    " coaching move is <em>routines installed</em>, not <em>personality changed</em>. The Impact list"
    " is not a call to soften the founder-CEO; it is a short list of specific, observable behaviors"
    " installed against the grain &mdash; and the leverage is enormous because your Teach-item"
    " fluencies (the proactivity engine, the decision-clarity, the difficult-conversation habit)"
    " multiply every unit of work you do on the Impact side.</p>"
    "<p>One synthesis. The reason Provable is at $75MM in April 2025 is the <em>engine</em> &mdash;"
    " your tempo, your decision quality under pressure, your willingness to remove a co-founder when"
    " the company required it. The practice frontier is the <em>conductor</em> &mdash; the ability"
    " to step back from the baton while the engine still runs, and to install belief in the team"
    " before they have proved it out. Same leader, higher altitude. Especially the best get better.</p>"
)

CAREER_TIMELINE_TITLE = "Career Timeline &mdash; Matthew Cohen"

CAREER_TIMELINE_HTML = """
            <div class="timeline">
                <div class="timeline-block" style="flex: 1.8; background: #95a5a6;">Earlier career &middot; banking / trading arc<br><span style="font-size:9px; opacity:0.8;">Pre-2017 &middot; Securities-finance / special-situations operator</span></div>
                <div class="timeline-block" style="flex: 1.5; background: #34495e;">Jefferies &middot; MD Head of Securities Finance Trading<br><span style="font-size:9px; opacity:0.8;">2017&ndash;2018 &middot; $65MM &rarr; $90MM revenue growth</span></div>
                <div class="timeline-block" style="flex: 1.5; background: #2c3e50;">Nomura &middot; MD Head of Special Situations<br><span style="font-size:9px; opacity:0.8;">2018&ndash;2019 &middot; $125MM &rarr; $150MM revenue growth</span></div>
                <div class="timeline-block" style="flex: 3.0; background: #1a2332;">Provable Markets &middot; CEO &amp; Co-Founder <span style="background:#d4a84b; color:#1a2332; padding:0 4px; border-radius:2px; font-size:9px;">$18MM Seed &rarr; ~$75MM Series A Ext</span><br><span style="font-size:9px; opacity:0.8;">Jan 2021&ndash;present &middot; FinTech platform &middot; <strong>current seat</strong></span></div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item">
                    <div class="legend-dot" style="background: #95a5a6;"></div>
                    <span><strong>Earlier career (pre-2017):</strong> Securities-finance and special-situations operator. Foundation for the MD-level moves that followed at Jefferies and Nomura.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #34495e;"></div>
                    <span><strong>Jefferies (2017&ndash;2018):</strong> MD &mdash; Head of Securities Finance Trading. Self-reported revenue growth from $65MM to $90MM over 1.5 years. Demonstrates the per-seat revenue-generation engine at an MD altitude on the sell-side.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #2c3e50;"></div>
                    <span><strong>Nomura (2018&ndash;2019):</strong> MD &mdash; Head of Special Situations. Self-reported growth $125MM to $150MM over 1.5 years. Q9927 note: boss Mike Caperonis &mdash; &ldquo;extremely even keel in good and bad, a prerequisite for a risk manager/trader and rare to achieve consistently.&rdquo; The exit framing names a political-dynamics growth area.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #1a2332;"></div>
                    <span><strong>Provable Markets (Jan 2021&ndash;present):</strong> CEO &amp; Co-Founder. FinTech platform for securities lending. Seed $18MM (Jan 2021) &rarr; Series A Extension ~$75MM (April 2025). Self-reported KPI (Q76): 10K avg daily executions, 40% of top-20 clients signed on. Removed original CEO / co-founder Thomer Gil when the operating approach diverged; Q9917 self-assessment names patience, delegation, aggressiveness, work/life balance, and temper as personal areas the separation surfaced.</span>
                </div>
            </div>
            <div class="timeline-banner" style="background:#e8f4ea; color:#1e5a30;">
                Two MD-level sell-side seats with verified per-seat revenue growth, followed by founding and running a FinTech through the co-founder-separation stress test and the Seed&rarr;Series&nbsp;A&nbsp;Ext arc to $75MM valuation. The signature is an operator who builds with discipline and makes the hardest people-calls when the company requires it. <em>Note: LinkedIn public view is partially constrained &mdash; Awards/Honors sections unverifiable in session; Talent-axis read is anchored on the instrument file, Non-Scorable career-history answers (Q9911&ndash;Q9927), and publicly-searchable Provable Markets funding records.</em>
            </div>
"""

CLOSING_NOTE_HTML = (
    "<p>One last frame. You are already one of the best &mdash; Not Pleasing +1.75, Deliberate"
    " Urgency +1.60, Org Decision Making +1.20, Investing in Others +0.83; four L1 strengths that"
    " anchor the founder-CEO operating system. A FinTech built from $18MM Seed to $75MM Series A"
    " Extension through a co-founder separation is the live evidence of the engine. This guide is"
    " how to get even better &mdash; because <em>especially the best get better</em>.</p>"
    "<p>Two effectiveness levers, in sequence.</p>"
    "<p><strong>First, the prerequisite: HoldsAvgDownChain + Condit Belief &mdash; depth and belief.</strong>"
    " Before the in-the-room Initiating Accountability work fully pays out, the layer below your"
    " directs needs to be at a standard where the directs can push decisions down and have them"
    " held. Condit Belief Sev + Upgrade Team + Satisfied with Gripes | Low all point at the same"
    " foundation work &mdash; extend belief earlier, invest developmentally before the proof lands,"
    " and move standing gripes to decisions. The Impact items #2 (fight to keep) and #3 (crew not"
    " basketball) and #6 (no irreplaceables) are the operating expressions of this lever.</p>"
    "<p><strong>Second, on that foundation: the in-the-room Initiating Accountability frontier.</strong>"
    " A small set of routines that move you from lead guitarist to conductor &mdash; question-first"
    " dialogue, tasking only your directs, building the organization to run for a week without you."
    " The accountability machinery (Pushing Accountability, Clarity of Accountability, Extreme"
    " Proactivity) is already strong; this is the layer on top that lets the next $75M&rarr;$200M"
    " run happen without you as the routing node. Impact #9 (task only your directs) is the"
    " gateway routine.</p>"
    "<p>Neither frontier is good-vs-bad; both are effectiveness-vs-effectiveness &mdash; the levers"
    " that move this file further into the territory of the admired and successful founder-CEO.</p>"
    "<p>Especially the best get better.</p>"
)

# ============================================================================
# DISC NOTES
# ============================================================================

DISC_NOTE_TEXT = (
    "Conducting Implementor (D=69, C=61, S=48, I=32). Adapted shift: +9D, −4I, −13S, +3C."
)
DISC_NOTE_DETAIL = (
    "Adapted profile pulls harder toward commanding posture and further away from steadiness and"
    " warmth in the current founder-CEO seat. Objective #1 Primary + Structured #2 Primary + Resourceful"
    " #3 Primary + Commanding #4 Primary is the wiring of a data-driven, disciplined builder"
)
DISC_ANNOTATION_CODE = ""

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    # 1. No unreplaced tokens
    leaks = re.findall(r'\{\{([A-Z_0-9]+)\}\}', html)
    if leaks:
        failures.append(f"Unreplaced tokens: {sorted(set(leaks))}")

    # 2. Canvases
    for cid in ['distChart1', 'distChart2', 'distChart3', 'discChart', 'excstdsChart']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    # 3. Brand lockup
    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("Brand lockup count < 2")

    # 4. Practice items — at least 10 teach and 13 impact (4 flag + 9 per-answer)
    teach_items = html.split('Part 1 — What You Teach')[1].split('Part 2 — What to Work On')[0]
    impact_items = html.split('Part 2 — What to Work On')[1].split('How the Two Lists Connect')[0]
    n_teach = teach_items.count('class="practice-item"')
    n_impact_std = impact_items.count('class="practice-item"')
    n_impact_flag = impact_items.count('class="practice-item flag-driven"')
    if n_teach < 10:
        failures.append(f"Teach items: got {n_teach}, need ≥10")
    if n_impact_flag < 3:
        failures.append(f"Flag-driven impact items: got {n_impact_flag}, need ≥3")
    if n_impact_std < 9:
        failures.append(f"Per-answer impact items: got {n_impact_std}, need ≥9")

    # 5. Dimensional scorecard has L1 UPPERCASE + L2 indent
    m_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_isl1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_labels and m_isl1:
        labels = json.loads(m_labels.group(1))
        isl1 = json.loads(m_isl1.group(1))
        if len(labels) < 12:
            failures.append(f"Scorecard has only {len(labels)} rows (min 12)")
        l2_rows = sum(1 for v in isl1 if not v)
        if l2_rows < 6:
            failures.append(f"Scorecard has only {l2_rows} L2 rows (min 6)")
        for lbl, is_l1 in zip(labels, isl1):
            if is_l1 and lbl != lbl.upper():
                failures.append(f"L1 label '{lbl}' must be UPPERCASE")
                break
            if not is_l1 and not lbl.startswith('    '):
                failures.append(f"L2 label '{lbl}' must be 4-space indented")
                break

    # 6. Chart 3 reversed — high flags should be on LEFT
    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first and int(first.group(1)) < 30:
            failures.append(f"Chart 3 axis not reversed — first label starts with {first.group(1)}")

    # 7. Motivators section rendered
    if '.ma-section' not in html or 'class="ma-section"' not in html:
        failures.append("Motivators section not injected")

    # 8. Candidate data surfaces
    if 'Matthew Cohen' not in html:
        failures.append("Candidate name missing")
    if 'Provable Markets' not in html:
        failures.append("Candidate role / company missing")

    # 9. Fingerprint narrative present
    if 'Conducting Implementor' not in html:
        failures.append("Fingerprint narrative missing wedge label")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} issue(s) ***")
        for f in failures:
            print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)} issue(s)")
    else:
        print("*** QA GATE PASSED ***")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("Loading respondent data...")
    respondent_data = load_respondent_data()

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution chart tokens...")
    dist_tokens = build_distribution_tokens(
        zalgo_rows, flag_rows,
        respondent_data['z_algo_overall'],
        respondent_data['z_human_overall'],
        respondent_data['rf_num'],
    )

    print("Building motivators section...")
    respondent = build_respondent_dict(respondent_data)
    motivators_html = build_section(respondent, include_css=True)

    print("Building scorecard...")
    scorecard = build_excstds_scorecard(respondent_data)

    print("Building impact / teach items...")
    impact_html = build_impact_items_html()
    teach_html = build_teach_items_html()

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    # Neutralize in-comment {{EXCSTDS_COLOR_OVERRIDES}} reference if present
    template_html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}}",
        "Substitute EXCSTDS_COLOR_OVERRIDES"
    )

    # Inject motivators section (contains raw braces — do this before token-replace loop)
    template_html = template_html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    replacements = {
        # Header / meta
        'CANDIDATE_NAME': 'Matthew Cohen',
        'CANDIDATE_CREDS': 'CEO &amp; Co-Founder &middot; Provable Markets',
        'CANDIDATE_ROLE': 'FinTech &middot; Securities Lending Platform &middot; $18MM Seed &rarr; ~$75MM Series A Ext',
        'REPORT_DATE': 'April 21, 2026',

        # Signature + fingerprint + driving forces
        'SIGNATURE_PATTERN': SIGNATURE_PATTERN,
        'FINGERPRINT_NARRATIVE': FINGERPRINT_NARRATIVE,
        'DRIVING_FORCES_PRIMARY_HTML': DRIVING_FORCES_PRIMARY_HTML,
        'DRIVING_FORCES_INDIFFERENT_HTML': DRIVING_FORCES_INDIFFERENT_HTML,
        'DRIVING_FORCES_IMPLICATIONS_HTML': DRIVING_FORCES_IMPLICATIONS_HTML,

        # Headline metrics
        'ZALGO_OVERALL': f'{respondent_data["z_algo_overall"]:+.2f}',
        'COHORT_AVG': '+0.24',
        'TEACH_ITEMS': '10/10',
        'IMPACT_ITEMS': '13',
        'FLAGS_LIT': str(len(respondent_data['flags_lit'])),
        'REVERSE_FLAGS': str(respondent_data['rf_num']),

        # DISC
        'DISC_D_NAT': '69', 'DISC_I_NAT': '32', 'DISC_S_NAT': '48', 'DISC_C_NAT': '61',
        'DISC_D_ADP': '78', 'DISC_I_ADP': '28', 'DISC_S_ADP': '35', 'DISC_C_ADP': '64',
        'DISC_NOTE_TEXT': DISC_NOTE_TEXT,
        'DISC_NOTE_DETAIL': DISC_NOTE_DETAIL,
        'DISC_ANNOTATION_CODE': DISC_ANNOTATION_CODE,

        # Wiring-Fit
        'WIRING_FIT_ITEMS': WIRING_FIT_ITEMS,

        # Excellence Standards scorecard
        **scorecard,

        # Distribution charts
        **dist_tokens,

        # Teach & Impact
        'TEACH_ITEMS_HTML': teach_html,
        'IMPACT_ITEMS_HTML': impact_html,

        # Connection / timeline / closing
        'CONNECTION_NARRATIVE_HTML': CONNECTION_NARRATIVE_HTML,
        'CAREER_TIMELINE_TITLE': CAREER_TIMELINE_TITLE,
        'CAREER_TIMELINE_HTML': CAREER_TIMELINE_HTML,
        'CLOSING_NOTE_HTML': CLOSING_NOTE_HTML,
    }

    html = template_html
    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    # Run QA gate
    qa_gate(html)

    # Write output
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")

if __name__ == '__main__':
    main()
