"""Build the Excellence Standards hiring report for Stacey LoBreglio (Senior HR Manager).

Forked from build_armstrong_hiring.py 2026-04-27. Currently Senior HR Manager at Comcast
(May 2025-present, Englewood CO). Previously 19y9m at Verizon as HR Manager / Senior HR
Manager (2006-2024, no internal promotion across nearly two decades). Education: Bachelors,
Human Resources Management, Ottawa University. Self-report Q107 names "Quantum" as previous
company; LinkedIn shows Verizon — likely "Quantum" was an internal team/program name at
Verizon. Evaluated for a VP/Director step-up at a Public/Mature company.

Run from repo root:
    python _pipeline/scripts/build_lobreglio_hiring.py
Outputs to:
    _reports/Lobreglio_Stacey_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section, compute_intensity_from_disc

RESPONDENT_XLSX = ROOT / '_respondents' / '20260423.lobreglio71@gmail.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Lobreglio_Stacey_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title, 'z_algo': z_algo, 'z_human': z_human, 'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flag_names = []
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        if flag_name:
            flag_names.append((col, flag_name))
    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data, 'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall, 'z_human_overall': z_human_overall,
        'rf_num': rf_num, 'questions_answered': questions_answered,
        'flags_lit': flags_lit, 'non_scorable': non_scorable,
        'name': name, 'email': email, 'date_str': date_str,
    }

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]
    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo, z_human, sf = row['z_algo'], row['z_human'], row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1; continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1; break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    rev_flag_bin = len(flag_counts) - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# NARRATIVE
# ============================================================================

def build_three_axes_narratives():
    talent_badge = 'badge-red'
    talent_badge_text = 'RED'
    talent_card_body = '''Talent is the capacity to grow, read holistically across the career record. The right questions: <strong>has this person been promoted consistently?</strong> <strong>Does the file show greatness in more than one area?</strong> <strong>Does the arc suggest capacity to grow into the next altitude?</strong> On Lobreglio&rsquo;s file, all three lenses land negative.

<strong>Promotions:</strong> The dominant career line is 19 years 9 months at Verizon (Jan 2006 - Sep 2024/2025), held continuously at HR Manager / Senior Human Resources Manager / HR Manager - Engagement. <strong>No internal promotion is visible across nearly two decades at the same employer.</strong> The May 2025 move to Comcast was a lateral &mdash; same title (Senior Human Resources Manager). Q107 self-report names &ldquo;Quantum&rdquo; as a prior company, but LinkedIn shows Verizon as the only prior seat over the 2006&ndash;2024 window; &ldquo;Quantum&rdquo; is most likely an internal team or program name at Verizon, not a separate employer.

<strong>Greatness in more than one area:</strong> One craft (HR), one school, one Fortune-100 lifer arc. Bachelors in Human Resources Management from Ottawa University &mdash; a small private mid-tier school; no MBA, no graduate credential, no SHRM/HRCI senior-certification visible. No varsity-sport / cross-domain pursuit signal. No awards, no board service, no advisory work, no published writing. The career record reads as competent execution in a single lane at a single company at a single altitude.

<strong>Capacity to grow:</strong> 20+ years held at the Senior-Manager altitude is the dominant fact. The seat we are evaluating for is a VP / Director step-up &mdash; one to two altitudes higher than the entire prior career has demonstrated. The instrument data deepens the concern: Z|Algo overall &minus;0.51 sits below cohort average; Z|Human &minus;1.27 (a 0.76-Z below the algorithmic read) suggests human reviewers see a meaningfully weaker file than pattern-matching alone would indicate. Three Sev flags (Conditional Belief, HandsOn, Driving Accountability), two Hi flags (Initiating Accountability, Clarity of Accountability), and the rare <em>Satisfied with Average</em> flag together describe a leader who does not hold the line on standards or push for excellence at depth &mdash; the exact muscles a step-up to VP/Director would require.

The badge is RED rather than amber-red because all three Talent lenses point the same direction; there is no offsetting evidence on the file (no high-prestige selector, no award density, no internal promotion track) that the instrument would have to override.'''

    judgment_badge = 'badge-red'
    judgment_badge_text = 'RED'
    judgment_card_body = '''The instrument and the TTI corroborate each other almost line for line. <strong>Z|Algo overall &minus;0.51 sits below the cohort average of +0.24</strong>; Z|Human is &minus;1.27 (a 0.76-Z further negative than the algorithmic read &mdash; human reviewers see meaningfully weaker judgment than pattern-matching does). Per-dimension, the gap is widest on <strong>Organizational Decision Making (Z|Algo &minus;0.41 / Z|Human &minus;2.96 &mdash; a 2.55-Z gap)</strong> &mdash; the largest single-L1 Algo/Human disagreement in any file in the working set.

The L1 mass is concentrated where executive judgment under pressure lives. <strong>Not Pleasing L1 &minus;1.93</strong> (the worst L1) with 5 reverse flags out of 11 questions answered. Pushing Extreme Accountability L1 &minus;0.83 with 5 reverse flags out of 15. Conducting &amp; Outvoted L1 &minus;0.62. Deliberate Urgency L1 &minus;0.62. The L2 evidence converges: Demonstrating Genuine Fanness &minus;2.52, Clarity of Accountability &minus;2.06, Discomfort For Team &minus;1.49, Cares About Others Not Their Approval &minus;1.44, Power &amp; Status Management &minus;1.13, Sublimating Ego &minus;0.91, Proactive Belief In People &minus;0.85.

<strong>Three Sev flags, two Hi flags, plus a rare Satisfied-with-Average lit.</strong> Conditional Belief Sev. HandsOn Sev. Driving Accountability Sev. Initiating Accountability Hi. Clarity of Accountability Hi. Satisfied with Gripes Medium. <em>Satisfied with Average</em> &mdash; this last flag fires only when the answer pattern explicitly tolerates mediocre performance, and it is the rarest lit-flag in the working set.

The TTI Behavioral Characteristics and Conflict sections name the exact same failure mode in plain language: <em>&ldquo;Stacey prefers not disciplining people. She may sidestep direct disciplinary action because she wants to maintain the friendly relationship,&rdquo; &ldquo;Because of her trust and willing acceptance of people, she may misjudge the abilities of others,&rdquo; &ldquo;Would rather move on to a new person than have a difficult conversation with a direct report,&rdquo; &ldquo;Has difficulty hearing what others are saying if it contradicts her beliefs.&rdquo;</em> No wiring-vs-behavior divergence to argue the instrument out of.

Counter-evidence on the file: Personal Reliability L1 +1.52 (top decile), Extreme Ownership L2 +1.78, Respects Collective Wisdom L2 +1.96. She owns her own work cleanly and respects collective input. The judgment failure mode is not personal-reliability; it is <em>standard-holding through a team</em>.'''

    skills_badge = 'badge-amber'
    skills_badge_text = 'AMBER-RED'
    skills_card_body = '''The instrument does not measure domain skill directly. Domain credentials: 19+ years in HR roles at Verizon (2006-2024) plus 1 year at Comcast (May 2025-present). Bachelors in Human Resources Management, Ottawa University. HR-specific skill (HRIS, employee engagement programs, HR business partner work) is implied by the long single-lane tenure, though without an internal-promotion track or documented certification (SHRM-SCP, HRCI SPHR/GPHR) the depth-vs-breadth read is harder to call.

<strong>Wheel position:</strong> Natural <strong>position 16, Promoting Relater</strong>; Adapted <strong>position 17, Supporting Relater</strong>. Both clearly anchored in the Relater wedge of the TTI wheel &mdash; bottom of the wheel, high-S / high-I / low-D / low-C. Neither marked ACROSS; this is a strongly-anchored Relater profile, not a center-of-wheel one. Natural DISC D=48, I=72, S=72, C=34. Adapted D=42, I=62, S=66, C=48. Modest adaptation.

<strong>Driving Forces are diagnostic:</strong> Harmonious 86 (<em>Extreme</em>, 3&sigma; above mean), Collaborative 82 (<em>Extreme</em>, 3&sigma; above mean), Intentional 68, Structured 56. Indifferent: Objective, Altruistic, <strong>Commanding</strong>. The Relater wiring + Harmonious/Collaborative-Extreme DF + Commanding-Indifferent stack describe a high-empathy, harmony-seeking, supportive-team-member operator. <strong>For an HR seat the wiring is partially seat-aligned</strong> &mdash; the people-care function rewards empathy. <strong>For a VP/Director step-up at a Fortune 100, the wiring is the wrong leadership posture</strong> &mdash; the seat requires holding the line on standards, having difficult conversations, and driving consequences across hundreds of people. The TTI itself names the gap: <em>&ldquo;She gets frustrated when she is chosen to lead involuntarily,&rdquo; &ldquo;has difficulty looking at situations objectively,&rdquo; &ldquo;may overestimate the impact she can have on driving results.&rdquo;</em>'''

    return {
        'TALENT_BADGE_CLASS': talent_badge,
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': judgment_badge,
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': skills_badge,
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    concerns_title = 'Two Targeted Concerns'
    concerns_items = """            <div class="concern-item">
                <div class="concern-number">1</div>
                <div class="concern-text">
                    <strong>Accountability architecture is broken at every layer of the basket &mdash; Conditional, Initiating, Driving, Clarity, AND Holding-the-line.</strong>
                    <strong>Five flags lit on accountability:</strong> Conditional Belief <strong>Sev</strong>, Driving Accountability <strong>Sev</strong>, HandsOn <strong>Sev</strong>, Initiating Accountability <strong>Hi</strong>, Clarity of Accountability <strong>Hi</strong>. Plus the rare <strong>Satisfied with Average</strong> flag and Satisfied with Gripes Medium &mdash; she explicitly tolerates mediocre performance and accumulated gripes. L1 #7 Not Pleasing &minus;1.93 (worst L1) with 5 reverse flags out of 11 questions answered (45% wrong-direction). L2 evidence: Demonstrating Genuine Fanness &minus;2.52, Clarity of Accountability &minus;2.06, Discomfort For Team &minus;1.49, Cares About Others Not Their Approval &minus;1.44, Drives Accountability &minus;0.63. The TTI conflict section names the in-the-room mechanism in her own wiring: <em>&ldquo;would rather move on to a new person than have a difficult conversation with a direct report,&rdquo; &ldquo;prefers not disciplining people,&rdquo; &ldquo;may sidestep direct disciplinary action because she wants to maintain the friendly relationship.&rdquo;</em> For a VP/Director-of-HR seat at a Fortune 100, this is the structural disqualifier &mdash; HR leadership IS the function that holds the line on people-standards. Probe Form 8 Talent Development and Leadership Deep-Dive directly on this.
                </div>
            </div>
            <div class="concern-item">
                <div class="concern-number">2</div>
                <div class="concern-text">
                    <strong>20 years at the Senior-Manager altitude with no internal promotion + a lateral exit + a step-up jump being asked of her now.</strong>
                    LinkedIn shows 19y9m at Verizon (2006-2024) all held at HR Manager / Senior HR Manager / HR Manager - Engagement &mdash; the same altitude. Then a lateral move to Comcast in May 2025 at the same title (Senior HR Manager). The seat being considered is a VP/Director step-up &mdash; one to two altitudes above her sustained career range. Combined with Concern 1 (the standard-holding muscle the seat requires is exactly what the instrument and the TTI both say she has not built), the asked-jump is large. Probe Form 8 Two-Sport Athlete (what is she most proud of in life beyond the HR lane?) and Leadership Deep-Dive (sketch the Comcast HR org and name what would change in her routines at VP/Director altitude). The strong answer either: (a) names a specific stretch project at Verizon she ran outside the Senior HR Manager remit, with documented results; or (b) names an external pursuit (board, certification, advisory) that built the standard-holding muscle. A weak answer narrates duties at the prior altitude rather than evidence of capacity for the next one.
                </div>
            </div>"""
    return {'CONCERNS_TITLE': concerns_title, 'CONCERNS_ITEMS': concerns_items}

def build_wiring_fit():
    wiring_fit_items = (
        '<strong>Wheel positions: Natural 16 (Promoting Relater), Adapted 17 (Supporting Relater).</strong> '
        'Both clearly anchored in the Relater wedge &mdash; bottom of the TTI wheel, high-S / high-I / low-D / '
        'low-C. NEITHER position is marked ACROSS; this is a strongly-anchored Relater profile, not center-of-'
        'wheel. The wiring is purpose-built for empathy, harmony, supportive-team-member work. <strong>For an HR '
        'seat at the Senior-Manager altitude this wiring is partially seat-aligned</strong> &mdash; HR is a people-'
        'care function. <strong>For a VP/Director step-up at a Fortune 100 the wiring is the wrong leadership posture</strong> '
        '&mdash; the seat requires standard-holding, difficult conversations, and driving consequences across '
        'hundreds of people. The DF profile sharpens the read: Harmonious 86 (Extreme, 3&sigma; above mean), '
        'Collaborative 82 (Extreme), Intentional 68. Indifferent: Objective, Altruistic, Commanding.'
        '<span class="wiring-flag">Targeted Concern</span><br>'
        '<strong>The TTI Conflict section names the in-the-room failure modes directly:</strong> '
        '<em>&ldquo;Would rather move on to a new person than have a difficult conversation with a direct report,&rdquo; '
        '&ldquo;Has difficulty looking at situations objectively,&rdquo; &ldquo;Has difficulty hearing what others '
        'are saying if it contradicts her beliefs,&rdquo; &ldquo;May overestimate the impact she can have on driving '
        'results.&rdquo;</em> Each line corroborates a different lit flag (HandsOn Sev, Conditional Belief Sev, '
        'Initiating Accountability Hi, Driving Accountability Sev). No wiring-vs-behavior divergence to argue '
        'the instrument out of.'
        '<span class="wiring-flag">Targeted Concern</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    # Lobreglio Hard-to-Learn count: Low Urgency (L1 9 = -0.62) — partial; Org DM (Algo -0.41/Human -2.96) — lit
    # at Human read; Conditional Belief Sev — lit; Satisfied with Gripes Medium — lit. 4/4 lit.
    return {'HARD_TO_LEARN': '4/4'}

FORM8_QUESTIONS = [
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]

def build_interview_probes():
    tailored = [
        ("TWO-SPORT ATHLETE (Talent axis &middot; capacity-to-grow probe)",
         "Central probe for Concern 2 / Talent axis. The career record is single-lane HR for 20+ years at the same altitude with no documented adjacent excellence (no varsity sports listed, no awards, no board service, no advanced credentials). Listen for whether the proudest result is a specific stretch project at Verizon she ran outside the Senior HR Manager remit, with documented results &mdash; or whether the answer goes to family / personal-life pride OR to a normal-duty deliverable. The weakest signal: defaults to Verizon HR-business-partner duties. The strongest signal: names a specific external pursuit (board, certification track, civic leadership) that demonstrates capacity-to-grow into the next altitude."),
        ("TALENT DEVELOPMENT (Concern 1 &middot; Pre-proof belief)",
         "Conditional Belief <strong>Sev</strong> + L2 1.4 Demonstrating Genuine Fanness &minus;2.52 + L2 9.3 Proactive Belief In People &minus;0.85 + Lower Standards for Others than Self flag + Upgrade Team flag + the TTI line <em>&ldquo;may misjudge the abilities of others&rdquo;</em> all converge: she trusts uncritically rather than developing intentionally. Listen for one specific case where she backed someone <em>before</em> they had earned it, with the developmental investment named (a stretch role, a defended hire, a sustained coaching cadence with someone struggling). Generic answers (&ldquo;everyone has potential,&rdquo; &ldquo;I believe in people&rdquo;) are negative signals."),
        ("TORC &mdash; TRUTH-OVER-COMFORT (Self-awareness)",
         "Anchor to her stated boss: Katie Staples, HR Director at Comcast. Her own Q9917 self-answer about Katie was warm and partner-shaped (&ldquo;easy to work with, partner as a team&rdquo;) and the &ldquo;area of improvement&rdquo; she named for her boss was workload-not-walking-away. That is itself diagnostic &mdash; her self-described boss-relationship leans more friendship than truth-over-comfort. Press for a real developmental area Katie would surface for HER. The TTI conflict section explicitly says <em>&ldquo;has difficulty hearing what others are saying if it contradicts her beliefs&rdquo;</em> &mdash; expect deflection or strengths-reframing on the first answer; press past it."),
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "20 years at one company with no internal promotion + a recent lateral move puts real adversity material on the table (the plateau itself, the move, the transformation work she described in Q76). Listen for whether she names a personal adversity or defaults to professional. If professional, listen for ownership location: internal (&ldquo;I should have done X differently&rdquo;) or external (&ldquo;Verizon&rsquo;s structure,&rdquo; &ldquo;the team that was in place&rdquo;). The healthy answer attributes learning to self even when the situation was hard."),
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior probe. With TTI Harmonious-Extreme + Collaborative-Extreme primary DF, the most natural strong answer would be a specific people-listening / curiosity / team-building routine that compounds &mdash; something that distinguishes her from the average HR-business-partner peer. Generic answers (&ldquo;I work harder,&rdquo; &ldquo;I care more about people&rdquo;) are negative signals. Press for the artifact: the listening cadence, the engagement-survey-driven action loop, the specific employee-program she designed and ran end-to-end."),
        ("FACILITATIVE MINDSET (Concern 1 &middot; Dialogue and dissent)",
         "Tests whether she will hold a belief firmly AND name conditions for updating it. The TTI Conflict section says <em>&ldquo;has difficulty hearing what others are saying if it contradicts her beliefs&rdquo;</em> &mdash; high-S + Harmonious-Extreme DF means she may sound facilitative (lots of empathy and listening) but may not actually update from dissent. A Facilitative answer cites a recent specific case where her team changed her mind on something material. A non-Facilitative answer holds the belief absolute under pressure. Follow-up: <em>&ldquo;When was the last time someone on your team convinced you to change your mind on something that mattered?&rdquo;</em>"),
        ("COMMITMENT (Diligence wiring &middot; Personal Reliability +1.52 strong)",
         "Personal Reliability L1 +1.52 (top decile) + L2 6.3 Commitment to Routine +0.91 + Q24 daily-routines-for-years answered TRUE = the personal-discipline machinery is wired in cleanly. Strong answer names a concrete daily practice (an early-morning routine, a specific reading or reflection block, a calendar-discipline). Vague answers would contradict the instrument; press for the artifact."),
        ("LEADERSHIP DEEP-DIVE (Concerns 1 &amp; 2 &middot; Org-chart conversation)",
         "Ask her to sketch the Comcast HR org she sits inside (or the Verizon HR org she just left). Tests both Targeted Concerns directly. A leader running the accountability routine names single-point owners for each function and identifies open seats / upgrade candidates in the same breath. With <strong>Satisfied with Average + Lower Standards for Others than Self + Upgrade Team flags all lit</strong>, follow up with: <em>&ldquo;Which of these seats would you not re-hire today?&rdquo;</em> &mdash; this is the diagnostic question. Hesitation, generic warmth, or &ldquo;everyone has potential&rdquo; are the flags. Then: <em>&ldquo;What would change in your routines at VP/Director altitude?&rdquo;</em> &mdash; the strong answer names specific routines she would install (skip-level cadences, talent-review structure, performance-calibration practice); the weak answer narrates more of the same Senior-Manager-altitude work."),
        ("PASSION &amp; MISSION CONSTRUCTION",
         "Tests mission-construction. The career arc is single-lane HR. Listen for whether she roots passion in a specific kind of HR problem (engagement programs, transformation work, talent development) or defaults to scale-language. Q76 self-report mentions transformation work at Comcast (&ldquo;help the business unify after a large transformation&rdquo;) &mdash; press for the specific mission framing she carries into the work. Bridges into &ldquo;Why VP / Director, why now?&rdquo; given the 20-year plateau."),
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes Concern 1. Listen for whether she counts time spent holding the line, having difficult conversations, or running performance-calibration loops as core HR work &mdash; or only programs, engagement initiatives, and partnership work. With Satisfied-with-Average + HandsOn Sev + Driving Accountability Sev all lit, the diagnostic answer pattern is: she names empathy / listening / partnership work as core but doesn&rsquo;t name standard-holding / consequence-driving as core. The strong answer explicitly counts performance-management and standards-enforcement as real HR work."),
    ]
    assert len(tailored) == len(FORM8_QUESTIONS) == 10
    probes = []
    for (category, coaching), (_n, canonical_q, _s) in zip(tailored, FORM8_QUESTIONS):
        probes.append((category, f'"{canonical_q}"', coaching))
    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    return {'INTERVIEW_PROBE_CARDS': "\n".join(parts)}

def build_teach_items():
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (7, 'NOT PLEASING', [
            'Cares About Others Not Their Approval',
            'Discomfort For Team',
            'Discomfort For Self',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Drives Accountability',
            'Stds = What Tolerate',
            'Basic Machinery Of Accountability',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Clarity Of Accountability',
            'Simplification Methods',
            'Respects Collective Wisdom',
        ]),
        (6, 'PERSONAL RELIABILITY', [
            'Extreme Ownership',
            'Commitment To Routine',
        ]),
    ]

    labels, scores, is_l1 = [], [], []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")

    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    # Two-Sport Athlete: 1 (no varsity sports, no cross-domain pursuit on file; 20-yr single-lane career)
    # Punctuates Differently: 2 (Personal Reliability +1.52 is real but Not Pleasing -1.93 catastrophic)
    # Facilitative Mindset: 2 (Dialogue +1.14 OK, but Sublimating Ego -0.91, Power & Status -1.13)
    # Wit: null
    # Deep Repertoire: 2 (single-lane HR career, single school, no MBA / certifications)
    # Discipline/Routine: 4 (Personal Reliability +1.52 top decile, Extreme Ownership +1.78,
    #     Commitment to Routine +0.91)
    # Understands Symbolism: null
    talent_radar_scores = json.dumps([1, 2, 2, None, 2, 4, None])
    talent_radar_profile = (
        "<strong>Discipline/Routine (4)</strong> is the sole standout dimension &mdash; Personal Reliability "
        "L1 +1.52 (top decile), Extreme Ownership L2 +1.78, Commitment to Routine L2 +0.91. The personal-"
        "discipline machinery is real. <strong>Two-Sport Athlete (1)</strong> &mdash; read holistically as "
        "<em>capacity to grow</em>: consistent promotions across roles? (No, 20 years at Senior HR Manager "
        "altitude.) Greatness in more than one area? (No, single career lane.) Cross-domain excellence? "
        "(None visible.) <strong>Punctuates Differently (2)</strong> &mdash; Personal Reliability is real but "
        "Not Pleasing L1 &minus;1.93 catastrophic; no rare-behavior signal jumps off the file. "
        "<strong>Facilitative Mindset (2)</strong> &mdash; Dialogue vs. Direction L2 +1.14 OK, but Sublimating "
        "Ego &minus;0.91 + Power &amp; Status &minus;1.13 + TTI <em>&ldquo;may feel her view is the only way&rdquo;</em> "
        "argue against. <strong>Deep Repertoire (2)</strong> &mdash; single-lane HR career, Bachelors only, "
        "no MBA / SHRM-SCP / SPHR certifications visible. <strong>Wit (&#8856;)</strong> and <strong>Understands "
        "Symbolism (&#8856;)</strong> interview-only."
    )
    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_role_fit():
    role_fit_title = 'Role-Fit Read &mdash; What Will Be Easy, What Will Be Hard'
    role_fit_seat = ('Evaluating for: VP / Director of HR (or functional-lead seat at a similar altitude) '
                     '&middot; step-up from current Senior HR Manager altitude &middot; Public / Mature company '
                     '&middot; multi-thousand-employee scale')
    role_fit_easy = (
        '<strong>Personal reliability and ownership are top-decile.</strong> Personal Reliability L1 '
        '+1.52, Extreme Ownership L2 +1.78, Commitment to Routine L2 +0.91. She follows through on her '
        'own commitments and takes ownership of her own work. The TTI Areas of Awareness corroborate: '
        'Structured DF 56 (mainstream-positive) means she works inside established processes and '
        'protocols comfortably.<br><br>'
        '<strong>HR-business-partner skill at the Senior-Manager altitude is real.</strong> 19 years '
        'at Verizon executing HR-Manager / Senior-HR-Manager work plus 1 year in the equivalent '
        'Comcast seat means the FP&amp;A-equivalent of HR work &mdash; engagement programs, partnership '
        'with line leaders, transformation execution work like the one she described in Q76 (Comcast '
        'job-title-changes program with a June 1 deadline) &mdash; will land naturally.<br><br>'
        '<strong>Empathy, harmony, and team-orientation are wired in cleanly.</strong> DF Harmonious '
        '86 (Extreme), Collaborative 82 (Extreme), Intentional 68. L2 Respects Collective Wisdom +1.96 '
        'and Developmental Mindset +1.10. She will listen well, build psychological safety, and partner '
        'across functions. For an HR role those are genuine strengths.'
    )
    role_fit_hard = (
        '<strong>Step back: the wiring is anchored in the wrong wedge for a leadership-altitude HR seat.</strong> '
        'For an HR seat at the Senior-Manager altitude, Relater wiring (high-S, high-I, low-D, low-C) is '
        'partially seat-aligned &mdash; the people-care function rewards empathy. <em>For a VP/Director '
        'step-up at a Fortune 100, Relater wiring is the wrong leadership posture.</em> Wheel positions: '
        'Natural <strong>position 16 (Promoting Relater)</strong>, Adapted <strong>position 17 (Supporting '
        'Relater)</strong> &mdash; both clearly anchored in the bottom-of-wheel Relater wedge. Neither is '
        'ACROSS. The DF profile sharpens it: Harmonious 86 (Extreme, 3&sigma; above mean), Collaborative '
        '82 (Extreme), with <strong>Commanding Indifferent</strong>. The wiring will resist exactly the '
        'standard-holding, hard-conversation, drive-consequences-across-the-org work the seat requires.<br><br>'
        '<strong>Five accountability flags lit + Satisfied with Average.</strong> Conditional Belief Sev, '
        'HandsOn Sev, Driving Accountability Sev, Initiating Accountability Hi, Clarity of Accountability '
        'Hi, plus the rare Satisfied-with-Average flag. L1 #7 Not Pleasing &minus;1.93 (the worst L1 in the '
        'profile) with 5 reverse flags out of 11 questions answered. She tolerates mediocre performance, '
        'avoids difficult conversations (TTI: <em>&ldquo;would rather move on to a new person than have a '
        'difficult conversation with a direct report&rdquo;</em>), trusts uncritically (TTI: <em>&ldquo;may '
        'misjudge the abilities of others&rdquo;</em>), and does not drive consequences. For a VP/Director-of-'
        'HR seat, those are not coachable frontiers &mdash; they are the core of the work.<br><br>'
        '<strong>20-year plateau + lateral move + step-up jump being asked.</strong> 19y9m at Verizon held '
        'continuously at HR Manager / Senior HR Manager / HR Manager - Engagement &mdash; no internal '
        'promotion across nearly two decades. Lateral move to Comcast in May 2025 at the same title. The '
        'seat being considered is one to two altitudes above her sustained career range. The instrument '
        'evidence (Z|Algo &minus;0.51 below cohort, Z|Human &minus;1.27 deeper, three Sev flags, two Hi '
        'flags) does not support the asked-jump.<br><br>'
        '<strong>The recommendation is therefore not a binary diligence question.</strong> The default '
        'lean is <em>no</em>. The amount of onsite evidence required to override the convergent instrument + '
        'TTI + LinkedIn pattern is high. If the hiring company decides to advance, onsite must produce: '
        '(a) a specific stretch project at Verizon she ran outside the Senior HR Manager remit with '
        'documented results; (b) at least one specific case where she had a hard conversation with a '
        'direct report and held the line on standards; (c) a clear answer to the Form 8 Leadership Deep-'
        'Dive probe naming the routines she would install at VP/Director altitude. Without all three, '
        'default to no.'
    )
    return {
        'ROLE_FIT_TITLE': role_fit_title,
        'ROLE_FIT_SEAT': role_fit_seat,
        'ROLE_FIT_EASY': role_fit_easy,
        'ROLE_FIT_HARD': role_fit_hard,
    }

def build_career_timeline():
    timeline_title = 'Career Timeline &mdash; Single-Lane HR at Two Fortune 100s'
    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Education<br>Ottawa University<br>Bachelors HR Mgmt</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">Verizon &mdash; HR Manager / Senior HR Manager<br>Jan 2006 - Sep 2024 &middot; 18+ yrs<br>Denver, CO</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">Verizon &mdash; HR Manager - Engagement<br>2006 - Sep 2025 &middot; 19y9m total<br>(concurrent role / role transition)</div>
                <div class="timeline-block" style="background:#c0392b;color:#fff;">Comcast &mdash; Senior HR Manager<br>May 2025 - Present &middot; 1y on file<br>Englewood, CO &middot; Hybrid</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item"><div class="legend-dot" style="background:#3498db;"></div><span><strong>Education:</strong> Ottawa University, Bachelors in Human Resources Management and Services. No MBA / graduate degree on file. No SHRM-SCP / HRCI SPHR / GPHR certification visible.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e40af;"></div><span><strong>Verizon &mdash; HR Manager / Senior Human Resources Manager (Jan 2006 - Sep 2024):</strong> 18y11m. Denver, CO. The dominant career line. Held continuously at the Manager / Senior Manager altitude across nearly two decades. <strong>No internal promotion event visible on the LinkedIn record across the entire span.</strong></span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#16a085;"></div><span><strong>Verizon &mdash; HR Manager - Engagement (2006 - Sep 2025):</strong> 19y9m total Verizon tenure, with the most recent stretch as HR Manager - Engagement. Skills: HRIS, employee engagement, partnership work. Q107 self-report names &ldquo;Quantum&rdquo; as previous company &mdash; most likely an internal team/program name within Verizon, not a separate employer.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#c0392b;"></div><span><strong>Comcast &mdash; Senior Human Resources Manager (May 2025 - Present):</strong> 1y on file. Englewood, CO, Hybrid. Lateral move from Verizon (same title). Q76 self-report names current work: &ldquo;help the business unify after a large transformation&rdquo; with a job-title-changes program targeted for June 1, 2026 completion. Boss: Katie Staples, HR Director (per Q9917 self-report).</span></div>
            </div>
            <div class="timeline-banner">
                Career pattern: 20 years of HR work at two Fortune 100 telecoms (Verizon then Comcast), held continuously at the Senior-Manager altitude. <strong>No internal promotion event visible across the entire span.</strong> Single career lane, single school (Ottawa Bachelors), no graduate credential, no professional certification visible, no awards / board service / cross-domain pursuit. The career record reads as competent execution at one altitude in one craft at one type of company. The seat being evaluated for (VP / Director step-up) is one to two altitudes above the sustained range.
            </div>"""
    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

def build_respondent_dict(respondent_data):
    """TTI: Natural D=48 I=72 S=72 C=34 → Promoting Relater (16);
            Adapted D=42 I=62 S=66 C=48 → Supporting Relater (17).
       Both clearly anchored in the Relater wedge (NOT ACROSS).
       DF Primary: Harmonious 86 (Extreme), Collaborative 82 (Extreme), Intentional 68, Structured 56.
       DF Indifferent: Objective, Altruistic, Commanding."""
    nat_pos = 16
    nat_label = 'Promoting Relater'
    nat_disc = [48, 72, 72, 34]
    nat_intensity = compute_intensity_from_disc(nat_disc)  # 62/200 → 0.31

    adp_pos = 17
    adp_label = 'Supporting Relater'
    adp_disc = [42, 62, 66, 48]
    adp_intensity = compute_intensity_from_disc(adp_disc)  # 38/200 → 0.19

    shift_note = ('Adapted shift: &minus;6D, &minus;10I, &minus;6S, +14C. Modest adaptation overall &mdash; '
                  'both Natural and Adapted positions are clearly anchored in the Relater wedge of the wheel. '
                  'The C up-shift is the only meaningful change.')

    respondent = {
        'name': 'Stacey LoBreglio',
        'first_name': 'Stacey',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [48, 72, 72, 34],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }
    return respondent
# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    n_concern_item = count_class('concern-item')
    n_concern_number = count_class('concern-number')
    n_concern_text = count_class('concern-text')
    if n_concern_item < 2:
        failures.append(f"S1: concern-item count {n_concern_item} < 2")
    if n_concern_number != n_concern_item:
        failures.append(f"S1: concern-number != concern-item")
    if n_concern_text != n_concern_item:
        failures.append(f"S1: concern-text != concern-item")
    concern_nums = re.findall(r'<div class="concern-number">(\d+)</div>', html)
    expected_seq = [str(i) for i in range(1, n_concern_item + 1)]
    if concern_nums != expected_seq:
        failures.append(f"S1: concern sequence {concern_nums} != {expected_seq}")

    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: {cls} expected {expected}, got {got}")

    probe_questions = re.findall(r'<div class="probe-question">(.*?)</div>', html, re.DOTALL)
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched = []
        for idx, q in enumerate(probe_questions, 1):
            qlow = q.lower()
            hit = False
            for (_n, _c, substr) in FORM8_QUESTIONS:
                if substr.lower() in qlow:
                    matched_substrs.add(substr); hit = True; break
            if not hit: unmatched.append(idx)
        if len(matched_substrs) < 10:
            failures.append(f"S9-Form8: {len(matched_substrs)}/10 canonical strings; unmatched probes {unmatched}")

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: {cls} expected min {expected}, got {got}")

    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be N/M, got '{teach_val}'")
    if 'To be populated' in html:
        failures.append("S2: Forbidden placeholder 'To be populated'")

    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:
        failures.append("S4-C1: zLabels2 not nested 2-row")

    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/fail/success")
    else:
        sf_labels = json.loads(m_sf_l.group(1))
        fail_data = json.loads(m_fail.group(1))
        succ_data = json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append("S4-C2: length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with +, got {sf_labels[-1]}")

    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first:
            v = int(first.group(1))
            if v < 30:
                failures.append(f"S4-C3: flagLabels3 first = {v} — axis not reversed")

    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        exc_labels = json.loads(m_exc_labels.group(1))
        is_l1_arr = json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: scorecard rows {len(exc_labels)} < 12")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: L2 rows {l2_rows} < 6")
        for lbl, l1f in zip(exc_labels, is_l1_arr):
            if l1f and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' not UPPERCASE"); break
            if not l1f and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' not indented"); break
    else:
        failures.append("S6: excLabels/isL1 not found")

    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    for cls, mn in [('role-fit-box', 1), ('role-fit-grid', 1),
                     ('role-fit-col easy', 1), ('role-fit-col hard', 1)]:
        if count_class(cls) < mn:
            failures.append(f"S-RF: {cls} expected min {mn}")
    if 'What Will Be Easy' not in html or 'What Will Be Hard' not in html:
        failures.append("S-RF: column labels missing")
    seat_stage_keywords = ['Series A', 'Series B', 'Series C', 'Series D',
                           'Seed', 'late-stage', 'Late-stage', 'pre-IPO', 'Pre-IPO',
                           'Public', 'public/mature', 'Mature', 'Fortune', 'IPO']
    if not any(kw in html for kw in seat_stage_keywords):
        failures.append(f"S-RF: no seat-stage keyword found (one of {seat_stage_keywords})")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    # S10 — Recommendation badge length (added 2026-04-27, Schott build)
    rec_m = re.search(r'<div class="recommendation-badge">(.*?)</div>', html, re.S)
    if rec_m:
        rec_text = re.sub(r'&[a-zA-Z]+;', 'x', rec_m.group(1))
        rec_text = re.sub(r'<[^>]+>', '', rec_text).strip()
        if len(rec_text) > 300:
            failures.append(f"S10: recommendation-badge too long ({len(rec_text)} > 300 chars)")

    # S11b — Role-Fit Hard step-back content (added 2026-04-27, Schott build)
    rfh_m = re.search(r'<div class="role-fit-col hard">(.*?)</div>\s*</div>\s*</div>', html, re.S)
    if rfh_m:
        rfh = rfh_m.group(1).lower()
        has_quadrant = any(q in rfh for q in ['top-left', 'top-right', 'bottom-left', 'bottom-right',
                                              'across', 'center-of-wheel', 'center of wheel'])
        wedges = ['implementor', 'conductor', 'persuader', 'promoter',
                  'relater', 'supporter', 'coordinator', 'analyzer']
        has_wedge = any(w in rfh for w in wedges)
        if not has_quadrant:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wiring-shape reference (quadrant or ACROSS)")
        if not has_wedge:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wedge name")

    # S11b.1 — Wheel-position-required rule (added 2026-04-27, Armstrong build):
    # any narrative that claims a canonical-CFO / top-left / right-wiring / Implementor-quadrant
    # read MUST also cite the actual wheel position number. Reading DISC scores in isolation
    # without checking the wheel has produced shipping-quality errors. The check looks across
    # all wiring-relevant sections (Role-Fit, Wiring-Fit, Skills card, DISC notes) for any
    # canonical-CFO claim and verifies a wheel-position number (1-60) is also present nearby.
    wiring_claim_phrases = [
        'top-left', 'canonical strong-cfo', 'canonical cfo wiring', 'right wiring',
        'wiring is right', 'wiring fits the seat', 'implementor quadrant',
    ]
    html_lower = html.lower()
    has_wiring_claim = any(p in html_lower for p in wiring_claim_phrases)
    has_position_number = bool(re.search(r'\bposition\s+\d{1,2}\b', html_lower))
    if has_wiring_claim and not has_position_number:
        failures.append(
            "S11b.1: wiring narrative makes a canonical-CFO / top-left / right-wiring claim "
            "but does NOT cite a TTI wheel position number. Reading DISC scores in isolation "
            "is insufficient; cite the actual wheel position (Natural / Adapted) before making "
            "wiring-fit claims. See METHODOLOGY 'TTI wheel position is REQUIRED reading'."
        )

    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} ***")
        for f in failures: print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)}")
    print("*** QA GATE PASSED ***")


def main():
    print("Loading respondent data...")
    rd = load_respondent_data()
    print(f"  name={rd['name']}, Z|Algo={rd['z_algo_overall']:+.3f}, RF={rd['rf_num']}")
    print(f"  flags: {list(rd['flags_lit'].keys())}")

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution tokens...")
    dist = build_distribution_tokens(zalgo_rows, flag_rows,
                                     rd['z_algo_overall'], rd['z_human_overall'], rd['rf_num'])

    print("Building respondent dict for motivators_section...")
    respondent = build_respondent_dict(rd)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()
    role_fit = build_role_fit()
    scorecard = build_excstds_scorecard(rd)

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    replacements = {
        'CANDIDATE_NAME': 'Stacey LoBreglio',
        'CANDIDATE_CREDS': 'Candidate &mdash; VP / Director of HR Step-Up Seat',
        'CANDIDATE_ROLE': 'Candidate for VP / Director of HR &mdash; Public / Mature Company &middot; Step-Up from Senior HR Manager',
        'REPORT_DATE': 'April 27, 2026',

        'ZALGO_OVERALL': f'{rd["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{rd["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(rd['rf_num']),
        'FLAGS_LIT': f'{len(rd["flags_lit"])}/16',
        'TEACH_ITEMS': '10/10',
        'COHORT_AVG': '+0.24',
        'COHORT_AVG_NUM': '0.2440',

        **axes, **concerns, **wiring, **htl, **probes, **teach,
        **radar, **timeline, **role_fit, **dist, **scorecard,

        'DISC_D_NAT': '48',
        'DISC_I_NAT': '72',
        'DISC_S_NAT': '72',
        'DISC_C_NAT': '34',
        'DISC_D_ADP': '42',
        'DISC_I_ADP': '62',
        'DISC_S_ADP': '66',
        'DISC_C_ADP': '48',
        'DISC_NOTE_TEXT': 'Natural position 16 (Promoting Relater) &rarr; Adapted position 17 (Supporting Relater). Both clearly anchored in the bottom-of-wheel Relater wedge &mdash; NEITHER position is ACROSS. High-S, high-I, low-D, low-C wiring &mdash; the empathy / harmony / supportive-team-member quadrant.',
        'DISC_NOTE_DETAIL': 'TTI Driving Forces &mdash; Primary: Harmonious 86 (Extreme, 3&sigma; above mean), Collaborative 82 (Extreme, 3&sigma; above mean), Intentional 68, Structured 56. Indifferent: Objective, Altruistic, <strong>Commanding</strong>. The Relater wiring + Harmonious/Collaborative-Extreme + Commanding-Indifferent stack describes a high-empathy, harmony-seeking, supportive-team-member operator. For an HR seat at the Senior-Manager altitude this wiring is partially seat-aligned. <em>For a VP/Director step-up at a Fortune 100, this wiring is the wrong leadership posture</em> &mdash; the seat requires standard-holding, hard conversations, and driving consequences that the wiring resists.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        'RECOMMENDATION_TEXT': 'NO HIRE &middot; HIGH BAR FOR ANY RECONSIDERATION &mdash; Z|Algo &minus;0.51 below cohort; Z|Human &minus;1.27 deeper still; 3 Sev + 2 Hi flags + Satisfied with Average; 20-year plateau at Senior-Manager altitude; Relater wiring is wrong leadership quadrant for VP/Director step-up.',
    }

    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    html = re.sub(
        r'<title>.*?</title>',
        '<title>Stacey LoBreglio &mdash; VP/Director HR Candidate | HALE GLOBAL SUCCESS DIAGNOSTICS</title>',
        html, count=1
    )
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    qa_gate(html)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")


if __name__ == '__main__':
    main()
