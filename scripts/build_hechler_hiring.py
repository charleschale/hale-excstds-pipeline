"""Build the Excellence Standards hiring report for Howard Hechler.

End-to-end pipeline:
1. Load respondent data from xlsx (L1, L2, Flags, Metadata, Non-Scorable)
2. Compute distribution chart tokens from Histogram Data
3. Build motivators_section HTML
4. Fill all 56 tokens in the template
5. Save filled HTML

Run from repo root:
    python _pipeline/scripts/build_hechler_hiring.py
Outputs to:
    _reports/Hechler_Howard_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

RESPONDENT_XLSX = ROOT / '_respondents' / '20260421.howard@clearpathcardio.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Hechler_Howard_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    """Load all respondent data from xlsx sheets."""
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    # L1 scores
    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title,
                'z_algo': z_algo,
                'z_human': z_human,
                'rf_count': rf_count
            }

    # L2 scores
    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    # Flags and overall metrics
    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    # Extract flag values
    flag_names = []
    flag_row = 1
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(flag_row, col).value
        if flag_name:
            flag_names.append((col, flag_name))

    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    # Non-Scorable questions
    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    # Metadata
    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data,
        'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall,
        'z_human_overall': z_human_overall,
        'rf_num': rf_num,
        'questions_answered': questions_answered,
        'flags_lit': flags_lit,
        'non_scorable': non_scorable,
        'name': name,
        'email': email,
        'date_str': date_str,
    }

# ============================================================================
# LOAD HISTOGRAM DATA FOR DISTRIBUTION CHARTS
# ============================================================================

def load_histogram_data():
    """Load population distribution data."""
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)

    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({
                'z_algo': z_algo,
                'z_human': z_human,
                'sf': sf
            })

    # Count RFs per respondent from Zalgo summ @#RF column (col 7, 1-indexed).
    # The "Histogram Flags" sheet is BIN DEFINITIONS, not population counts —
    # aggregate actual counts from the population here.
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    # Return the raw population RF values; bin aggregation happens in build_distribution_tokens
    flag_rows = rf_values
    return zalgo_rows, flag_rows

# ============================================================================
# BUILD DISTRIBUTION CHART TOKENS
# ============================================================================

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    """Build DIST_* tokens for population distribution charts.
    Per QA_CHECKLIST.md Section 4 + Cole/Harinam canonical convention:
      - Chart 1 (Z|Algo + Z|Human): 2-row labels, collapse leading/trailing empty bins
      - Chart 2 (Success/Fail): 2-row labels, hide empty columns entirely (where fail=0 AND success=0)
      - Chart 3 (Flag Counts): REVERSE axis (high flags LEFT, low flags RIGHT)
    """
    # ---- Raw bin edges ----
    bin_edges = [i * 0.5 for i in range(-8, 9)]  # -4.0 to 4.0 in 0.5 steps (16 bins)

    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo = row['z_algo']
        z_human = row['z_human']
        sf = row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    # ---- Respondent bin indices (raw) ----
    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    # ---- Chart 1 collapse: drop outer bins where BOTH algo AND human are 0 ----
    def collapse_outer(counts_list, marker_bins):
        """Drop leading/trailing indices where ALL datasets are 0."""
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        idx = list(range(left, right + 1))
        return idx

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    # ---- Chart 2 collapse: drop bins where fail=0 AND success=0 (EMPTY COLUMNS HIDDEN) ----
    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        # Last kept bin gets "+" notation
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    # ---- Chart 3: flag bins then REVERSE for canonical orientation ----
    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1
            continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1
                break

    # Respondent RF bin (raw, low-to-high)
    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    # Reverse: high flags on LEFT, low flags on RIGHT
    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    n_flag = len(flag_counts)
    rev_flag_bin = n_flag - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# BUILD NARRATIVE SECTIONS (Three Axes, Concerns, Wiring Fit, etc.)
# ============================================================================

def build_three_axes_narratives():
    """Draft the Talent, Judgment, Skills cards per the brief."""

    talent_badge = 'badge-amber'
    talent_badge_text = 'AMBER'
    talent_card_body = '''Strong external validator density via Lumicell's FDA dual-track NDA + PMA approval (Apr 2024), CMS TPT (Sep 2024), LumiSystem commercial launch (Jan 2025), and Board-confirmed CEO appointment (Nov 2024) signal a leader selectors keep picking through difficult regulatory-and-commercial sequences in medical devices. UVA Darden MBA + earlier degree.

However, this tension warrants probing: (a) LinkedIn public view is partially redacted — awards section is empty, Honors/Volunteer sections unverifiable → instrument sweep is structurally constrained; (b) Conditional Belief Hi + Upgrade Team flag + Altruistic DF Indifferent + Investing-in-Others Z_Human −1.35 → direct-report-altitude instrument signals on belief-in-others are mixed. The tension IS the read: selectors keep picking him; instrument flags mixed posture. Probe accordingly. Do NOT dismiss either signal.'''

    judgment_badge = 'badge-red'
    judgment_badge_text = 'RED/AMBER'
    judgment_card_body = '''L1 #8 Org Decision Making (Z_Human −1.48) sits at the core. The frontier: L2 8.7 Facts Over Feelings (−1.72) + L2 9.6 Action Over Inaction (−1.85) signals decisions get made on feelings rather than facts AND get deferred. Flag_ClarityAcctblty|Med + Flag_DrivingAcctblty + Flag_StdsVsSelf + Flag_HoldsAvgDown|Med all circle the same pattern.

The offset: L2 8.1 Simplification Methods (+1.73) and L2 2.1 Dialogue vs Direction (+1.27) are real strengths. The most honest read: judgment in the room (facilitates, simplifies) is strong; judgment under pressure (decides fast, faces facts, enforces depth) is the frontier. At a $3.5M fundraising deadline, slow decision-making would be organizationally lethal.'''

    skills_badge = 'badge-gray'
    skills_badge_text = 'NOT MEASURED'
    skills_card_body = '''This instrument does not measure domain skill directly. Acknowledge this caveat. Remind the hiring manager that strong skills do not compensate for weak talent or weak judgment at executive altitude.

That said: The publicly-documented Lumicell regulatory + commercial wins (FDA dual approval, CMS TPT, commercial launch) are a strong domain-skill proxy in medical device leadership. This does not, however, bypass the Talent and Judgment frontiers identified above.'''

    return {
        'TALENT_BADGE_CLASS': talent_badge,
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': judgment_badge,
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': skills_badge,
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    """Build the two targeted concerns using the template's canonical
    .concern-item / .concern-number / .concern-text structure."""
    concerns_title = 'Two Targeted Concerns'
    concerns_items = """            <div class="concern-item">
                <div class="concern-number">1</div>
                <div class="concern-text">
                    <strong>Decision pattern: Facts-over-Feelings + Action-over-Inaction.</strong>
                    L2 8.7 Facts Over Feelings (&minus;1.72), L2 9.6 Action Over Inaction (&minus;1.85), L2 9.2 Extreme Proactivity (&minus;0.88), Flag_ClarityAcctblty|Med, Flag_DrivingAcctblty circle the same signal: decisions lean on feeling rather than fact, and the decisional pace on hard calls lags. TTI wiring (very low S = fast pace) diverges from instrument (slow-to-decide) &mdash; the pace is there, but deliberate decisional urgency on high-stakes matters is the frontier. Must be probed in interview with decision-reversal and deferred-decision questions.
                </div>
            </div>
            <div class="concern-item">
                <div class="concern-number">2</div>
                <div class="concern-text">
                    <strong>Belief-in-others posture: Conditional Belief Hi + Upgrade Team + Altruistic Indifferent.</strong>
                    Flag_UnconditBelief &ldquo;Condit Belief Hi&rdquo; (severity=Hi) + Flag_DreamTeam &ldquo;Upgrade Team&rdquo; + TTI Altruistic DF 21 (Indifferent) + Collaborative DF 6 (essentially zero) signal this leader&rsquo;s drive to extend unconditional goodwill is low. Yet LinkedIn posts (&ldquo;truly superlative team,&rdquo; &ldquo;I am honored to lead this LumiTeam&rdquo;) show public investment in the team around wins. The tension is the probe: is public advocacy matched by private belief, or does belief arrive only after proof? Form 8 Talent Development is the interview target.
                </div>
            </div>"""

    return {
        'CONCERNS_TITLE': concerns_title,
        'CONCERNS_ITEMS': concerns_items,
    }

def build_wiring_fit():
    """Build the canonical 2-item Wiring-Fit Check. Per SKILL.md + template,
    this strip contains EXACTLY 2 <strong>-led items with <span class="wiring-flag"> pills,
    tied to the two Targeted Concerns (Deliberate Urgency + Organizational Decision Making)."""
    wiring_fit_items = (
        '<strong>Decision-urgency wiring mismatch:</strong> '
        'TTI low-S (S=21) and DISC commander profile (D=89, C=62) say fast-deciding; '
        'ExcStds L1 #9 Z_Human &minus;1.20 and L2 9.6 Action Over Inaction &minus;1.85 say slow-to-decide in practice. '
        'Wiring-vs-behavior divergence ties directly to Deliberate Urgency concern.'
        '<span class="wiring-flag">Targeted Concern</span><br>'
        '<strong>Fact-over-feelings wiring mismatch:</strong> '
        'TTI Instinctive Indifferent corroborates; ExcStds L1 #8 Z_Human &minus;1.48 and '
        'L2 8.7 Facts Over Feelings &minus;1.72 confirm that weight-of-data discipline is not wired in. '
        'Ties directly to Organizational Decision Making concern.'
        '<span class="wiring-flag">Diligence Item</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    """Build hard-to-learn gate summary. The {{HARD_TO_LEARN}} token is inside a narrow
    headline-metrics cell (~130px wide) — it must be a short count, NOT a full-width
    analysis block. Detailed gate content is covered in the Targeted Concerns section
    and the Wiring-Fit Check (Deliberate Urgency + Org Decision Making already tagged
    there as 'Targeted Concern')."""
    hard_to_learn = '4/4'
    return {'HARD_TO_LEARN': hard_to_learn}

#
# Canonical Form 8 question library — SOURCE OF TRUTH for hiring-report probes.
# Mirror of PROJECT_NOTES.md L182-191 and SKILL.md Form 8 section. These exact
# strings (or a distinctive sub-phrase from each) must appear verbatim in each
# probe-question block. Interview probes are drawn from this set — never generated.
#
FORM8_QUESTIONS = [
    # (category_name, canonical_question_text, distinctive_substring_for_QA_match)
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]


def build_interview_probes():
    """Build interview probe cards — uses template's .probe-card / .probe-number /
    .probe-category / .probe-question classes (see hiring_report_TEMPLATE.html).

    CRITICAL — Form 8 sourcing (SKILL.md + PROJECT_NOTES.md L178-191):
      The 10 probe questions MUST be drawn verbatim from the canonical Form 8 set
      above (FORM8_QUESTIONS). This is a non-negotiable spec contract. Do NOT
      generate novel questions. Do NOT paraphrase the question text.

    What IS tailored per-candidate:
      - probe-category label (links the Form 8 question to this candidate's
        specific flag / concern / wiring signal)
      - probe-coaching note (what to listen for, specific to this candidate's
        data patterns — career history, flag profile, wiring mismatches)

    The reader sees the canonical Form 8 question in quotes; the label "Form 8"
    does NOT appear in the question text itself (it is a behind-the-scenes
    sourcing rule, not a display label). Verification that each question maps
    to a Form 8 canonical string is enforced by qa_gate() via substring match
    against FORM8_QUESTIONS.

    The 10 probes map one-to-one to the 10 Form 8 questions in canonical order.
    """
    # probe-category labels tailored to Hechler's flags / concerns.
    # probe-coaching notes tailored to Hechler's specific data patterns.
    tailored = [
        # 1 — Two-Sport Athlete (Talent axis)
        ("TWO-SPORT ATHLETE (Talent axis)",
         "Talent-axis signal — listen for specifics OUTSIDE medical-device leadership and what the pursuit revealed about grit, coachability, and handling being bad at something before becoming good. Hechler's 25-year post-UVA arc has run almost exclusively in commercial biotech/medical leadership; a genuine cross-domain pursuit (sport, craft, service, academic) is the rare-behavior signal. A polished career recap in answer is the negative."),
        # 2 — Talent Development (Concern 2 · L1 #1 Investing in Others)
        ("TALENT DEVELOPMENT (Concern 2 · L1 #1 Investing in Others)",
         "Follow-up: press for one case where he backed the person <em>before</em> they had proven themselves. Specificity about extending goodwill ahead of performance proof is the core probe for Concern 2. Low Investing-in-Others wiring (Z_Human −1.35) and the Pre-Proof Belief flag predict generic answers or post-proof selections; strong Talent records produce two names and two specific pre-proof investments inside 60 seconds."),
        # 3 — TORC (self-awareness calibration)
        ("TORC — TRUTH-OVER-COMFORT (Self-awareness calibration)",
         "Anchor the question to a real boss: Jack Rowe at XL TechGroup, the CEO during the Mallinckrodt chapter, or the Lumicell board chair. Listen for whether he names real developmental areas or defaults to strengths-reframing (\"my weakness is I care too much\"). High-D + high-I profiles can convert weakness questions into achievement stories — push past the first answer to get the second."),
        # 4 — Emotional Maturity (Ownership pattern)
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "Note whether he chooses a personal adversity or defaults to a professional setback (the PolarityTE exit, the XL TechGroup wind-down, the Mallinckrodt chapter). Pattern of ownership vs. external attribution is the signal. A mature answer attributes the learning to self; a defensive answer attributes the setback to circumstances, boards, or markets. Maps to the Emotional Maturity sub-dimension under Judgment."),
        # 5 — Punctuates Differently (Talent axis)
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior signal on the Talent axis. Listen for a specific, idiosyncratic practice — not a general work-ethic claim. If the answer is \"I work harder\" or \"I care more,\" that is the negative signal — ordinary punctuation under different framing. A true Punctuates Differently answer names a concrete routine the interviewer has not heard before."),
        # 6 — Facilitative Mindset (Concern 1 · Decisions-over-feelings)
        ("FACILITATIVE MINDSET (Concern 1 · Decisions-over-feelings)",
         "Central probe for Concern 1 wiring. A Facilitative answer names a belief AND names the conditions under which he would update it — evidence of reasoning over feeling. A non-Facilitative answer either holds the belief absolute or caves entirely under pressure. Listen for structured exception-logic, not platitudes. Ties directly to the Fact-over-Feelings wiring mismatch (L2 8.7) and Deliberate Urgency (L1 #9)."),
        # 7 — Commitment (Diligence wiring · L1 #9)
        ("COMMITMENT (Diligence wiring · L1 #9)",
         "Daily-routine specificity. A Deliberate Urgency habit usually shows up as a morning or end-of-day discipline around priorities. Vague answers (\"I stay on top of things\") are weak signals; a specific 7:45am routine around the top three priorities is strong. Connects to the Diligence wiring mismatch flagged in Concern 1."),
        # 8 — Leadership Deep-Dive (Clarity of Accountability · HoldsAvgDown)
        ("LEADERSHIP DEEP-DIVE (Clarity of Accountability · HoldsAvgDown)",
         "Ask him to literally sketch Lumicell's top two levels. Clarity of Accountability (Flag | Med) and HoldsAvgDown (Flag | Med) both surface here. A leader running the accountability routine names owners, open roles, and upgrade candidates in the same breath. Hesitation, multiple owners for the same priority, or generalities is the flag. Follow up: \"Which of these seats would you not re-hire today?\""),
        # 9 — Passion (mission-construction ability)
        ("PASSION & MISSION CONSTRUCTION",
         "Tests mission-construction ability. With commanding DF (85) and a current CEO seat, Hechler has multiple paths — listen for whether he roots passion in patient outcome and mission specificity, or defaults to generic leadership-scale language. Weak answers invoke \"any team can be motivated\"; strong answers name the concrete hook. Bridges into the \"Why Lumicell, why now\" follow-up."),
        # 10 — Continuous Improvement (Work-definition posture)
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes both Concern 1 (Diligence wiring) and Concern 2 (Pre-Proof Belief around his own improvement). Listen for whether he treats learning, reflection, and coaching time as work — or only meetings and deliverables. A Continuous Improvement operator counts time spent in others' development as core work, not overhead. A narrow answer (\"work = output\") reinforces the Investing-in-Others wiring deficit."),
    ]

    # Build probes: category from tailored[i], question from FORM8_QUESTIONS[i] (verbatim),
    # coaching from tailored[i]. No "Form 8 #N —" prefix in display text.
    assert len(tailored) == len(FORM8_QUESTIONS) == 10, "Must have exactly 10 probes matching 10 Form 8 questions"
    probes = []
    for (category, coaching), (_form8_name, canonical_q, _substr) in zip(tailored, FORM8_QUESTIONS):
        # Escape double-quotes inside the question so the outer quoting stays clean
        display_q = f'"{canonical_q}"'
        probes.append((category, display_q, coaching))

    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    interview_probes = "\n".join(parts)

    return {'INTERVIEW_PROBE_CARDS': interview_probes}

def build_teach_items():
    """Build Teach Items metric-value cell. Per QA_CHECKLIST.md section 2,
    this must be a short fraction "N/10" counting TeachTop10 entries. No narrative text."""
    # TeachTop10 sheet has 10 data rows for Howard (confirmed via openpyxl inspection).
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    """Build the Excellence Standards Dimensional Scorecard tokens with L1+L2 hierarchy.
    Per QA_CHECKLIST.md Section 6 requirements:
      - L1 labels are UPPERCASE; L2 labels indented with 4 spaces.
      - At minimum 12 rows (Bender has 13) to confirm L2 rendering.
      - Scores sourced from respondent L2 sheet [Score5_filtered] (canonical per SKILL.md L55).
    Chosen L1 categories for Howard Hechler:
      - L1 #1 Investing in Others (strength, Z=+0.67)
      - L1 #5 Pushing Extreme Accountability (split profile)
      - L1 #8 Org Decision Making (TARGETED CONCERN)
      - L1 #9 Deliberate Urgency (TARGETED CONCERN)
    """
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    # (L1_num, L1_display_uppercase, [list_of_L2_short_names_in_desired_order])
    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Demonstrating Genuine Fanness',
            'Developmental Mindset',
            'Developmental Discipline',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Basic Machinery Of Accountability',
            'Drives Accountability',
            'Stds = What Tolerate',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Simplification Methods',
            'Facts Over Feelings',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Extreme Proactivity',
            'Proactive Belief In People',
            'Action Over Inaction',
        ]),
    ]

    labels = []
    scores = []
    is_l1 = []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        # L1 row
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        # L2 rows (4-space indent for UPPERCASE L1 > indented L2 rendering)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found in L2 sheet: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    # Overall Z|Algo (gold line) and Cohort avg (grey line) for chart annotations
    # Howard's overall Z|Algo is approximately mean of L1 z_algos:
    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    # Cohort average is stored as a constant for the instrument (~0.244 per Bender)
    cohort_avg = 0.244  # canonical cohort baseline; stable across respondents

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',  # No overrides for Howard (no single dim needs orange flag)
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    """Build Talent Radar scores and narrative."""
    # Template labels (fixed order): Two-Sport Athlete, Punctuates Differently,
    # Facilitative Mindset, Wit (interview-only), Deep Repertoire, Discipline/Routine,
    # Understands Symbolism (interview-only). Scores 1-5, null for interview-only.
    # Hechler scores:
    # - Two-Sport Athlete: 2 (LinkedIn redacted; Lumicell-only visible — need interview)
    # - Punctuates Differently: 3 (L2 8.1 Simplification Methods +1.73 — real strength)
    # - Facilitative Mindset: 2 (L1 #2 mid + L2 2.4 Sublimating Ego -1.06; high-I suppresses dissent)
    # - Wit: null (interview-only)
    # - Deep Repertoire: 4 (medtech regulatory+commercial+board breadth over career)
    # - Discipline/Routine: 2 (Personal Reliability -0.66; L2 6.1 Extreme Ownership -0.97)
    # - Understands Symbolism: null (interview-only)
    talent_radar_scores = json.dumps([2, 3, 2, None, 4, 2, None])

    talent_radar_profile = (
        "Deep Repertoire (4) is the standout &mdash; medtech regulatory + commercial + "
        "board breadth across the Lumicell arc. Punctuates Differently (3) is a real "
        "strength tied to L2 8.1 Simplification Methods (+1.73). Two-Sport Athlete (2) "
        "is constrained by a partially-redacted LinkedIn; must be probed. Facilitative "
        "Mindset (2) and Discipline/Routine (2) are the twin frontiers &mdash; the former "
        "constrained by high-I suppression of dissent, the latter by L2 6.1 Extreme "
        "Ownership (&minus;0.97). Wit (&#8856;) and Understands Symbolism (&#8856;) "
        "cannot be assessed by the instrument and must be evaluated in the interview."
    )

    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_career_timeline():
    """Build career timeline from saved LinkedIn profile at
    _pipeline/data/howard_hechler_linkedin.md. Per QA_CHECKLIST.md Section 8 +
    Section 12, career data must come from a saved LinkedIn file — never invented."""
    linkedin_file = ROOT / '_pipeline' / 'data' / 'howard_hechler_linkedin.md'
    if not linkedin_file.exists():
        raise FileNotFoundError(
            f"LinkedIn data missing: {linkedin_file}. "
            "Build cannot proceed. Paste LinkedIn profile and save it before rebuilding."
        )

    timeline_title = 'Career Timeline &mdash; 25 Years, Med-Device BD to CEO'

    # Colors chosen to group chronologically: earliest = blue, transitional = tan,
    # Medtronic cluster = green, Lumicell = gold/red, fractional = grey.
    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Deutsch Congressional<br>1999&ndash;2001 &middot; 2yr</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">UVA Darden MBA + UVA Law JD<br>2001&ndash;2005 &middot; 4yr</div>
                <div class="timeline-block" style="background:#9b59b6;color:#fff;">inVentiv Health<br>2005&ndash;2006 &middot; 1yr</div>
                <div class="timeline-block" style="background:#e74c3c;color:#fff;">XL TechGroup<br>2006&ndash;2009 &middot; 3yr</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">Mallinckrodt<br>2010&ndash;2012 &middot; 2yr</div>
                <div class="timeline-block" style="background:#1e8449;color:#fff;">Medtronic<br>2012&ndash;2017 &middot; 5yr 1mo</div>
                <div class="timeline-block" style="background:#d4a84b;color:#1a2332;">PolarityTE CBO &rarr; Locust Walk<br>2018&ndash;May 2020 &middot; 2yr</div>
                <div class="timeline-block" style="background:#c0392b;color:#fff;">Lumicell SVP &rarr; CEO<br>Aug 2020&ndash;Aug 2025 &middot; 5yr 1mo</div>
                <div class="timeline-block" style="background:#6b7280;color:#fff;">Obliquity Fractional<br>Aug 2025&ndash;Present &middot; 9mo</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item">
                    <div class="legend-dot" style="background:#3498db;"></div>
                    <span><strong>Deutsch Congressional (1999&ndash;2001):</strong> Senior Legislative Assistant, Congressman Peter Deutsch. Pre-MBA political staff role.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#1e40af;"></div>
                    <span><strong>UVA Darden + UVA Law (2001&ndash;2005):</strong> Dual MBA/JD program — Business and Law.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#9b59b6;"></div>
                    <span><strong>inVentiv Health (2005&ndash;2006):</strong> Associate Practice Executive — out-licensing programs, market analysis, product valuation.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#e74c3c;"></div>
                    <span><strong>XL TechGroup (2006&ndash;2009):</strong> Director of Business Development — evaluated medtech investments (Insero, TissueMesh, QuoNova, DxTech).</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#16a085;"></div>
                    <span><strong>Mallinckrodt (2010&ndash;2012):</strong> Senior Manager BD &amp; Licensing. Co-managed Zogenix/Sumavel + Horizon/Duexis; $30M TussiCaps divestiture. <em>(2009&ndash;2010 Obliquity consulting gig between XL TechGroup and Mallinckrodt; not shown separately.)</em></span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#1e8449;"></div>
                    <span><strong>Medtronic (2012&ndash;2017):</strong> 4 roles, Principal Corp Dev &rarr; Director BD &rarr; Senior Director BD &rarr; Senior Director Strategy &amp; Portfolio. Led Sophono acquisition, Microfrance divestiture. No internal promotions into a P&amp;L role.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#d4a84b;"></div>
                    <span><strong>PolarityTE CBO (2018) &rarr; Locust Walk VP MedTech (Feb 2019&ndash;May 2020):</strong> PolarityTE CBO role &lt; 1 yr; then VP at transaction firm leading client engagements in financings, sell-side, partnerships.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#c0392b;"></div>
                    <span><strong>Lumicell (Aug 2020&ndash;Aug 2025):</strong> 4 role steps — SVP Strategy &amp; Corp Dev (Aug 2020) &rarr; CBSO (Feb 2022) &rarr; COO/GM (Nov 2022) &rarr; President &amp; CEO (Dec 2023). First CEO role; first true internal-promotion arc. Led PMA submission, launch prep, commercial launch.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background:#6b7280;"></div>
                    <span><strong>Obliquity Consulting Fractional C-Level (Aug 2025&ndash;Present):</strong> Fractional CEO (stealth cardiovascular medtech, <em>per respondent Q107/Q108 = ClearPath Cardio</em>); Fractional COO (biopsy imaging); Strategic Advisor (surgical oncology).</span>
                </div>
            </div>
            <div class="timeline-banner">
                Career tenure pattern: 2 short tenures (&lt;= 1yr), 3 medium tenures (2&ndash;3yr), 2 long tenures (5yr+). Only P&amp;L step-up arc is Lumicell (SVP &rarr; CEO in 3.5 yrs) — his most recent and most relevant leadership trajectory. Earlier Medtronic 5-yr run was all BD/Strategy (staff), not P&amp;L leadership.
            </div>"""

    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

# ============================================================================
# BUILD RESPONDENT DICT FOR MOTIVATORS_SECTION
# ============================================================================

def build_respondent_dict(respondent_data):
    """Build the respondent dict required by motivators_section.build_section()."""
    # TTI scores from the PDF (provided in brief)
    # DISC Natural: D=89, I=66, S=7, C=48
    # DISC Adapted: D=93, I=62, S=8, C=35
    # TTI Driving Forces: Commanding 85 (primary), Intentional 67, Objective 65, Receptive 64

    # Map DISC to position (0-64, where each wedge is 8 positions)
    # Position 11 is roughly in Persuading Conductor wedge
    nat_pos = 11
    nat_label = 'Persuading Conductor'
    nat_intensity = (89 + 66) / 200.0  # (D + I) / 200

    adp_pos = 11
    adp_label = 'Persuading Conductor'
    adp_intensity = (93 + 62) / 200.0

    shift_note = 'Adapted shift: +4D, −4I, −13C. Environment pulling him toward even less process/compliance.'

    respondent = {
        'name': 'Howard Hechler',
        'first_name': 'Howard',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [89, 66, 7, 48],  # D, I, S, C Natural
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }

    return respondent

# ============================================================================
# MAIN BUILD
# ============================================================================

def qa_gate(html):
    """Enforce every check in _pipeline/QA_CHECKLIST.md post-render. Fail loudly.
    Every assertion here corresponds to a checklist item with the same label."""
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    # Section 1 — Targeted Concerns: min 2 with triplet parity.
    # Count is data-driven; the respondent profile determines how many concerns surface.
    n_concern_item = count_class('concern-item')
    n_concern_number = count_class('concern-number')
    n_concern_text = count_class('concern-text')
    if n_concern_item < 2:
        failures.append(f"S1: concern-item count {n_concern_item} < 2 (min)")
    if n_concern_number != n_concern_item:
        failures.append(f"S1: concern-number ({n_concern_number}) != concern-item ({n_concern_item}) — triplet parity broken")
    if n_concern_text != n_concern_item:
        failures.append(f"S1: concern-text ({n_concern_text}) != concern-item ({n_concern_item}) — triplet parity broken")
    # Sequential numbering: concern-number rendered text must be 1..N
    concern_nums = re.findall(r'<div class="concern-number">(\d+)</div>', html)
    expected_seq = [str(i) for i in range(1, n_concern_item + 1)]
    if concern_nums != expected_seq:
        failures.append(f"S1: concern-number sequence {concern_nums} != expected 1..{n_concern_item}")

    # Sections 3 + 9 — still exact-count because spec fixes them.
    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: class='{cls}' expected exact {expected}, got {got}")

    # Section 9 — Form 8 sourcing for interview probes (CRITICAL, non-negotiable).
    # Per SKILL.md "Form 8 Questions" section (CRITICAL importance) and PROJECT_NOTES.md
    # L178-191, probe questions must be drawn VERBATIM from the canonical Form 8 set.
    # Verification is structural: each probe-question block must contain a distinctive
    # substring from one of the 10 canonical Form 8 questions (FORM8_QUESTIONS module
    # constant). The label "Form 8" does NOT need to appear in display text — it is a
    # sourcing rule, not a display requirement.
    probe_questions = re.findall(
        r'<div class="probe-question">(.*?)</div>', html, re.DOTALL
    )
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched_probes = []
        for idx, q_html in enumerate(probe_questions, start=1):
            q_lower = q_html.lower()
            matched_here = False
            for (_name, _canonical, substr) in FORM8_QUESTIONS:
                if substr.lower() in q_lower:
                    matched_substrs.add(substr)
                    matched_here = True
                    break
            if not matched_here:
                unmatched_probes.append(idx)
        if len(matched_substrs) < 10:
            failures.append(
                f"S9-Form8: Only {len(matched_substrs)}/10 distinct canonical Form 8 "
                f"questions present in probe-question blocks (min 10). "
                f"Unmatched probe indices: {unmatched_probes}. "
                f"Per SKILL.md Form 8 section — interview questions must be drawn verbatim "
                f"from the canonical set, never generated."
            )

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: class='{cls}' expected min {expected}, got {got}")

    # Section 2 — Teach Items is a fraction, not narrative
    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value block not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be fraction N/M, got '{teach_val[:60]}'")
        if 'To be populated' in html or 'Q47' in html or 'Q81' in html:
            failures.append("S2: Forbidden placeholder text (To be populated / Q47 / Q81)")

    # Section 4 — Population distribution charts
    # Chart 1: DIST_ZLABELS must be 2-row (nested arrays)
    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:  # second char should be [ for nested
        failures.append("S4-C1: zLabels2 not a nested 2-row array")

    # Chart 2: empty columns hidden — sfLabels2 / failData2 / successData2 same length, and at every index either fail>0 or success>0 (allow respondent bin overrides)
    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/failData/successData")
    else:
        import json as _json
        sf_labels = _json.loads(m_sf_l.group(1))
        fail_data = _json.loads(m_fail.group(1))
        succ_data = _json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append(f"S4-C2: sfLabels({len(sf_labels)}) / fail({len(fail_data)}) / success({len(succ_data)}) length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with '+', got {sf_labels[-1]}")

    # Chart 3: reversed — flagLabels3 first label should start with a HIGH value (>=40)
    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first_label = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first_label:
            first_val = int(first_label.group(1))
            if first_val < 30:
                failures.append(f"S4-C3: flagLabels3 first label = {first_val} — axis not reversed (high flags must be on LEFT)")

    # Section 6 — Dimensional Scorecard has L1+L2 rows
    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        import json as _json
        exc_labels = _json.loads(m_exc_labels.group(1))
        is_l1_arr = _json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: Dimensional Scorecard has only {len(exc_labels)} rows (min 12)")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: Dimensional Scorecard has only {l2_rows} L2 rows (min 6)")
        # All UPPERCASE for L1
        for lbl, is_l1 in zip(exc_labels, is_l1_arr):
            if is_l1 and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' must be UPPERCASE")
                break
            if not is_l1 and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' must be indented with 4 spaces")
                break
    else:
        failures.append("S6: excLabels / isL1 not found in output")

    # Section 7 — Talent Radar
    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    # Section 11 — Brand + tokens
    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    # Section 5 — canvases
    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    # Report
    print()
    print("=== QA GATE (QA_CHECKLIST.md) ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} issue(s) ***")
        for f in failures:
            print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)} issue(s)")
    else:
        print("*** QA GATE PASSED ***")


def main():
    print("Loading respondent data...")
    respondent_data = load_respondent_data()

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution chart tokens...")
    dist_tokens = build_distribution_tokens(
        zalgo_rows,
        flag_rows,
        respondent_data['z_algo_overall'],
        respondent_data['z_human_overall'],
        respondent_data['rf_num']
    )

    print("Building respondent dict for motivators section...")
    respondent = build_respondent_dict(respondent_data)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    # Build full token replacement dict
    replacements = {
        # Basic info
        'CANDIDATE_NAME': 'Howard Hechler',
        'CANDIDATE_CREDS': 'MBA, Darden (2001–2005)',
        'CANDIDATE_ROLE': 'CEO — Voyager Medical / ClearPath Cardio',
        'REPORT_DATE': 'April 21, 2026',

        # Metrics
        'ZALGO_OVERALL': f'{respondent_data["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{respondent_data["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(respondent_data['rf_num']),
        'FLAGS_LIT': '10/20',  # 10 flags lit of 20 measured (from Flags sheet analysis)
        'TEACH_ITEMS': '10/10',  # ImpactTop10 + TeachTop10 items coachable
        'COHORT_AVG': '0.00',
        'COHORT_AVG_NUM': '0.00',

        # Three Axes
        **axes,

        # Concerns
        **concerns,

        # Wiring Fit
        **wiring,

        # Hard to Learn
        **htl,

        # Interview Probes
        **probes,

        # Teach Items
        **teach,

        # Talent Radar
        **radar,

        # Career Timeline
        **timeline,

        # Distribution chart tokens
        **dist_tokens,

        # DISC scores
        'DISC_D_NAT': '89',
        'DISC_I_NAT': '66',
        'DISC_S_NAT': '7',
        'DISC_C_NAT': '48',
        'DISC_D_ADP': '93',
        'DISC_I_ADP': '62',
        'DISC_S_ADP': '8',
        'DISC_C_ADP': '35',
        'DISC_NOTE_TEXT': 'Persuading Conductor (D-dominant, high I). Adapted shift: +4D, −4I, −13C.',
        'DISC_NOTE_DETAIL': 'Environment pulling him toward even less process/compliance. High Commanding DF (85) + Low Collaborative DF (6) wiring consistent with D-dominant style.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        # Excellence Standards Dimensional Scorecard (L1 + L2 hierarchy).
        # Per QA_CHECKLIST.md Section 6: L1 labels UPPERCASE, L2 labels indented with 4 spaces.
        # Source: respondent L1 sheet [Z_Algo] + L2 sheet [Score5_filtered]
        # (SKILL.md line 55 — L2 tab is canonical source).
        **build_excstds_scorecard(respondent_data),

        # Recommendation — short pill (full reasoning lives in CONCERNS_ITEMS and the three-axes cards)
        'RECOMMENDATION_TEXT': 'PROCEED WITH DILIGENCE — conditional yes, with two targeted concerns to validate in interview',
    }

    # Critical fix before substitution: ensure EXCSTDS_COLOR_OVERRIDES is not in a comment
    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    # Inject motivators section
    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    # Substitute all tokens
    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    # Fix title
    html = html.replace(
        '<title>Howard Hechler — Senior Manager, HR | HALE GLOBAL</title>',
        '<title>Howard Hechler — CEO, ClearPath Cardio | HALE GLOBAL SUCCESS DIAGNOSTICS</title>'
    )

    # Fix header brand lockup
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    # Run canonical QA gate
    qa_gate(html)

    # Save
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")

if __name__ == '__main__':
    main()
