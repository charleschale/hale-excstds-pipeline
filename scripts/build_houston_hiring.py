"""Build the Excellence Standards hiring report for Megan Houston (Director, Global HRBP).

Forked from build_lobreglio_hiring.py 2026-04-27. Currently Director, Global Business
Partner at Pax8 (Jul 2023-Present). Promoted internally 4 times across an 8-year career:
McAfee (Provisioning Support → Sales Operations Analyst), then Pax8 (HR Generalist →
People Operations Business Partner → Manager of People Operations → Director, Global
Business Partner). Education: B.A. Communication Studies, Colorado State University 2007-
2011. Evaluated for a VP/Director step-up at a Public/Mature company.

Run from repo root:
    python _pipeline/scripts/build_houston_hiring.py
Outputs to:
    _reports/Houston_Megan_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section, compute_intensity_from_disc

RESPONDENT_XLSX = ROOT / '_respondents' / '20260409.meganhouston19@gmail.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Houston_Megan_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title, 'z_algo': z_algo, 'z_human': z_human, 'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flag_names = []
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        if flag_name:
            flag_names.append((col, flag_name))
    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data, 'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall, 'z_human_overall': z_human_overall,
        'rf_num': rf_num, 'questions_answered': questions_answered,
        'flags_lit': flags_lit, 'non_scorable': non_scorable,
        'name': name, 'email': email, 'date_str': date_str,
    }

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]
    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo, z_human, sf = row['z_algo'], row['z_human'], row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1; continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1; break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    rev_flag_bin = len(flag_counts) - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================

# ============================================================================
# NARRATIVE
# ============================================================================

def build_three_axes_narratives():
    talent_badge_text = 'AMBER-GREEN'
    talent_card_body = '''Talent is read holistically on three lenses. <strong>Consistent promotions</strong> is the strongest line on the file: 4 internal promotion events in 8 years (McAfee Provisioning &rarr; Sales Ops; Pax8 HR Generalist &rarr; People Ops BP &rarr; Manager &rarr; Director, Global Business Partner). The Pax8 record is exec-altitude evidence &mdash; $6M+ enterprise-wide RIF strategy, Support-at-Scale transformation, global job-architecture design across AMER/EMEA/APAC during a 1,122% revenue-growth period.

<strong>Greatness in more than one area</strong> is light: single craft (HR since 2018), single school (Colorado State, B.A. Communication Studies 2007&ndash;2011 &mdash; non-HR major), no MBA, no SHRM-SCP / SPHR certification, no awards / board service / cross-domain pursuit on the file.

<strong>Capacity to grow into the next altitude</strong> is the tension: the promotion record is genuine evidence of demonstrated competence at progressively higher altitudes, but the instrument data (Z|Algo &minus;1.75, lowest in the working set) argues that the muscles a VP step-up most needs &mdash; decisive judgment, standard-holding, urgency &mdash; are precisely the muscles she has not yet built.'''

    judgment_badge_text = 'RED'
    judgment_card_body = '''Judgment is anchored on <strong>L1 #8 Organizational Decision Making</strong>: <strong>&minus;2.22 Algo / &minus;3.12 Human</strong> &mdash; the most negative L1 reading on her file and one of the worst in the working set. Both Algo and Human reviewers agree (no disagreement to argue with). The L2 mass shows where the failure actually lives: <strong>8.2 Clarity of Accountability &minus;1.71</strong>, <strong>8.1 Simplification Methods &minus;1.41</strong>, <strong>8.7 Facts Over Feelings &minus;1.11</strong> &mdash; every supporting L2 is deeply negative.

The decisional posture this paints: she does not crisply assign single-point ownership, does not simplify decisions to one prioritized question, does not separate facts from feelings under pressure. Direct judgment-architecture flag corroboration: <strong>Clarity of Accountability Hi</strong>, <strong>Initiating Accountability</strong> lit, <strong>Driving Accountability Med</strong>. TTI Conflict section adds <em>&ldquo;difficulty making ambiguous choices when confronted with changing the system&rdquo;</em>.

Other instrument concerns (Deliberate Urgency &minus;2.51, Conducting &amp; Outvoted &minus;1.65, Not Pleasing &minus;1.39) are real and addressed in the Targeted Concerns and Role-Fit sections below; they are separate failure modes from the judgment axis.'''

    skills_badge_text = 'AMBER'
    skills_card_body = '''Domain credentials are real but light on certification. 8 years of HR practice (HR Generalist 2018 &rarr; Director, Global Business Partner 2023+), multi-region scope at Pax8 (AMER/EMEA/APAC), real exec-altitude outputs ($6M+ enterprise-wide RIF strategy, global org architecture, Support-at-Scale transformation). B.A. Communication Studies, Colorado State 2007&ndash;2011. <strong>No MBA, no SHRM-SCP / HRCI SPHR / GPHR certification</strong> on the LinkedIn record. Skills endorsements modest (1&ndash;9 per skill).

Wiring-fit (full read in the Wiring-Fit panel below): wheel position 36 Supporting Coordinator (FLEXIBLE) is partially seat-aligned for HR-Operations / People-Ops work but lacks the decisiveness a VP-of-HR step-up requires when she IS the standard-holder.'''

    return {
        'TALENT_BADGE_CLASS': 'badge-amber',
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': 'badge-red',
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': 'badge-amber',
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    """Concerns moved to Interview section (see build_concerns_intro).
    The legacy concerns-box in the template auto-hides via CSS :empty rule when these
    tokens are empty strings."""
    return {'CONCERNS_TITLE': '', 'CONCERNS_ITEMS': ''}

def build_wiring_fit():
    wiring_fit_items = (
        '<strong>Wheel positions: Natural and Adapted both at position 36, Supporting Coordinator (FLEXIBLE).</strong> '
        'Anchored in the Coordinator/Supporter overlap zone of the wheel; <em>neither position is ACROSS</em> '
        '(not center-of-wheel). The FLEXIBLE marker indicates she can move between Coordinator (high-C, '
        'process-and-systems) and Supporter (high-S, people-care) modes within the wedge depending on '
        'situation. Effectively <strong>no Natural-to-Adapted shift</strong> &mdash; D=17 &rarr; 25, I=62 &rarr; 58, '
        'S=74 &rarr; 72, C=71 &rarr; 64. She is operating in her natural wiring at Pax8.'
        '<span class="wiring-flag">Diagnostic</span><br>'
        '<strong>The TTI corroborates the instrument with rare line-by-line clarity.</strong> '
        'Time-Wasters page names <em>&ldquo;Not Exercising Authority&rdquo;</em> explicitly. Conflict section: '
        '<em>&ldquo;struggles with delivering a tough message,&rdquo; &ldquo;may not always stand up for oneself or '
        'others,&rdquo; &ldquo;keeps to herself resulting in unexpressed viewpoints.&rdquo;</em> Behavioral '
        'Characteristics: <em>&ldquo;may not project a sense of urgency.&rdquo;</em> DF Indifferent: <strong>Resourceful '
        '17 (near 2&sigma; below mean)</strong>, Objective 28, Intellectual 32 &mdash; not driven by results-and-'
        'efficiency or by objectivity. The wiring is genuinely supportive and process-oriented, NOT decisive.'
        '<span class="wiring-flag">Targeted Concern</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    # Houston: Urgency -2.51 (lit, severe); Org DM -2.22/-3.12 (lit, catastrophic); Conditional Belief Sev (lit);
    # Satisfied with Gripes — NOT lit. Total 3/4.
    return {'HARD_TO_LEARN': '3/4'}

FORM8_QUESTIONS = [
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]

def build_interview_probes():
    tailored = [
        ("TWO-SPORT ATHLETE (Talent axis &middot; capacity-to-grow)",
         "The Talent record is real &mdash; 4 internal promotions in 8 years. Listen for whether the proudest result is one of those promotion / responsibility-expansion moments at Pax8 (the $6M RIF strategy, the Support-at-Scale transformation, the global job architecture rollout) OR something further afield. The cleanest strong answer points to a personal pursuit, a stretch project outside the HR remit, a non-work achievement, OR a specific teammate she invested in who then succeeded. Defaulting to her current Director title or a single Pax8 program is the average answer."),
        ("TALENT DEVELOPMENT (Concern 1 &middot; Pre-proof belief)",
         "L1 #9 Proactive Belief In People <strong>&minus;1.96</strong> + L1 #1 Investing in Others &minus;0.54 + Conditional Belief Sev + Lower Standards for Others than Self lit + DF Selfless-primary describe a leader who supports people warmly but may not extend developmental capital before they have proven themselves. Listen for one specific case where she backed someone <em>before</em> they earned it &mdash; a stretch role for someone struggling, a defended hire, a specific coaching-cadence with someone underperforming. The Q9918 referral letter named her <em>&ldquo;listening, building trust, employee relations&rdquo;</em> as strengths &mdash; the gap between being a trusted listener and being a developmental backer is the read."),
        ("TORC &mdash; TRUTH-OVER-COMFORT (Self-awareness)",
         "She has already shown unusual self-awareness here. Q9917 self-report names Lori Fraiser&rsquo;s explicit feedback to her: <em>&ldquo;be more explicit in my point of view and recommendations.&rdquo;</em> Q9918 names the separate referral feedback: <em>&ldquo;compassionate and affirming listening skills can sometimes be interpreted as agreement.&rdquo;</em> Anchor the question to a real boss (Lori, or her current Pax8 leadership). The strong answer goes <em>past</em> what she has already disclosed in writing &mdash; names a specific developmental edge her current boss is working with her on right now. The weak answer reaches back to the Lori feedback again."),
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "8 years at Pax8 spanning the 1,122% growth period AND the Support-at-Scale RIF transformation puts real adversity material on the table. Listen for whether she names a personal adversity or defaults to professional. If professional, listen for ownership location: internal (&ldquo;I should have done X differently&rdquo;) or external. The healthy answer attributes learning to self even when the context was genuinely outside her control."),
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior probe. Wiring is Coordinator/Supporter; DF Selfless/Structured/Harmonious. The most natural strong answer is a specific listening / empathy-driven routine that compounds &mdash; an engagement program structure, a specific 1:1 cadence with leaders she partners with, a process she designed and ran end-to-end. Generic answers (&ldquo;I work harder,&rdquo; &ldquo;I care more about people&rdquo;) are negative signals."),
        ("FACILITATIVE MINDSET (Concern 1 &middot; Dialogue and dissent)",
         "Central probe for Concern 1. The TTI conflict and behavioral sections explicitly flag the gap: <em>&ldquo;may not always stand up for oneself or others,&rdquo; &ldquo;keeps to herself resulting in unexpressed viewpoints,&rdquo; &ldquo;struggles with delivering a tough message even if it&rsquo;s to the person&rsquo;s benefit.&rdquo;</em> A Facilitative answer names a belief AND the conditions under which she would update it AND ideally cites a recent specific case where she pushed back / held the line / named her position when it would have been easier to defer. Follow-up: <em>&ldquo;When was the last time you disagreed with a senior stakeholder publicly and kept disagreeing?&rdquo;</em>"),
        ("COMMITMENT (Diligence wiring)",
         "Personal Reliability L1 &minus;0.77 and Commitment to Routine L2 &minus;0.84 are <em>both negative</em> &mdash; unusual for a 4-promotion career. The instrument argues that her routine-machinery is not where her professional success has come from; success has come from relationship-building and process-following inside Pax8 specifically. Listen for whether she names a concrete daily practice. Vague answers consistent with the instrument; press for the specific artifact."),
        ("LEADERSHIP DEEP-DIVE (Concerns 1 &amp; 2 &middot; Org-chart conversation)",
         "Ask her to sketch the Pax8 HRBP org she runs today. Tests both Targeted Concerns directly. With Conditional Belief Sev + Lower Standards for Others than Self + Upgrade Team flag lit, follow up: <em>&ldquo;Which of these seats would you not re-hire today?&rdquo;</em> &mdash; this is the diagnostic question. Hesitation, generic warmth, or &ldquo;everyone has potential&rdquo; are the flags. Then: <em>&ldquo;What would change in your routines at VP altitude?&rdquo;</em> &mdash; the strong answer names specific routines she would install (skip-level cadences, talent calibration, performance-management discipline); the weak answer narrates more of the Director-altitude work."),
        ("PASSION &amp; MISSION CONSTRUCTION",
         "Tests mission-construction. The Pax8 work has a clear mission narrative (scaling SaaS through 1,122% growth, then doing it humanely through RIFs). Listen for whether she roots passion in the specific kind of HR problem (large-scale change management, leadership development, people-analytics) or defaults to scale-language. Strong answers name the concrete hook; weak answers invoke <em>&ldquo;help leaders be the best version of themselves.&rdquo;</em>"),
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes Concern 1. With Resourceful 17 (DF Indifferent, near 2&sigma; below mean) and DF Selfless/Altruistic primary, the pattern argues she counts <em>helping others</em> as work but may NOT count <em>holding the line / driving consequences</em> as work. The strong answer explicitly names performance-management, calibration, and standards-enforcement as core HR-leader work &mdash; not just process-and-partnership."),
    ]
    assert len(tailored) == len(FORM8_QUESTIONS) == 10
    probes = []
    for (category, coaching), (_n, canonical_q, _s) in zip(tailored, FORM8_QUESTIONS):
        probes.append((category, f'"{canonical_q}"', coaching))
    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    return {'INTERVIEW_PROBE_CARDS': "\n".join(parts)}

def build_teach_items():
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (9, 'DELIBERATE URGENCY', [
            'Action Over Inaction',
            'Proactive Belief In People',
            'Extreme Proactivity',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Clarity Of Accountability',
            'Simplification Methods',
            'Facts Over Feelings',
        ]),
        (3, 'CONDUCTING & OUTVOTED', [
            'Conductor > Lead Guitarist',
            'Empower Team Authority',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Stds = What Tolerate',
            'Drives Accountability',
            'Basic Machinery Of Accountability',
        ]),
    ]

    labels, scores, is_l1 = [], [], []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")

    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    # Two-Sport Athlete: 3 (4 internal promotions over 8 years is the strong line; held back by single
    #     craft / single school / no awards / no varsity)
    # Punctuates Differently: 2 (no rare-behavior signal jumps off the file; Personal Reliability -0.77)
    # Facilitative Mindset: 2 (Dialogue +0.10 mid; Sublimating Ego -0.38; Conducting -1.65)
    # Wit: null
    # Deep Repertoire: 3 (8-yr HR career with global multi-regional scope at Pax8, real exec-level
    #     outputs; held back by single-company HR-only context)
    # Discipline/Routine: 2 (Personal Reliability L1 -0.77, Commitment to Routine L2 -0.84 — unusual
    #     for a 4-promotion career, but the instrument is what it is)
    # Understands Symbolism: null
    talent_radar_scores = json.dumps([3, 2, 2, None, 3, 2, None])
    talent_radar_profile = (
        "<strong>Two-Sport Athlete (3)</strong> &mdash; the strongest dimension on the radar, anchored in the "
        "<em>consistent-promotions</em> lens: 4 internal promotion events across an 8-year career (McAfee "
        "Provisioning &rarr; Sales Ops; Pax8 HR Generalist &rarr; People Ops BP &rarr; Manager &rarr; Director). "
        "Held back from 4+ by the greatness-in-more-than-one-area lens: single craft (HR), single school "
        "(Colorado State Communication Studies), no certifications, no awards, no varsity, no cross-domain "
        "pursuit on the file. <strong>Deep Repertoire (3)</strong> &mdash; real exec-level outputs at Pax8 "
        "($6M RIF strategy, global org design across AMER/EMEA/APAC, Support-at-Scale transformation), but "
        "single-company HR-only context. <strong>Punctuates Differently (2)</strong> &mdash; Personal Reliability "
        "L1 &minus;0.77 (negative for once); no rare-behavior signal jumps off the file; interview required. "
        "<strong>Facilitative Mindset (2)</strong> &mdash; Dialogue vs Direction L2 +0.10 only mid; Conducting &amp; "
        "Outvoted L1 &minus;1.65; the boss + referral + TTI + instrument all name the standard-holding gap. "
        "<strong>Discipline/Routine (2)</strong> &mdash; Personal Reliability L1 &minus;0.77, Commitment to Routine "
        "L2 &minus;0.84 are both negative &mdash; unusual for a 4-promotion career, but the instrument is what "
        "it is. <strong>Wit (&#8856;)</strong> and <strong>Understands Symbolism (&#8856;)</strong> interview-only."
    )
    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_role_fit():
    role_fit_title = 'Role-Fit Read &mdash; What Will Be Easy, What Will Be Hard'
    role_fit_seat = ('Evaluating for: VP / Director step-up at a Public / Mature company &middot; '
                     'currently Director, Global Business Partner at Pax8 (~2,000 employees)')
    role_fit_easy = (
        '<strong>The promotion record is real and rare.</strong> 4 internal promotions across 8 years '
        '(McAfee Provisioning &rarr; Sales Ops; Pax8 HR Generalist &rarr; People Ops BP &rarr; Manager &rarr; '
        'Director, Global Business Partner). The Pax8 work itself is exec-altitude evidence: $6M+ enterprise-'
        'wide RIF strategy, Support-at-Scale transformation with multimillion-dollar cost savings, global '
        'job-architecture redesign, executive-leadership talent reviews and succession planning across AMER/'
        'EMEA/APAC.<br><br>'
        '<strong>HR-business-partner skill is the strongest part of the file.</strong> 8 years of HR practice, '
        'multi-region (AMER/EMEA/APAC), through a 1,122% revenue-growth period at a 2,000-employee SaaS &mdash; '
        'with a high-trust developmental relationship with her boss (Lori Fraiser, per Q9917 self-report) and '
        'positive referral-letter content (Q9918) describing her as a <em>&ldquo;servant leader,&rdquo;</em> <em>&ldquo;trusted '
        'advisor to the business,&rdquo;</em> and <em>&ldquo;stabilizing force amidst chaos.&rdquo;</em><br><br>'
        '<strong>Wiring is partially seat-aligned for an HR-Operations / People-Ops role.</strong> Position 36 '
        '(Supporting Coordinator, FLEXIBLE), DISC C=71 + S=74 Natural, DF Selfless / Structured / Harmonious '
        'primary cluster &mdash; the wiring carries process-and-systems work and people-care work cleanly. '
        'Effectively no Natural-to-Adapted gap; she is operating in her natural wiring.'
    )
    role_fit_hard = (
        '<strong>Step back: the wiring is anchored in the Coordinator/Supporter overlap, not in a leadership-'
        'altitude wedge.</strong> Position 36, Supporting Coordinator (FLEXIBLE) &mdash; clearly anchored in '
        'the lower-right quadrant of the wheel (NOT ACROSS, not center-of-wheel). High-S, high-C, low-D '
        'wiring is <em>partially</em> seat-aligned for HR work but lacks the decisiveness and standard-holding '
        'edge that a VP step-up at a Fortune 100 requires. DF Indifferent: Resourceful 17 (near 2&sigma; '
        'below mean), Objective 28, Intellectual 32 &mdash; she is not driven by results-and-efficiency, by '
        'objectivity, or by knowledge-acquisition. The Coordinator-Supporter wedge is process-and-people; '
        'the seat needs Conductor / Implementor decisiveness.<br><br>'
        '<strong>Z|Algo overall &minus;1.75 is the lowest in the working set.</strong> Z|Human &minus;1.66 (no '
        'Algo/Human disagreement). RF=20 reverse flags out of 90 questions answered. The mass concentrates '
        'in the dimensions a VP step-up most needs: Deliberate Urgency &minus;2.51 / &minus;3.00 (bottom decile), '
        'Org Decision Making &minus;2.22 / &minus;3.12 (catastrophic both reads), Conducting &amp; Outvoted '
        '&minus;1.65 / &minus;2.22.<br><br>'
        '<strong>Three independent voices outside the instrument corroborate the standard-holding gap.</strong> '
        'Boss Lori Fraiser told her to <em>&ldquo;be more explicit in my point of view and recommendations&rdquo;</em> '
        '(Q9917 self-report). Separate referral letter (Q9918) said <em>&ldquo;compassionate and affirming listening '
        'skills can sometimes be interpreted as agreement.&rdquo;</em> TTI Time-Wasters page lists <em>&ldquo;Not '
        'Exercising Authority&rdquo;</em> as a named time-waster. Plus the instrument: Conditional Belief Sev, '
        'HandsOn Sev, Initiating Accountability lit, Lower Standards for Others than Self lit, Clarity of '
        'Accountability Hi.<br><br>'
        '<strong>The binary diligence question.</strong> What is the actual seat shape? <em>Option A: VP / Director '
        'of People Operations, HR Operations, or HRBP under a stronger CHRO who carries the standard-holding '
        'work.</em> Houston&rsquo;s wiring + skills + promotion record fit this seat well; the instrument-flagged '
        'gaps stay on the CHRO&rsquo;s plate. <em>Option B: VP-of-HR seat where she IS the standard-holding '
        'leader at scale.</em> The convergent evidence (instrument + TTI + boss + referral) argues against '
        'her readiness for this version of the seat. Onsite must surface which version applies.'
    )
    return {
        'ROLE_FIT_TITLE': role_fit_title,
        'ROLE_FIT_SEAT': role_fit_seat,
        'ROLE_FIT_EASY': role_fit_easy,
        'ROLE_FIT_HARD': role_fit_hard,
    }

def build_career_timeline():
    timeline_title = 'Career Timeline &mdash; Sales Ops at McAfee, Then 4 Promotions in 8 Years at Pax8'
    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Education<br>Colorado State<br>BA Communication Studies<br>2007-2011</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">McAfee &mdash; Provisioning &amp; Sales Support<br>Aug 2011 - Nov 2015 &middot; 4y4m<br>Denver Metro</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">McAfee &mdash; Sales Operations Analyst<br>Nov 2015 - Dec 2017 &middot; 2y2m<br>(internal promotion)</div>
                <div class="timeline-block" style="background:#1e8449;color:#fff;">Pax8 &mdash; HR Generalist<br>Feb 2018 - Jan 2020 &middot; 2y0m<br>(career pivot to HR)</div>
                <div class="timeline-block" style="background:#0e7490;color:#fff;">Pax8 &mdash; People Ops BP<br>Jan 2020 - Jun 2021 &middot; 1y6m<br>(promotion)</div>
                <div class="timeline-block" style="background:#7c3aed;color:#fff;">Pax8 &mdash; Manager of People Operations<br>Jul 2023 - Present &middot; 2y10m<br>(promotion)</div>
                <div class="timeline-block" style="background:#d4a84b;color:#1a2332;">Pax8 &mdash; DIRECTOR, Global Business Partner<br>Jul 2023 - Present &middot; 2y10m<br>(concurrent / promotion)</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item"><div class="legend-dot" style="background:#3498db;"></div><span><strong>Education:</strong> Colorado State University, B.A. Communication Studies (2007&ndash;2011). No MBA. No SHRM-SCP / HRCI SPHR / GPHR certification visible. Mid-tier state school; non-HR-major undergraduate degree.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e40af;"></div><span><strong>McAfee &mdash; Provisioning &amp; Sales Support (Aug 2011 - Nov 2015):</strong> 4y4m. Career-start sales-operations seat in Denver Metro.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#16a085;"></div><span><strong>McAfee &mdash; Sales Operations Analyst (Nov 2015 - Dec 2017):</strong> 2y2m. <strong>First internal promotion event.</strong> Remote.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e8449;"></div><span><strong>Pax8 &mdash; HR Generalist (Feb 2018 - Jan 2020):</strong> 2y0m. <strong>Career pivot from Sales Operations into HR.</strong> Managed payroll / benefits / 401(k) for 300+ employees, owned onboarding and offboarding processes.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#0e7490;"></div><span><strong>Pax8 &mdash; People Operations Business Partner (Jan 2020 - Jun 2021):</strong> 1y6m. <strong>Second internal promotion at Pax8.</strong> Designed and implemented employee engagement strategies; led HRIS implementation with ADP; partnered on voluntary and involuntary separations.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#7c3aed;"></div><span><strong>Pax8 &mdash; Manager of People Operations (Jul 2023 - Present):</strong> 2y10m. <strong>Third internal promotion at Pax8.</strong> Managed the People Operations Business Partner team overseeing HR systems, compliance, benefits administration, and end-to-end employee lifecycle. Note: ~2-year gap between Jun 2021 and Jul 2023 not detailed on LinkedIn.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#d4a84b;"></div><span><strong>Pax8 &mdash; Director, Global Business Partner (Jul 2023 - Present):</strong> 2y10m, concurrent / role expansion. <strong>Fourth promotion event &mdash; current role.</strong> Strategic HR leader partnering with C-Suite across AMER/EMEA/APAC. Owned enterprise-wide RIF strategy ($6M+ savings); led Support-at-Scale transformation; designed global job architecture; facilitated executive talent reviews and succession planning. ~2,000+ employee SaaS during 1,122% revenue-growth period.</span></div>
            </div>
            <div class="timeline-banner">
                Career pattern: 6 years of sales-operations work at McAfee (one internal promotion), then a career pivot to HR at Pax8 in 2018, where she has been promoted three additional times to her current Director, Global Business Partner role. <strong>4 internal promotion events across an 8-year career &mdash; the strongest internal-promotion velocity on any file in the working set.</strong> Single craft (HR since 2018), single school (Colorado State Communication Studies), no MBA, no professional certifications, no awards / board service / cross-domain pursuit visible. The Pax8 record is impressive; the question is whether it transfers to a VP step-up at a different company without the trusted Pax8 relationships and proven processes carrying her.
            </div>"""
    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

def build_respondent_dict(respondent_data):
    """TTI: Natural D=17 I=62 S=74 C=71 → Supporting Coordinator (36, FLEXIBLE);
            Adapted D=25 I=58 S=72 C=64 → Supporting Coordinator (36, FLEXIBLE).
       Both at the same wheel position (36), both FLEXIBLE (not ACROSS).
       Effectively no Natural-to-Adapted gap.
       DF Primary: Selfless 62, Structured 57, Harmonious 54, Altruistic 47.
       DF Indifferent: Receptive 32, Intellectual 32, Objective 28, Resourceful 17 (near 2σ below mean)."""
    nat_pos = 36
    nat_label = 'Supporting Coordinator'
    nat_disc = [17, 62, 74, 71]
    nat_intensity = compute_intensity_from_disc(nat_disc)  # 90/200 → 0.45

    adp_pos = 36
    adp_label = 'Supporting Coordinator'
    adp_disc = [25, 58, 72, 64]
    adp_intensity = compute_intensity_from_disc(adp_disc)  # 69/200 → 0.345

    shift_note = ('Natural-to-Adapted shift: +8D, &minus;4I, &minus;2S, &minus;7C. Effectively '
                  '<strong>no adaptation</strong> &mdash; both Natural and Adapted at wheel position 36 '
                  '(Supporting Coordinator, FLEXIBLE). She is operating in her natural wiring at Pax8.')

    respondent = {
        'name': 'Megan Houston',
        'first_name': 'Megan',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [17, 62, 74, 71],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }
    return respondent
def qa_gate(html):
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    # S1 — Top 2 Concerns now live inside the Interview section as a numbered ordered list
    # (per METHODOLOGY "Concerns are validation framing for the Interview" + QA_CHECKLIST §1).
    # The legacy concerns-box auto-hides via CSS when CONCERNS_TITLE/CONCERNS_ITEMS are empty.
    n_concerns_intro = count_class('concerns-intro')
    if n_concerns_intro < 1:
        failures.append(f"S1: concerns-intro wrapper expected min 1, got {n_concerns_intro}")
    ol_match = re.search(r'<div class="concerns-intro">(.*?)</div>\s*<div class="probes-subtitle">', html, re.DOTALL)
    if not ol_match:
        failures.append("S1: concerns-intro structure not found (expected concerns-intro followed by probes-subtitle)")
    else:
        n_li = len(re.findall(r'<li[^>]*>', ol_match.group(1)))
        if n_li < 2:
            failures.append(f"S1: concerns-intro must contain ≥ 2 <li> items, got {n_li}")
    if '2 Top Concerns' not in html:
        failures.append("S1: '2 Top Concerns' sub-header not found in Interview section")
    if 'Excellence Standards Interview Questions' not in html:
        failures.append("S1/S9: 'Excellence Standards Interview Questions' sub-header not found")
    if 'Validating the Targeted Concerns' not in html:
        failures.append("S9: Interview section title 'Validating the Targeted Concerns' not found")

    # S11f — Standards-beat-the-interview principle: no banned phrasings in Concerns/Interview block
    interview_match = re.search(r'<div class="probes-section">.*?<div class="footer">', html, re.DOTALL)
    if interview_match:
        interview_block = interview_match.group(0).lower()
        banned = [
            'validate whether',
            'interview behavior is the test',
            'test whether the standard',
        ]
        for phrase in banned:
            if phrase in interview_block:
                failures.append(f"S11f: banned phrasing '{phrase}' found in Concerns/Interview section")

    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: {cls} expected {expected}, got {got}")

    probe_questions = re.findall(r'<div class="probe-question">(.*?)</div>', html, re.DOTALL)
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched = []
        for idx, q in enumerate(probe_questions, 1):
            qlow = q.lower()
            hit = False
            for (_n, _c, substr) in FORM8_QUESTIONS:
                if substr.lower() in qlow:
                    matched_substrs.add(substr); hit = True; break
            if not hit: unmatched.append(idx)
        if len(matched_substrs) < 10:
            failures.append(f"S9-Form8: {len(matched_substrs)}/10 canonical strings; unmatched probes {unmatched}")

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: {cls} expected min {expected}, got {got}")

    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be N/M, got '{teach_val}'")
    if 'To be populated' in html:
        failures.append("S2: Forbidden placeholder 'To be populated'")

    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:
        failures.append("S4-C1: zLabels2 not nested 2-row")

    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/fail/success")
    else:
        sf_labels = json.loads(m_sf_l.group(1))
        fail_data = json.loads(m_fail.group(1))
        succ_data = json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append("S4-C2: length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with +, got {sf_labels[-1]}")

    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first:
            v = int(first.group(1))
            if v < 30:
                failures.append(f"S4-C3: flagLabels3 first = {v} — axis not reversed")

    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        exc_labels = json.loads(m_exc_labels.group(1))
        is_l1_arr = json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: scorecard rows {len(exc_labels)} < 12")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: L2 rows {l2_rows} < 6")
        for lbl, l1f in zip(exc_labels, is_l1_arr):
            if l1f and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' not UPPERCASE"); break
            if not l1f and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' not indented"); break
    else:
        failures.append("S6: excLabels/isL1 not found")

    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    for cls, mn in [('role-fit-box', 1), ('role-fit-grid', 1),
                     ('role-fit-col easy', 1), ('role-fit-col hard', 1)]:
        if count_class(cls) < mn:
            failures.append(f"S-RF: {cls} expected min {mn}")
    if 'What Will Be Easy' not in html or 'What Will Be Hard' not in html:
        failures.append("S-RF: column labels missing")
    seat_stage_keywords = ['Series A', 'Series B', 'Series C', 'Series D',
                           'Seed', 'late-stage', 'Late-stage', 'pre-IPO', 'Pre-IPO',
                           'Public', 'public/mature', 'Mature', 'Fortune', 'IPO']
    if not any(kw in html for kw in seat_stage_keywords):
        failures.append(f"S-RF: no seat-stage keyword found (one of {seat_stage_keywords})")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    # S10 — Recommendation badge length (added 2026-04-27, Schott build)
    rec_m = re.search(r'<div class="recommendation-badge">(.*?)</div>', html, re.S)
    if rec_m:
        rec_text = re.sub(r'&[a-zA-Z]+;', 'x', rec_m.group(1))
        rec_text = re.sub(r'<[^>]+>', '', rec_text).strip()
        if len(rec_text) > 300:
            failures.append(f"S10: recommendation-badge too long ({len(rec_text)} > 300 chars)")

    # S11b — Role-Fit Hard step-back content (added 2026-04-27, Schott build)
    rfh_m = re.search(r'<div class="role-fit-col hard">(.*?)</div>\s*</div>\s*</div>', html, re.S)
    if rfh_m:
        rfh = rfh_m.group(1).lower()
        has_quadrant = any(q in rfh for q in ['top-left', 'top-right', 'bottom-left', 'bottom-right',
                                              'across', 'center-of-wheel', 'center of wheel'])
        wedges = ['implementor', 'conductor', 'persuader', 'promoter',
                  'relater', 'supporter', 'coordinator', 'analyzer']
        has_wedge = any(w in rfh for w in wedges)
        if not has_quadrant:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wiring-shape reference (quadrant or ACROSS)")
        if not has_wedge:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wedge name")

    # S11b.1 — Wheel-position-required rule (added 2026-04-27, Armstrong build):
    # any narrative that claims a canonical-CFO / top-left / right-wiring / Implementor-quadrant
    # read MUST also cite the actual wheel position number. Reading DISC scores in isolation
    # without checking the wheel has produced shipping-quality errors. The check looks across
    # all wiring-relevant sections (Role-Fit, Wiring-Fit, Skills card, DISC notes) for any
    # canonical-CFO claim and verifies a wheel-position number (1-60) is also present nearby.
    wiring_claim_phrases = [
        'top-left', 'canonical strong-cfo', 'canonical cfo wiring', 'right wiring',
        'wiring is right', 'wiring fits the seat', 'implementor quadrant',
    ]
    html_lower = html.lower()
    has_wiring_claim = any(p in html_lower for p in wiring_claim_phrases)
    has_position_number = bool(re.search(r'\bposition\s+\d{1,2}\b', html_lower))
    if has_wiring_claim and not has_position_number:
        failures.append(
            "S11b.1: wiring narrative makes a canonical-CFO / top-left / right-wiring claim "
            "but does NOT cite a TTI wheel position number. Reading DISC scores in isolation "
            "is insufficient; cite the actual wheel position (Natural / Adapted) before making "
            "wiring-fit claims. See METHODOLOGY 'TTI wheel position is REQUIRED reading'."
        )

    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} ***")
        for f in failures: print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)}")
    print("*** QA GATE PASSED ***")



def build_concerns_intro():
    """Returns the INNER HTML of the Concerns intro block at the top of the Interview section.
    The .concerns-intro wrapper lives in the template; this function returns only the content inside.
    See METHODOLOGY.md "Concerns are validation framing for the Interview" + QA_CHECKLIST Section 1."""
    return """
                <p class="ci-subhdr">2 Top Concerns</p>
                <ol>
                    <li>The signature pattern (Conditional Belief Sev, Lower Standards for Others, Not Pleasing L1 &minus;1.39, Initiating Accountability lit) plus boss and referral corroboration name a clear standard-holding gap. Use <strong>Facilitative Mindset</strong> (question #6) and <strong>Talent Development</strong> (question #2) to surface the behavioral detail: one specific case where she held the line against pressure, and one pre-proof developmental investment rather than retrospective admiration.</li>
                    <li>Bottom-decile Deliberate Urgency (&minus;2.51) and catastrophic Organizational Decision Making (&minus;2.22) describe a decisional-velocity gap. The Pax8 record is competent execution at Director altitude with proven processes; the open seat-fit question is whether the same execution travels to a new company without those processes. Use <strong>Leadership Deep-Dive</strong> (question #8) to surface a recent decision that needed to happen fast and what would change in her routines at VP altitude.</li>
                </ol>
        """

def build_signature_pattern():
    """Returns the INNER HTML of the Signature Pattern block. The .signature-pattern-box
    wrapper lives in the template; this function returns only the content inside it.
    See METHODOLOGY.md "Signature Pattern block" + QA_CHECKLIST Section 11d."""
    return """
            <h3>Signature Pattern &mdash; Excellence Standards Read</h3>
            <p class="headline">The instrument shows a leader who does not hold the line on talent and does not act with urgency.</p>
            <p class="sp-subhdr">Where the L1 standards land negatively</p>
            <ul class="sp-list">
                <li><strong>#9 Deliberate Urgency &minus;2.51</strong> (the lowest L1 in her file) &mdash; she does not move on hard things quickly.</li>
                <li><strong>#8 Organizational Decision Making &minus;2.22</strong> &mdash; decisions are not crisply assigned, simplified, or grounded in facts over feelings.</li>
                <li><strong>#3 Conducting &amp; Outvoted &minus;1.65</strong> &mdash; she runs meetings as a participant rather than drawing out the best thinking from the team.</li>
                <li><strong>#7 Not Pleasing &minus;1.39</strong> &mdash; she defers to approval over holding the line on standards.</li>
                <li><strong>#4 Replacing Self &minus;0.93</strong> &mdash; she does not actively build a team that can run without her in the room.</li>
                <li><strong>#6 Personal Reliability &minus;0.77</strong> &mdash; her own follow-through sits below cohort.</li>
                <li><strong>#1 Investing in Others &minus;0.54</strong> &mdash; she does not actively invest in people&rsquo;s development.</li>
                <li class="muted">(#5 Pushing Extreme Accountability +0.46 is the only positive L1.)</li>
            </ul>
            <p class="sp-subhdr">Flags lit</p>
            <ul class="sp-list">
                <li><strong>Conditional Belief Sev</strong> &mdash; she believes in people after they&rsquo;ve proven themselves, not before.</li>
                <li><strong>HandsOn Sev</strong> &mdash; she stays involved in the work rather than letting people own it.</li>
                <li><strong>Clarity of Accountability Hi</strong> + <strong>Driving Accountability Med</strong> + <strong>Initiating Accountability</strong> lit &mdash; she does not crisply assign single-point owners, does not drive consequences when standards are missed, and accountability does not flow up to her from the team.</li>
                <li><strong>Lower Standards for Others than Self</strong> &mdash; she holds herself to a higher bar than she enforces on the team.</li>
            </ul>
            <p class="sp-seat-reflection">For an HR leader, this profile is <strong>MORE concerning</strong> than the same profile would be in a finance, sales, or operations seat. HR is the function whose primary remit is exactly the dimensions the instrument flags: holding the talent bar across the org, extending developmental belief before performance proof, having difficult conversations cleanly, and modeling the personal reliability the function asks of others. A leader in another function carries these standards at one remove from daily work; an HR leader is hired to do this work directly.</p>
        """

def main():
    print("Loading respondent data...")
    rd = load_respondent_data()
    print(f"  name={rd['name']}, Z|Algo={rd['z_algo_overall']:+.3f}, RF={rd['rf_num']}")
    print(f"  flags: {list(rd['flags_lit'].keys())}")

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution tokens...")
    dist = build_distribution_tokens(zalgo_rows, flag_rows,
                                     rd['z_algo_overall'], rd['z_human_overall'], rd['rf_num'])

    print("Building respondent dict for motivators_section...")
    respondent = build_respondent_dict(rd)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building signature pattern block...")
    signature_pattern_html = build_signature_pattern()
    print("Building concerns intro block...")
    concerns_intro_html = build_concerns_intro()

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()
    role_fit = build_role_fit()
    scorecard = build_excstds_scorecard(rd)

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    replacements = {
        'CANDIDATE_NAME': 'Megan Houston',
        'CANDIDATE_CREDS': 'Candidate &mdash; VP / Director Step-Up Seat',
        'CANDIDATE_ROLE': 'Candidate for VP / Director of HR &mdash; Public / Mature Company &middot; Step-Up from Director, Global Business Partner',
        'REPORT_DATE': 'April 27, 2026',

        'ZALGO_OVERALL': f'{rd["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{rd["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(rd['rf_num']),
        'FLAGS_LIT': f'{len(rd["flags_lit"])}/16',
        'TEACH_ITEMS': '10/10',
        'COHORT_AVG': '+0.24',
        'COHORT_AVG_NUM': '0.2440',

        **axes, **concerns, **wiring, **htl, **probes, **teach,
        **radar, **timeline, **role_fit, **dist, **scorecard,

        'SIGNATURE_PATTERN_BLOCK': signature_pattern_html,
        'CONCERNS_INTRO': concerns_intro_html,

        'DISC_D_NAT': '17',
        'DISC_I_NAT': '62',
        'DISC_S_NAT': '74',
        'DISC_C_NAT': '71',
        'DISC_D_ADP': '25',
        'DISC_I_ADP': '58',
        'DISC_S_ADP': '72',
        'DISC_C_ADP': '64',
        'DISC_NOTE_TEXT': 'Natural and Adapted both at wheel position 36, Supporting Coordinator (FLEXIBLE). Anchored in the Coordinator/Supporter overlap zone. <em>Neither position is ACROSS</em> &mdash; not center-of-wheel. Effectively no Natural-to-Adapted gap; she is operating in her natural wiring.',
                'DISC_NOTE_DETAIL': 'TTI Driving Forces &mdash; Primary: Selfless 62, Structured 57, Harmonious 54, Altruistic 47. Indifferent: Receptive 32, Intellectual 32, Objective 28, <strong>Resourceful 17 (near 2&sigma; below mean)</strong>. The wiring carries process-and-people-care work cleanly but lacks the results-and-decisiveness orientation a VP step-up typically requires. The TTI Time-Wasters page lists <em>&ldquo;Not Exercising Authority&rdquo;</em> as one of her named time-wasters &mdash; an unusually explicit corroboration of the instrument&rsquo;s standard-holding flags.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        'RECOMMENDATION_TEXT': 'CONDITIONAL HIRE &middot; SEAT-DEPENDENT &mdash; Z|Algo &minus;1.75 (worst in the working set); 4 promotions in 8 years is genuinely strong but instrument + TTI + boss + referral all converge on standard-holding gap; advance ONLY if seat is People-Ops/HRBP under stronger CHRO carrying the standard-holding work.',
    }

    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    html = re.sub(
        r'<title>.*?</title>',
        '<title>Megan Houston &mdash; VP/Director HR Candidate | HALE GLOBAL SUCCESS DIAGNOSTICS</title>',
        html, count=1
    )
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    qa_gate(html)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")


if __name__ == '__main__':
    main()
