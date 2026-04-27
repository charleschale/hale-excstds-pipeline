"""Build the Excellence Standards hiring report for Timothy Schott (CFO).

Forked from build_cohen_hiring.py 2026-04-27. Same Provable Markets context (per
respondent's self-report Q109), but evaluated for a Series B CFO seat at a different
company. LinkedIn full-profile sweep complete (Charles supplied 2026-04-27): Deloitte
1992-2011 (Partner Aug 2007, Chicago), Lazard, Tiptree CAO, Associated Capital Group
(NYSE:AC) EVP/CFO, Yieldstreet CFO, 1970 Group CFO (just ended Apr 2026), hey freya
Board Advisor concurrent. Fairfield Golf Team 1987-91 (Two-Sport Athlete signal).

Run from repo root:
    python _pipeline/scripts/build_schott_hiring.py
Outputs to:
    _reports/Schott_Timothy_hiring_report.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

RESPONDENT_XLSX = ROOT / '_respondents' / '20260420.tsnyc21@yahoo.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Schott_Timothy_hiring_report.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title, 'z_algo': z_algo, 'z_human': z_human, 'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flag_names = []
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        if flag_name:
            flag_names.append((col, flag_name))
    flags_lit = {}
    for col, flag_name in flag_names:
        flag_val = ws_flags.cell(2, col).value
        if flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[q_num] = answer

    ws_meta = wb["Metadata"]
    name = ws_meta.cell(2, 5).value
    email = ws_meta.cell(2, 4).value
    date_str = ws_meta.cell(2, 6).value

    return {
        'l1_data': l1_data, 'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall, 'z_human_overall': z_human_overall,
        'rf_num': rf_num, 'questions_answered': questions_answered,
        'flags_lit': flags_lit, 'non_scorable': non_scorable,
        'name': name, 'email': email, 'date_str': date_str,
    }

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        rf_val = ws_zalgo.cell(row, 7).value
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]
    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo, z_human, sf = row['z_algo'], row['z_human'], row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1; continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1; break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]
        upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    rev_flag_bin = len(flag_counts) - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# NARRATIVE
# ============================================================================

def build_three_axes_narratives():
    talent_badge = 'badge-amber'
    talent_badge_text = 'AMBER'
    talent_card_body = '''Real but unspectacular external-validator density across a 35-year career arc. The headline strengths are clear: 19 years at Deloitte ending as <strong>Partner in Chicago (Aug 2007)</strong> on a single internal promotion track (Staff &rarr; Senior Manager &rarr; Partner) with the Midwest Regional Leader role for the financial-instruments SMR network. CPA / AICPA credential. Subsequent moves to Lazard (Director of Accounting Policy, 4y9m), Tiptree Inc. (CAO, 4y9m), Associated Capital Group NYSE:AC (EVP Finance &amp; CFO with public press-release), Yieldstreet, and 1970 Group are real seats &mdash; but they are <em>lateral-to-modest-step-up</em> moves rather than progressive sponsor-bets from competitive selectors at increasing altitude. Two-Sport Athlete signal exists but is light: Fairfield University Golf Team 1987&ndash;1991 alongside BS Accounting &mdash; small-Jesuit-school varsity, no documented all-conference / scholar-athlete callout. <strong>No awards</strong> on the LinkedIn record.

Instrument corroborates the personal-discipline layer cleanly: Extreme Ownership L2 +2.71 (top of file), Personal Reliability L1 +1.88, Developmental Discipline L2 +1.77, Handling Daily Difficulties With Dignity L2 +1.18, Dialogue vs. Direction L2 +2.00, Facts Over Feelings L2 +2.10, Simplification Methods L2 +2.38. He is reliable, ownership-strong, and intellectually curious (TTI Intellectual Primary 71 + Receptive Primary 69).

<strong>The Talent-axis tension that pulls the badge to amber.</strong> Three CFO seats in five years (AC 1y7m, Yieldstreet 2y2m, 1970 Group 1y6m), each ending without crossing the three-year mark, is itself a Talent-axis signal &mdash; not just a diligence item to verify. The instrument names a coherent reading: HandsOn|Sev + Replacing Self &minus;1.51 (worst L1) + L2 1.4 Demonstrating Genuine Fanness &minus;2.52 + Conditional Belief Hi + 3-of-5 self-graded exec peers at D-grade describe the leader who comes in, becomes the operating engine, can&rsquo;t extend belief or build the team below directs, and the seat ends. The award-density and competitive-selector evidence required to override that reading is not on the file. Onsite must produce the candidate&rsquo;s own framing of the three exits and at least one specific pre-proof developmental investment in a current direct report.'''

    judgment_badge = 'badge-amber'
    judgment_badge_text = 'AMBER'
    judgment_card_body = '''Two readings sit on top of each other and the gap between them is the diagnostic feature. Algorithmically he passes Organizational Decision Making (Z|Algo +0.76); on the same answers, human reviewers score him &minus;1.31 &mdash; <strong>a 2.07-Z gap that is the largest Algo/Human disagreement in his profile</strong>. Same pattern shows up in Not Pleasing (Algo +0.86 / Human &minus;0.05) and Deliberate Urgency (Algo +0.57 / Human &minus;0.50). The signature of someone whose answers technically pass the pattern-matching but whose underlying judgment, read by humans seeing the full texture, is materially weaker.

Where the L2s say it lives: <strong>Clarity of Accountability L2 &minus;2.54 is the worst score in his entire profile.</strong> Drives Accountability L2 &minus;1.08 corroborates. He answered the canonical clarity-of-accountability questions (Q381, Q382, Q384) in the wrong direction &mdash; agreeing that multiple parties accountable to the same outcome produces <em>better</em> results, when the standard says it produces worse results. Two flags lit at <strong>Sev</strong>: Flag_ClarityAcctblty and Flag_DrivingAcctblty. TTI corroborates exactly: &ldquo;Make decisions based on surface analysis,&rdquo; &ldquo;May overlook vital details,&rdquo; &ldquo;May not recognize increased risk associated with bigger rewards.&rdquo; Receptive Primary (69) + Structured Indifferent (19, second-lowest DF) + Natural C=18 = wiring that actively resists installing the structural systems that fix this.

Strong-judgment-signal counter-evidence on the same file: Facts Over Feelings L2 +2.10, Simplification Methods L2 +2.38, Q68 (50% odds of 50% growth over 80% odds of 25% growth) answered correctly &mdash; he does the math when the math is offered. The judgment failure mode is not analytical; it is structural &mdash; defining who-decides-what.'''

    skills_badge = 'badge-amber'
    skills_badge_text = 'AMBER'
    skills_card_body = '''The instrument does not measure domain skill directly. Domain credentials are strong: CPA (AICPA), Deloitte Partner, Director of Accounting Policy at Lazard, CAO at Tiptree, three sequential CFO seats including a NYSE-listed firm. Public-company finance, technical accounting, and capital-markets posture are all in the file.

The wiring concern under Skills is CFO-specific and worth naming explicitly. <strong>Schott&rsquo;s wiring is much closer to a high-I CRO/strategic-finance lead than to a high-C controls-and-mechanics CFO.</strong> Natural C=18 (very low). Adapted C=42 (still low for the seat). Receptive Primary + Structured Indifferent (19) + Instinctive Indifferent (11, lowest DF) describe an operator who resists structure being imposed and who does not naturally rely on past methodologies. His own LinkedIn activity reposts the &ldquo;Your CFO should not be your Controller&rdquo; thesis &mdash; he leans hard into the strategic-CFO end of the seat and away from the close-the-books mechanics. For a Series B CFO seat the diligence question is direct: <em>does the company have a strong Controller / VP Finance below the CFO who owns the close, the audit, and the FP&amp;A discipline, OR does the CFO need to own those personally?</em> If the former, the wiring fits and the strategic-finance posture is real. If the latter, this is the wrong wiring &mdash; and the tenure pattern in the Talent diligence note may itself be a downstream signal of that mismatch.'''

    return {
        'TALENT_BADGE_CLASS': talent_badge,
        'TALENT_BADGE_TEXT': talent_badge_text,
        'TALENT_CARD_BODY': talent_card_body,
        'JUDGMENT_BADGE_CLASS': judgment_badge,
        'JUDGMENT_BADGE_TEXT': judgment_badge_text,
        'JUDGMENT_CARD_BODY': judgment_card_body,
        'SKILLS_BADGE_CLASS': skills_badge,
        'SKILLS_BADGE_TEXT': skills_badge_text,
        'SKILLS_CARD_BODY': skills_card_body,
    }

def build_concerns_section():
    concerns_title = 'Two Targeted Concerns'
    concerns_items = """            <div class="concern-item">
                <div class="concern-number">1</div>
                <div class="concern-text">
                    <strong>Accountability architecture &mdash; Clarity AND Drives both severe.</strong>
                    Two Sev flags &mdash; Flag_ClarityAcctblty and Flag_DrivingAcctblty &mdash; sit on top of L2 8.2 Clarity Of Accountability <strong>&minus;2.54</strong> (the worst single score in the file) and L2 5.3 Drives Accountability &minus;1.08. The instrument is not subtle here: he answered Q381, Q382, and Q384 in the wrong direction &mdash; agreeing that holding multiple parties accountable to the same outcome produces <em>better</em> results, when the standard reads worse. Combined with Q33 answered toward TRUE (the &ldquo;trade bottom 10%&rdquo; basketball frame instead of the crew frame) and HoldsAvgDownChain lit, this is the leader who tolerates ambiguity about who-owns-what and absorbs the consequences personally rather than redistributing them. TTI corroborates exactly: &ldquo;Make decisions based on surface analysis,&rdquo; &ldquo;May overlook vital details.&rdquo; And the wiring resists the fix: Receptive Primary 69 + Structured Indifferent 19 + Natural C=18 say he will <em>resist</em> installing the RACI/single-point-ownership systems his weakest L2 needs. Probe Form 8 Leadership Deep-Dive (sketch the org chart and name single-point owners for each priority), then Facilitative Mindset.
                </div>
            </div>
            <div class="concern-item">
                <div class="concern-number">2</div>
                <div class="concern-text">
                    <strong>Three CFO seats in five years + the bottleneck pattern.</strong>
                    Associated Capital Group 1y7m &rarr; Yieldstreet 2y2m &rarr; 1970 Group 1y6m (just ended Apr 2026, same month as the survey). Tenure pattern needs onsite explanation. The instrument names a coherent shape that fits: Flag_HandsOn at <strong>Sev</strong> + L1 #4 Replacing Self &minus;1.51 (the worst L1) + L2 9.6 Action Over Inaction &minus;1.90 + Q81 (&ldquo;not the one creating urgency&rdquo;) answered toward FALSE (he IS) + Q31 (resists involvement until needed) answered toward FALSE (he doesn&rsquo;t resist) + Q7 (tasks below directs) answered TRUE. Plus the team self-grades in the Non-Scorable tab: 3 of 5 named exec peers (CRO, CPO, COO) graded D-grade on dimension 1; only CMO graded A. He sees the team as mediocre, becomes the operating engine, can&rsquo;t extend belief (Conditional Belief Hi + L2 1.4 Demonstrating Genuine Fanness &minus;2.52 + TTI Altruistic Indifferent 28), and the seat ends. Probe Form 8 Leadership Deep-Dive (the org chart conversation, with the &ldquo;which seats would you not re-hire today&rdquo; follow-up) and Talent Development (one pre-proof investment, not retrospective admiration).
                </div>
            </div>"""
    return {'CONCERNS_TITLE': concerns_title, 'CONCERNS_ITEMS': concerns_items}

def build_wiring_fit():
    wiring_fit_items = (
        '<strong>The wiring corroborates the accountability-architecture concern (not contradicts it):</strong> '
        'TTI Receptive Primary 69 + Structured Indifferent 19 (second-lowest DF) + Natural C=18 + '
        'Instinctive Indifferent 11 (lowest DF) all point at an operator who resists imposed '
        'structure and does not rely on proven methodologies. ExcStds Clarity of Accountability '
        '&minus;2.54 + Drives Accountability &minus;1.08 + two Sev flags say the same: he does not '
        'install single-point-ownership systems naturally. No wiring-vs-behavior divergence.'
        '<span class="wiring-flag">Targeted Concern</span><br>'
        '<strong>Adapted DISC and the &ldquo;urgency engine&rdquo; pattern:</strong> '
        'Natural D=63 / I=72 / S=72 / C=18 (Promoting Persuader) shifts to Adapted D=72 / I=66 / '
        'S=25 / C=42 (Conducting Persuader). The &minus;47-point S compression is the loudest signal '
        'in the wiring panel: he is grinding hard against natural steadiness/people-orientation '
        'to act decisively in the current seat. Compounds HandsOn|Sev + Replacing Self &minus;1.51.'
        '<span class="wiring-flag">Diligence Item</span>'
    )
    return {'WIRING_FIT_ITEMS': wiring_fit_items}

def build_hard_to_learn():
    # Schott Hard-to-Learn count: Urgency partial (Algo+0.57 but L2 9.6 Action Over Inaction
    # -1.90 lit), Org DM Algo+0.76 / Human-1.31 (split — counts as lit at Human read),
    # Conditional Belief Hi (lit), Satisfied with Gripes Low (lit). 4/4 lit (with severities
    # ranging from partial to severe).
    return {'HARD_TO_LEARN': '4/4'}

FORM8_QUESTIONS = [
    ("Two-Sport Athlete",
     "Of all the things you've done in life, tell me what results you're most proud of.",
     "most proud of"),
    ("Talent Development",
     "What people over your career have you nurtured who have gone on to do great things?",
     "nurtured who have gone on"),
    ("TORC",
     "What was your boss's name? What will they say your strengths and areas for improvement were?",
     "strengths and areas for improvement"),
    ("Emotional Maturity",
     "What's the greatest adversity you've faced in life?",
     "greatest adversity"),
    ("Punctuates Differently",
     "What do you do to achieve excellence that others don't?",
     "achieve excellence that others don"),
    ("Facilitative Mindset",
     "What's something you really believe in? When is it okay to make exceptions?",
     "okay to make exceptions"),
    ("Commitment",
     "Tell me something important to you that you do every day.",
     "important to you that you do every day"),
    ("Leadership Deep-Dive",
     "Draw the org chart you're responsible for today.",
     "draw the org chart"),
    ("Passion",
     "What is the worst job you could imagine? How would you create passion around it?",
     "worst job you could imagine"),
    ("Continuous Improvement",
     "What counts as work? When do you work, when don't you?",
     "what counts as work"),
]

def build_interview_probes():
    tailored = [
        ("TWO-SPORT ATHLETE (Talent axis &middot; Fairfield Golf 1987-91)",
         "Strong opening probe: he has a documented Two-Sport Athlete signal (Fairfield Golf Team during BS Accounting). Listen for whether the proudest results he names tie back to that pattern of competitive-pursuit-plus-academic-rigor or to a specific discipline that compounded over the 35-year career arc. The Deloitte Partner promotion in Chicago is a real selector-validation moment; the cleaner answer goes to a personal pursuit (a build, a craft, a relationship) rather than the most obvious career milestone."),
        ("TALENT DEVELOPMENT (Concern 2 &middot; Pre-proof belief)",
         "Central probe for Concern 2. L2 1.4 Demonstrating Genuine Fanness &minus;2.52 + Conditional Belief Hi + TTI Altruistic Indifferent 28 + 3-of-5 self-graded exec peers at D-grade all predict the same answer pattern: post-proof admiration of performers who already delivered, not pre-proof investment in those who had not. Listen for one specific case where he extended developmental capital (time, assignment, exposure) <em>before</em> the person had earned it. A strong Talent-Development record produces two such names inside 60 seconds. Vague answers, retrospective-only stories, or &ldquo;I challenge people every day&rdquo; without naming a specific person are the negative signals."),
        ("TORC &mdash; TRUTH-OVER-COMFORT (Self-awareness)",
         "Anchor to a real boss: Mario Gabelli at Associated Capital Group (NYSE:AC), the Lazard partners, the Yieldstreet CEO during the 2022&ndash;2024 fundraising window, or his Deloitte partner-class peers. High-I + Adapted I=66 wiring tends to convert weakness questions into &ldquo;my weakness is I care too much&rdquo; or &ldquo;I work too hard.&rdquo; Push past the first answer. Listen for a real area his most recent boss would name &mdash; especially around the transition out of 1970 Group, since that seat ended in the same month as this assessment."),
        ("EMOTIONAL MATURITY (Ownership pattern)",
         "Three short CFO tenures (AC 1y7m, Yieldstreet 2y2m, 1970 Group 1y6m) put real adversity material on the table. Listen for whether he names a personal adversity or defaults to a professional setback &mdash; and if professional, whether ownership is internal (&ldquo;I should have done X differently&rdquo;) or external (&ldquo;the board changed direction,&rdquo; &ldquo;market conditions&rdquo;). The healthy answer attributes learning to self even when the situation was genuinely difficult; the unhealthy answer accumulates external attributions across the three short seats."),
        ("PUNCTUATES DIFFERENTLY (Talent axis)",
         "Rare-behavior probe. With TTI Intellectual Primary 71 + Receptive Primary 69 + Adapted DISC Conducting Persuader, the most likely strong answer is a specific intellectual-pursuit routine &mdash; reading discipline, technical-accounting-research habit, or a recurring deep-work ritual &mdash; not a generic work-ethic claim. &ldquo;I work harder&rdquo; or &ldquo;I outprepare people&rdquo; without specifics is the negative signal. Press for the artifact (the journal, the calendar block, the deep-work routine) if the first answer is general."),
        ("FACILITATIVE MINDSET (Concern 1 &middot; Dialogue and decisional clarity)",
         "Central probe for Concern 1. The instrument shows Dialogue vs. Direction L2 +2.00 (strong) but Sublimating Ego L2 &minus;0.43 and Power &amp; Status L2 +0.17 (mid). High-I wiring + S&minus;47 adaptation creates a leader who runs <em>conversational</em> meetings but may not actually update from dissent. A Facilitative answer names a belief AND the conditions under which he would update it &mdash; and ideally cites a recent case where his team changed his mind on something material. A non-Facilitative answer holds the belief absolute or caves on personality. Q119 (the &ldquo;woke test&rdquo;) is also worth a careful walk-through if the conversation goes there."),
        ("COMMITMENT (Diligence wiring &middot; Personal Reliability +1.88 strong)",
         "Daily-routine specificity. Personal Reliability L1 +1.88 + Extreme Ownership L2 +2.71 + L2 6.3 Commitment To Routine +0.88 + Q24 (daily routines for years) answered TRUE = the habit machinery is wired in. A strong answer names a concrete morning routine, a specific physical-discipline practice, or a daily reading/research block. Vague answers (&ldquo;I&rsquo;m disciplined,&rdquo; &ldquo;I stay on top of things&rdquo;) would contradict the instrument read &mdash; press for the artifact."),
        ("LEADERSHIP DEEP-DIVE (Concerns 1 &amp; 2 &middot; Ownership clarity)",
         "Ask him to literally sketch his most recent finance org (Provable Markets per self-report, or 1970 Group looking back). Tests both Targeted Concerns. A leader running the accountability routine names single-point owners for each function and identifies open seats / upgrade candidates in the same breath. Q9961-99103 self-grades show CRO, CPO, COO at D-grade and CMO at A-grade &mdash; ask him to walk that delta. Follow-ups: <em>which of these seats would you not re-hire today?</em> + <em>name one person two levels down you backed before they had proven themselves.</em> Hesitation, multiple owners for the same priority, or post-proof-only names are the flags."),
        ("PASSION &amp; MISSION CONSTRUCTION",
         "Tests mission-construction. The 35-year career arc is mostly technical accounting and public-company finance &mdash; substantively serious work without obvious passion-narrative. Listen for whether he roots passion in a specific kind of finance problem (private credit / regulated FinTech / investor relations) or defaults to scale-language. Strong answers name the concrete hook; weak answers invoke &ldquo;any team can be motivated.&rdquo; Bridges into &ldquo;Why this seat, why now?&rdquo; given the 1970 Group exit timing."),
        ("CONTINUOUS IMPROVEMENT (Work-definition posture)",
         "Probes Concern 2. Listen for whether he counts time spent in others&rsquo; development as core work, or only metrics and deliverables. The Continuous Improvement operator at CFO altitude treats coaching the FP&amp;A bench, the Controller, and the deal-finance team as core finance work &mdash; not overhead. A narrow answer (&ldquo;work = output, board prep, close&rdquo;) reinforces the Conditional Belief Hi + Demonstrating Genuine Fanness &minus;2.52 wiring. A strong answer names a specific coaching cadence with directs."),
    ]
    assert len(tailored) == len(FORM8_QUESTIONS) == 10
    probes = []
    for (category, coaching), (_n, canonical_q, _s) in zip(tailored, FORM8_QUESTIONS):
        probes.append((category, f'"{canonical_q}"', coaching))
    parts = []
    for i, (category, question, listen_for) in enumerate(probes, start=1):
        parts.append(f"""                <div class="probe-card">
                    <div class="probe-number">{i}</div>
                    <div class="probe-category">{category}</div>
                    <div class="probe-question">{question}</div>
                    <div class="probe-coaching">{listen_for}</div>
                </div>""")
    return {'INTERVIEW_PROBE_CARDS': "\n".join(parts)}

def build_teach_items():
    return {'TEACH_ITEMS': '10/10'}

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Demonstrating Genuine Fanness',
            'Developmental Discipline',
            'Handling Daily Difficulties With Dignity',
        ]),
        (4, 'RISKING BY REPLACING SELF', [
            'Ability To Disappear',
            'Urgency Down Chain Of Command',
            'CEO gets outside exec',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Clarity Of Accountability',
            'Simplification Methods',
            'Facts Over Feelings',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Action Over Inaction',
            'Extreme Proactivity',
        ]),
    ]

    labels, scores, is_l1 = [], [], []
    missing_l2 = []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = sum(l1_data[i]['z_algo'] for i in range(1, 10)) / 9.0
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'EXCSTDS_OVERALL_Z': f"{overall_z:.3f}",
        'EXCSTDS_COHORT_AVG': f"{cohort_avg:.3f}",
    }

def build_talent_radar():
    # Two-Sport Athlete: 3 (Fairfield Golf small-Jesuit varsity, no all-conference/scholar-
    #     athlete callout; signal exists but is light)
    # Punctuates Differently: 3 (Personal Reliability +1.88 + Extreme Ownership +2.71 top-decile
    #     but Not Pleasing only +0.86; mid-tier rare-behavior signal, interview required)
    # Facilitative Mindset: 3 (Dialogue +2.00 strong, Sublimating Ego -0.43 weak, Power & Status
    #     +0.17 mid; ratio not yet conductor-quality)
    # Wit: null
    # Deep Repertoire: 3 (Deloitte Partner is one promotion track; subsequent moves are lateral
    #     to modest step-ups, not progressive sponsor-bets at increasing altitude; no awards)
    # Discipline/Routine: 4 (Personal Reliability +1.88, Extreme Ownership +2.71, Commitment to
    #     Routine +0.88, daily-routines TRUE; Adapted C=42 caps it below 5)
    # Understands Symbolism: null
    talent_radar_scores = json.dumps([3, 3, 3, None, 3, 4, None])
    talent_radar_profile = (
        "<strong>Discipline/Routine (4)</strong> is the strongest single dimension &mdash; "
        "Personal Reliability L1 +1.88, Extreme Ownership L2 +2.71, Commitment to Routine "
        "+0.88, Q24 daily-routines-for-years answered TRUE; Adapted C=42 caps it below 5. "
        "<strong>Two-Sport Athlete (3)</strong> &mdash; Fairfield Golf Team 1987-91 is a "
        "documented varsity signal during college, but small-Jesuit-school golf without a "
        "documented all-conference / scholar-athlete callout keeps it mid-tier. "
        "<strong>Deep Repertoire (3)</strong> &mdash; Deloitte Partner (one promotion track over "
        "19 years), Lazard, Tiptree CAO, three sequential CFO seats; real breadth, but the "
        "moves are lateral-to-modest-step-up rather than progressive sponsor-bets at increasing "
        "altitude. No awards. <strong>Punctuates Differently (3)</strong> &mdash; Not Pleasing "
        "L1 only +0.86 keeps the rare-behavior signal mid-tier; interview probe required. "
        "<strong>Facilitative Mindset (3)</strong> &mdash; Dialogue vs. Direction L2 +2.00 strong, "
        "but Sublimating Ego &minus;0.43 and Power &amp; Status +0.17 keep the ratio below "
        "conductor-quality. <strong>Wit (&#8856;)</strong> and <strong>Understands Symbolism "
        "(&#8856;)</strong> not measurable from the instrument &mdash; interview-only."
    )
    return {
        'TALENT_RADAR_SCORES': talent_radar_scores,
        'TALENT_RADAR_PROFILE_TEXT': talent_radar_profile,
    }

def build_role_fit():
    role_fit_title = 'Role-Fit Read &mdash; What Will Be Easy, What Will Be Hard'
    role_fit_seat = ('Evaluating for: Series B growth-stage CFO seat &middot; same functional '
                     'altitude as today &middot; ~50&ndash;300 employees &middot; building out '
                     'the finance org under the CFO')
    role_fit_easy = (
        '<strong>Strategic-CFO posture and board-facing presence.</strong> Personal Reliability '
        'L1 +1.88, Extreme Ownership L2 +2.71, Simplification Methods L2 +2.38, Facts Over '
        'Feelings L2 +2.10, Dialogue vs. Direction L2 +2.00, Developmental Discipline L2 +1.77 '
        'are all top-decile or near-it. TTI Intellectual Primary 71 + Receptive Primary 69 + '
        'Resourceful 58 = a curious, adaptive, ROI-oriented strategic finance lead. DISC Adapted '
        'Conducting Persuader (D=72 / I=66) is a board-room-friendly profile.<br><br>'
        '<strong>Domain depth and selector validation.</strong> CPA (AICPA), Deloitte Partner '
        '2007 (Midwest financial-instruments SMR network leader), Lazard, Tiptree CAO, Associated '
        'Capital Group (NYSE:AC) EVP/CFO, Yieldstreet CFO, 1970 Group CFO. Public-company '
        'finance, technical accounting, fintech, private credit, and investment-management '
        'cycles are all in the file. Counterparty credibility and audit/regulator interface '
        'are genuinely covered.<br><br>'
        '<strong>Capital-raising and investor-relations posture.</strong> Self-reported skill '
        'set (Roadshows, Investor Relations) plus the Yieldstreet seat during a difficult capital '
        'cycle and the AC NYSE-listed-firm tenure mean this seat will land naturally for a '
        'Series B that is preparing for the next raise.'
    )
    role_fit_hard = (
        '<strong>Step back: the wiring profile itself is the headline concern.</strong> Strong CFOs '
        'almost always sit in the <em>top-left</em> quadrant of the TTI wheel &mdash; Implementor '
        'and Conductor (high-D, high-C, low-I, low-S) &mdash; the wiring of structured directive '
        'execution that the close, the audit, the FP&amp;A discipline, and the regulator/auditor '
        'interface all reward. Schott sits in the <em>bottom-right</em> quadrant: Natural Promoting '
        'Persuader (I=72 / S=72 / D=63 / C=18) shifting to Adapted Conducting Persuader (D=72 / I=66 '
        '/ S=25 / C=42). For any CFO seat, that I-dominant Persuader/Promoter signature is the '
        'inverse of the typical strong-CFO wiring. The follow-up signal is the size of the Natural-'
        'to-Adapted gap: <strong>S=72 collapsing to S=25 is a &minus;47-point compression</strong>, '
        'one of the loudest adaptation signals on the wheel. He is grinding hard against natural '
        'steadiness and people-orientation to perform the seat. Sustained large adaptations are '
        'how leaders burn out of seats, and three CFO tenures under three years apiece are a '
        'pattern consistent with that mechanism.<br><br>'
        '<strong>The controls-and-mechanics half of the CFO seat.</strong> Downstream of the '
        'wiring shape: Natural C=18 + Adapted C=42 + Receptive Primary 69 + Structured Indifferent '
        '19 (second-lowest DF) + Instinctive Indifferent 11 (lowest DF) describe an operator who '
        'resists imposed structure and does not rely on proven methodologies. ExcStds Clarity of '
        'Accountability L2 &minus;2.54 (worst score in the profile) + Drives Accountability L2 '
        '&minus;1.08 + two Sev flags (Flag_ClarityAcctblty, Flag_DrivingAcctblty) corroborate. '
        'The wiring resists exactly the controls work a growing finance team needs from its CFO.<br><br>'
        '<strong>Building belief in the layer below.</strong> L2 1.4 Demonstrating Genuine Fanness '
        '&minus;2.52 + Conditional Belief Hi + TTI Altruistic Indifferent 28 + 3-of-5 self-graded '
        'exec peers at D-grade describe the post-proof-admiration, not pre-proof-investment '
        'pattern. At Series B, the Controller, FP&amp;A lead, and finance-systems owner below the '
        'CFO need active developmental capital before they have earned it.<br><br>'
        '<strong>The binary diligence question.</strong> Does the hiring company already have a '
        'strong Controller / VP Finance below the CFO who owns close / audit / FP&amp;A mechanics, '
        'OR is the CFO required to own them personally? If the former, this could land as a '
        'strategic-CFO hire with a real diligence frontier on team-development. If the latter, '
        'the wiring mismatch will compound &mdash; and the three short CFO tenures over the last '
        'five years (AC 1y7m, Yieldstreet 2y2m, 1970 Group 1y6m) may already be a downstream '
        'signal of that exact mismatch.'
    )
    return {
        'ROLE_FIT_TITLE': role_fit_title,
        'ROLE_FIT_SEAT': role_fit_seat,
        'ROLE_FIT_EASY': role_fit_easy,
        'ROLE_FIT_HARD': role_fit_hard,
    }

def build_career_timeline():
    timeline_title = 'Career Timeline &mdash; Big-4 Audit to Public-Company CFO'
    timeline_html = """            <div class="timeline">
                <div class="timeline-block" style="background:#3498db;color:#fff;">Education<br>Fairfield Univ. BS Accounting<br>Golf Team 1987-91</div>
                <div class="timeline-block" style="background:#1e40af;color:#fff;">Deloitte &mdash; Staff/Manager<br>Oct 1992 - Aug 2001 &middot; NYC</div>
                <div class="timeline-block" style="background:#16a085;color:#fff;">Deloitte &mdash; Senior Manager<br>Sep 2001 - Aug 2007 &middot; NYC</div>
                <div class="timeline-block" style="background:#1e8449;color:#fff;">Deloitte &mdash; PARTNER<br>Aug 2007 - Jul 2011 &middot; Chicago<br>Midwest FI SMR Leader</div>
                <div class="timeline-block" style="background:#0e7490;color:#fff;">Lazard &mdash; Dir. Accounting Policy<br>Jul 2011 - Mar 2016 &middot; NYC</div>
                <div class="timeline-block" style="background:#7c3aed;color:#fff;">Tiptree Inc. &mdash; CAO<br>Apr 2016 - Dec 2020 &middot; NYC</div>
                <div class="timeline-block" style="background:#d4a84b;color:#1a2332;">Associated Capital Group (NYSE:AC) &mdash; EVP Finance &amp; CFO<br>Jan 2021 - Jul 2022 &middot; Greenwich, CT</div>
                <div class="timeline-block" style="background:#c0392b;color:#fff;">Yieldstreet &mdash; CFO<br>Jul 2022 - Aug 2024 &middot; NYC<br>Private-Markets Investing</div>
                <div class="timeline-block" style="background:#9333ea;color:#fff;">1970 Group &mdash; CFO<br>Nov 2024 - Apr 2026 &middot; NYC</div>
                <div class="timeline-block" style="background:#6b7280;color:#fff;">hey freya &mdash; Board Advisor<br>Aug 2024 - Present (concurrent)<br>Provable Markets per survey self-report</div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item"><div class="legend-dot" style="background:#3498db;"></div><span><strong>Education:</strong> Fairfield University, BS Accounting, 1987&ndash;1991. <strong>Two-Sport Athlete signal:</strong> varsity Golf Team alongside the BSAccounting program.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e40af;"></div><span><strong>Deloitte &mdash; Staff/Manager (Oct 1992 - Aug 2001):</strong> 8y11m. Audit client service to large financial-services companies in Insurance, Asset Management, and Broker/Dealer industries. NYC.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#16a085;"></div><span><strong>Deloitte &mdash; Senior Manager (Sep 2001 - Aug 2007):</strong> 6y. Client service in financial services across Audit, Mergers and Acquisitions, and FARS service functions. NYC.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#1e8449;"></div><span><strong>Deloitte &mdash; Partner (Aug 2007 - Jul 2011):</strong> 4y. Practice leadership and client service in the Financial Accounting Advisory Services (FARS) practice. <strong>Midwest Regional Leader of the financial-instruments accounting and valuation SMR network.</strong> Greater Chicago Area. Top-tier internal-promotion signal: Staff &rarr; Senior Manager &rarr; Partner over 18+ years.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#0e7490;"></div><span><strong>Lazard &mdash; Director of Accounting Policy (Jul 2011 - Mar 2016):</strong> 4y9m. Top-tier white-shoe investment-bank seat. NYC.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#7c3aed;"></div><span><strong>Tiptree Inc. &mdash; Chief Accounting Officer (Apr 2016 - Dec 2020):</strong> 4y9m. Public-company financial-services holding company. NYC.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#d4a84b;"></div><span><strong>Associated Capital Group (NYSE:AC) &mdash; EVP Finance &amp; CFO (Jan 2021 - Jul 2022):</strong> 1y7m. <em>Public press-release announcement.</em> NYSE-listed firm CFO seat. Greenwich, CT.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#c0392b;"></div><span><strong>Yieldstreet &mdash; CFO (Jul 2022 - Aug 2024):</strong> 2y2m. Private-markets investing platform. NYC. Roadshows / Investor Relations.</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#9333ea;"></div><span><strong>1970 Group &mdash; CFO (Nov 2024 - Apr 2026):</strong> 1y6m. Insurance-collateral-funding firm; he reposts 1970 Group content frequently on LinkedIn. NYC. <strong>Seat ended in the same month as this assessment</strong> (Apr 2026).</span></div>
                <div class="legend-item"><div class="legend-dot" style="background:#6b7280;"></div><span><strong>hey freya &mdash; Board Advisor / Self-employed (Aug 2024 - Present):</strong> 1y9m, concurrent with 1970 Group. NYC, Remote. Financial Oversight, Shepherding. Survey self-report (Q109) names Provable Markets as current company &mdash; possible new role beginning Apr 2026 not yet reflected on LinkedIn.</span></div>
            </div>
            <div class="timeline-banner">
                Career pattern: 19 years at Deloitte ending as Partner in Chicago (top-tier internal-promotion signal); transition to Lazard, Tiptree, then a CFO arc starting at NYSE-listed Associated Capital Group. <strong>Three CFO seats over the last five years</strong> (AC 1y7m, Yieldstreet 2y2m, 1970 Group 1y6m) is the tenure shape that needs onsite explanation; instrument names a coherent reading (HandsOn|Sev + Replacing Self &minus;1.51 + 3-of-5 self-graded exec peers at D-grade), but the candidate&rsquo;s own framing is the diligence answer.
            </div>"""
    return {
        'CAREER_TIMELINE_TITLE': timeline_title,
        'CAREER_TIMELINE_HTML': timeline_html,
    }

def build_respondent_dict(respondent_data):
    """TTI: Natural D=63 I=72 S=72 C=18 → Promoting Persuader (45);
            Adapted D=72 I=66 S=25 C=42 → Conducting Persuader (12).
       DF Primary: Intellectual 71, Receptive 69, Resourceful 58, Objective 57.
       DF Indifferent: Selfless 28, Structured 19, Harmonious 18, Instinctive 11."""
    nat_pos = 45
    nat_label = 'Promoting Persuader'
    nat_intensity = (72 + 72) / 200.0  # I + S = 0.72

    adp_pos = 12
    adp_label = 'Conducting Persuader'
    adp_intensity = (72 + 66) / 200.0  # D + I = 0.69

    shift_note = ('Adapted shift: +9D, &minus;6I, &minus;47S, +24C. The current seat is dialing '
                  'D up and S massively down &mdash; he is grinding hard against natural '
                  'steadiness and people-orientation to act decisively.')

    respondent = {
        'name': 'Timothy Schott',
        'first_name': 'Timothy',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': [63, 72, 72, 18],
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }
    return respondent

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    def count_class(cls):
        return html.count(f'class="{cls}"') + html.count(f'class="{cls} ')

    n_concern_item = count_class('concern-item')
    n_concern_number = count_class('concern-number')
    n_concern_text = count_class('concern-text')
    if n_concern_item < 2:
        failures.append(f"S1: concern-item count {n_concern_item} < 2")
    if n_concern_number != n_concern_item:
        failures.append(f"S1: concern-number != concern-item")
    if n_concern_text != n_concern_item:
        failures.append(f"S1: concern-text != concern-item")
    concern_nums = re.findall(r'<div class="concern-number">(\d+)</div>', html)
    expected_seq = [str(i) for i in range(1, n_concern_item + 1)]
    if concern_nums != expected_seq:
        failures.append(f"S1: concern sequence {concern_nums} != {expected_seq}")

    checks_exact = {
        'wiring-fit-content': 1, 'wiring-flag': 2,
        'probe-card': 10, 'probe-number': 10, 'probe-category': 10,
        'probe-question': 10, 'probe-coaching': 10,
    }
    for cls, expected in checks_exact.items():
        got = count_class(cls)
        if got != expected:
            failures.append(f"S3,S9: {cls} expected {expected}, got {got}")

    probe_questions = re.findall(r'<div class="probe-question">(.*?)</div>', html, re.DOTALL)
    if len(probe_questions) == 10:
        matched_substrs = set()
        unmatched = []
        for idx, q in enumerate(probe_questions, 1):
            qlow = q.lower()
            hit = False
            for (_n, _c, substr) in FORM8_QUESTIONS:
                if substr.lower() in qlow:
                    matched_substrs.add(substr); hit = True; break
            if not hit: unmatched.append(idx)
        if len(matched_substrs) < 10:
            failures.append(f"S9-Form8: {len(matched_substrs)}/10 canonical strings; unmatched probes {unmatched}")

    checks_min = {
        'timeline-block': 4, 'timeline-legend': 1, 'timeline-banner': 1,
        'legend-item': 4, 'recommendation-badge': 1,
    }
    for cls, expected in checks_min.items():
        got = count_class(cls)
        if got < expected:
            failures.append(f"S8,S10: {cls} expected min {expected}, got {got}")

    teach_m = re.search(r'<div class="metric-label">Teach Items</div>\s*<div class="metric-value">([^<]*)</div>', html)
    if not teach_m:
        failures.append("S2: Teach Items metric-value not found")
    else:
        teach_val = teach_m.group(1).strip()
        if not re.fullmatch(r'\d+/\d+', teach_val):
            failures.append(f"S2: Teach Items must be N/M, got '{teach_val}'")
    if 'To be populated' in html:
        failures.append("S2: Forbidden placeholder 'To be populated'")

    m = re.search(r'const zLabels2 = (\[.*?\]);', html, re.DOTALL)
    if not m or '[' not in m.group(1)[1:3]:
        failures.append("S4-C1: zLabels2 not nested 2-row")

    m_sf_l = re.search(r'const sfLabels2 = (\[.*?\]);', html, re.DOTALL)
    m_fail = re.search(r'const failData2 = (\[.*?\]);', html)
    m_succ = re.search(r'const successData2 = (\[.*?\]);', html)
    if not (m_sf_l and m_fail and m_succ):
        failures.append("S4-C2: missing sfLabels/fail/success")
    else:
        sf_labels = json.loads(m_sf_l.group(1))
        fail_data = json.loads(m_fail.group(1))
        succ_data = json.loads(m_succ.group(1))
        if len(fail_data) != len(sf_labels) or len(succ_data) != len(sf_labels):
            failures.append("S4-C2: length mismatch")
        if sf_labels and sf_labels[-1][-1] != '+':
            failures.append(f"S4-C2: last sfLabel must end with +, got {sf_labels[-1]}")

    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first:
            v = int(first.group(1))
            if v < 30:
                failures.append(f"S4-C3: flagLabels3 first = {v} — axis not reversed")

    m_exc_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_exc_isL1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_exc_labels and m_exc_isL1:
        exc_labels = json.loads(m_exc_labels.group(1))
        is_l1_arr = json.loads(m_exc_isL1.group(1))
        if len(exc_labels) < 12:
            failures.append(f"S6: scorecard rows {len(exc_labels)} < 12")
        l2_rows = sum(1 for v in is_l1_arr if not v)
        if l2_rows < 6:
            failures.append(f"S6: L2 rows {l2_rows} < 6")
        for lbl, l1f in zip(exc_labels, is_l1_arr):
            if l1f and lbl != lbl.upper():
                failures.append(f"S6: L1 label '{lbl}' not UPPERCASE"); break
            if not l1f and not lbl.startswith('    '):
                failures.append(f"S6: L2 label '{lbl}' not indented"); break
    else:
        failures.append("S6: excLabels/isL1 not found")

    if 'Two-Sport Athlete' not in html or 'Understands Symbolism' not in html:
        failures.append("S7: Talent Radar canonical labels missing")

    for cls, mn in [('role-fit-box', 1), ('role-fit-grid', 1),
                     ('role-fit-col easy', 1), ('role-fit-col hard', 1)]:
        if count_class(cls) < mn:
            failures.append(f"S-RF: {cls} expected min {mn}")
    if 'What Will Be Easy' not in html or 'What Will Be Hard' not in html:
        failures.append("S-RF: column labels missing")
    if 'Series B' not in html:
        failures.append("S-RF: Series B reference missing")

    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("S11: Brand lockup count < 2")
    leaks = re.findall(r'\{\{([A-Z_]+)\}\}', html)
    if leaks:
        failures.append(f"S11: Unreplaced tokens: {sorted(set(leaks))}")

    # S10 — Recommendation badge length (added 2026-04-27, Schott build):
    # the badge is a 12px gold inline pill, NOT a paragraph block. Long narratives
    # belong in Role-Fit Hard or Targeted Concerns, not the badge.
    rec_m = re.search(r'<div class="recommendation-badge">(.*?)</div>', html, re.S)
    if rec_m:
        rec_text = re.sub(r'&[a-zA-Z]+;', 'x', rec_m.group(1))  # collapse entities to 1 char
        rec_text = re.sub(r'<[^>]+>', '', rec_text).strip()
        if len(rec_text) > 300:
            failures.append(f"S10: recommendation-badge too long ({len(rec_text)} > 300 chars) — should be summary pill")

    # S11b — Role-Fit Hard step-back content (added 2026-04-27, Schott build):
    # the Hard column must take a step back to the wiring quadrant before per-dimension
    # concerns. Required for any seat where wiring-quadrant matters (CFO, controls roles).
    rfh_m = re.search(r'<div class="role-fit-col hard">(.*?)</div>\s*</div>\s*</div>', html, re.S)
    if rfh_m:
        rfh = rfh_m.group(1).lower()
        has_quadrant = any(q in rfh for q in ['top-left', 'top-right', 'bottom-left', 'bottom-right'])
        wedges = ['implementor', 'conductor', 'persuader', 'promoter',
                  'relater', 'supporter', 'coordinator', 'analyzer']
        has_wedge = any(w in rfh for w in wedges)
        if not has_quadrant:
            failures.append("S11b: ROLE_FIT_HARD missing TTI quadrant reference (top-left / top-right / bottom-left / bottom-right)")
        if not has_wedge:
            failures.append("S11b: ROLE_FIT_HARD missing TTI wedge name (Implementor / Conductor / Persuader / etc.)")

    for cid in ['distChart1','distChart2','distChart3','discChart','excstdsChart','talentRadar']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    print()
    print("=== QA GATE ===")
    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} ***")
        for f in failures: print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)}")
    print("*** QA GATE PASSED ***")


def main():
    print("Loading respondent data...")
    rd = load_respondent_data()
    print(f"  name={rd['name']}, Z|Algo={rd['z_algo_overall']:+.3f}, RF={rd['rf_num']}")
    print(f"  flags: {list(rd['flags_lit'].keys())}")

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution tokens...")
    dist = build_distribution_tokens(zalgo_rows, flag_rows,
                                     rd['z_algo_overall'], rd['z_human_overall'], rd['rf_num'])

    print("Building respondent dict for motivators_section...")
    respondent = build_respondent_dict(rd)

    print("Calling motivators_section.build_section()...")
    motivators_html = build_section(respondent, include_css=True)

    print("Building narrative sections...")
    axes = build_three_axes_narratives()
    concerns = build_concerns_section()
    wiring = build_wiring_fit()
    htl = build_hard_to_learn()
    probes = build_interview_probes()
    teach = build_teach_items()
    radar = build_talent_radar()
    timeline = build_career_timeline()
    role_fit = build_role_fit()
    scorecard = build_excstds_scorecard(rd)

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    replacements = {
        'CANDIDATE_NAME': 'Timothy Schott',
        'CANDIDATE_CREDS': 'Candidate &mdash; Series B Growth-Stage CFO Seat',
        'CANDIDATE_ROLE': 'Candidate for Series B Growth-Stage CFO &mdash; Same Functional Altitude as Today',
        'REPORT_DATE': 'April 27, 2026',

        'ZALGO_OVERALL': f'{rd["z_algo_overall"]:+.2f}',
        'ZALGO_OVERALL_NUM': f'{rd["z_algo_overall"]:.4f}',
        'REVERSE_FLAGS': str(rd['rf_num']),
        'FLAGS_LIT': f'{len(rd["flags_lit"])}/16',
        'TEACH_ITEMS': '10/10',
        'COHORT_AVG': '+0.24',
        'COHORT_AVG_NUM': '0.2440',

        **axes, **concerns, **wiring, **htl, **probes, **teach,
        **radar, **timeline, **role_fit, **dist, **scorecard,

        'DISC_D_NAT': '63',
        'DISC_I_NAT': '72',
        'DISC_S_NAT': '72',
        'DISC_C_NAT': '18',
        'DISC_D_ADP': '72',
        'DISC_I_ADP': '66',
        'DISC_S_ADP': '25',
        'DISC_C_ADP': '42',
        'DISC_NOTE_TEXT': 'Promoting Persuader (Natural, position 45) &rarr; Conducting Persuader (Adapted, position 12). The &minus;47-point S compression is the loudest signal in the wiring panel.',
                'DISC_NOTE_DETAIL': 'TTI Driving Forces &mdash; Primary: Intellectual 71, Receptive 69, Resourceful 58, Objective 57. Indifferent: Selfless 28, Structured 19, Harmonious 18, Instinctive 11. Receptive Primary + Structured Indifferent + Instinctive Indifferent describe an operator who actively resists imposed structure and does not rely on past methodologies &mdash; the wiring tell that corroborates Concern 1 (accountability architecture) and the Skills-axis CFO-fit question.',
        'DISC_ANNOTATION_CODE': '',
        'DISC_ANNOTATION': '',

        'RECOMMENDATION_TEXT': 'CONDITIONAL HIRE &middot; HIGH BAR FOR ONSITE &mdash; three amber axes; default lean is no without onsite evidence on Controller bench depth, candidate&rsquo;s own framing of three short CFO exits, and one pre-proof developmental investment.',
    }

    html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}} with raw JS",
        "Substitute EXCSTDS_COLOR_OVERRIDES with raw JS"
    )

    html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    html = re.sub(
        r'<title>.*?</title>',
        '<title>Timothy Schott &mdash; CFO Candidate | HALE GLOBAL SUCCESS DIAGNOSTICS</title>',
        html, count=1
    )
    html = html.replace(
        '<div class="hale-logo">HALE GLOBAL</div>',
        '<div class="hale-logo">HALE GLOBAL SUCCESS DIAGNOSTICS</div>'
    )

    qa_gate(html)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")
    print(f"Report ready for PDF rendering.")


if __name__ == '__main__':
    main()
