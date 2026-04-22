"""Generate a TEST hiring report for Bill George with the Motivators &
Anti-Motivators section integrated.

Most non-Motivators tokens are filled with placeholder values so the section
under test (the new {{MOTIVATORS_ANTIMOTIVATORS_SECTION}} token) renders in
context. For a full production report, run the /report skill with a real Key3.

Run from repo root:
    python _pipeline/scripts/build_test_hiring_report.py
Outputs to:
    _reports/Bill_George_TEST_hiring_report.html
"""
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section

TEMPLATE = ROOT / '_templates' / 'hiring_report_TEMPLATE.html'
OUT = ROOT / '_reports' / 'Bill_George_TEST_hiring_report.html'

# ---- Build Bill's respondent dict ----
wb = load_workbook(ROOT / 'bill_george' / 'L1 L2 Skinny Flags.xlsx', data_only=True)
ws = wb['L2']
bill_l2 = {ws.cell(row=r, column=4).value: ws.cell(row=r, column=5).value
           for r in range(2, ws.max_row + 1)
           if ws.cell(row=r, column=4).value and isinstance(ws.cell(row=r, column=5).value, (int, float))}

bill = {
    'name': 'Bill George',
    'first_name': 'Bill',
    'nat_pos': 11,
    'nat_label': 'Persuading Conductor',
    'nat_intensity': 0.92,
    'adp_pos': 11,
    'adp_label': 'Persuading Conductor',
    'adp_intensity': 0.85,
    'disc': [92, 72, 32, 14],
    'l2_scores': bill_l2,
    'shift_note': 'No Natural→Adapted shift — no stress-masking signal.',
    'partial_data': False,
}

motivators_html = build_section(bill)

# ---- Load template and substitute ----
html = TEMPLATE.read_text(encoding='utf-8')

# Inject the Motivators section
html = html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

# Placeholder values for OTHER tokens so the page renders without errors.
# These would normally be computed by the /report workflow.
PLACEHOLDERS = {
    'CANDIDATE_NAME': 'Bill George',
    'CANDIDATE_CREDS': 'TEST RENDER · 2026-04-17',
    'CANDIDATE_ROLE': 'Senior Manager, HR — Test Build',
    'RECOMMENDATION_BADGE_TEXT': 'TEST RENDER',
    'RECOMMENDATION_BADGE_CLASS': 'badge-yellow',
    'RECOMMENDATION_TEXT': '<em>This is a test render to verify the Motivators &amp; Anti-Motivators section flows correctly through the hiring template. Other sections show placeholder content. For a full production report run /report with a real Key3.</em>',
    'AXIS1_BADGE_TEXT': 'TEST', 'AXIS1_BADGE_CLASS': 'badge-grey',
    'AXIS1_HEADER': 'Talent', 'AXIS1_BODY': 'Placeholder.',
    'AXIS2_BADGE_TEXT': 'TEST', 'AXIS2_BADGE_CLASS': 'badge-grey',
    'AXIS2_HEADER': 'Judgment', 'AXIS2_BODY': 'Placeholder.',
    'AXIS3_BADGE_TEXT': 'TEST', 'AXIS3_BADGE_CLASS': 'badge-grey',
    'AXIS3_HEADER': 'Skills', 'AXIS3_BODY': 'Placeholder.',
    'CONCERN1_TITLE': 'Test concern 1', 'CONCERN1_BODY': 'Placeholder.',
    'CONCERN2_TITLE': 'Test concern 2', 'CONCERN2_BODY': 'Placeholder.',
    'METRIC_ZALGO': '+0.50', 'METRIC_REVERSE_FLAGS': '12',
    'METRIC_FLAGS_LIT': '4 of 16', 'METRIC_TEACH_ITEMS': '8 of 10',
    'METRIC_HARD_TO_LEARN': '0',
    'DIST_CHART1_LABELS': '[]', 'DIST_CHART1_ALGO': '[]', 'DIST_CHART1_HUMAN': '[]',
    'DIST_CHART1_ALGO_TRI': '[]', 'DIST_CHART1_HUMAN_DIA': '[]',
    'DIST_CHART2_LABELS': '[]', 'DIST_CHART2_FAIL': '[]', 'DIST_CHART2_SUCCESS': '[]',
    'DIST_CHART2_ALGO_TRI': '[]', 'DIST_CHART2_HUMAN_DIA': '[]',
    'DIST_CHART3_LABELS': '[]', 'DIST_CHART3_DATA': '[]', 'DIST_CHART3_ALGO_TRI': '[]',
    'DISC_D_NAT': 92, 'DISC_I_NAT': 72, 'DISC_S_NAT': 32, 'DISC_C_NAT': 14,
    'DISC_D_ADP': 85, 'DISC_I_ADP': 70, 'DISC_S_ADP': 35, 'DISC_C_ADP': 18,
    'DISC_NOTE_TEXT': 'Persuading Conductor (D-dominant)',
    'DISC_NOTE_DETAIL': 'no significant Natural→Adapted shift',
    'DISC_ANNOTATION': '', 'DISC_ANNOTATION_CODE': '',
    'WIRING_FIT_ITEMS': '<em>Placeholder — wiring-fit items computed by /report.</em>',
    'ZALGO_OVERALL': '+0.50',
    'TALENT_RADAR_PROFILE_TEXT': 'Placeholder profile · 2.0',
    'CAREER_TIMELINE_TITLE': 'Career Timeline (placeholder)',
    'CAREER_TIMELINE_HTML': '<div class="timeline-block" style="background:#888; padding:10px;">Placeholder</div>',
    'INTERVIEW_PROBE_CARDS': '<div class="probe-card"><div class="probe-question">Placeholder probe.</div></div>',
}

# Substitute remaining {{TOKEN}} placeholders
def sub_token(match):
    token = match.group(1)
    if token in PLACEHOLDERS:
        return str(PLACEHOLDERS[token])
    return f'<span style="background:#fef3cf; padding:1px 5px; border-radius:3px; font-size:11px;">[unfilled: {token}]</span>'

html = re.sub(r'\{\{(\w+)\}\}', sub_token, html)

OUT.write_text(html, encoding='utf-8')
print(f'Wrote: {OUT}')
print(f'Size: {OUT.stat().st_size} bytes')
unfilled = re.findall(r'\[unfilled: (\w+)\]', html)
if unfilled:
    print(f'Unfilled tokens (rendered with yellow markers): {sorted(set(unfilled))}')
