"""Build the Excellence Standards Coaching Guide for Jody Bender.

Respondent: 20260415.jodybenderhr@gmail.com
Role: CEO & Co-Founder, Provable Markets (FinTech — securities lending platform)
Deliverable: Integrated Coaching Guide (Graphical HTML + PDF)

Run from repo root:
    python _pipeline/scripts/build_bender_coaching.py
Outputs:
    _reports/cohen_coaching/Bender_Jody_coaching_guide.html
"""
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / '_pipeline' / 'src'))

from pipeline.motivators_section import build_section, compute_intensity_from_disc

RESPONDENT_XLSX = ROOT / '_respondents' / '20260415.jodybenderhr@gmail.com' / 'data.xlsx'
HISTOGRAM_XLSX = ROOT / 'Histogram Data.xlsx'
TEMPLATE = ROOT / '_templates' / 'coaching_guide_TEMPLATE.html'
OUT_DIR = ROOT / '_reports'
OUT = OUT_DIR / 'Bender_Jody_coaching_guide.html'

# ============================================================================
# LOAD RESPONDENT DATA
# ============================================================================

def load_respondent_data():
    wb = load_workbook(RESPONDENT_XLSX, data_only=True)

    ws_l1 = wb["L1"]
    l1_data = {}
    for row in range(2, ws_l1.max_row + 1):
        l1_num = ws_l1.cell(row, 1).value
        l1_title = ws_l1.cell(row, 2).value
        z_algo = ws_l1.cell(row, 3).value
        z_human = ws_l1.cell(row, 4).value
        rf_count = ws_l1.cell(row, 5).value
        if l1_num and z_algo is not None:
            l1_data[l1_num] = {
                'title': l1_title,
                'z_algo': z_algo,
                'z_human': z_human,
                'rf_count': rf_count
            }

    ws_l2 = wb["L2"]
    l2_scores = {}
    for row in range(2, ws_l2.max_row + 1):
        l2_short = ws_l2.cell(row, 4).value
        z_algo = ws_l2.cell(row, 5).value
        if l2_short and isinstance(z_algo, (int, float)):
            l2_scores[l2_short] = z_algo

    ws_flags = wb["Flags"]
    z_algo_overall = ws_flags.cell(2, 1).value
    z_human_overall = ws_flags.cell(2, 2).value
    rf_num = ws_flags.cell(2, 3).value
    questions_answered = ws_flags.cell(2, 4).value

    flags_lit = {}
    for col in range(5, ws_flags.max_column + 1):
        flag_name = ws_flags.cell(1, col).value
        flag_val = ws_flags.cell(2, col).value
        if flag_name and flag_val:
            flags_lit[flag_name] = flag_val

    ws_nonscore = wb["Non-Scorable"]
    non_scorable = {}
    for row in range(2, ws_nonscore.max_row + 1):
        q_num = ws_nonscore.cell(row, 2).value
        answer = ws_nonscore.cell(row, 3).value
        if q_num and answer:
            non_scorable[str(q_num)] = answer

    return {
        'l1_data': l1_data,
        'l2_scores': l2_scores,
        'z_algo_overall': z_algo_overall,
        'z_human_overall': z_human_overall,
        'rf_num': rf_num,
        'questions_answered': questions_answered,
        'flags_lit': flags_lit,
        'non_scorable': non_scorable,
    }

# ============================================================================
# LOAD HISTOGRAM DATA FOR DISTRIBUTION CHARTS
# ============================================================================

def load_histogram_data():
    wb = load_workbook(HISTOGRAM_XLSX, data_only=True)
    ws_zalgo = wb["Zalgo summ"]
    zalgo_rows = []
    rf_values = []
    for row in range(2, ws_zalgo.max_row + 1):
        z_algo = ws_zalgo.cell(row, 8).value
        z_human = ws_zalgo.cell(row, 9).value
        sf = ws_zalgo.cell(row, 2).value
        rf_val = ws_zalgo.cell(row, 7).value
        if z_algo is not None and z_human is not None:
            zalgo_rows.append({'z_algo': z_algo, 'z_human': z_human, 'sf': sf})
        if rf_val is not None:
            try:
                rf_values.append(int(rf_val))
            except (TypeError, ValueError):
                pass
    return zalgo_rows, rf_values

# ============================================================================
# BUILD DISTRIBUTION CHART TOKENS
# ============================================================================

def build_distribution_tokens(zalgo_rows, flag_rows, respondent_z_algo, respondent_z_human, respondent_rf):
    bin_edges = [i * 0.5 for i in range(-8, 9)]

    algo_counts = [0] * (len(bin_edges) - 1)
    human_counts = [0] * (len(bin_edges) - 1)
    success_counts = [0] * (len(bin_edges) - 1)
    fail_counts = [0] * (len(bin_edges) - 1)

    for row in zalgo_rows:
        z_algo = row['z_algo']
        z_human = row['z_human']
        sf = row['sf']
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= z_algo < bin_edges[i+1]:
                algo_counts[i] += 1
                if sf is True: success_counts[i] += 1
                elif sf is False: fail_counts[i] += 1
            if bin_edges[i] <= z_human < bin_edges[i+1]:
                human_counts[i] += 1

    def bin_to_pair(i):
        return [f"{bin_edges[i]:.1f}", f"{bin_edges[i+1]:.1f}"]

    def find_bin(val):
        for i in range(len(bin_edges) - 1):
            if bin_edges[i] <= val < bin_edges[i+1]:
                return i
        return len(bin_edges) - 2 if val >= bin_edges[-1] else 0

    raw_algo_bin = find_bin(respondent_z_algo)
    raw_human_bin = find_bin(respondent_z_human)

    # Chart 1 collapse
    def collapse_outer(counts_list, marker_bins):
        n = len(counts_list[0])
        left = 0
        while left < n and all(c[left] == 0 for c in counts_list) and left not in marker_bins:
            left += 1
        right = n - 1
        while right > left and all(c[right] == 0 for c in counts_list) and right not in marker_bins:
            right -= 1
        return list(range(left, right + 1))

    idx1 = collapse_outer([algo_counts, human_counts], [raw_algo_bin, raw_human_bin])
    zlabels_pairs = [bin_to_pair(i) for i in idx1]
    algo_collapsed = [algo_counts[i] for i in idx1]
    human_collapsed = [human_counts[i] for i in idx1]
    jalgo_bin = idx1.index(raw_algo_bin) if raw_algo_bin in idx1 else -1
    jhuman_bin = idx1.index(raw_human_bin) if raw_human_bin in idx1 else -1

    # Chart 2: hide empty bins
    sf_keep_idx = [i for i in range(len(bin_edges) - 1)
                   if fail_counts[i] > 0 or success_counts[i] > 0
                   or i == raw_algo_bin or i == raw_human_bin]
    sf_labels_pairs = []
    for j, i in enumerate(sf_keep_idx):
        if j == len(sf_keep_idx) - 1:
            sf_labels_pairs.append([f"{bin_edges[i]:.1f}", "+"])
        else:
            sf_labels_pairs.append(bin_to_pair(i))
    success_kept = [success_counts[i] for i in sf_keep_idx]
    fail_kept = [fail_counts[i] for i in sf_keep_idx]
    sf_algo_bin = sf_keep_idx.index(raw_algo_bin) if raw_algo_bin in sf_keep_idx else -1
    sf_human_bin = sf_keep_idx.index(raw_human_bin) if raw_human_bin in sf_keep_idx else -1

    # Chart 3: flag bins — reversed axis (high flags LEFT, low flags RIGHT)
    flag_edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
    flag_counts = [0] * (len(flag_edges) - 1)
    for rf_val in flag_rows:
        try:
            rf_int = int(rf_val)
        except (TypeError, ValueError):
            continue
        if rf_int >= flag_edges[-1]:
            flag_counts[-1] += 1
            continue
        for i in range(len(flag_edges) - 1):
            if flag_edges[i] <= rf_int < flag_edges[i+1]:
                flag_counts[i] += 1
                break

    raw_flag_bin = 0
    for i in range(len(flag_edges) - 1):
        if flag_edges[i] <= respondent_rf < flag_edges[i+1]:
            raw_flag_bin = i; break
    if respondent_rf >= flag_edges[-1]:
        raw_flag_bin = len(flag_edges) - 2

    flag_labels_pairs_raw = []
    for i in range(len(flag_edges) - 1):
        lower = flag_edges[i]; upper = flag_edges[i+1]
        if i == len(flag_edges) - 2:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}+"])
        else:
            flag_labels_pairs_raw.append([f"{lower}", f"{upper}"])
    flag_labels_rev = list(reversed(flag_labels_pairs_raw))
    flag_counts_rev = list(reversed(flag_counts))
    n_flag = len(flag_counts)
    rev_flag_bin = n_flag - 1 - raw_flag_bin

    return {
        'DIST_ZLABELS': json.dumps(zlabels_pairs),
        'DIST_ALGO_COUNTS': json.dumps(algo_collapsed),
        'DIST_HUMAN_COUNTS': json.dumps(human_collapsed),
        'DIST_JALGO_BIN': json.dumps(jalgo_bin),
        'DIST_JHUMAN_BIN': json.dumps(jhuman_bin),
        'DIST_SF_LABELS': json.dumps(sf_labels_pairs),
        'DIST_SUCCESS_COUNTS': json.dumps(success_kept),
        'DIST_FAIL_COUNTS': json.dumps(fail_kept),
        'DIST_SF_ALGO_BIN': json.dumps(sf_algo_bin),
        'DIST_SF_HUMAN_BIN': json.dumps(sf_human_bin),
        'DIST_FLAG_LABELS': json.dumps(flag_labels_rev),
        'DIST_FLAG_COUNTS': json.dumps(flag_counts_rev),
        'DIST_FLAG_BIN': json.dumps(rev_flag_bin),
    }

# ============================================================================
# RESPONDENT DICT FOR MOTIVATORS_SECTION
# ============================================================================

def build_respondent_dict(respondent_data):
    """TTI: Natural D=15 I=66 S=66 C=85 → Relating Supporter (34, FLEXIBLE);
            Adapted D=18 I=74 S=74 C=83 → Supporting Relater (17).
       Both clearly anchored in the Relater/Supporter people-care quadrant (NOT ACROSS).
       DF Primary: Selfless 96 (Extreme — 4σ above mean), Intellectual 65, Harmonious 60, Receptive 56.
       DF Indifferent: Resourceful 0 (Extreme — at the floor), Instinctive 17, Objective 19, Collaborative 25."""
    nat_pos = 34
    nat_label = 'Relating Supporter'
    nat_disc = [15, 66, 66, 85]
    nat_intensity = compute_intensity_from_disc(nat_disc)

    adp_pos = 17
    adp_label = 'Supporting Relater'
    adp_disc = [18, 74, 74, 83]
    adp_intensity = compute_intensity_from_disc(adp_disc)

    shift_note = ('Adapted shift: +3D, +8I, +8S, &minus;2C. Modest adaptation. Both Natural and '
                  'Adapted are anchored in the Relater/Supporter people-care zone of the wheel; the '
                  'current seat is dialing up I (people-engagement) and S (steadiness) slightly '
                  'while the C (controls/standards) wiring stays high.')

    return {
        'name': 'Jody Bender',
        'first_name': 'Jody',
        'nat_pos': nat_pos,
        'nat_label': nat_label,
        'nat_intensity': nat_intensity,
        'adp_pos': adp_pos,
        'adp_label': adp_label,
        'adp_intensity': adp_intensity,
        'disc': nat_disc,
        'l2_scores': respondent_data['l2_scores'],
        'shift_note': shift_note,
        'partial_data': False,
    }

# ============================================================================
# NARRATIVE SECTIONS — hand-authored prose for Cohen
# ============================================================================

SIGNATURE_PATTERN = (
    "<p><strong>Leadership is making people and situations better.</strong> That is the scoreboard. "
    "Not the title, not the size of the org. A leader who delivers compliance and engagement metrics while "
    "shrinking the people around them has not led; they have administered. A leader who has held standards "
    "for the talent bar across multiple companies, multiple acquisitions, and multiple regions while keeping "
    "the people they led intact and growing &mdash; that is the work.</p>"
    "<p>Your file shows you are doing this work at a high level already. <strong>Investing in Others L1 +2.51 "
    "(top decile)</strong>. Eleven written recommendations from people who reported to you and from peers, "
    "with consistent themes &mdash; supportive, advocate, clear instructions without micromanaging, grace "
    "under pressure, retention-focused. Three active certifications (SPHR, SHRM-SCP, PHR) that compound. A "
    "decade-and-a-half career arc with consistent internal promotions: Generalist &rarr; HR Manager &rarr; "
    "Senior HR Manager &rarr; HR Director &rarr; Global HR Operations Leader. Concrete results that travel: "
    "<strong>56% reduction in voluntary turnover at 9Gauge, 85% talent retention across 10+ M&amp;A "
    "integrations at E78</strong>, currently building People Operations from the ground up at Atlas Technica "
    "across 35+ U.S. states and 17 countries.</p>"
    "<p><strong>Especially the best get better.</strong> The instrument identifies a small number of specific "
    "routines where deliberate practice will produce the largest lift in the next chapter. They are not character "
    "concerns. They are not hidden weaknesses. They are the canonical second-decade frontiers for a senior HR "
    "leader: extending developmental belief before performance proof (Conditional Belief Sev), driving "
    "consequences when standards are missed (Driving Accountability Hi), and caring about team results over "
    "team approval (L2 7.2 Cares About Others Not Their Approval &minus;1.44). Your TTI corroborates the "
    "instrument almost line-by-line, which is unusual and useful: it means the routines that follow have a "
    "clean target. The whole guide is the answer to <em>how do you get better at the few specific things the "
    "instrument identifies, given the wiring you actually have?</em></p>"
)

FINGERPRINT_NARRATIVE = (
    "<p><strong>Z|Algo overall +0.77, above the cohort average of +0.24.</strong> Z|Human &minus;0.51 "
    "(an Algo/Human gap of 1.28 &mdash; modest, attributable mostly to a single L1). 14 reverse flags out of "
    "91 questions answered &mdash; among the lower flag-counts in the working set. The flag basket is small "
    "but specific: <strong>Dream Team Close</strong> (positive &mdash; you are within reach of one), "
    "<strong>Conditional Belief Sev</strong>, <strong>Driving Accountability Hi</strong>, "
    "<strong>Clarity of Accountability</strong> (lit, no severity), Team Qualif Mixed, and the rare "
    "<strong>NotEyesOn / HandsOff</strong> read on the Eyes flag &mdash; the opposite extreme from the more "
    "common HandsOn over-involvement pattern, suggesting you have learned to delegate and may at times be "
    "<em>under</em>-engaged with what the team is doing rather than over-involved.</p>"
    "<p>The L1 mass is mostly positive. <strong>Investing in Others L1 +2.51 (top decile)</strong>. "
    "<strong>Replacing Self L1 +1.34 Algo</strong>. Personal Reliability +0.60 in both reads. The negatives "
    "cluster around Not Pleasing (&minus;0.76) and Conducting &amp; Outvoted (&minus;0.31) &mdash; modest, "
    "and concentrated where the wiring would predict. The Algo/Human split worth noting is on Deliberate "
    "Urgency (+0.74 Algo / &minus;1.50 Human) and Org Decision Making (&minus;0.04 Algo / &minus;1.07 Human): "
    "human reviewers see your decisional-velocity and judgment-architecture as weaker than the algorithm does. "
    "Worth probing in your own reflection &mdash; not in interview, in your own routine.</p>"
)

DRIVING_FORCES_PRIMARY_HTML = (
    "<p>Your Primary cluster is unusually pure for the function you are in. <strong>Selfless 96 "
    "(Extreme &mdash; 4&sigma; above the population mean)</strong> sits at the very top of the population "
    "distribution. Intellectual 65, Harmonious 60, Receptive 56 fill out the cluster. The DF picture: "
    "you are deeply driven by completing tasks for the sake of completion (Selfless), by acquiring "
    "knowledge (Intellectual, supported by your three active HR certifications), by harmony in your "
    "surroundings (Harmonious), and by openness to new methods (Receptive).</p>"
    "<p>This is the canonical senior-HR-practitioner DF cluster. The work pulls you because the work "
    "matters to the people in it &mdash; and you do not need recognition or credit for the work to be "
    "the work. The risk inside this strength is the corollary: when the work itself becomes the only "
    "feedback signal, it can crowd out the harder feedback signal of <em>holding the line on standards "
    "even when it is uncomfortable for the people you are serving</em>. The Conditional Belief and "
    "Driving Accountability flags both live downstream of that risk.</p>"
)

DRIVING_FORCES_INDIFFERENT_HTML = (
    "<p>The Indifferent cluster is as diagnostic as the Primary. <strong>Resourceful 0 (Extreme &mdash; "
    "literally at the population floor)</strong> means you are not driven by practical results, "
    "efficiency, or maximum return on your own time, talent, and energy. Combined with Selfless 96, "
    "the DF picture is consistent: you spend yourself for the work and the people, and you do not "
    "optimize for self-return. Instinctive 17 (your past experiences and intuition are not what you "
    "lean on first), Objective 19 (functionality and objectivity are not your primary lens), "
    "Collaborative 25 (a supporting role with little need for individual recognition is not what "
    "drives you).</p>"
    "<p>The Resourceful 0 score is the most diagnostic feature of the wiring panel. It explains why "
    "the work is not transactional for you. It also explains why the moments where the standards "
    "demand <em>uncomfortable efficiency</em> (a hard performance conversation, a no-equal-trade "
    "removal, a quick decision that protects a budget) feel against the grain. Those moments are real "
    "leadership and they will not stop appearing.</p>"
)

DRIVING_FORCES_IMPLICATIONS_HTML = (
    "<p>The pragmatic implication for the work ahead is concentrated in three places where deliberate "
    "practice produces the largest lift, given who you are.</p>"
    "<p><strong>1. Extend developmental belief before the proof arrives.</strong> Selfless 96 + Receptive "
    "56 means you give people the benefit of effort and ideas already. The instrument flags Conditional "
    "Belief Sev because in the moments that matter most &mdash; backing a struggling direct report through "
    "a stretch role, defending a hire your boss is uncomfortable with, naming someone&rsquo;s next promotion "
    "before they have earned it &mdash; the wiring may default to waiting for the proof. The routine: name "
    "one developmental investment per direct report in advance of the proof, write it down, and budget the "
    "specific coaching time against it.</p>"
    "<p><strong>2. Drive consequences when the standard is missed.</strong> Resourceful 0 + Harmonious 60 "
    "means efficiency-and-discomfort is not a Primary driver, and harmony is. The instrument flags Driving "
    "Accountability at <em>Hi</em> because the moments the role demands an uncomfortable consequence &mdash; "
    "the deferred performance conversation, the role change you have been working around, the candidate you "
    "would not re-hire today &mdash; the wiring may not produce the consequence on its own cadence. The "
    "routine: pick the one hardest standard you have been quietly carrying for the team, and act on it this "
    "week.</p>"
    "<p><strong>3. Care about results over approval.</strong> L2 7.2 Cares About Others Not Their Approval "
    "&minus;1.44 is the worst L2 in your file. The wiring is consistent (Relating Supporter, high-S, "
    "high-I, low-D) and the TTI Conflict section corroborates: <em>&ldquo;not willing to share opinions "
    "until comfortable about how others will receive them.&rdquo;</em> The routine: in your next executive "
    "meeting, name your position on a contested issue first, before listening to the room. Do this once a "
    "week until it stops feeling unfamiliar.</p>"
)

WIRING_FIT_ITEMS = (
    '<strong>Wheel positions: Natural 34, Relating Supporter (FLEXIBLE); Adapted 17, Supporting Relater.</strong> '
    'Both clearly anchored in the Relater/Supporter people-care quadrant of the wheel. Neither marked ACROSS. '
    'High-C (85/83), high-S (66/74), high-I (66/74), low-D (15/18). The wiring is purpose-built for the work '
    'a senior HR practitioner does: detail-and-process discipline, sustained relational engagement, '
    'team-care and harmony. <em>Modest adaptation</em> &mdash; you are operating in your natural wiring, '
    'with the Adapted DISC dialing up I and S slightly to engage more actively in the current Atlas Technica '
    'seat.'
    '<span class="wiring-flag">Seat-aligned</span><br>'
    '<strong>The TTI Time-Wasters page lists &ldquo;Not Exercising Authority&rdquo; among your named time-'
    'wasters</strong> with causes <em>&ldquo;want to be seen as supportive, fear offending others, fear '
    'creating conflict between team members.&rdquo;</em> The TTI Conflict section adds <em>&ldquo;not '
    'willing to share opinions until comfortable about how others will receive them,&rdquo; &ldquo;may have '
    'difficulty breaking habits that hinder accomplishments,&rdquo; &ldquo;sees unwarranted change as an '
    'obstacle.&rdquo;</em> Each line corroborates a specific instrument flag (Conditional Belief Sev, '
    'Driving Accountability Hi, Cares About Approval). The wiring picture and the instrument picture agree '
    '&mdash; which means the routines ahead have a clean target.'
    '<span class="wiring-flag">Coaching frontier</span>'
)

# ============================================================================
# EXCSTDS SCORECARD
# ============================================================================

def build_excstds_scorecard(respondent_data):
    l1_data = respondent_data['l1_data']
    l2 = respondent_data['l2_scores']

    plan = [
        (1, 'RISKING BY INVESTING IN OTHERS', [
            'Handling Daily Difficulties With Dignity',
            'Developmental Mindset',
            'Developmental Discipline',
        ]),
        (2, 'RISKING BY FACILITATIVE MINDSET', [
            'Dialogue Vs. Direction',
            'Power & Status Management',
            'Sublimating Ego',
        ]),
        (3, 'CONDUCTING & OUTVOTED', [
            'Conductor > Lead Guitarist',
            'Empower Team Authority',
        ]),
        (4, 'RISKING BY REPLACING SELF', [
            'Ability To Disappear',
            'Urgency Down Chain Of Command',
        ]),
        (5, 'RISKING BY PUSHING EXTREME ACCOUNTABILITY', [
            'Basic Machinery Of Accountability',
            'Drives Accountability',
            'Stds = What Tolerate',
        ]),
        (7, 'NOT PLEASING', [
            'Cares About Others Not Their Approval',
            'Discomfort For Self',
            'Discomfort For Team',
        ]),
        (8, 'ORGANIZATIONAL DECISION MAKING', [
            'Simplification Methods',
            'Clarity Of Accountability',
            'Respects Collective Wisdom',
            'Facts Over Feelings',
        ]),
        (9, 'DELIBERATE URGENCY', [
            'Extreme Proactivity',
            'Proactive Belief In People',
            'Action Over Inaction',
        ]),
    ]

    labels, scores, is_l1, missing_l2 = [], [], [], []
    for l1_num, l1_display, l2_names in plan:
        labels.append(l1_display)
        scores.append(round(l1_data[l1_num]['z_algo'], 3))
        is_l1.append(1)
        for l2_name in l2_names:
            if l2_name in l2:
                labels.append('    ' + l2_name)
                scores.append(round(l2[l2_name], 3))
                is_l1.append(0)
            else:
                missing_l2.append(l2_name)
    if missing_l2:
        print(f"WARN: L2 names not found: {missing_l2}")
        print(f"  Available L2: {sorted(l2.keys())}")

    overall_z = respondent_data['z_algo_overall']
    cohort_avg = 0.244

    return {
        'EXCSTDS_LABELS': json.dumps(labels),
        'EXCSTDS_SCORES': json.dumps(scores),
        'EXCSTDS_ISL1': json.dumps(is_l1),
        'EXCSTDS_COLOR_OVERRIDES': '',
        'ZALGO_OVERALL_NUM': f"{overall_z:.4f}",
        'COHORT_AVG_NUM': f"{cohort_avg:.3f}",
    }

# ============================================================================
# IMPACT ITEMS — 4 flag-driven + 9 algorithm-ranked per-answer
# ============================================================================

def build_impact_items_html():
    flag_subheader = (
        '<div class="practice-subsection-hdr">'
        '<div class="practice-subsection-hdr-title">Flag-Driven Items</div>'
        '<div class="practice-subsection-hdr-blurb">Lit flags from the cohort research &mdash;'
        ' pattern-level effectiveness levers that sit across multiple L1s, which is why they are'
        ' flags and not single-L2 routines. Addressed before per-answer lifts.</div>'
        '</div>'
    )

    flag_cards = []

    # FLAG 1 — Conditional Belief Sev (the recurring HR-relevant flag)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">1</span><div style="flex:1;">'
        '<div class="practice-item-title">Conditional Belief Sev</div>'
        '<div class="practice-qref">Belief earned by proof, not extended in advance &middot; severity Severe</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.4">1.4 Demonstrating Genuine Fanness</div>'
        '</div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> Conditional Belief watches whether a leader '
        'extends developmental belief <em>before</em> a team member has earned it, or only after the '
        'evidence is in. Admired leaders extend belief early &mdash; they back the person before the '
        'proof, because the belief itself is often what enables the proof. Leaders who fire the flag at '
        'Sev tend to withhold belief until the work is delivered, which keeps the team in a permanent '
        'audition posture.</p>'
        '<p><strong>Where your file lands.</strong> Investing in Others L1 is <strong>top decile (+2.51)</strong>; '
        'L2 1.4 Demonstrating Genuine Fanness +0.93; L2 9.3 Proactive Belief In People +0.87. The '
        'machinery that extends warmth and admiration in the moment is wired in. The flag fires at Sev '
        'because the more demanding form &mdash; <em>extending belief in someone before the work proves '
        'them out, and resourcing it</em> &mdash; is the harder muscle. Your TTI corroborates: Selfless 96 '
        '(Extreme) and Receptive 56 mean you give people the benefit of effort and ideas; what does not '
        'come naturally is naming a person&rsquo;s capability gap, telling them, and budgeting your time '
        'against helping them grow into it before the proof.</p>'
        '<p><strong>Why this lever compounds for your seat specifically.</strong> The HR function is the '
        'function that holds the talent bar across the org. Every other function gets to extend belief '
        'inside its own remit; HR is the only function that does it across <em>all</em> remits, which '
        'means the absence of the routine compounds.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: for each direct report on your team of 7, write one '
        'sentence naming a capability you would like them to grow into that they have not yet '
        'demonstrated. Tell them. Budget a specific coaching investment (a stretch assignment, a '
        'co-presentation slot, an exposure to a senior conversation, a regular coaching block) against '
        'it this quarter. Belief becomes observable when it is named and resourced before the proof.</div>'
        '</div>'
    )

    # FLAG 2 — Driving Accountability Hi
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">2</span><div style="flex:1;">'
        '<div class="practice-item-title">Driving Accountability Hi</div>'
        '<div class="practice-qref">Driving consequences when standards are missed &middot; severity Hi</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="5.3">5.3 Drives Accountability</div>'
        '</div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> Driving Accountability watches whether the leader '
        'follows through with a consequence when a standard is missed. The standard is straightforward: '
        'when the team agreed to a delivery, a quality bar, or a behavior, and the agreement was not '
        'kept &mdash; what happens next? Admired leaders close the loop with a clear, calm, '
        'commensurate consequence. Leaders who fire the flag at Hi tend to absorb the miss, work '
        'around it, or have one more conversation when the consequence was already due.</p>'
        '<p><strong>Where your file lands.</strong> Pushing Extreme Accountability L1 is &minus;0.20 '
        '(slightly below cohort) with 3 reverse flags. L2 5.3 Drives Accountability is &minus;0.09 '
        '(roughly cohort-average). L2 5.8 Stds = What You Tolerate is +0.18. The machinery is mostly '
        'present; what fires the flag at Hi is the <em>follow-through-with-the-consequence</em> step '
        'specifically. The TTI Time-Wasters page is unusually explicit on this: <em>&ldquo;Not '
        'Exercising Authority&rdquo;</em> is named directly with causes <em>&ldquo;want to be seen as '
        'supportive, fear offending others, fear creating conflict between team members.&rdquo;</em> '
        'Your DF Resourceful 0 and Harmonious 60 explain the wiring: the cost of acting on the '
        'consequence (disrupting harmony, expending political capital) is felt loudly, and the '
        'reward of acting on it (efficiency, results) does not register on a Primary driver.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: at the start of the week, identify the one standard you '
        'have been carrying for the team that should have produced a consequence by now. The deferred '
        'performance conversation. The role change you have been working around. The candidate whose '
        'continued presence you would not re-hire today. Pick one. Act on it this week. Write '
        'down what made it harder than it should have been &mdash; that is the fuel for the next one.</div>'
        '</div>'
    )

    # FLAG 3 — NotEyesOn,HandsOff (the unusual Eyes-flag inversion)
    flag_cards.append(
        '<div class="practice-item flag-driven"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">3</span><div style="flex:1;">'
        '<div class="practice-item-title">Not Eyes-On, Hands-Off</div>'
        '<div class="practice-qref">The opposite extreme of the typical micromanagement flag &middot; the under-engagement read</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="4.1">4.1 Ability To Disappear</div>'
        '</div>'
        '<div class="practice-body">'
        '<p><strong>What the flag measures.</strong> The Excellence Standard for an experienced leader '
        'is <em>Eyes On, Hands Off</em> &mdash; aware of what is happening across the team and not '
        'intervening in the work itself. Leaders who fire the Eyes flag typically fire it on the '
        '<em>Hands On</em> extreme &mdash; over-involvement, can&rsquo;t-disappear. Yours fires on the '
        '<em>Not Eyes On, Hands Off</em> read: detached AND disengaged. That is the opposite extreme. '
        'It is what happens when the discipline of <em>letting people own the work</em> &mdash; which '
        'you have built (Replacing Self L1 +1.34 Algo, L2 4.1 Ability To Disappear +0.17) &mdash; is '
        'paired with insufficient cadence on staying informed about what the work is producing.</p>'
        '<p><strong>Where your file lands.</strong> The Replacing Self / delegation muscles are real. '
        'You have built HR functions from scratch and stepped back from the operational work as the '
        'org scaled (E78, Atlas Technica). Q138 (every direct report fully capable without supervision) '
        'answered correctly. The flag is reading the cadence question: <em>are you in the room often '
        'enough to know what is actually happening with the team beyond what they tell you?</em></p>'
        '</div>'
        '<div class="practice-fuel">Routine: install a weekly skip-level cadence with the layer below '
        'your direct reports. Two 30-minute conversations per week, rotating through the team of 7&rsquo;s '
        'reports. The job in those conversations is to listen, not to fix. The signal you are watching '
        'for: where does what you are hearing diverge from what your directs are telling you? That gap '
        'is the work.</div>'
        '</div>'
    )

    peranswer_subheader = (
        '<div class="practice-subsection-hdr" style="margin-top:32px;">'
        '<div class="practice-subsection-hdr-title">Per-Answer Impact Items</div>'
        '<div class="practice-subsection-hdr-blurb">Specific question-level lifts ranked by impact. '
        'Each one is a single-L2 effectiveness lever that compounds when paired with the flag-driven '
        'items above.</div>'
        '</div>'
    )

    impact_cards = []

    # IMPACT 1 — Q123 — irreversible decisions / optionality
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">4</span><div style="flex:1;">'
        '<div class="practice-item-title">Make irreversible decisions later, not earlier</div>'
        '<div class="practice-qref">Q123 &middot; Org Decision Making &middot; you answered 4 (FALSE-leaning); the standard is 1 (TRUE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="8.7">8.7 Facts Over Feelings</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;I don&rsquo;t make an irreversible decision until I have to: '
        'I prize the optionality inherent in studying a problem and not acting until it is necessary.&rdquo;</em> '
        'You answered 4 (close to FALSE), meaning the wiring leans toward making the call earlier than '
        'strictly necessary. The standard is to hold optionality.</p>'
        '<p>The HR-leadership texture: this shows up as deciding the structure, the comp band, the policy '
        'change, or the role definition <em>before</em> the team has stress-tested the alternatives. Your '
        'L2 8.7 Facts Over Feelings is &minus;0.01 (cohort-average) and L2 8.1 Simplification Methods is '
        '&minus;1.00 (below cohort). The risk is committing the org to a path before the easier-to-reverse '
        'path has been tested. The routine: every time you find yourself about to commit to a structural '
        'decision, ask <em>can this be done as a pilot, a 90-day trial, or a reversible test?</em> If yes, '
        'do that first.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: every time you find yourself about to commit to a structural HR decision (a comp band, a policy, a role definition, an org-shape change), ask &mdash; <em>can this be done as a pilot, a 90-day trial, or a reversible test first?</em> If yes, do that first. Optionality is the discipline.</div>'
        '</div>'
    )

    # IMPACT 2 — Q119 — woke test / stance on non-mission issues
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">5</span><div style="flex:1;">'
        '<div class="practice-item-title">Hold the line on mission-relevance for organizational stances</div>'
        '<div class="practice-qref">Q119 &middot; Org Decision Making &middot; you answered 4 (FALSE-leaning); the standard is 1 (TRUE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="8.3">8.3 Respects Collective Wisdom</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;When leaders take a stance and attempt to do the &lsquo;right&rsquo; '
        'thing about issues unconnected to the core mission of the organization, everyone loses.&rdquo;</em> '
        'You answered 4, meaning you lean toward the leader-should-take-stances side of the question.</p>'
        '<p>For an HR leader, this is a real and current frontier. Your remit is the work and the people '
        'doing the work. The HR seat naturally surfaces issues that are tangentially-but-not-mission-'
        'connected (cause campaigns, public-affairs questions, executive-amplification of personal views). '
        'The standard says: when the issue is unconnected to the core mission, organizational stances cost '
        'more than they earn, including for the people the stance was meant to support. The routine: when '
        'a stance-taking opportunity surfaces, ask <em>does this directly serve the mission, the work, or '
        'the people doing the work?</em> If the answer requires three reframings to get to yes, the answer '
        'is no.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: when a stance-taking opportunity surfaces, ask <em>does this directly serve the mission, the work, or the people doing the work?</em> If the answer requires three reframings to get to yes, the answer is no.</div>'
        '</div>'
    )

    # IMPACT 3 — Q65 — smart vs work hard
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">6</span><div style="flex:1;">'
        '<div class="practice-item-title">Broadcast that genius is 99% perspiration</div>'
        '<div class="practice-qref">Q65 &middot; Facilitative Mindset &middot; you answered 2 (smart-leaning); the standard is 5 (work hard)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="2.1">2.1 Dialogue Vs. Direction</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: prefer broadcasting that excellence comes from <em>work hard</em>, not '
        'from being <em>smart</em>. You answered 2, leaning toward the smart-side. This is a Gateway '
        'Belief item per the methodology.</p>'
        '<p>Why it matters for your seat: HR leadership shapes the culture-of-effort across the org. When '
        'the implicit message is that excellence is about being smart, the team that is not the smartest '
        'opts out of stretch work. When the implicit message is that excellence is about hard-but-honest '
        'effort that compounds, the bar lifts for everyone. Your work at 9Gauge and E78 already produced '
        'this culture in places (56% turnover reduction, 85% retention). The routine: in the next staff '
        'meeting, name a recent excellence moment from the team and attribute it explicitly to <em>effort '
        'that compounded</em>, not to <em>natural talent</em>. Do this until it is the default frame.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in the next staff meeting, name a recent excellence moment from the team and attribute it explicitly to <em>effort</em> &mdash; the late nights, the iteration, the thoroughness &mdash; not to <em>smartness</em>. Repeat in the next 1:1 with each direct.</div>'
        '</div>'
    )

    # IMPACT 4 — Q12 — final call
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">7</span><div style="flex:1;">'
        '<div class="practice-item-title">Stop being the one who makes the final call</div>'
        '<div class="practice-qref">Q12 &middot; Conducting &amp; Outvoted &middot; you answered 1 (TRUE); the standard is 5 (FALSE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="3.1">3.1 Conductor > Lead Guitarist</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;I am relied upon to make the final call.&rdquo;</em> The '
        'standard answer is FALSE &mdash; admired leaders build teams that make most calls without them. '
        'You answered TRUE.</p>'
        '<p>The routine here is the same one that anchors Flag 3 (Not Eyes-On, Hands-Off) but pointed at '
        'a different muscle: increasing the number of decisions your team of 7 makes without you in the '
        'room. The instrument named the conductor frontier, and your wiring (Receptive 56, Selfless 96) '
        'will tolerate this well once the routine is started. The routine: name two recurring decisions '
        'this month that you are currently making and that your team should be making instead. Hand them '
        'over with the decision rights AND the accountability for outcomes. Your L2 3.2 Empower Team '
        'Authority is +0.05 &mdash; the muscle is at cohort-average; a few deliberate handoffs will move '
        'it.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: name two recurring decisions this month that you are currently making and that your team should be making instead. Hand them over with the decision rights AND the accountability for outcomes. Do not take them back when the call goes a way you would not have called it.</div>'
        '</div>'
    )

    # IMPACT 5 — Q75 — team knows the #1 priority
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">8</span><div style="flex:1;">'
        '<div class="practice-item-title">Your team should know the #1 priority, the quantification, and the date</div>'
        '<div class="practice-qref">Q75 &middot; Not Pleasing &middot; you answered 5 (FALSE); the standard is 1 (TRUE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="5.8">5.8 Stds = What Tolerate</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;My team knows the #1 overall corporate priority, the '
        'quantification of it, and the date we will hit it.&rdquo;</em> You answered FALSE, meaning the '
        'team does not currently have that line of sight from you.</p>'
        '<p>This is the most actionable item on the file. The current Atlas Technica work you described &mdash; '
        'building HRIS, total rewards revamp, AI competencies framework &mdash; is multi-stream, and each '
        'stream has its own milestones. The standard says the team should know <em>the one thing</em> that '
        'matters most, what counts as success, and when. Then everything else gets prioritized against it. '
        'The routine: name the one thing this week. Quantify it. Date it. Tell the team. Repeat at the '
        'start of every staff meeting until they can recite it back to you in their own words.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: name the one thing this week. Quantify it. Date it. Tell the team. Repeat at the start of every staff meeting until they can recite it back to you in their own words.</div>'
        '</div>'
    )

    # IMPACT 6 — Q82 — knowledge-as-irreplaceable
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">9</span><div style="flex:1;">'
        '<div class="practice-item-title">Don&rsquo;t carry seats because the knowledge feels too critical to risk losing</div>'
        '<div class="practice-qref">Q82 &middot; Pushing Accountability &middot; you answered 2 (TRUE-leaning); the standard is 5 (FALSE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#d4a84b; color:#1a2332; border-color:#d4a84b;" data-l2="5.2">5.2 Jrs Extreme Proactivity</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;I have some people in my org I&rsquo;d replace except their '
        'experience and knowledge are too critical to risk it.&rdquo;</em> You answered 2 (close to TRUE), '
        'meaning there are people in the team you would replace if knowledge-criticality were not a '
        'constraint.</p>'
        '<p>This is the canonical knowledge-as-hostage trap, and it pairs directly with Flag 2 (Driving '
        'Accountability Hi). The routine has two parts. First, write down the names of the seats you '
        'would re-evaluate today if knowledge transfer were free. Second, for each one, design a 90-day '
        'documented-transfer plan: what does the person need to teach the team or systematize so that the '
        'role becomes replaceable? Then act on the original re-evaluation. Knowledge that lives only in '
        'one head is operational risk; knowledge that has been captured into the team is leverage.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: write down the names of the seats you would re-evaluate today if knowledge transfer were free. For each one, design a 90-day documented-transfer plan &mdash; what must the person teach the team or systematize so that the role becomes replaceable? Then act on the original re-evaluation.</div>'
        '</div>'
    )

    # IMPACT 7 — Q7 — tasking below directs
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">10</span><div style="flex:1;">'
        '<div class="practice-item-title">Stop tasking below your direct reports</div>'
        '<div class="practice-qref">Q7 &middot; Pushing Accountability &middot; you answered 4 (FALSE-leaning); the standard is 1 (TRUE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="4.3">4.3 Urgency Down Chain Of Command</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;In the last 6 months, I haven&rsquo;t tasked anyone except my '
        'direct reports.&rdquo;</em> You answered 4 (close to FALSE), meaning you have at times tasked '
        'people below your direct-report layer.</p>'
        '<p>This pairs with Flag 3 in the opposite direction: the cure for under-engagement is NOT '
        'jumping the chain of command and tasking the layer below your directs. That undermines the '
        'middle layer. The cure is the skip-level <em>listening</em> cadence (Flag 3 routine) plus the '
        'discipline of routing tasking through your directs even when it would be faster to ask the '
        'subordinate directly. The routine: when you catch yourself about to message someone below your '
        'direct, redirect through the direct. The latency cost is real and worth it.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: when you catch yourself about to message someone below your direct, redirect through the direct. The latency cost (one extra hop) is real and worth it &mdash; it protects the middle layer&rsquo;s authority and the chain of accountability.</div>'
        '</div>'
    )

    # IMPACT 8 — Q18 — debate-by-taking-positions vs. dialogue-by-asking-questions
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">11</span><div style="flex:1;">'
        '<div class="practice-item-title">Foster dialogue by asking questions, not by taking positions</div>'
        '<div class="practice-qref">Q18 &middot; Facilitative Mindset &middot; you answered 2 (TRUE-leaning); the standard is 5 (FALSE)</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="2.1">2.1 Dialogue Vs. Direction</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads: <em>&ldquo;I foster debate by taking positions and challenging people to '
        'think critically more than I foster dialogue by asking questions.&rdquo;</em> The standard answer '
        'is FALSE &mdash; admired leaders foster dialogue by asking questions. You answered 2, leaning '
        'toward TRUE.</p>'
        '<p>This pairs directly with Impact 4 (stop being the one who makes the final call). Your L2 2.1 '
        'Dialogue vs. Direction is &minus;0.33 (slightly below cohort) and L2 2.4 Sublimating Ego is '
        '&minus;0.59. The TTI conflict section corroborates: <em>&ldquo;not willing to share opinions '
        'until comfortable about how others will receive them&rdquo;</em> &mdash; an interesting nuance '
        'because the question pattern is the opposite extreme: stating positions firmly when the room '
        'is comfortable, instead of asking the questions that draw out the team&rsquo;s own positions. '
        'The routine: in the next staff meeting, run an entire agenda item where every contribution from '
        'you is reframed as a question. The team will surface positions you did not know they held.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: in the next staff meeting, run an entire agenda item where every contribution from you is reframed as a question. The team will surface positions you did not know they held. Repeat weekly until the muscle is wired in for group settings the way it is for 1:1s.</div>'
        '</div>'
    )

    # IMPACT 9 — Q1 — talent redundancy
    impact_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">12</span><div style="flex:1;">'
        '<div class="practice-item-title">Build talent redundancy &mdash; people can rise to fill in for each other</div>'
        '<div class="practice-qref">Q1 &middot; Replacing Self &middot; you answered 2 (close to TRUE-1); the standard is 5</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="4.4">4.4 CEO gets outside exec</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The standard reads, on talent redundancy: <em>&ldquo;My team&rsquo;s talent is such that '
        'people can rise to fill in for other people.&rdquo;</em> The standard answer (5) describes a '
        'team where any role can be covered by an adjacent teammate. You answered 2, meaning your team '
        'currently has more single-points-of-failure than the standard would set.</p>'
        '<p>For a People Operations function building globally across 17 countries, this is operationally '
        'urgent. Your team is 7. Each role probably has at least one knowledge-domain it owns alone. '
        'The routine: pick the two highest-risk single-points-of-failure on the team. Design a 90-day '
        'cross-training plan for each &mdash; not full mastery transfer, but enough that an adjacent '
        'teammate can carry the role through a vacation, an unexpected leave, or a transition. Pair '
        'this with Impact 6 (knowledge-as-irreplaceable) &mdash; the routines reinforce each other.</p>'
        '</div>'
        '<div class="practice-fuel">Routine: pick the two highest-risk single-points-of-failure on the team. Design a 90-day cross-training plan for each &mdash; not full mastery transfer, but enough that an adjacent teammate can carry the role through a vacation, an unexpected leave, or a transition.</div>'
        '</div>'
    )

    return flag_subheader + '\n'.join(flag_cards) + peranswer_subheader + '\n'.join(impact_cards)

def build_teach_items_html():
    teach_cards = []

    teach_subheader = (
        '<div class="practice-subsection-hdr">'
        '<div class="practice-subsection-hdr-title">Teach Items</div>'
        '<div class="practice-subsection-hdr-blurb">Standards you answered correctly. These are the '
        'muscles already in place &mdash; worth naming explicitly because the work ahead depends on '
        'protecting them as the seat scales.</div>'
        '</div>'
    )

    # TEACH 1 — Q134 daily challenge
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T1</span><div style="flex:1;">'
        '<div class="practice-item-title">You challenge your direct reports daily</div>'
        '<div class="practice-qref">Q134 &middot; Investing in Others &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="1.5">1.5 Developmental Mindset</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You named that you challenge your direct reports every day or almost every day. This is the '
        'foundational daily-rep that produces the development outcomes your file reports: the 56% turnover '
        'reduction at 9Gauge, the 85% retention across 10+ M&amp;A integrations at E78, the consistent theme '
        'in your written recommendations that you elevated the people who reported to you. The work to '
        'protect: as the team grows (now 7) and the geography stretches (35+ U.S. states, 17 countries), '
        'the daily-challenge cadence has to compound through asynchronous channels too. A weekly written '
        'challenge per direct &mdash; one specific stretch ask, in writing &mdash; will preserve the '
        'discipline at distributed scale.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: as the team grows and the geography stretches, install a weekly written challenge per direct &mdash; one specific stretch ask, in writing &mdash; so the daily-challenge cadence compounds through asynchronous channels too.</div>'
        '</div>'
    )

    # TEACH 2 — Q137 enable full potential
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T2</span><div style="flex:1;">'
        '<div class="practice-item-title">Your goal is to enable your leaders to reach their full potential</div>'
        '<div class="practice-qref">Q137 &middot; Investing in Others &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="1.6">1.6 Developmental Discipline</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You hold the canonical developmental-leader frame. Your DF Selfless 96 (Extreme) is the wiring '
        'that produces this orientation naturally &mdash; you do the work because the work matters to the '
        'people in it. Worth pairing with Impact item 1 (Conditional Belief): the goal of enabling '
        'leaders to reach their full potential lands hardest when the belief is extended <em>before</em> '
        'the proof, not after. Your daily-challenge muscle (Teach 1) plus a pre-proof-belief routine '
        '(Impact 1) is the full developmental loop.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: pair the daily-challenge muscle (T1) with a pre-proof-belief routine (Impact 1, Conditional Belief). Together they form the full developmental loop &mdash; belief extended before the work proves them out, plus the daily reps that build the proof.</div>'
        '</div>'
    )

    # TEACH 3 — Q138 capable without supervision
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T3</span><div style="flex:1;">'
        '<div class="practice-item-title">Each direct report is fully capable of doing their job without supervision</div>'
        '<div class="practice-qref">Q138 &middot; Replacing Self &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="4.1">4.1 Ability To Disappear</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You answered that each of your direct reports is fully capable without supervision or guidance. '
        'This is the muscle that built your Replacing Self L1 +1.34 Algo, and it is the underlying reason '
        'your file shows the unusual <em>Not Eyes-On, Hands-Off</em> read on the Eyes flag rather than the '
        'more common HandsOn over-involvement. The work to protect: this strength is what enables Flag 3&rsquo;s '
        'routine (skip-level listening cadence) to actually work &mdash; the directs are capable, so the '
        'job is to stay informed about how they are doing it, not to do it.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: stay informed about <em>how</em> directs are doing the work, not <em>what</em> they are doing. Skip-level listening (Flag 3 cadence) is the discipline that protects this strength as the seat scales.</div>'
        '</div>'
    )

    # TEACH 4 — Q83 difficult conversation daily
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T4</span><div style="flex:1;">'
        '<div class="practice-item-title">You have a difficult conversation every day or almost every day</div>'
        '<div class="practice-qref">Q83 &middot; Investing in Others &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="1.3">1.3 Handling Daily Difficulties With Dignity</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>This is the most important teach item on your file. It is also the strongest counter-evidence '
        'to the surface read of the Conditional Belief and Driving Accountability flags. You are NOT a '
        'leader who avoids difficult conversations; you have them on a daily cadence. The flags are firing '
        'on a more specific question: whether the difficult conversations <em>produce a consequence</em> '
        'when the standard is missed (Driving Accountability), and whether they extend belief <em>before '
        'the proof</em> (Conditional Belief). The Q83 daily-cadence is the foundation; the flag-driven '
        'routines are the next layer of refinement on top of it.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: keep the daily-difficult-conversation cadence in place. The flag-driven routines on consequence (Flag 2) and pre-proof-belief (Flag 1) refine the conversations from there &mdash; they are not a replacement for the cadence.</div>'
        '</div>'
    )

    # TEACH 5 — Q50 zero voluntary attrition
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T5</span><div style="flex:1;">'
        '<div class="practice-item-title">Zero voluntary attrition; team members turn down higher offers</div>'
        '<div class="practice-qref">Q50 &middot; Investing in Others &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="1.1">1.1 Reciprocal Followership</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You named that team members turn down higher offers from other companies to stay on your team. '
        'This is the strongest external-validator signal in the data. The 11 written recommendations on '
        'your LinkedIn corroborate it &mdash; multiple from former direct reports describing why they wanted '
        'to keep working with you. The pattern is rare and worth naming explicitly: leaders who produce '
        'this read create the conditions money cannot buy &mdash; growth, exposure, and the daily challenge '
        'that makes external offers feel like a step sideways. The work to protect: as the seat scales to '
        'distributed teams across 17 countries, the conditions that produce this retention have to be '
        'engineered (deliberate growth conversations, deliberate exposure, deliberate stretch) rather than '
        'happening organically through proximity. Your Atlas Technica work is the test of whether this '
        'discipline scales.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: as the seat scales to 17-country distributed teams, the conditions that produced this retention have to be engineered deliberately &mdash; named growth conversations, named exposure, named stretch &mdash; rather than relying on proximity to do the work.</div>'
        '</div>'
    )

    # TEACH 6 — Q24 daily routines for years
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T6</span><div style="flex:1;">'
        '<div class="practice-item-title">You have routines you have done every single day for years</div>'
        '<div class="practice-qref">Q24 &middot; Personal Reliability &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="6.3">6.3 Commitment To Routine</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The personal-discipline machinery is wired in. Personal Reliability L1 +0.60 in both reads, '
        'L2 6.3 Commitment to Routine +0.79. This is what compounds across a 14-year HR career arc with '
        'three active certifications and consistent internal promotions. The work to protect: each new '
        'altitude (Manager &rarr; Sr Manager &rarr; Director &rarr; Global HR Operations Leader) requires '
        'a small routine refresh. The routines that got you to this seat are not necessarily the routines '
        'this seat needs. Worth a quarterly audit: which daily routines are still earning their cost, and '
        'which have become muscle memory that no longer produces a result?</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: quarterly audit. Which daily routines are still earning their cost, and which have become muscle memory that no longer produces a result? The routines that got you to this seat are not necessarily the routines this seat needs.</div>'
        '</div>'
    )

    # TEACH 7 — Q381 multiple parties accountable
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T7</span><div style="flex:1;">'
        '<div class="practice-item-title">When multiple parties are accountable, no one truly is</div>'
        '<div class="practice-qref">Q381 &middot; Org Decision Making &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#22c55e; color:#ffffff; border-color:#22c55e;" data-l2="8.2">8.2 Clarity Of Accountability</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You answered the canonical clarity-of-accountability question correctly. This is worth naming '
        'because it is one of the questions other senior leaders in the working set have gotten wrong &mdash; '
        'and it is the foundational principle behind L2 8.2 Clarity of Accountability (your score: +0.64, '
        'above cohort). The principle compounds at every altitude: as the team grows from 7 to a larger '
        'org, the diffusion-of-accountability risk grows quadratically. Worth pairing with Impact 7 (stop '
        'tasking below directs) &mdash; the routine that protects clarity of accountability at distributed '
        'scale is single-point ownership routed through the appropriate layer.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: pair this principle with Impact 7 (stop tasking below directs). Single-point ownership routed through the appropriate layer is the discipline that protects clarity of accountability at distributed scale.</div>'
        '</div>'
    )

    # TEACH 8 — Q44 emotional discipline on email
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T8</span><div style="flex:1;">'
        '<div class="practice-item-title">You don&rsquo;t send emails (or messages) when upset</div>'
        '<div class="practice-qref">Q44 &middot; Personal Reliability &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="6.1">6.1 Extreme Ownership</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You answered the canonical emotional-discipline question correctly. This is the muscle that '
        'protects relationships across the org under pressure &mdash; particularly relevant for an HR '
        'leader, who absorbs the most emotionally-charged moments in the company by job description. '
        'Your L2 6.1 Extreme Ownership (&minus;0.16) is roughly cohort-average; this discipline-around-'
        'communication routine is what keeps that score from falling under stress. The work to protect: '
        'as the seat scales to 17-country distributed teams, the asynchronous-communication channels '
        'multiply (Slack across timezones, async video, written reviews). The same discipline applies '
        'to all of them.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: the same emotional discipline applies to every asynchronous channel &mdash; Slack across timezones, async video, written reviews. Audit yourself weekly: did anything ship under heat?</div>'
        '</div>'
    )

    # TEACH 9 — Q120 remove distasteful team members quickly
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T9</span><div style="flex:1;">'
        '<div class="practice-item-title">When a team member can&rsquo;t surrender their distaste, they must be removed quickly</div>'
        '<div class="practice-qref">Q120 &middot; Org Decision Making &middot; you answered 1 (TRUE); the standard is 1</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dbeafe; color:#1a2332; border-color:#2563eb;" data-l2="8.3">8.3 Respects Collective Wisdom</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>You answered the canonical remove-the-corrosive-team-member question correctly. This is '
        'a counter-evidence to the surface read of the Driving Accountability flag &mdash; you '
        'understand the principle, and your record at E78 (multi-company RIFs handled with care for '
        'both impacted employees and remaining team morale) corroborates that you can act on it. The '
        'flag fires not on the catastrophic case (corrosive team member) but on the more common case '
        '(missed-standard, deferred-consequence). The principle in this teach item is the foundation '
        'for the Flag 2 routine.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: this teach is the foundation for Flag 2&rsquo;s routine. The catastrophic case (corrosive team member) is wired right; the next layer is the missed-standard, deferred-consequence case &mdash; the more common form of the same muscle.</div>'
        '</div>'
    )

    # TEACH 10 — Q20 answer questions with questions
    teach_cards.append(
        '<div class="practice-item"><div style="display:flex; align-items:baseline; gap:12px;">'
        '<span class="practice-num">T10</span><div style="flex:1;">'
        '<div class="practice-item-title">You answer questions with questions</div>'
        '<div class="practice-qref">Q20 &middot; Facilitative Mindset &middot; you answered 5 (FALSE = you DO ask back); the standard is 5</div>'
        '</div>'
        '<div class="practice-l2-tag" style="background:#dcfce7; color:#1a2332; border-color:#22c55e;" data-l2="2.1">2.1 Dialogue Vs. Direction</div>'
        '</div>'
        '<div class="practice-body">'
        '<p>The Facilitative Mindset standard rejects the framing that &ldquo;the leader&rsquo;s job is '
        'to tell people what to do.&rdquo; You answered correctly: you DO answer questions with '
        'questions, drawing out the team&rsquo;s own thinking. This is the muscle that pairs with Impact '
        '8 (foster dialogue by asking questions). The teach card and the impact item are complementary: '
        'you have the answer-with-questions muscle in 1:1 conversations; the impact item is the harder '
        'extension of the same muscle into group settings (staff meetings, executive forums) where the '
        'wiring may default to taking a position first.</p>'
        '</div>'
        '<div class="practice-fuel">Routine to protect: extend the answer-with-questions muscle from 1:1s into group settings. In the next staff meeting, run an item where every contribution from you is a question &mdash; this pairs with Impact 8 (foster dialogue by asking questions).</div>'
        '</div>'
    )

    return teach_subheader + '\n'.join(teach_cards)

CONNECTION_NARRATIVE_HTML = (
    "<p>The most useful frame for the work ahead is the alignment between your wiring and the standards "
    "the instrument flags. The wiring is anchored in the Relater/Supporter people-care quadrant of the "
    "TTI wheel (Natural 34 Relating Supporter, Adapted 17 Supporting Relater). The DF profile is dominated "
    "by Selfless 96 (Extreme &mdash; 4&sigma; above the population mean) and balanced by Resourceful 0 "
    "(Extreme &mdash; at the population floor). The picture: you are deeply driven by completing the work "
    "for the people in it, and you are not driven by efficiency-for-self.</p>"
    "<p><strong>This wiring is the reason your file shows what it shows.</strong> The 56% turnover "
    "reduction at 9Gauge, the 85% retention across 10+ M&amp;A integrations at E78, the 11 written "
    "recommendations describing supportive leadership and clear instruction without micromanaging, the "
    "current Atlas Technica scope across 17 countries &mdash; all of that is what Selfless 96 + Relater/"
    "Supporter wiring produces when paired with three active HR certifications and a decade of practice. "
    "Investing in Others L1 +2.51 (top decile) is the instrument confirming that the wiring is "
    "manifesting in the work.</p>"
    "<p><strong>The wiring is also the reason for the small number of specific frontiers the instrument "
    "flags.</strong> Conditional Belief Sev fires because Selfless 96 + Receptive 56 means you give "
    "people the benefit of effort and ideas naturally &mdash; what does not come naturally is the more "
    "demanding form: extending developmental belief in someone <em>before</em> they have proven themselves, "
    "and resourcing it. Driving Accountability Hi fires because Resourceful 0 + Harmonious 60 means the "
    "cost of acting on a missed-standard consequence (disrupting harmony) is felt loudly, while the reward "
    "of acting on it (efficiency, results) does not register on a Primary driver. Cares About Others Not "
    "Their Approval at &minus;1.44 fires because Relating Supporter wiring with low-D (15) means caring "
    "about how others receive your position is wired into the natural posture, even when the seat requires "
    "naming the position first.</p>"
    "<p><strong>None of these are character concerns.</strong> They are the canonical second-decade "
    "frontiers for someone with your wiring in your function. The whole guide is the answer to <em>what "
    "specific routines, given who you actually are, will close these gaps without burning out the parts of "
    "the wiring that produce the strengths the file celebrates?</em></p>"
)


CAREER_TIMELINE_TITLE = "Career Timeline &mdash; Jody Bender"

CAREER_TIMELINE_HTML = """
            <div class="timeline">
                <div class="timeline-block" style="flex: 1.8; background: #95a5a6;">Earlier career &middot; banking / trading arc<br><span style="font-size:9px; opacity:0.8;">Pre-2017 &middot; Securities-finance / special-situations operator</span></div>
                <div class="timeline-block" style="flex: 1.5; background: #34495e;">Jefferies &middot; MD Head of Securities Finance Trading<br><span style="font-size:9px; opacity:0.8;">2017&ndash;2018 &middot; $65MM &rarr; $90MM revenue growth</span></div>
                <div class="timeline-block" style="flex: 1.5; background: #2c3e50;">Nomura &middot; MD Head of Special Situations<br><span style="font-size:9px; opacity:0.8;">2018&ndash;2019 &middot; $125MM &rarr; $150MM revenue growth</span></div>
                <div class="timeline-block" style="flex: 3.0; background: #1a2332;">Provable Markets &middot; CEO &amp; Co-Founder <span style="background:#d4a84b; color:#1a2332; padding:0 4px; border-radius:2px; font-size:9px;">$18MM Seed &rarr; ~$75MM Series A Ext</span><br><span style="font-size:9px; opacity:0.8;">Jan 2021&ndash;present &middot; FinTech platform &middot; <strong>current seat</strong></span></div>
            </div>
            <div class="timeline-legend">
                <div class="legend-item">
                    <div class="legend-dot" style="background: #95a5a6;"></div>
                    <span><strong>Earlier career (pre-2017):</strong> Securities-finance and special-situations operator. Foundation for the MD-level moves that followed at Jefferies and Nomura.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #34495e;"></div>
                    <span><strong>Jefferies (2017&ndash;2018):</strong> MD &mdash; Head of Securities Finance Trading. Self-reported revenue growth from $65MM to $90MM over 1.5 years. Demonstrates the per-seat revenue-generation engine at an MD altitude on the sell-side.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #2c3e50;"></div>
                    <span><strong>Nomura (2018&ndash;2019):</strong> MD &mdash; Head of Special Situations. Self-reported growth $125MM to $150MM over 1.5 years. Q9927 note: boss Mike Caperonis &mdash; &ldquo;extremely even keel in good and bad, a prerequisite for a risk manager/trader and rare to achieve consistently.&rdquo; The exit framing names a political-dynamics growth area.</span>
                </div>
                <div class="legend-item">
                    <div class="legend-dot" style="background: #1a2332;"></div>
                    <span><strong>Provable Markets (Jan 2021&ndash;present):</strong> CEO &amp; Co-Founder. FinTech platform for securities lending. Seed $18MM (Jan 2021) &rarr; Series A Extension ~$75MM (April 2025). Self-reported KPI (Q76): 10K avg daily executions, 40% of top-20 clients signed on. Removed original CEO / co-founder Thomer Gil when the operating approach diverged; Q9917 self-assessment names patience, delegation, aggressiveness, work/life balance, and temper as personal areas the separation surfaced.</span>
                </div>
            </div>
            <div class="timeline-banner" style="background:#e8f4ea; color:#1e5a30;">
                Two MD-level sell-side seats with verified per-seat revenue growth, followed by founding and running a FinTech through the co-founder-separation stress test and the Seed&rarr;Series&nbsp;A&nbsp;Ext arc to $75MM valuation. The signature is an operator who builds with discipline and makes the hardest people-calls when the company requires it. <em>Note: LinkedIn public view is partially constrained &mdash; Awards/Honors sections unverifiable in session; Talent-axis read is anchored on the instrument file, Non-Scorable career-history answers (Q9911&ndash;Q9927), and publicly-searchable Provable Markets funding records.</em>
            </div>
"""

CLOSING_NOTE_HTML = (
    "<p><strong>Especially the best get better.</strong> That is the Hale Global tagline and it is the "
    "right closing frame for your file. The instrument identified you as one of the best, and this guide "
    "is the answer to how you get even better &mdash; not by becoming someone you are not, but by "
    "refining the few specific routines the instrument flagged, given the wiring you actually have.</p>"
    "<p>The three things to come back to:</p>"
    "<ol style='font-size:14px; line-height:1.6; margin:12px 0 18px 24px;'>"
    "<li><strong>Extend belief before the proof.</strong> One developmental investment per direct report, "
    "named, written down, and resourced before the work proves them out. (Flag 1 routine.)</li>"
    "<li><strong>Drive the consequence when the standard is missed.</strong> One hardest-deferred-standard "
    "this week, acted on. (Flag 2 routine.)</li>"
    "<li><strong>Care about results over approval.</strong> Name your position first in the next executive "
    "meeting, before listening to the room. (DF Implication 3 routine.)</li>"
    "</ol>"
    "<p>The work the instrument celebrates is real. The 11 recommendations are real. The retention work, "
    "the M&amp;A integrations, the global build at Atlas Technica are all real. This guide is not a "
    "counterargument to that record. It is the next chapter of practice that builds on it.</p>"
)


# ============================================================================
# DISC NOTES
# ============================================================================

DISC_NOTE_TEXT = (
    "Conducting Implementor (D=69, C=61, S=48, I=32). Adapted shift: +9D, −4I, −13S, +3C."
)
DISC_NOTE_DETAIL = (
    "Adapted profile pulls harder toward commanding posture and further away from steadiness and"
    " warmth in the current founder-CEO seat. Objective #1 Primary + Structured #2 Primary + Resourceful"
    " #3 Primary + Commanding #4 Primary is the wiring of a data-driven, disciplined builder"
)
DISC_ANNOTATION_CODE = ""

# ============================================================================
# QA GATE
# ============================================================================

def qa_gate(html):
    failures = []

    # 1. No unreplaced tokens
    leaks = re.findall(r'\{\{([A-Z_0-9]+)\}\}', html)
    if leaks:
        failures.append(f"Unreplaced tokens: {sorted(set(leaks))}")

    # 2. Canvases
    for cid in ['distChart1', 'distChart2', 'distChart3', 'discChart', 'excstdsChart']:
        if f'id="{cid}"' not in html:
            failures.append(f"Canvas missing: {cid}")

    # 3. Brand lockup
    if html.count('HALE GLOBAL SUCCESS DIAGNOSTICS') < 2:
        failures.append("Brand lockup count < 2")

    # 4. Practice items — at least 10 teach and 13 impact (4 flag + 9 per-answer)
    teach_items = html.split('Part 1 — What You Teach')[1].split('Part 2 — What to Work On')[0]
    impact_items = html.split('Part 2 — What to Work On')[1].split('How the Two Lists Connect')[0]
    n_teach = teach_items.count('class="practice-item"')
    n_impact_std = impact_items.count('class="practice-item"')
    n_impact_flag = impact_items.count('class="practice-item flag-driven"')
    if n_teach < 10:
        failures.append(f"Teach items: got {n_teach}, need ≥10")
    if n_impact_flag < 3:
        failures.append(f"Flag-driven impact items: got {n_impact_flag}, need ≥3")
    if n_impact_std < 9:
        failures.append(f"Per-answer impact items: got {n_impact_std}, need ≥9")

    # 5. Dimensional scorecard has L1 UPPERCASE + L2 indent
    m_labels = re.search(r'var excLabels = (\[.*?\]);', html, re.DOTALL)
    m_isl1 = re.search(r'var isL1 = (\[.*?\]);', html, re.DOTALL)
    if m_labels and m_isl1:
        labels = json.loads(m_labels.group(1))
        isl1 = json.loads(m_isl1.group(1))
        if len(labels) < 12:
            failures.append(f"Scorecard has only {len(labels)} rows (min 12)")
        l2_rows = sum(1 for v in isl1 if not v)
        if l2_rows < 6:
            failures.append(f"Scorecard has only {l2_rows} L2 rows (min 6)")
        for lbl, is_l1 in zip(labels, isl1):
            if is_l1 and lbl != lbl.upper():
                failures.append(f"L1 label '{lbl}' must be UPPERCASE")
                break
            if not is_l1 and not lbl.startswith('    '):
                failures.append(f"L2 label '{lbl}' must be 4-space indented")
                break

    # 6. Chart 3 reversed — high flags should be on LEFT
    m_flabels = re.search(r'const flagLabels3 = (\[.*?\]);', html, re.DOTALL)
    if m_flabels:
        first = re.match(r'\[\[?"?(\d+)', m_flabels.group(1))
        if first and int(first.group(1)) < 30:
            failures.append(f"Chart 3 axis not reversed — first label starts with {first.group(1)}")

    # 7. Motivators section rendered
    if '.ma-section' not in html or 'class="ma-section"' not in html:
        failures.append("Motivators section not injected")

    # 8. Candidate data surfaces
    if 'Jody Bender' not in html:
        failures.append("Candidate name missing")
    if 'Provable Markets' not in html:
        failures.append("Candidate role / company missing")

    # 9. Fingerprint narrative present
    if 'Relating Supporter' not in html and 'Supporting Relater' not in html:
        failures.append("Fingerprint narrative missing wedge label")

    print()
    print("=== QA GATE ===")
    # Section 11g — Coaching-guide L2 color tags (added 2026-04-28, Bender coaching build)
    # Every Impact and Teach card MUST have a practice-l2-tag with motivators-wheel colors.
    n_practice_items = html.count('class="practice-item"') + html.count('class="practice-item flag-driven"')
    n_l2_tags = html.count('class="practice-l2-tag"')
    if n_l2_tags < n_practice_items:
        failures.append(
            f"S11g: practice-l2-tag count ({n_l2_tags}) < practice-item count ({n_practice_items}) "
            f"— missing L2 tag on {n_practice_items - n_l2_tags} card(s). "
            f"Every Impact and Teach card must end with a <div class=\"practice-l2-tag\" style=\"background:#XXX; "
            f"color:#YYY; border-color:#ZZZ;\" data-l2=\"X.Y\">X.Y Name</div> using motivators-wheel colors."
        )

    # Section 11h - Coaching-guide practice-fuel routine box (added 2026-04-28, Bender coaching build)
    # Every Impact and Teach card MUST end with a <div class="practice-fuel"> Routine: ... </div>.
    n_fuel = html.count('class="practice-fuel"')
    if n_fuel != n_practice_items:
        failures.append(
            f"S11h: practice-fuel count ({n_fuel}) != practice-item count ({n_practice_items}). "
            f"Every Impact and Teach card MUST end with <div class=\"practice-fuel\">Routine: ...</div>. "
            f"Bender shipped without practice-fuel on 19 of 22 cards in v6 - this is the regression that "
            f"drove the rule. See METHODOLOGY 'Coaching-guide practice-fuel - every Impact and Teach card MUST carry one'."
        )

    # All practice-fuel content begins with the word 'Routine' (or 'Routine to protect' for teach cards).
    fuel_content_re = re.compile(r'<div class="practice-fuel">([^<]*)')
    bad_fuel = [m.group(1)[:60] for m in fuel_content_re.finditer(html) if not m.group(1).lstrip().lower().startswith('routine')]
    if bad_fuel:
        failures.append(
            f"S11h: {len(bad_fuel)} practice-fuel block(s) do not start with 'Routine': "
            f"{[s[:40] for s in bad_fuel[:3]]}"
        )

    # Section 11h structural-integrity check - practice-item open/close balance.
    # For each <div class="practice-item..."> opener, find the next <div class="practice-item..." opener
    # and assert the depth between them returns to the section level (i.e., equal opens and closes between them).
    item_open_re = re.compile(r'<div class="practice-item(?:"|\s+flag-driven")')
    closes_re = re.compile(r'</div>')
    opens_re = re.compile(r'<div\b')
    item_positions = [m.start() for m in item_open_re.finditer(html)]
    for i in range(len(item_positions) - 1):
        between = html[item_positions[i]:item_positions[i+1]]
        opens = len(opens_re.findall(between))
        closes = len(closes_re.findall(between))
        # Each card should be self-contained: opens == closes between two adjacent practice-item starts
        if opens != closes:
            failures.append(
                f"S11h: practice-item structural imbalance between cards {i+1} and {i+2}: "
                f"{opens} opens vs {closes} closes between them. "
                f"Each card MUST close its practice-item div before the next opens. "
                f"See METHODOLOGY 'Implementation pattern - EXACT structural placement'."
            )
            break

    # Section 11i - Print-CSS pagination contract is enforced in the render script,
    # not here. The render script must inject the print stylesheet via page.addStyleTag and
    # run a post-paint canvas health-check. See QA_CHECKLIST 11i for the full contract.

    if failures:
        print(f"*** QA GATE FAILED: {len(failures)} issue(s) ***")
        for f in failures:
            print(f"  - {f}")
        raise AssertionError(f"QA gate failed with {len(failures)} issue(s)")
    else:
        print("*** QA GATE PASSED ***")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("Loading respondent data...")
    respondent_data = load_respondent_data()

    print("Loading histogram data...")
    zalgo_rows, flag_rows = load_histogram_data()

    print("Building distribution chart tokens...")
    dist_tokens = build_distribution_tokens(
        zalgo_rows, flag_rows,
        respondent_data['z_algo_overall'],
        respondent_data['z_human_overall'],
        respondent_data['rf_num'],
    )

    print("Building motivators section...")
    respondent = build_respondent_dict(respondent_data)
    motivators_html = build_section(respondent, include_css=True)

    print("Building scorecard...")
    scorecard = build_excstds_scorecard(respondent_data)

    print("Building impact / teach items...")
    impact_html = build_impact_items_html()
    teach_html = build_teach_items_html()

    print("Loading template...")
    template_html = TEMPLATE.read_text(encoding='utf-8')

    # Neutralize in-comment {{EXCSTDS_COLOR_OVERRIDES}} reference if present
    template_html = template_html.replace(
        "Substitute {{EXCSTDS_COLOR_OVERRIDES}}",
        "Substitute EXCSTDS_COLOR_OVERRIDES"
    )

    # Inject motivators section (contains raw braces — do this before token-replace loop)
    template_html = template_html.replace('{{MOTIVATORS_ANTIMOTIVATORS_SECTION}}', motivators_html)

    replacements = {
        # Header / meta
        'CANDIDATE_NAME': 'Jody Bender',
        'CANDIDATE_CREDS': 'Director, Global Business Partner &middot; Atlas Technica',
        'CANDIDATE_ROLE': 'People Operations &middot; SaaS &middot; ~2,000 employees across 35+ U.S. states and 17 countries',
        'REPORT_DATE': 'April 21, 2026',

        # Signature + fingerprint + driving forces
        'SIGNATURE_PATTERN': SIGNATURE_PATTERN,
        'FINGERPRINT_NARRATIVE': FINGERPRINT_NARRATIVE,
        'DRIVING_FORCES_PRIMARY_HTML': DRIVING_FORCES_PRIMARY_HTML,
        'DRIVING_FORCES_INDIFFERENT_HTML': DRIVING_FORCES_INDIFFERENT_HTML,
        'DRIVING_FORCES_IMPLICATIONS_HTML': DRIVING_FORCES_IMPLICATIONS_HTML,

        # Headline metrics
        'ZALGO_OVERALL': f'{respondent_data["z_algo_overall"]:+.2f}',
        'COHORT_AVG': '+0.24',
        'TEACH_ITEMS': '10/10',
        'IMPACT_ITEMS': '13',
        'FLAGS_LIT': str(len(respondent_data['flags_lit'])),
        'REVERSE_FLAGS': str(respondent_data['rf_num']),

        # DISC
        'DISC_D_NAT': '69', 'DISC_I_NAT': '32', 'DISC_S_NAT': '48', 'DISC_C_NAT': '61',
        'DISC_D_ADP': '78', 'DISC_I_ADP': '28', 'DISC_S_ADP': '35', 'DISC_C_ADP': '64',
        'DISC_NOTE_TEXT': DISC_NOTE_TEXT,
        'DISC_NOTE_DETAIL': DISC_NOTE_DETAIL,
        'DISC_ANNOTATION_CODE': DISC_ANNOTATION_CODE,

        # Wiring-Fit
        'WIRING_FIT_ITEMS': WIRING_FIT_ITEMS,

        # Excellence Standards scorecard
        **scorecard,

        # Distribution charts
        **dist_tokens,

        # Teach & Impact
        'TEACH_ITEMS_HTML': teach_html,
        'IMPACT_ITEMS_HTML': impact_html,

        # Connection / timeline / closing
        'CONNECTION_NARRATIVE_HTML': CONNECTION_NARRATIVE_HTML,
        'CAREER_TIMELINE_TITLE': CAREER_TIMELINE_TITLE,
        'CAREER_TIMELINE_HTML': CAREER_TIMELINE_HTML,
        'CLOSING_NOTE_HTML': CLOSING_NOTE_HTML,
    }

    html = template_html
    for token, value in replacements.items():
        html = html.replace(f'{{{{{token}}}}}', str(value))

    # Run QA gate
    qa_gate(html)

    # Write output
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT.write_text(html, encoding='utf-8')

    print(f"\nSUCCESS: {OUT}")
    print(f"Size: {OUT.stat().st_size:,} bytes")

if __name__ == '__main__':
    main()
